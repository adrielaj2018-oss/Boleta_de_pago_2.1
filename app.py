# -*- coding: utf-8 -*-
"""
TAREO MÓVIL – GRUPO DE COSECHA | PRIZE PRO
Flask + SQLite/PostgreSQL + PWA. Listo para GitHub + Render.

Mejoras incluidas:
- Login con formato móvil verde/blanco como referencia.
- Pantalla principal: Soporte, Configuraciones, Sincronización y Hojas de Tareo.
- Creación de hoja: fecha, grupo, subgrupo, labor, responsable, turno y tipo.
- Detalle de hoja con tabs: Labores, Trabajadores, Rend./Avance por Labor, con iconos funcionales.
- Registro por labor-consumidor, detalle de trabajador por labor y lecturas por balde.
- Adaptado a desktop y celular con diseño responsive tipo app.
"""
import os, re, sqlite3, base64, json
from datetime import datetime, date
from functools import wraps
from io import BytesIO

from flask import Flask, request, redirect, url_for, session, flash, jsonify, send_file, render_template_string, Response
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash

try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PERSIST_DIR = os.getenv("PERSIST_DIR", "/data" if os.path.isdir("/data") else BASE_DIR)
FOTO_DIR = os.path.join(PERSIST_DIR, "fotos_marcacion")
FIRMA_DIR = os.path.join(PERSIST_DIR, "firmas_documentos")
os.makedirs(FOTO_DIR, exist_ok=True)
os.makedirs(FIRMA_DIR, exist_ok=True)
DB_PATH = os.path.join(PERSIST_DIR, "asistencia_tareo.db")
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "cambiar-clave-en-render")
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024

# ========================= DB =========================
def is_pg(): return USE_POSTGRES and psycopg2 is not None

def get_conn():
    if is_pg():
        return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def qmark(sql): return sql.replace("?", "%s") if is_pg() else sql

def row_to_dict(r): return dict(r) if r else None

def rows_to_dict(rows): return [row_to_dict(r) for r in (rows or [])]

def execute(sql, params=(), fetchone=False, fetchall=False, commit=False):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(qmark(sql), params)
    data = None
    if fetchone: data = cur.fetchone()
    if fetchall: data = cur.fetchall()
    if commit: conn.commit()
    cur.close(); conn.close()
    return data

def scalar(sql, params=()):
    r = row_to_dict(execute(sql, params, fetchone=True))
    return list(r.values())[0] if r else 0

def now_str(): return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
def today_str(): return date.today().strftime("%Y-%m-%d")


def calcular_horas_laborales(hora_inicio, hora_fin, ref_inicio=None, ref_fin=None):
    """Calcula horas netas entre inicio/fin, restando refrigerio si cruza el rango.
    Soporta turno noche cuando la hora fin es menor o igual que inicio.
    """
    def to_min(v):
        try:
            hh, mm = str(v or "00:00")[:5].split(":")
            return int(hh) * 60 + int(mm)
        except Exception:
            return 0
    ini = to_min(hora_inicio)
    fin = to_min(hora_fin)
    if fin <= ini:
        fin += 24 * 60
    total = max(0, fin - ini)
    if ref_inicio and ref_fin:
        ri = to_min(ref_inicio)
        rf = to_min(ref_fin)
        if rf <= ri:
            rf += 24 * 60
        # Si la jornada cruza medianoche y el refrigerio quedó antes del inicio, moverlo al día siguiente.
        if fin > 24 * 60 and ri < ini:
            ri += 24 * 60
            rf += 24 * 60
        cruce = max(0, min(fin, rf) - max(ini, ri))
        total -= cruce
    return round(max(0, total) / 60, 2)


def calcular_horas_nocturnas(hora_inicio, hora_fin, ref_inicio=None, ref_fin=None):
    """Horas nocturnas dentro de 22:00 a 06:00, restando refrigerio si cruza ese rango."""
    def to_min(v):
        try:
            hh, mm = str(v or "00:00")[:5].split(":")
            return int(hh) * 60 + int(mm)
        except Exception:
            return 0
    ini = to_min(hora_inicio)
    fin = to_min(hora_fin)
    if fin <= ini:
        fin += 24 * 60
    intervals = [(22*60, 30*60), (46*60, 54*60)]
    noct = 0
    for a,b in intervals:
        noct += max(0, min(fin, b) - max(ini, a))
    if ref_inicio and ref_fin:
        ri, rf = to_min(ref_inicio), to_min(ref_fin)
        if rf <= ri:
            rf += 24*60
        if fin > 24*60 and ri < ini:
            ri += 24*60; rf += 24*60
        for a,b in intervals:
            noct -= max(0, min(rf, b, fin) - max(ri, a, ini))
    return round(max(0, noct) / 60, 2)

def _minutos_hora(v):
    try:
        hh, mm = str(v or '00:00')[:5].split(':')
        return int(hh) * 60 + int(mm)
    except Exception:
        return 0

def horario_coherente(hora_inicio, hora_fin, ref_inicio, ref_fin):
    """Valida que el refrigerio esté dentro de la jornada, incluso si cruza medianoche."""
    hi = _minutos_hora(hora_inicio); hf = _minutos_hora(hora_fin)
    ri = _minutos_hora(ref_inicio); rf = _minutos_hora(ref_fin)
    if hf <= hi:
        hf += 24 * 60
    if rf <= ri:
        rf += 24 * 60
    if hf > 24 * 60 and ri < hi:
        ri += 24 * 60; rf += 24 * 60
    if not (hi <= ri < rf <= hf):
        return False, 'El refrigerio debe estar dentro de la jornada de inicio y fin.'
    if (rf - ri) > (hf - hi):
        return False, 'El refrigerio no puede ser mayor que la jornada.'
    # No se bloquea por duración total de jornada.
    # La validación solicitada es únicamente que el refrigerio quede dentro
    # del rango de inicio y fin de trabajo, incluso si cruza medianoche.
    return True, ''

def hoja_enviada(hoja_id):
    h = row_to_dict(execute('SELECT estado FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    return bool(h and str(h.get('estado') or '').upper() == 'ENVIADA')

def _add_column_if_missing(cur, table, column, ddl):
    """Migración segura para SQLite/PostgreSQL."""
    try:
        if is_pg():
            cur.execute(qmark(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {ddl}"))
        else:
            cur.execute(f"PRAGMA table_info({table})")
            cols = [r[1] for r in cur.fetchall()]
            if column not in cols:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {ddl}")
    except Exception as e:
        print(f"No se pudo agregar columna {table}.{column}:", e)

def init_db():
    conn = get_conn(); cur = conn.cursor()
    idtype = "SERIAL PRIMARY KEY" if is_pg() else "INTEGER PRIMARY KEY AUTOINCREMENT"
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS usuarios(
        id {idtype}, usuario TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
        nombres TEXT, rol TEXT DEFAULT 'operador', estado TEXT DEFAULT 'ACTIVO', creado_en TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS trabajadores(
        id {idtype}, dni TEXT UNIQUE NOT NULL, trabajador TEXT, empresa TEXT,
        area TEXT, cargo TEXT, actividad TEXT, planilla TEXT, estado TEXT DEFAULT 'ACTIVO', fecha_carga TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS asistencia(
        id {idtype}, dni TEXT NOT NULL, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        tipo TEXT NOT NULL, fecha TEXT NOT NULL, hora TEXT NOT NULL, fecha_hora TEXT NOT NULL,
        metodo TEXT, foto_path TEXT, latitud TEXT, longitud TEXT, registrado_por TEXT, observacion TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS documentos_firma(
        id {idtype}, dni TEXT NOT NULL, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        documento TEXT NOT NULL, estado TEXT DEFAULT 'FIRMADO', fecha TEXT NOT NULL, hora TEXT NOT NULL, fecha_hora TEXT NOT NULL,
        metodo TEXT, firma_path TEXT, foto_path TEXT, registrado_por TEXT, observacion TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS transporte_conductores(
        id {idtype}, dni TEXT UNIQUE, nombres TEXT, telefono TEXT, licencia TEXT, categoria TEXT,
        venc_licencia TEXT, venc_cert_medico TEXT, venc_sctr TEXT, estado TEXT DEFAULT 'APTO',
        observacion TEXT, creado_en TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS transporte_vehiculos(
        id {idtype}, placa TEXT UNIQUE NOT NULL, tipo TEXT, capacidad INTEGER DEFAULT 0,
        empresa_transportista TEXT, soat_venc TEXT, revision_tecnica_venc TEXT, gps_codigo TEXT,
        estado TEXT DEFAULT 'ACTIVO', observacion TEXT, creado_en TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS transporte_rutas(
        id {idtype}, fecha TEXT NOT NULL, nombre TEXT, origen TEXT, destino TEXT, sede TEXT,
        hora_salida TEXT, hora_retorno TEXT, vehiculo_id INTEGER, conductor_id INTEGER,
        estado TEXT DEFAULT 'PROGRAMADA', latitud TEXT, longitud TEXT, ultima_ubicacion TEXT,
        creado_por TEXT, creado_en TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS transporte_pasajeros(
        id {idtype}, ruta_id INTEGER NOT NULL, dni TEXT NOT NULL, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        fecha TEXT NOT NULL, hora TEXT NOT NULL, fecha_hora TEXT NOT NULL, metodo TEXT,
        latitud TEXT, longitud TEXT, registrado_por TEXT, observacion TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS transporte_gps(
        id {idtype}, ruta_id INTEGER NOT NULL, latitud TEXT, longitud TEXT, fecha_hora TEXT, registrado_por TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS transporte_ruta_esperados(
        id {idtype}, ruta_id INTEGER NOT NULL, dni TEXT NOT NULL, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        estado TEXT DEFAULT 'ESPERADO', creado_en TEXT, creado_por TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS tareos(
        id {idtype}, hoja_id INTEGER, labor_id INTEGER, dni TEXT NOT NULL, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        fecha TEXT NOT NULL, labor TEXT, lote TEXT, fundo TEXT, horas REAL DEFAULT 0,
        cantidad REAL DEFAULT 0, unidad TEXT, observacion TEXT, registrado_por TEXT, creado_en TEXT,
        hora_inicio TEXT, hora_fin TEXT, ref_inicio TEXT, ref_fin TEXT, turno TEXT, tipo_tareo TEXT, horas_nocturnas REAL DEFAULT 0)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS hojas_tareo(
        id {idtype}, fecha TEXT NOT NULL, grupo TEXT, subgrupo TEXT, labor TEXT, responsable TEXT,
        turno TEXT DEFAULT 'DIA', tipo_tareo TEXT DEFAULT 'JORNAL',
        estado TEXT DEFAULT 'ABIERTA', registros INTEGER DEFAULT 0, horas_total REAL DEFAULT 0, rendimiento_total REAL DEFAULT 0,
        creado_por TEXT, creado_en TEXT, horario_fijado INTEGER DEFAULT 0, hora_inicio_default TEXT, hora_fin_default TEXT, ref_inicio_default TEXT, ref_fin_default TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS hoja_labores(
        id {idtype}, hoja_id INTEGER NOT NULL, grupo TEXT, subgrupo TEXT, labor TEXT,
        turno TEXT DEFAULT 'DIA', tipo_tareo TEXT DEFAULT 'JORNAL', responsable TEXT, creado_en TEXT, creado_por TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS actividades_maestras(
        id {idtype}, cod_actividad TEXT, desc_actividad TEXT, cod_labor TEXT, desc_labor TEXT,
        cod_consumidor TEXT, desc_consumidor TEXT, estado TEXT DEFAULT 'ACTIVO', fecha_carga TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS lecturas_balde(
        id {idtype}, hoja_id INTEGER, labor_id INTEGER, dni TEXT, trabajador TEXT, fecha_hora TEXT,
        a_diurno REAL DEFAULT 0, a_noct REAL DEFAULT 0, metodo TEXT, registrado_por TEXT)"""))

    # Migraciones sobre bases ya existentes en Render
    for col, ddl in [('turno', "TEXT DEFAULT 'DIA'"), ('tipo_tareo', "TEXT DEFAULT 'JORNAL'"), ('horario_fijado','INTEGER DEFAULT 0'), ('hora_inicio_default','TEXT'), ('hora_fin_default','TEXT'), ('ref_inicio_default','TEXT'), ('ref_fin_default','TEXT')]:
        _add_column_if_missing(cur, 'hojas_tareo', col, ddl)
    for col, ddl in [('labor_id','INTEGER'),('hora_inicio','TEXT'),('hora_fin','TEXT'),('ref_inicio','TEXT'),('ref_fin','TEXT'),('turno','TEXT'),('tipo_tareo','TEXT'),('horas_nocturnas','REAL DEFAULT 0')]:
        _add_column_if_missing(cur, 'tareos', col, ddl)
    for col, ddl in [('movil_usuario','TEXT'),('movil_pin','TEXT'),('movil_estado',"TEXT DEFAULT 'ACTIVO'"),('ultimo_gps','TEXT'),('ultima_latitud','TEXT'),('ultima_longitud','TEXT')]:
        _add_column_if_missing(cur, 'transporte_conductores', col, ddl)
    for col, ddl in [('conductor_id','INTEGER'),('placa','TEXT'),('ruta_nombre','TEXT')]:
        _add_column_if_missing(cur, 'transporte_gps', col, ddl)
    for col, ddl in [('hora_llegada','TEXT'),('cerrado_por','TEXT'),('cerrado_en','TEXT'),('km_inicio','REAL DEFAULT 0'),('km_fin','REAL DEFAULT 0')]:
        _add_column_if_missing(cur, 'transporte_rutas', col, ddl)

    cur.execute(qmark("SELECT id FROM usuarios WHERE usuario=?"), ("admin",))
    if not cur.fetchone():
        cur.execute(qmark("INSERT INTO usuarios(usuario,password_hash,nombres,rol,estado,creado_en) VALUES(?,?,?,?,?,?)"),
                    ("admin", generate_password_hash("admin123"), "ADMINISTRADOR", "admin", "ACTIVO", now_str()))
    conn.commit(); cur.close(); conn.close()

# ========================= UTIL =========================
def normalizar_columna(c):
    c = str(c or "").strip().upper()
    for a,b in {"Á":"A","É":"E","Í":"I","Ó":"O","Ú":"U","Ñ":"N"}.items(): c = c.replace(a,b)
    return re.sub(r"\s+", " ", c)

def limpiar_dni(v):
    solo = re.sub(r"\D", "", str(v or ""))
    return solo[-8:] if len(solo) >= 8 else solo

def limpiar_texto(v, upper=True):
    s = "" if v is None else str(v).strip()
    return s.upper() if upper else s

def login_required(f):
    @wraps(f)
    def w(*args, **kwargs):
        if not session.get("usuario"): return redirect(url_for("login"))
        return f(*args, **kwargs)
    return w

def admin_required(f):
    @wraps(f)
    def w(*args, **kwargs):
        if not session.get("usuario"): return redirect(url_for("login"))
        if session.get("rol") != "admin":
            flash("Solo administrador puede ingresar a esta opción.", "danger")
            return redirect(url_for("home"))
        return f(*args, **kwargs)
    return w

def excel_response(headers, rows, filename, sheet="DATOS"):
    wb = Workbook(); ws = wb.active; ws.title = sheet[:31]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h.lower(), r.get(h, "")) for h in headers])
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)
    out = BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ========================= UI =========================
BASE_HTML = r"""
<!doctype html><html lang="es"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#2f773b"><link rel="manifest" href="{{ url_for('manifest') }}">
<title>{{ title or 'Tareo Móvil' }}</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<script src="https://unpkg.com/html5-qrcode" type="text/javascript"></script>
<style>
:root{--verde:#2f773b;--verde2:#3f8748;--verde3:#276a33;--verdeClaro:#eaf5eb;--line:#e8ece8;--txt:#2d5b35;--gris:#6b7d6d;--amarillo:#ffc20e}
*{box-sizing:border-box}body{margin:0;background:#fff;font-family:Inter,Segoe UI,Arial,sans-serif;color:#21472a}.app-bg{min-height:100vh;background:linear-gradient(180deg,#fff 0%,#fff 62%,#fbfbfb 100%)}.shell{width:min(1120px,100%);margin:0 auto;padding:18px}.phone-wrap{max-width:430px;margin:0 auto}.desktop-grid{display:grid;grid-template-columns:360px 1fr;gap:28px;align-items:start}.home-desktop-list .worker-card{margin:10px 0}.home-desktop-list{max-width:640px;margin:0 auto}.header-title{text-align:center;color:#166534;font-family:Georgia,serif;font-weight:900;letter-spacing:.5px;font-size:23px;line-height:1.13;margin:4px 0 22px;text-transform:uppercase}.green-hero{background:var(--verde);border-radius:0 0 18px 18px;min-height:145px;padding:12px 16px 22px;color:white;text-align:center;position:relative;overflow:visible}.tareo-hero{min-height:124px!important;padding-bottom:42px!important}.tareo-toolbar{margin:12px 12px 8px!important;position:relative;z-index:4}.tareo-list-page .worker-card:first-of-type{margin-top:12px}.back-mini{display:inline-grid;place-items:center;width:36px;height:36px;border-radius:999px;color:var(--verde);text-decoration:none;font-size:24px}.green-top{display:flex;justify-content:space-between;align-items:center;font-size:11px;font-weight:800}.avatar{width:78px;height:78px;border-radius:999px;background:white;color:var(--verde);display:grid;place-items:center;margin:10px auto 2px;font-size:43px;box-shadow:0 8px 20px rgba(0,0,0,.13)}.login-name{font-size:11px;font-weight:800}.white-input{height:36px;background:white;border-radius:10px;box-shadow:0 5px 13px rgba(0,0,0,.18);border:0}.floating-card{background:white;border-radius:10px;box-shadow:0 8px 18px rgba(0,0,0,.15);padding:12px}.tile{width:74px;height:70px;border-radius:8px;background:white;box-shadow:0 7px 17px rgba(0,0,0,.14);display:flex;flex-direction:column;align-items:center;justify-content:center;color:var(--verde);font-weight:900;font-size:10px;text-align:center}.tile i{font-size:26px;margin-bottom:5px}.bottom-sync{position:fixed;left:10px;bottom:8px;color:#317a3e;font-size:10px;font-weight:700}.bottom-out{position:fixed;right:14px;bottom:8px;color:#c84c4c;font-size:20px}.tab-main{display:flex;gap:0;background:#fafafa;padding-left:0;border-top:4px solid #d7d7d7}.tab-main a{flex:1;text-align:center;text-decoration:none;color:#508557;font-weight:900;font-size:13px;padding:14px 8px;border-radius:7px 7px 0 0;background:#fff}.tab-main a.active{background:var(--verde);color:white;box-shadow:0 3px 7px rgba(0,0,0,.18)}.subtabs{display:flex;background:#fff}.subtabs a{flex:1;text-align:center;padding:13px 5px;text-decoration:none;color:#4b8a54;font-weight:900;font-size:12px}.subtabs a.active{background:var(--verde);color:white}.panel-green{background:var(--verde);color:white;text-align:center;padding:21px 12px 42px}.panel-green i{font-size:38px}.panel-green h4{font-size:11px;font-weight:900;margin:5px 0 0}.toolstrip{background:white;margin:-25px 9px 5px;border-radius:9px;min-height:49px;box-shadow:0 5px 13px rgba(0,0,0,.22);display:flex;align-items:center;gap:20px;padding:7px 14px;color:var(--verde);font-size:24px}.toolstrip button,.toolstrip a{border:0;background:transparent;color:var(--verde);font-size:24px;text-decoration:none}.info-bar{margin:0 9px;background:var(--verde);color:white;border-radius:2px;display:grid;grid-template-columns:1fr 1fr 1fr 1fr 22px;align-items:center;font-size:10px;font-weight:900;height:23px}.info-bar div{text-align:center;border-right:1px solid rgba(255,255,255,.28)}.worker-card{background:white;margin:10px 12px;border-radius:10px;border:1px solid #e3e8e3;box-shadow:0 3px 12px rgba(0,0,0,.20);padding:11px 13px;color:#397443;position:relative}.worker-title{display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:9px;font-weight:900;text-transform:uppercase}.worker-title b{font-size:10px;color:var(--verde)}.worker-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:7px;margin-top:7px}.worker-grid label{font-size:8px;font-weight:900;color:#6c8a6f;margin-bottom:1px}.mini-input{height:25px;border:1px solid #9ebaa0;border-radius:3px;font-size:10px;padding:3px 5px;width:100%;font-weight:800;color:#315f39}.mini-badge{border-radius:3px;color:white;font-size:8px;font-weight:900;text-align:center;padding:4px 3px;text-transform:uppercase}.bg-y{background:var(--amarillo)!important;color:white}.bg-g{background:#42b852!important}.person-dot{width:42px;height:42px;border-radius:999px;background:#378145;color:white;display:grid;place-items:center;font-size:26px;float:left;margin-right:9px}.small-label{font-size:8px;color:#79937b;font-weight:900}.small-value{font-size:9px;color:#466a49;font-weight:900}.leaf{width:120px;height:120px;border-radius:70% 30% 70% 30%;background:linear-gradient(135deg,#ffd8bd,#eef4c7,#cbd9b6);opacity:.72;margin:20px auto 0;transform:rotate(-20deg)}.card-pro{background:white;border:1px solid var(--line);border-radius:18px;box-shadow:0 8px 20px rgba(0,0,0,.10)}.btn-green{background:var(--verde);border-color:var(--verde);color:white;font-weight:900;border-radius:9px}.btn-green:hover{background:var(--verde3);color:white}.form-control,.form-select{border-radius:9px;border:1px solid #dfe7df;font-weight:700;font-size:13px}.form-label{font-size:12px;font-weight:900;color:#3e7545}.page-card{border-radius:13px;overflow:hidden;border:1px solid #e5e7e5;background:white}.list-table th{font-size:11px;color:#497550}.list-table td{font-size:12px;vertical-align:middle}.status-pill{display:inline-block;background:#39b54a;color:white;border-radius:4px;padding:4px 8px;font-size:9px;font-weight:900}.top-actions{display:flex;gap:12px;flex-wrap:wrap;justify-content:center;margin-top:14px;position:relative;z-index:5}.top-actions .tile{width:82px;height:76px}.login-page .shell{padding:0}.login-form{margin:-7px auto 0;width:92%;max-width:360px}.login-form .floating-card{padding:13px 14px 18px}.alert{border-radius:12px;font-size:13px}.desk-panel{display:block}.mobile-only{display:none}.clock-box{width:116px;height:116px;border:5px solid var(--verde);border-radius:999px;margin:8px auto;display:grid;place-items:center;color:var(--verde);font-weight:900;background:#fff;box-shadow:0 4px 14px rgba(47,119,59,.18)}.clock-box i{font-size:38px}.scan-box{border:2px dashed #8dbf93;border-radius:12px;padding:10px;background:#f8fff9}.toolstrip .hint{font-size:9px;font-weight:900;color:#2f773b;margin-left:-18px;margin-right:0}.splash-card{height:94vh;max-height:760px;background:#23773f;border-radius:10px;box-shadow:0 4px 12px rgba(0,0,0,.25);display:flex;flex-direction:column;align-items:center;justify-content:center;color:white;position:relative}.splash-logo{width:145px;height:145px;border-radius:999px;background:#fff;border:6px solid #92bd33;display:grid;place-items:center;color:#23773f;font-size:66px;box-shadow:0 3px 10px rgba(0,0,0,.22)}.splash-title{font-weight:900;margin-top:18px;letter-spacing:.5px}.splash-foot{position:absolute;bottom:26px;text-align:center;font-size:11px;color:#d9f2df;font-weight:700}.role-toggle{display:grid;grid-template-columns:1fr 1fr;gap:8px}.role-toggle label{border:1px solid #dce7dc;border-radius:9px;padding:9px;text-align:center;font-size:12px;font-weight:900;color:#2f773b}.role-toggle input{display:none}.role-toggle input:checked+span{background:#2f773b;color:white;border-radius:7px;padding:7px 9px;display:block}.bottom-nav{position:sticky;bottom:0;background:white;border-top:1px solid #e8ece8;display:flex;justify-content:space-around;padding:7px 0;color:#477b4d;font-size:10px;font-weight:800}.bottom-nav a{text-decoration:none;color:#477b4d;text-align:center}.bottom-nav i{display:block;font-size:17px}.copy-list{max-height:260px;overflow:auto;border:1px solid #e5ede5;border-radius:9px;padding:8px;background:#fbfffb}.clock-face{width:180px;height:180px;border-radius:999px;background:#e7e0ef;margin:8px auto;position:relative;display:grid;place-items:center;color:#5d44aa;touch-action:none;cursor:pointer;user-select:none}.clock-hand{width:65px;height:4px;background:#6b4eb8;position:absolute;transform-origin:left center;transform:rotate(-35deg);left:90px;top:90px;pointer-events:none}.clock-hand:after{content:'';position:absolute;right:-13px;top:-13px;width:30px;height:30px;border-radius:999px;background:#6b4eb8;box-shadow:0 2px 8px rgba(0,0,0,.18)}.clock-dot{width:12px;height:12px;border-radius:999px;background:#6b4eb8;position:absolute;left:84px;top:84px}.clock-num{position:absolute;font-size:12px;color:#37303c}.clock-bubble{position:absolute;right:20px;top:50px;background:#6b4eb8;color:white;border-radius:999px;padding:9px;font-weight:900}.field-required{box-shadow:inset 4px 0 0 var(--verde)}.big-plus{font-size:34px!important;line-height:1}.big-plus .bi-plus{font-size:22px!important;margin-left:-12px;font-weight:900}.labor-card-compact{padding:13px 16px}.labor-card-compact .worker-title{font-size:8px}.labor-card-compact .worker-title b{font-size:9px;line-height:1.15}.labor-card-compact .labor-main{font-size:15px!important;line-height:1.1;color:#146c35}.labor-card-compact .resp-main{font-size:13px!important;line-height:1.1;color:#146c35}.worker-queue{border:1px dashed #8cc79b;border-radius:12px;background:#f7fff8;padding:9px;margin-top:10px;max-height:155px;overflow:auto}.queue-item{display:flex;justify-content:space-between;gap:8px;align-items:center;border-bottom:1px solid #e2f3e5;padding:6px 0;font-size:12px}.queue-item:last-child{border-bottom:0}.scan-ok{background:#d1fae5;border:1px solid #86efac;color:#166534;border-radius:10px;padding:8px;font-size:12px;font-weight:800}.scan-bad{background:#fee2e2;border:1px solid #fecaca;color:#991b1b;border-radius:10px;padding:8px;font-size:12px;font-weight:800}.time-click{cursor:pointer;background:#fbfffb}.time-click:focus{outline:2px solid #2f773b}.report-wrap{max-width:540px;margin:0 auto}.config-header{display:flex;align-items:center;gap:8px;justify-content:center;position:relative}.config-header .back-mini{position:absolute;left:0}.btn-plus-fab{display:inline-flex!important;align-items:center;gap:2px}.btn-plus-fab i:first-child{font-size:30px!important}.btn-plus-fab i:last-child{font-size:20px!important;margin-left:-12px;margin-top:12px}.field-help{font-size:10px;color:#5f7d65;font-weight:800} .swipe-wrap{position:relative;margin:10px 12px;overflow:hidden;border-radius:12px;border:1px solid #e3e8e3;background:#fff;box-shadow:0 2px 8px rgba(0,0,0,.06)}.swipe-actions{position:absolute;right:0;top:0;bottom:0;display:flex;align-items:stretch;transform:translateX(100%);transition:.22s ease;z-index:1}.swipe-wrap.show-actions .swipe-actions{transform:translateX(0)}.swipe-actions a{display:flex;align-items:center;justify-content:center;min-width:74px;color:white;text-decoration:none;font-size:11px;font-weight:900}.act-edit{background:#2563eb}.act-send{background:#16a34a}.act-del{background:#dc2626}.swipe-wrap .worker-card{margin:0;transition:.22s ease;position:relative;z-index:2}.swipe-wrap.show-actions .worker-card{transform:translateX(-222px)}.locked-input{background:#f8fff9!important;cursor:pointer}.edit-hint{font-size:9px;color:#2f773b;font-weight:900}.clock-field-pills{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin:4px 0 10px}.clock-field-pills button{border:1px solid #cfe6d4;background:#fff;border-radius:9px;padding:7px;font-size:11px;font-weight:900;color:#2f773b}.clock-field-pills button.active{background:#2f773b;color:white}.clock-mode{display:flex;justify-content:center;gap:8px;margin-top:4px}.clock-mode button{border:0;border-radius:999px;padding:5px 12px;font-weight:900;background:#e7e0ef;color:#5d44aa}.clock-mode button.active{background:#6b4eb8;color:white}.modal-suggest{max-height:125px;overflow:auto;border:1px solid #e5ede5;border-radius:8px;margin-top:-5px;margin-bottom:8px;background:#fbfffb;display:none}.modal-suggest div{padding:8px 10px;border-bottom:1px solid #edf5ee;font-size:12px;font-weight:800;cursor:pointer}.modal-suggest div:hover{background:#eaf5eb}.queue-title{font-size:11px;font-weight:900;color:#166534;margin-top:8px}.scan-ok.flash{animation:flashOk .5s ease}@keyframes flashOk{0%{transform:scale(.98)}50%{transform:scale(1.02)}100%{transform:scale(1)}} 

/* ===== MEJORAS 246: mayúsculas, reloj táctil y layout horas ===== */
input[type="text"], input:not([type]), textarea { text-transform:uppercase; }
.time-worker-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px}
.time-worker-grid label{font-size:8px;font-weight:900;color:#55745a;margin-bottom:2px;display:block}
.time-box{background:#fbfffb;border:1px solid #9ebaa0;border-radius:5px;padding:5px 6px;font-size:11px;font-weight:900;color:#245a31;min-height:29px}
.time-metrics{display:grid;grid-template-columns:1fr;gap:7px;margin-top:8px;background:#eef8ef;border:1px solid #b9d7bd;border-radius:8px;padding:8px}
.time-metrics label{font-size:8px;font-weight:900;color:#176a35;margin-bottom:2px;display:block}
.metric-box{background:#dff3e2;border:1px solid #81ba8a;border-radius:6px;padding:7px 8px;font-size:13px;font-weight:900;color:#0f6b2c;min-height:32px}
.ref-worker-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px}
.ref-worker-grid label{font-size:8px;font-weight:900;color:#55745a;margin-bottom:2px;display:block}
.clock-face.pro-clock{width:210px;height:210px;background:#f4f8f4;border:7px solid #2f773b;color:#2f773b;box-shadow:0 6px 18px rgba(47,119,59,.2)}
.clock-face.pro-clock .clock-hand{left:105px;top:105px;background:#2f773b;width:76px;height:5px}
.clock-face.pro-clock .clock-hand:after{background:#2f773b}
.clock-face.pro-clock .clock-dot{left:99px;top:99px;background:#2f773b}
.clock-face.pro-clock .clock-bubble{background:#2f773b}
.clock-24-hint{text-align:center;font-size:11px;font-weight:900;color:#2f773b;margin-top:-4px;margin-bottom:6px}


/* === AJUSTE FINAL OMAR: modal compacto, tarjetas como referencia y controles táctiles/PC === */
.modal-dialog{max-width:520px!important;margin:.7rem auto!important}.modal-content{border-radius:12px!important}.modal-header{padding:10px 16px!important}.modal-body{padding:14px 16px!important}.modal-footer{padding:10px 16px!important}.modal .form-label{font-size:11px!important;margin-bottom:4px!important}.modal .form-control,.modal .form-select{height:41px!important;font-size:13px!important;border-radius:9px!important}.modal .btn-green{height:47px!important;border-radius:9px!important}.modal .alert{padding:10px 12px!important;margin-bottom:10px!important}
#modalLabor .modal-dialog{max-width:500px!important}#modalLabor .modal-body{padding-top:12px!important}#modalLabor .form-control,#modalLabor .form-select{height:40px!important}#modalLabor .modal-suggest{position:relative;z-index:2000;max-height:160px;overflow:auto;background:#fff;border:1px solid #b8d9bf;border-radius:8px;margin:2px 0 8px;box-shadow:0 8px 20px rgba(0,0,0,.10);display:none}#modalLabor .modal-suggest div{padding:9px 10px;font-size:12px;font-weight:800;color:#245c31;border-bottom:1px solid #edf5ee;cursor:pointer}#modalLabor .modal-suggest div:hover{background:#eaf5eb}.master-status{font-size:10px;font-weight:900;color:#2f773b;margin:-2px 0 6px}.master-status.bad{color:#b42318}
#modalHora .modal-dialog{max-width:490px!important}.touch-clock-panel{padding:10px!important;border-radius:12px!important}.time-display{height:42px!important;font-size:28px!important;line-height:39px!important}.time-slider{width:100%!important;height:34px!important;display:block!important;cursor:pointer!important;touch-action:pan-x!important;accent-color:#2f773b!important}.time-slider::-webkit-slider-thumb{width:28px!important;height:28px!important;cursor:pointer!important}.touch-clock-picks{grid-template-columns:1fr 1fr!important;gap:7px!important}.touch-clock-picks button{height:35px!important;border-radius:8px!important;font-size:10px!important}.locked-input{pointer-events:auto!important;cursor:pointer!important;background:#f8fff9!important}.trabajador-card-ref{padding:20px 25px!important;border-radius:13px!important;max-width:790px;margin:10px auto!important}.trabajador-card-ref .worker-title{grid-template-columns:1fr 1fr!important;font-size:12px!important;gap:35px!important}.trabajador-card-ref .worker-title b{font-size:18px!important;line-height:1.15!important}.trabajador-grid-ref{display:grid;grid-template-columns:1fr 1fr 1fr;gap:26px 20px;margin-top:34px}.trabajador-grid-ref label{font-size:13px;font-weight:900;color:#6b7d6d;margin-bottom:8px;display:block}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:48px!important;display:flex;align-items:center;border:1px solid #9ebaa0;border-radius:4px;background:#fbfffb;color:#006b2e;font-weight:900;font-size:22px;padding:8px 12px}.trabajador-grid-ref .mini-badge{height:48px;display:flex;align-items:center;justify-content:center;font-size:15px;border-radius:4px;max-width:236px!important}.editable-tareo{cursor:pointer}.editable-tareo:hover{outline:2px solid #2f773b}
@media(max-width:860px){.modal-dialog{max-width:94%!important;margin:.65rem auto!important}.trabajador-card-ref{padding:12px 14px!important;max-width:505px!important}.trabajador-card-ref .worker-title{font-size:9px!important;gap:10px!important}.trabajador-card-ref .worker-title b{font-size:13px!important}.trabajador-grid-ref{gap:12px 8px;margin-top:18px}.trabajador-grid-ref label{font-size:9px;margin-bottom:5px}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:35px!important;font-size:15px!important;padding:5px 8px}.trabajador-grid-ref .mini-badge{height:36px;font-size:11px;max-width:185px!important}}

@media(max-width:860px){.shell{padding:0}.desktop-grid{display:block}.desk-panel{display:none}.mobile-only{display:block}.header-title{font-size:17px;margin:16px 7px 20px}.page-card{border-radius:0;border-left:0;border-right:0}.phone-wrap{max-width:100%}.green-hero{border-radius:0}.worker-card{margin-left:9px;margin-right:9px}.toolstrip{gap:15px}.info-bar{font-size:8.5px}.bottom-sync,.bottom-out{position:fixed}.desktop-pad{padding:0 0 28px}.tab-main a,.subtabs a{font-size:11px}.worker-grid{gap:5px}.floating-card{border-radius:9px}.top-actions .tile{width:72px;height:70px}}

/* ===== AJUSTE FINAL: horario táctil compacto y edición de tareo ===== */
.clock-face.pro-clock{display:none!important}
.touch-clock-panel{background:#f8fff9;border:1px solid #cfe6d4;border-radius:10px;padding:8px;margin:6px 0 10px}
.touch-clock-picks{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:8px}
.touch-clock-picks button{border:1px solid #b9d7bd;background:#fff;border-radius:8px;padding:7px 4px;font-size:10px;font-weight:900;color:#2f773b}
.touch-clock-picks button.active{background:#2f773b;color:white}
.touch-clock-actions{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:6px}
.touch-clock-actions button{border:0;background:#2f773b;color:white;border-radius:8px;padding:8px 2px;font-size:11px;font-weight:900}
.touch-clock-value{text-align:center;font-size:24px;font-weight:900;color:#166534;background:#fff;border:1px solid #cfe6d4;border-radius:10px;padding:6px;margin-bottom:8px}
.modal-dialog{max-width:390px}.modal-content{border-radius:13px}.modal-header{padding:10px 14px}.modal-body{padding:12px 14px}.modal-footer{padding:10px 14px}.modal-title{font-size:18px}.modal .form-label{font-size:11px;margin-bottom:3px}.modal .form-control,.modal .form-select{height:37px;font-size:12px;padding:6px 10px}.modal .alert{font-size:12px;padding:9px 11px;margin-bottom:10px}
.time-slider{width:100%;accent-color:#2f773b;touch-action:pan-y;margin:6px 0 3px}.time-display{height:33px;border:1px solid #cfe6d4;border-radius:9px;background:#fff;display:grid;place-items:center;font-weight:900;color:#166534;font-size:20px;margin-bottom:7px}.touch-clock-panel{padding:7px!important;margin:4px 0 8px!important}.touch-clock-picks{gap:5px!important;margin-bottom:6px!important}.touch-clock-picks button{padding:6px 3px!important;font-size:9px!important}.touch-clock-actions{display:none!important}
.time-worker-grid,.ref-worker-grid{gap:6px}.time-metrics{padding:7px;gap:6px}.worker-card{padding:9px 11px}.metric-box{padding:5px 7px;min-height:28px}.time-box{min-height:27px;padding:4px 6px}
.worker-card.editable-tareo{cursor:pointer}.worker-card.editable-tareo:after{content:'Tocar para editar';position:absolute;right:10px;bottom:8px;font-size:8px;font-weight:900;color:#2f773b}


/* === PARCHE REAL 247: tarjeta trabajador en 3 columnas compacto === */
.trabajador-card-ref{max-width:790px!important;padding:18px 24px!important}
.trabajador-grid-ref{display:grid!important;grid-template-columns:repeat(3,1fr)!important;gap:22px 18px!important;margin-top:28px!important}
.trabajador-grid-ref label{font-size:12px!important;margin-bottom:7px!important}
.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:46px!important;font-size:20px!important;background:#fbfffb!important;border:1px solid #9ebaa0!important;border-radius:4px!important;display:flex!important;align-items:center!important;color:#006b2e!important;font-weight:900!important}
@media(max-width:860px){.trabajador-card-ref{max-width:505px!important;padding:10px 13px!important}.trabajador-grid-ref{grid-template-columns:repeat(3,1fr)!important;gap:10px 7px!important;margin-top:16px!important}.trabajador-grid-ref label{font-size:8.5px!important;margin-bottom:4px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:32px!important;font-size:13px!important;padding:4px 7px!important}.trabajador-card-ref .worker-title b{font-size:12px!important}}

/* ===== PATCH 248 OMAR: sin doble desplegable, labor obligatoria, tarjeta compacta y slider tipo manija ===== */
#modalLabor datalist{display:none!important}
#modalLabor .modal-dialog{max-width:390px!important}
#modalLabor .modal-body{padding:10px 14px!important}
#modalLabor .form-control,#modalLabor .form-select{height:36px!important;font-size:12px!important}
#modalLabor .modal-suggest{position:absolute!important;left:20px!important;right:20px!important;z-index:4000!important;max-height:142px!important;overflow:auto!important;background:#fff!important;box-shadow:0 10px 24px rgba(0,0,0,.18)!important}
#modalHora .modal-dialog{max-width:430px!important}.touch-clock-panel{padding:7px!important}.time-display{height:36px!important;font-size:22px!important}.time-slider{height:38px!important;cursor:pointer!important;touch-action:none!important;appearance:none!important;-webkit-appearance:none!important;background:transparent!important}.time-slider::-webkit-slider-runnable-track{height:8px;background:#dfe7df;border:1px solid #9ebaa0;border-radius:999px}.time-slider::-webkit-slider-thumb{-webkit-appearance:none!important;width:38px!important;height:38px!important;margin-top:-16px!important;border-radius:999px!important;background:#2f773b!important;border:4px solid #fff!important;box-shadow:0 3px 9px rgba(0,0,0,.28)!important;cursor:grab!important}.time-slider::-webkit-slider-thumb:active{cursor:grabbing!important}.time-slider::-moz-range-track{height:8px;background:#dfe7df;border:1px solid #9ebaa0;border-radius:999px}.time-slider::-moz-range-thumb{width:34px!important;height:34px!important;border-radius:999px!important;background:#2f773b!important;border:4px solid #fff!important;box-shadow:0 3px 9px rgba(0,0,0,.28)!important;cursor:grab!important}.touch-clock-panel:after{content:'◷';display:block;text-align:center;color:#2f773b;font-size:18px;font-weight:900;margin-top:-4px}.trabajador-card-ref{max-width:100%!important;margin:7px 8px!important;padding:10px 14px!important;border-radius:12px!important}.trabajador-card-ref .worker-title{gap:8px!important;font-size:8px!important}.trabajador-card-ref .worker-title b{font-size:13px!important;line-height:1.12!important}.trabajador-grid-ref{grid-template-columns:repeat(3,1fr)!important;gap:10px 8px!important;margin-top:14px!important}.trabajador-grid-ref label{font-size:8.5px!important;margin-bottom:4px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:31px!important;font-size:14px!important;padding:4px 7px!important;border-radius:4px!important}.trabajador-grid-ref .mini-badge{height:35px!important;font-size:11px!important;max-width:100%!important}.worker-card.editable-tareo:after{right:10px!important;bottom:5px!important;font-size:8px!important}.phone-wrap{max-width:560px!important}@media(max-width:860px){.phone-wrap{max-width:100%!important}.trabajador-card-ref{margin:7px 5px!important;padding:9px 12px!important}.trabajador-card-ref .worker-title b{font-size:12px!important}.trabajador-grid-ref{gap:9px 7px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:30px!important;font-size:13px!important}}


/* PATCH URGENTE: coherencia horario, avance auto, registros compactos */
.trabajador-card-ref{padding:7px 10px!important;margin:5px 4px!important;border-width:2px!important}.trabajador-card-ref .worker-title b{font-size:12px!important}.trabajador-grid-ref{gap:7px 6px!important;margin-top:9px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:28px!important;justify-content:center!important;text-align:center!important;font-size:12px!important}.trabajador-grid-ref .metric-box{font-size:14px!important;background:#daf5dc!important}.trabajador-grid-ref label{font-size:7.5px!important}.avance-card-ref{padding:8px 10px!important;margin:6px 10px!important}.avance-card-ref .worker-title b{font-size:12px!important}.avance-grid .mini-badge{height:26px!important;display:flex;align-items:center;justify-content:center}.labor-card-compact{padding:10px 13px!important}.labor-card-compact .labor-main{font-size:16px!important}.labor-card-compact .resp-main{font-size:14px!important}.page-card{max-width:620px!important;margin:0 auto!important}.phone-wrap{max-width:650px!important}#modalHora input.locked-input{background:#fff!important;cursor:text!important}.scan-ok{background:#e7f7ea!important;border-color:#8ad092!important}.scan-bad{background:#fee2e2!important;border-color:#fca5a5!important;color:#991b1b!important}
@media(max-width:720px){.phone-wrap{max-width:100%!important}.page-card{max-width:100%!important}.trabajador-card-ref{padding:6px 8px!important}.trabajador-grid-ref{gap:6px 5px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:26px!important;font-size:11px!important}.trabajador-grid-ref .metric-box{font-size:13px!important}.avance-card-ref{margin:5px 8px!important}}


/* === PATCH 249 FINAL: ancho tipo celular + tarjetas compactas + flecha acciones === */
html, body{max-width:100%;overflow-x:hidden!important;background:#fff!important;}
.shell{width:100%!important;max-width:430px!important;margin:0 auto!important;padding:6px 8px!important;}
.phone-wrap{width:100%!important;max-width:390px!important;margin:0 auto!important;}
.page-card{width:100%!important;max-width:390px!important;margin:0 auto!important;border-radius:12px!important;overflow:hidden!important;}
.header-title{font-size:16px!important;line-height:1.08!important;margin:10px 4px 12px!important;letter-spacing:.2px!important;}
.green-hero,.panel-green{padding:13px 10px 28px!important;min-height:104px!important;border-radius:0!important;}
.panel-green i{font-size:28px!important}.panel-green h4{font-size:10px!important;line-height:1.15!important;margin-top:4px!important;}
.tab-main a{font-size:11px!important;padding:10px 4px!important}.subtabs a{font-size:10px!important;padding:8px 4px!important;line-height:1.15!important;}
.toolstrip{margin:-22px 8px 5px!important;min-height:42px!important;padding:6px 11px!important;gap:16px!important;border-radius:8px!important;}
.toolstrip button,.toolstrip a{font-size:21px!important}.info-bar{margin:0 8px!important;height:22px!important;font-size:9px!important;grid-template-columns:1fr 1fr 1fr 1fr 18px!important;}
.leaf{width:80px!important;height:80px!important;margin:14px auto 0!important;opacity:.55!important;}
.worker-card{margin:7px 7px!important;padding:8px 9px!important;border-radius:9px!important;box-shadow:0 4px 11px rgba(0,0,0,.13)!important;}
.labor-card-compact{padding:8px 10px!important}.labor-card-compact .worker-title{font-size:7.5px!important}.labor-card-compact .worker-title b{font-size:9px!important}.labor-card-compact .labor-main{font-size:14px!important}.labor-card-compact .resp-main{font-size:12px!important}.labor-card-compact .mini-badge{height:22px!important;font-size:8px!important;padding:4px!important;}
.trabajador-card-ref{width:auto!important;max-width:100%!important;margin:6px 5px!important;padding:7px 9px 18px!important;border-radius:10px!important;border-width:2px!important;}
.trabajador-card-ref .worker-title{grid-template-columns:1.25fr .85fr!important;gap:5px!important;font-size:7px!important;line-height:1.1!important;}
.trabajador-card-ref .worker-title b{font-size:10.5px!important;line-height:1.12!important;}
.trabajador-grid-ref{display:grid!important;grid-template-columns:repeat(3,1fr)!important;gap:6px 6px!important;margin-top:9px!important;}
.trabajador-grid-ref label{font-size:7px!important;margin-bottom:2px!important;line-height:1!important;}
.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:27px!important;font-size:11px!important;padding:3px 5px!important;justify-content:center!important;text-align:center!important;border-radius:4px!important;}
.trabajador-grid-ref .metric-box{font-size:13px!important;background:#dff6df!important;border:1px solid #42b852!important;color:#006b2e!important;}
.trabajador-grid-ref div:has(label+ .metric-box):last-child .metric-box,.trabajador-grid-ref div:nth-child(6) .metric-box{background:#dbeafe!important;border-color:#60a5fa!important;color:#0057a8!important;}
.trabajador-grid-ref .mini-badge{height:30px!important;font-size:10px!important;max-width:100%!important;}
.avance-card-ref{margin:6px 7px!important;padding:8px 9px!important;border-radius:9px!important;min-height:82px!important;}
.avance-card-ref .person-dot{width:34px!important;height:34px!important;font-size:20px!important;margin-right:8px!important}.avance-card-ref .worker-title{font-size:7.5px!important}.avance-card-ref .worker-title b{font-size:10px!important}.avance-grid{grid-template-columns:1fr 1fr 32px!important;align-items:end!important;gap:6px!important}.avance-grid .mini-badge{height:24px!important;font-size:9px!important;display:flex!important;align-items:center!important;justify-content:center!important;}
.card-action-chevron{position:absolute;right:8px;top:50%;transform:translateY(-50%);width:28px;height:38px;display:grid;place-items:center;color:#08713b;font-size:28px;font-weight:900;cursor:pointer;z-index:10;border-radius:8px;background:rgba(255,255,255,.75)}
.card-action-chevron:hover{background:#eaf7ed}.card-action-chevron::before{content:'‹';}
.card-menu{display:none;position:absolute;right:36px;top:50%;transform:translateY(-50%);background:#fff;border:1px solid #b9d7bd;border-radius:9px;box-shadow:0 8px 20px rgba(0,0,0,.18);z-index:20;overflow:hidden;min-width:112px;}
.card-menu.show{display:block}.card-menu button,.card-menu a{display:block;width:100%;padding:9px 10px;border:0;background:#fff;text-decoration:none;text-align:left;color:#166534;font-size:11px;font-weight:900}.card-menu .danger{color:#b42318}.card-menu button:hover,.card-menu a:hover{background:#eef8ef;}
.worker-card.editable-tareo:after{display:none!important;content:''!important;}
/* formulario crear hoja y modales encajados */
body .modal-dialog{max-width:365px!important;margin:.55rem auto!important;}.modal-content{border-radius:12px!important}.modal-header{padding:8px 12px!important}.modal-body{padding:10px 12px!important}.modal-footer{padding:8px 12px!important}.modal-title{font-size:17px!important}.modal .alert{font-size:11px!important;padding:8px 10px!important}.modal .form-label{font-size:10px!important;margin-bottom:3px!important}.modal .form-control,.modal .form-select{height:34px!important;font-size:12px!important;border-radius:8px!important;padding:5px 9px!important}.modal .btn-green{height:40px!important;font-size:13px!important;border-radius:8px!important;}
#createHojaCompact .panel-green{padding:14px 10px 24px!important;min-height:90px!important}#createHojaCompact .floating-card{margin:-18px 7px 10px!important;padding:10px!important}#createHojaCompact .form-control,#createHojaCompact .form-select{height:36px!important;font-size:12px!important}#createHojaCompact .btn{height:38px!important;font-size:13px!important;padding:6px 10px!important}.desktop-pad{padding:0 0 18px!important;}
/* horario: recuadros de hora deslizables y digitables */
#modalHora .touch-clock-panel{padding:7px!important;margin:4px 0 8px!important}#modalHora .clock-face{width:165px!important;height:165px!important}#modalHora .time-display{height:34px!important;font-size:21px!important}.time-drag-help{font-size:9px;color:#166534;font-weight:900;margin-top:3px}.time-draggable{cursor:ew-resize!important;touch-action:none!important;background:#fff!important;}
@media(max-width:420px){.shell{padding:4px 6px!important}.phone-wrap,.page-card{max-width:360px!important}.tab-main a{font-size:10px!important}.subtabs a{font-size:9px!important}.toolstrip{gap:13px!important}.trabajador-card-ref .worker-title b{font-size:9.5px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:25px!important;font-size:10.5px!important}.trabajador-grid-ref .metric-box{font-size:12px!important}.modal-dialog{max-width:94%!important}}


/* PATCH 252: botón visible para apagar cámara */
.scanner-close-x{position:absolute!important;right:8px!important;top:8px!important;z-index:99999!important;width:36px!important;height:36px!important;border:0!important;border-radius:999px!important;background:#dc2626!important;color:#fff!important;font-size:24px!important;font-weight:900!important;line-height:34px!important;text-align:center!important;box-shadow:0 3px 10px rgba(0,0,0,.28)!important;}
.scanner-close-x:hover{background:#b91c1c!important;}
[id^="reader"]{position:relative!important;min-height:0;}
</style></head><body class="{{ 'login-page' if not session.get('usuario') else '' }}"><div class="app-bg"><main class="shell">
{% with messages=get_flashed_messages(with_categories=true) %}{% if messages %}<div class="phone-wrap mt-2">{% for cat,msg in messages %}<div class="alert alert-{{cat}} shadow-sm">{{msg}}</div>{% endfor %}</div>{% endif %}{% endwith %}
{{ body|safe }}</main></div>
<audio id="sndOk"><source src="data:audio/wav;base64,UklGRjQAAABXQVZFZm10IBAAAAABAAEAESsAACJWAAACABAAZGF0YRAAAAAAAP//AAD//wAA//8AAP//AAD//wAA//8=" type="audio/wav"></audio>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script><script>function beep(){try{let a=document.getElementById('sndOk');a.currentTime=0;a.play().catch(()=>{});}catch(e){}}if('serviceWorker' in navigator){navigator.serviceWorker.register('/sw.js').catch(()=>{});}function bindSwipeCards(){document.querySelectorAll('.swipe-wrap').forEach(w=>{let sx=0,dx=0;w.addEventListener('touchstart',e=>{sx=e.touches[0].clientX;});w.addEventListener('touchmove',e=>{dx=e.touches[0].clientX-sx;});w.addEventListener('touchend',()=>{if(dx<-35)w.classList.add('show-actions'); if(dx>35)w.classList.remove('show-actions'); dx=0;});w.addEventListener('contextmenu',e=>{e.preventDefault();w.classList.toggle('show-actions');});});}document.addEventListener('DOMContentLoaded',bindSwipeCards);
</script>

<script>
/* PARCHE FINAL DE LECTURA AUTOMÁTICA DNI/QR/BARRAS - no depende del script del modal */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  let timer=null, busy=false, last='';
  window.__dniQueueFinal = window.__dniQueueFinal || new Map();
  function dni(v){
    const raw=String(v||'');
    const m=raw.match(/(?:^|\D)(\d{8})(?:\D|$)/);
    const d=m?m[1]:raw.replace(/\D/g,'');
    return d.length>=8?d.slice(-8):d;
  }
  function sound(ok=true){
    try{
      if(typeof beep==='function'){beep();return;}
      const C=window.AudioContext||window.webkitAudioContext; if(!C)return;
      const ctx=new C(), o=ctx.createOscillator(), g=ctx.createGain();
      o.type='sine'; o.frequency.value=ok?880:220; g.gain.value=.08;
      o.connect(g); g.connect(ctx.destination); o.start(); setTimeout(()=>{try{o.stop();ctx.close();}catch(e){}},140);
    }catch(e){}
  }
  function st(kind,msg){
    const e=$('dniStatus'); if(!e)return;
    e.className=(kind==='ok'?'scan-ok mt-2 flash':kind==='bad'?'scan-bad mt-2 flash':'mt-2 field-help');
    e.innerHTML=msg;
  }
  function render(){
    const q=$('workerQueue'), h=$('dnisMasivos'); if(!q||!h)return;
    const arr=[...window.__dniQueueFinal.entries()];
    h.value=arr.map(x=>x[0]).join(',');
    if(!arr.length){q.innerHTML='<div class="text-muted small text-center">Aún no hay trabajadores detectados.</div>';return;}
    q.innerHTML=arr.map(([d,n])=>'<div class="queue-item"><div><b>'+d+'</b><br><span>'+String(n||'TRABAJADOR')+'</span></div><button type="button" class="btn btn-sm btn-outline-danger" onclick="window.__dniQueueFinal.delete(\''+d+'\');window.renderQueueFinal&&window.renderQueueFinal();">×</button></div>').join('');
  }
  window.renderQueueFinal=render;
  async function process(v, force=false){
    const input=$('dniTrab'); if(!input)return;
    const d=dni(v||input.value);
    if(d.length<8){ if(force)st('help','Escanee o digite DNI: al completar 8 dígitos se agregará al pre-registro con sonido.'); return; }
    input.value=d;
    if(!force && d===last)return;
    if(busy){clearTimeout(timer); timer=setTimeout(()=>process(d,true),90); return;}
    busy=true; last=d; st('ok','Buscando DNI <b>'+d+'</b>...');
    try{
      const r=await fetch('/api/trabajador/'+encodeURIComponent(d),{cache:'no-store',credentials:'same-origin'});
      let j={ok:false,msg:'Respuesta inválida'}; try{j=await r.json();}catch(e){}
      if(!j.ok){st('bad','✕ '+(j.msg||'DNI no encontrado en base trabajadores')+' <b>'+d+'</b>'); sound(false); input.select(); return;}
      const t=j.trabajador||{}, nombre=t.trabajador||t.nombres||t.nombre||'TRABAJADOR';
      if(!window.__dniQueueFinal.has(d)){window.__dniQueueFinal.set(d,nombre);}
      if(window.workerMap){ try{window.workerMap.set(d,nombre); window.renderQueue&&window.renderQueue();}catch(e){} }
      render(); st('ok','✓ Reconocido automáticamente: <b>'+nombre+'</b> · '+d); sound(true);
      setTimeout(()=>{input.value=''; last=''; input.focus();},160);
    }catch(e){st('bad','Error consultando trabajador. Revisa conexión/sesión.'); sound(false);}
    finally{busy=false;}
  }
  window.autoDetectarDniInline=function(el){clearTimeout(timer); timer=setTimeout(()=>process((el&&el.value)||($('dniTrab')&&$('dniTrab').value)||'',false),25);};
  document.addEventListener('input',e=>{if(e.target&&e.target.id==='dniTrab')window.autoDetectarDniInline(e.target);},true);
  document.addEventListener('keyup',e=>{if(e.target&&e.target.id==='dniTrab')window.autoDetectarDniInline(e.target);},true);
  document.addEventListener('paste',e=>{if(e.target&&e.target.id==='dniTrab')setTimeout(()=>process(e.target.value,true),60);},true);
  document.addEventListener('keydown',e=>{if(e.target&&e.target.id==='dniTrab'&&(e.key==='Enter'||e.key==='Tab')){process(e.target.value,true); if(e.key==='Enter')e.preventDefault();}},true);
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalRegistro'){setTimeout(()=>{const i=$('dniTrab'); if(i){i.focus(); if(dni(i.value).length>=8)process(i.value,true);}},80);}},true);
  document.addEventListener('submit',e=>{if(e.target&&e.target.id==='frmTrab'){render(); const h=$('dnisMasivos'), i=$('dniTrab'); if(h&&i&&dni(i.value).length===8){const d=dni(i.value); if(!h.value.includes(d)) h.value=(h.value?h.value+',':'')+d;}}},true);
  setInterval(()=>{const i=$('dniTrab'); if(i&&dni(i.value).length>=8)process(i.value,false);},220);

  async function loadQr(){
    if(window.Html5Qrcode)return true;
    return new Promise(res=>{const s=document.createElement('script');s.src='https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js';s.onload=()=>res(!!window.Html5Qrcode);s.onerror=()=>res(false);document.head.appendChild(s);});
  }
  let qr=null;
  window.abrirScanner=async function(readerId,inputId){
    const box=$(readerId), input=$(inputId); if(!box||!input)return;
    box.style.display='block'; box.innerHTML='<button type="button" class="btn btn-sm btn-danger" style="position:absolute;right:8px;top:8px;z-index:9999;border-radius:999px" onclick="cerrarScannerActivo()">×</button><div class="p-2 text-success fw-bold">Abriendo cámara...</div>';
    if(location.protocol!=='https:' && location.hostname!=='localhost' && location.hostname!=='127.0.0.1'){box.innerHTML='<div class="scan-bad">La cámara requiere HTTPS. Abre Render con https:// y permite cámara.</div>'; sound(false); return;}
    if(!navigator.mediaDevices){box.innerHTML='<div class="scan-bad">Este navegador no permite cámara. Usa Chrome actualizado.</div>'; sound(false); return;}
    if(!(await loadQr())){box.innerHTML='<div class="scan-bad">No cargó librería del lector. Revisa internet/CDN.</div>'; sound(false); return;}
    try{ if(qr){await qr.stop().catch(()=>{}); await qr.clear().catch(()=>{});} }catch(e){}
    try{
      qr=new Html5Qrcode(readerId);
      await qr.start({facingMode:{ideal:'environment'}},{fps:12,qrbox:{width:240,height:160}},async decoded=>{
        const d=dni(decoded); input.value=d; input.dispatchEvent(new Event('input',{bubbles:true}));
        if(inputId==='dniTrab') await process(d,true); else sound(true);
        try{await qr.stop(); await qr.clear();}catch(e){} box.style.display='none'; qr=null;
      },()=>{});
    }catch(e){box.innerHTML='<div class="scan-bad">No se pudo activar cámara. Permite cámara en el candado del navegador.</div>'; sound(false);}
  };
})();
</script>

<script>
/* ===== PATCH 248 JS: evitar doble lista nativa y hacer slider PC/táctil real por posición ===== */
(function(){
 const $=id=>document.getElementById(id);
 const pad=n=>String(Number(n)||0).padStart(2,'0');
 function minToTime(m){m=Math.max(0,Math.min(1435,parseInt(m||0,10)));return pad(Math.floor(m/60))+':'+pad(m%60);}
 function toMin(v){let p=String(v||'00:00').split(':'),h=parseInt(p[0]||0,10),m=parseInt(p[1]||0,10);return Math.max(0,Math.min(1435,(isNaN(h)?0:h)*60+(isNaN(m)?0:m)));}
 let active='horaInicioDefault';
 function setActive(id){active=id; ['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'].forEach(x=>{let e=$(x); if(e)e.classList.toggle('border-success',x===id);}); let s=$('timeSlider24'), v=$('touchClockValue'), e=$(id); if(e&&s){s.value=toMin(e.value); if(v)v.textContent=e.value;} let box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(b=>b.classList.toggle('active',b.dataset.target===id));}
 function sync(){let hi=$('horaInicioDefault')?.value||'06:30',hf=$('horaFinDefault')?.value||'16:30',ri=$('refInicioDefault')?.value||'12:00',rf=$('refFinDefault')?.value||'13:00'; [['horaInicioTrab',hi],['horaFinTrab',hf],['refInicioTrab',ri],['refFinTrab',rf]].forEach(([id,val])=>{let e=$(id); if(e)e.value=val;});}
 function apply(v){let e=$(active); if(!e)return; let t=minToTime(v); e.value=t; let d=$('touchClockValue'); if(d)d.textContent=t; sync();}
 function bindSlider(){let s=$('timeSlider24'); if(!s||s.dataset.patch248==='1')return; s.dataset.patch248='1';
   const calc=ev=>{let r=s.getBoundingClientRect(); let x=(ev.touches&&ev.touches[0]?ev.touches[0].clientX:ev.clientX); if(typeof x==='number'&&r.width){let pct=Math.max(0,Math.min(1,(x-r.left)/r.width)); let m=Math.round((pct*1435)/5)*5; s.value=m; apply(m);}else apply(s.value);};
   ['pointerdown','pointermove','mousedown','mousemove','touchstart','touchmove','click','input','change'].forEach(name=>s.addEventListener(name,ev=>{if(name.includes('move') && !(ev.buttons||ev.touches)) return; ev.preventDefault?.(); calc(ev);},{passive:false}));
 }
 function bindHorario(){let pills=$('clockPickFields'); if(pills)[...pills.querySelectorAll('button')].forEach(b=>{b.onclick=e=>{e.preventDefault();setActive(b.dataset.target);}; b.onpointerdown=e=>setActive(b.dataset.target);}); ['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'].forEach(id=>{let e=$(id); if(e){e.readOnly=true;e.onclick=()=>setActive(id);e.onpointerdown=()=>setActive(id);e.onfocus=()=>setActive(id);}}); bindSlider(); setActive(active); sync();}
 function fixModalLabor(){['modalActividad','modalLaborInput','modalConsumidor'].forEach(id=>{let e=$(id); if(e){e.removeAttribute('list'); e.setAttribute('autocomplete','off');}}); ['modal_actividad_list','modal_labor_list','modal_consumidor_list'].forEach(id=>{let d=$(id); if(d)d.innerHTML='';});}
 document.addEventListener('shown.bs.modal',ev=>{if(ev.target&&ev.target.id==='modalHora')setTimeout(bindHorario,80); if(ev.target&&ev.target.id==='modalLabor')setTimeout(fixModalLabor,30);});
 document.addEventListener('DOMContentLoaded',()=>{setTimeout(()=>{bindHorario();fixModalLabor();},120);});
 window.setCampoHorario=setActive; window.aplicarHorarioRegistro=sync;
})();
</script>

<script>
/* PATCH URGENTE: validación avance solo donde existe modalAvance, sin usar variables Jinja fuera de contexto */
(function(){
  const $=id=>document.getElementById(id);
  function onlyDni(v){const d=String(v||'').replace(/\D/g,'');return d.length>=8?d.slice(-8):d;}
  function setSt(ok,msg){const e=$('avanceTrabStatus'); if(!e)return; e.className=ok?'scan-ok mt-2':'scan-bad mt-2'; e.innerHTML=msg;}
  function hojaIdFromForm(){const f=$('frmAvance'); const m=f && String(f.getAttribute('action')||'').match(/\/hoja\/(\d+)\//); return m?m[1]:'';}
  async function validarTrabAvance(){
    const inp=$('dniAvance'), labor=$('avanceLaborId'); if(!inp||!labor)return;
    inp.dataset.trabLaborOk='0'; inp.dataset.trabLaborDni='';
    const dni=onlyDni(inp.value); if(dni.length<8){setSt(false,'Digite o escanee DNI de 8 dígitos.'); return;}
    inp.value=dni; const hoja=hojaIdFromForm();
    if(!hoja || !labor.value){setSt(false,'Primero seleccione una labor.'); return;}
    try{
      const r=await fetch('/api/trabajador-labor/'+encodeURIComponent(hoja)+'/'+encodeURIComponent(labor.value)+'/'+encodeURIComponent(dni),{cache:'no-store',credentials:'same-origin'});
      const j=await r.json();
      if(j.ok){inp.dataset.trabLaborOk='1'; inp.dataset.trabLaborDni=dni; inp.dataset.trabajadorNombre=(j.trabajador&&j.trabajador.trabajador)||''; setSt(true,'✓ Trabajador registrado en esta labor: <b>'+j.trabajador.trabajador+'</b>');}
      else{inp.dataset.trabLaborOk='0'; inp.dataset.trabLaborDni=dni; setSt(false,'✕ '+(j.msg||'Debe registrar primero al trabajador en el módulo Trabajadores de esta labor.'));}
    }catch(e){inp.dataset.trabLaborOk='0'; inp.dataset.trabLaborDni=dni; setSt(false,'No se pudo validar trabajador.');}
  }
  function detectarCantidad(){
    const c=$('codigoAvance'), q=$('cantidadAvance'), box=$('cantidadDetectada'); if(!c||!q)return;
    const txt=String(c.value||''); const m=txt.match(/(\d+(?:[\.,]\d+)?)(?!.*\d)/);
    if(m){const val=m[1].replace(',','.'); q.value=Number(val).toFixed(2); if(box){box.style.display='block'; box.innerHTML='Cantidad detectada: <b>'+q.value+'</b>';}}
  }
  document.addEventListener('input',e=>{if(e.target&&e.target.id==='dniAvance')validarTrabAvance(); if(e.target&&e.target.id==='codigoAvance')detectarCantidad();},true);
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalAvance')setTimeout(validarTrabAvance,80);},true);
})();
</script>


<script>
/* === PATCH 249 FINAL JS: flecha acciones + horas deslizables/digitables === */
(function(){
  const $=id=>document.getElementById(id);
  const ids=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'];
  const pad=n=>String(Number(n)||0).padStart(2,'0');
  function toMin(v){let m=String(v||'00:00').match(/^(\d{1,2}):(\d{2})$/); if(!m)return 0; let h=Math.max(0,Math.min(23,parseInt(m[1]||0,10))); let mi=Math.max(0,Math.min(59,parseInt(m[2]||0,10))); return h*60+mi;}
  function minToTime(m){m=((parseInt(m||0,10)%1440)+1440)%1440; return pad(Math.floor(m/60))+':'+pad(m%60);}
  function normTime(v){let s=String(v||'').replace(/[^0-9:]/g,''); if(/^\d{3,4}$/.test(s)){s=s.padStart(4,'0');s=s.slice(0,2)+':'+s.slice(2);} let m=s.match(/^(\d{1,2}):(\d{1,2})$/); if(!m)return null; let h=Math.min(23,Math.max(0,parseInt(m[1],10)||0)); let mi=Math.min(59,Math.max(0,parseInt(m[2],10)||0)); return pad(h)+':'+pad(mi);}
  function syncPreview(){try{ if(window.aplicarHorarioRegistro) window.aplicarHorarioRegistro(); }catch(e){} }
  window.toggleCardMenu=function(el){document.querySelectorAll('.card-menu.show').forEach(m=>{if(m!==el.nextElementSibling)m.classList.remove('show')}); const m=el.nextElementSibling; if(m)m.classList.toggle('show');};
  document.addEventListener('click',e=>{if(!e.target.closest('.card-menu')&&!e.target.closest('.card-action-chevron'))document.querySelectorAll('.card-menu.show').forEach(m=>m.classList.remove('show'));},true);
  function activate(id){try{ if(window.setCampoHorario) window.setCampoHorario(id); }catch(e){} ids.forEach(x=>{let e=$(x); if(e)e.classList.toggle('border-success',x===id);}); let v=$('touchClockValue'), s=$('timeSlider24'), el=$(id); if(el){if(v)v.textContent=el.value;if(s)s.value=toMin(el.value);} }
  function bindTimeBox(el){ if(!el||el.dataset.drag249==='1')return; el.dataset.drag249='1'; el.readOnly=false; el.classList.add('time-draggable'); el.setAttribute('inputmode','numeric'); el.setAttribute('placeholder','HH:MM');
    let startX=0,startVal=0,dragging=false,moved=false;
    const setVal=(mins)=>{el.value=minToTime(Math.round(mins/5)*5); activate(el.id); syncPreview();};
    el.addEventListener('pointerdown',ev=>{startX=ev.clientX;startVal=toMin(el.value);dragging=true;moved=false;activate(el.id); try{el.setPointerCapture(ev.pointerId)}catch(e){}}, {passive:true});
    el.addEventListener('pointermove',ev=>{if(!dragging)return; const dx=ev.clientX-startX; if(Math.abs(dx)>3)moved=true; setVal(startVal+Math.round(dx/6)*5); ev.preventDefault();}, {passive:false});
    el.addEventListener('pointerup',ev=>{dragging=false; if(moved){try{el.blur()}catch(e){}}}, {passive:true});
    el.addEventListener('click',()=>activate(el.id)); el.addEventListener('focus',()=>activate(el.id));
    el.addEventListener('change',()=>{const t=normTime(el.value); if(t)el.value=t; activate(el.id); syncPreview();});
    el.addEventListener('blur',()=>{const t=normTime(el.value); if(t)el.value=t; syncPreview();});
  }
  function bindAll(){ids.forEach(id=>bindTimeBox($(id))); let h=$('horarioCoherencia'); if(h&&!h.dataset.dragHelp){h.dataset.dragHelp='1';h.insertAdjacentHTML('beforeend','<div class="time-drag-help">Tip: toca un recuadro y arrastra a la derecha/izquierda para ajustar de 5 en 5 minutos. También puedes digitar HH:MM.</div>');}}
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalHora')setTimeout(bindAll,80);});
  document.addEventListener('DOMContentLoaded',()=>setTimeout(bindAll,150));
})();
</script>


<script>
/* === PATCH 250 OMAR FINAL: horario aislado, scanner con X, avance no se preregistra con solo DNI === */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  const pad=n=>String(Number(n)||0).padStart(2,'0');
  const ids=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'];
  function normTime(v){let s=String(v||'').trim().replace(/[^0-9:]/g,''); if(/^\d{3,4}$/.test(s)){s=s.padStart(4,'0');s=s.slice(0,2)+':'+s.slice(2)} let m=s.match(/^(\d{1,2}):(\d{1,2})$/); if(!m)return null; let h=Math.max(0,Math.min(23,parseInt(m[1]||0,10))); let mi=Math.max(0,Math.min(59,parseInt(m[2]||0,10))); return pad(h)+':'+pad(mi);}
  function toMin(v){let t=normTime(v)||'00:00',p=t.split(':');return parseInt(p[0],10)*60+parseInt(p[1],10)}
  function minToTime(m){m=((Math.round(Number(m||0))%1440)+1440)%1440;return pad(Math.floor(m/60))+':'+pad(m%60)}
  function cloneClean(id){const e=$(id); if(!e||!e.parentNode)return null; const c=e.cloneNode(true); e.parentNode.replaceChild(c,e); return c;}
  function installHorarioOmar(){
    const modal=$('modalHora'); if(!modal)return;
    ids.forEach(cloneClean); cloneClean('timeSlider24'); cloneClean('clockPickFields');
    let active='horaInicioDefault';
    const inp=()=>$(active)||$('horaInicioDefault');
    function sync(){
      const hi=$('horaInicioDefault')?.value||'06:30', hf=$('horaFinDefault')?.value||'16:30', ri=$('refInicioDefault')?.value||'12:00', rf=$('refFinDefault')?.value||'13:00';
      [['horaInicioTrab',hi],['horaFinTrab',hf],['refInicioTrab',ri],['refFinTrab',rf]].forEach(([id,v])=>{const e=$(id); if(e)e.value=v;});
      let a=toMin(hi),b=toMin(hf); if(b<=a)b+=1440; let c=toMin(ri),d=toMin(rf); if(d<=c)d+=1440; if(b>1440&&c<a){c+=1440;d+=1440;} const ref=Math.max(0,Math.min(b,d)-Math.max(a,c)); const total=Math.max(0,(b-a)-ref)/60;
      const h=$('horasTrab'); if(h)h.value=total.toFixed(2);
      const txt=$('horarioActivoTxt'); if(txt)txt.innerHTML='<b>Horario activo:</b> '+hi+' - '+hf+' / Refrigerio '+ri+' - '+rf+' / H.Normal '+total.toFixed(2)+'.';
    }
    function paint(){const e=inp(),sl=$('timeSlider24'),tv=$('touchClockValue'); if(e&&sl)sl.value=toMin(e.value); if(e&&tv)tv.textContent=e.value; const box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(b=>b.classList.toggle('active',b.dataset.target===active)); ids.forEach(id=>$(id)?.classList.toggle('border-success',id===active));}
    function setActive(id){if(ids.includes(id))active=id; paint();}
    function setVal(v){const e=inp(); if(!e)return; e.value=minToTime(v); sync(); paint();}
    ids.forEach(id=>{const e=$(id); if(!e)return; e.readOnly=false; e.classList.add('time-draggable'); e.setAttribute('inputmode','numeric');
      ['pointerdown','mousedown','touchstart','click','focus'].forEach(ev=>e.addEventListener(ev,evt=>{setActive(id); evt.stopPropagation();},true));
      e.addEventListener('change',()=>{const t=normTime(e.value); if(t)e.value=t; sync(); paint();}); e.addEventListener('blur',()=>{const t=normTime(e.value); if(t)e.value=t; sync(); paint();});
      let sx=0,sv=0,drag=false; e.addEventListener('pointerdown',ev=>{sx=ev.clientX;sv=toMin(e.value);drag=true;setActive(id);try{e.setPointerCapture(ev.pointerId)}catch(_){}}); e.addEventListener('pointermove',ev=>{if(!drag)return; const dx=ev.clientX-sx; if(Math.abs(dx)>2){setVal(sv+Math.round(dx/6)*5); ev.preventDefault();}}, {passive:false}); e.addEventListener('pointerup',()=>{drag=false;});
    });
    const box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(btn=>{['pointerdown','click','touchstart'].forEach(ev=>btn.addEventListener(ev,e=>{e.preventDefault();e.stopPropagation();setActive(btn.dataset.target);},true));});
    const sl=$('timeSlider24'); if(sl){const calc=ev=>{const r=sl.getBoundingClientRect(); const x=(ev.touches&&ev.touches[0])?ev.touches[0].clientX:ev.clientX; if(typeof x==='number'&&r.width){return Math.round(Math.max(0,Math.min(1,(x-r.left)/r.width))*1435);} return Number(sl.value||0);}; ['pointerdown','pointermove','touchstart','touchmove','click','input','change'].forEach(ev=>sl.addEventListener(ev,e=>{if(ev.includes('move')&&!(e.buttons||e.touches))return; e.preventDefault&&e.preventDefault(); e.stopPropagation&&e.stopPropagation(); setVal(calc(e));},{passive:false,capture:true}));}
    setActive(active); sync();
  }
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalHora')setTimeout(installHorarioOmar,260);},true);
  document.addEventListener('DOMContentLoaded',()=>setTimeout(installHorarioOmar,500));

  window.cerrarScannerActivo=async function(){try{if(window.__qr250){await window.__qr250.stop().catch(()=>{});await window.__qr250.clear().catch(()=>{});}}catch(e){} document.querySelectorAll('[id^="reader"]').forEach(b=>{b.style.display='none';b.innerHTML='';});};
  const oldOpen=window.abrirScanner;
  window.abrirScanner=async function(readerId,inputId){
    const box=$(readerId), input=$(inputId); if(!box||!input)return;
    box.style.display='block'; box.style.position='relative'; box.innerHTML='<button type="button" class="btn btn-sm btn-danger" style="position:absolute;right:8px;top:8px;z-index:9999;border-radius:999px" onclick="cerrarScannerActivo()">×</button><div class="p-2 text-success fw-bold">Abriendo cámara...</div>';
    try{
      if(!window.Html5Qrcode){await new Promise((res,rej)=>{const s=document.createElement('script');s.src='https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js';s.onload=res;s.onerror=rej;document.head.appendChild(s);});}
      await window.cerrarScannerActivo();
      box.style.display='block'; box.style.position='relative'; box.innerHTML='<button type="button" class="btn btn-sm btn-danger" style="position:absolute;right:8px;top:8px;z-index:9999;border-radius:999px">×</button>';
      box.querySelector('button').onclick=window.cerrarScannerActivo;
      window.__qr250=new Html5Qrcode(readerId);
      const cams=await Html5Qrcode.getCameras().catch(()=>[]); const camera=cams&&cams.length?{deviceId:{exact:cams[cams.length-1].id}}:{facingMode:{ideal:'environment'}};
      await window.__qr250.start(camera,{fps:12,qrbox:{width:240,height:160}},async decoded=>{let val=String(decoded||'').trim(); if(inputId==='dniTrab'||inputId==='dniAvance')val=String(val).replace(/\D/g,'').slice(-8); input.value=val; input.dispatchEvent(new Event('input',{bubbles:true})); input.dispatchEvent(new Event('change',{bubbles:true})); await window.cerrarScannerActivo();},()=>{});
    }catch(e){box.innerHTML='<button type="button" class="btn btn-sm btn-danger" style="float:right" onclick="cerrarScannerActivo()">×</button><div class="scan-bad mt-2">No se pudo activar cámara. Permite cámara en el candado del navegador o usa Chrome/Edge con HTTPS.</div>';}
  };

  let avanceAddTimer=null;
  document.addEventListener('input',e=>{
    if(e.target&&e.target.id==='dniAvance')setTimeout(()=>{const c=$('codigoAvance'); if(c&&String(e.target.value||'').replace(/\D/g,'').length>=8)c.focus();},250);
    if(e.target&&e.target.id==='codigoAvance'){clearTimeout(avanceAddTimer); avanceAddTimer=setTimeout(()=>{const ev=new KeyboardEvent('keydown',{key:'Enter',bubbles:true}); $('codigoAvance')?.dispatchEvent(ev);},500);}
  },true);
})();
</script>

<script>
/* === PATCH 251 OMAR: CANTIDAD + PRE-REGISTRO FINAL AVANCE + EDITAR HORARIO === */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  const dni=v=>{const d=String(v||'').replace(/\D/g,'');return d.length>=8?d.slice(-8):d;};
  const num=v=>{let m=String(v||'').replace(',','.').match(/(\d+(?:\.\d+)?)/);return m?Number(m[1]||0):0;};
  const key=()=>{const f=$('frmAvance');return 'avance_pre_'+(f?f.action:'')+'_'+($('avanceLaborId')?.value||'');};
  let avancePre=[];
  function load(){try{avancePre=JSON.parse(sessionStorage.getItem(key())||'[]')||[];}catch(e){avancePre=[];}}
  function save(){try{sessionStorage.setItem(key(),JSON.stringify(avancePre));}catch(e){} const h=$('avancePreJson'); if(h)h.value=JSON.stringify(avancePre);}
  function render(){const q=$('avanceQueue'); if(!q)return; save(); if(!avancePre.length){q.innerHTML='<div class="text-muted small text-center">Aún no hay avances pre-registrados.</div>';return;} q.innerHTML=avancePre.map((x,i)=>'<div class="queue-item"><div><b>'+x.dni+'</b><br><span>Cantidad: '+Number(x.cantidad||0).toFixed(2)+' · '+(x.metodo||'QR/CÓDIGO')+'</span></div><button type="button" class="btn btn-sm btn-outline-danger" data-del-avance="'+i+'">×</button></div>').join(''); q.querySelectorAll('[data-del-avance]').forEach(b=>b.onclick=()=>{avancePre.splice(Number(b.dataset.delAvance),1);render();});}
  function setMsg(ok,msg){const e=$('cantidadDetectada'); if(!e)return; e.style.display='block'; e.className=ok?'scan-ok mt-2':'scan-bad mt-2'; e.innerHTML=msg;}
  function addPre(){const daEl=$('dniAvance'); const d=dni(daEl?.value); const c=Number($('cantidadAvance')?.value||0); const labor=$('avanceLaborId')?.value||''; if(d.length!==8){setMsg(false,'Primero registre/escanee el DNI del trabajador.');return false;} if(!labor){setMsg(false,'Primero seleccione una labor.');return false;} if(!daEl || daEl.dataset.trabLaborOk!=='1' || daEl.dataset.trabLaborDni!==d){setMsg(false,'✕ Debe registrar primero al trabajador en el módulo Trabajadores de esta labor. No se agregó al pre-registro.'); const st=$('avanceTrabStatus'); if(st){st.className='scan-bad mt-2'; st.innerHTML='✕ Debe registrar primero al trabajador en el módulo Trabajadores de esta labor.';} try{if(daEl){daEl.focus(); daEl.select();}}catch(e){} return false;} if(!(c>0)){setMsg(false,'Ingrese una cantidad mayor a cero.');return false;} const met=$('frmAvance')?.querySelector('select[name="metodo"]')?.value||'QR/CÓDIGO'; const exists=avancePre.find(x=>x.dni===d); if(exists){exists.cantidad=c;exists.metodo=met;} else {avancePre.push({dni:d,cantidad:c,a_noct:0,metodo:met});} render(); setMsg(true,'✓ Pre-registro agregado: <b>'+d+'</b> · Cantidad <b>'+c.toFixed(2)+'</b>. Falta GUARDAR AVANCE FINAL.'); try{beep();}catch(e){} const ca=$('codigoAvance'); if(ca)ca.value=''; const da=$('dniAvance'); if(da){da.value=''; da.dataset.trabLaborOk='0'; da.dataset.trabLaborDni=''; da.dispatchEvent(new Event('input',{bubbles:true})); da.dispatchEvent(new Event('change',{bubbles:true})); setTimeout(()=>da.focus(),80);} const st=$('avanceTrabStatus'); if(st){st.className='field-help mt-2'; st.innerHTML='Listo. Escanee o digite el siguiente DNI.';} return true;}
  function installAvance(){const m=$('modalAvance'); if(!m)return; load(); render(); const cod=$('codigoAvance'), cant=$('cantidadAvance'); if(cod&&!cod.dataset.pre251){cod.dataset.pre251='1'; cod.addEventListener('input',()=>{const v=num(cod.value); if(v>0&&cant) {cant.value=v.toFixed(2); setMsg(true,'Cantidad detectada: <b>'+v.toFixed(2)+'</b>');}},true); cod.addEventListener('keydown',e=>{if(e.key==='Enter'||e.key==='Tab'){e.preventDefault(); addPre();}},true); cod.addEventListener('change',()=>{const v=num(cod.value); if(v>0&&cant) cant.value=v.toFixed(2); addPre();},true);} const form=$('frmAvance'); if(form&&!form.dataset.pre251){form.dataset.pre251='1'; form.addEventListener('submit',e=>{if(!avancePre.length){addPre();} if(!avancePre.length){e.preventDefault(); return false;} save(); sessionStorage.removeItem(key());},true);} }
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalAvance')setTimeout(installAvance,80);},true);
  document.addEventListener('DOMContentLoaded',()=>setTimeout(installAvance,300));
  window.abrirEditarAvance=function(id,cantidad){const f=$('frmEditAvance'); if(!f)return; f.action='/lectura/'+id+'/editar'; $('editCantAvance').value=Number(cantidad||0).toFixed(2); if($('editNoctAvance')) $('editNoctAvance').value='0.00'; new bootstrap.Modal($('modalEditAvance')).show();};
})();
</script>
<style>
.avance-grid.cantidad-only{grid-template-columns:1fr 1fr 32px!important}.avance-grid.cantidad-only label{font-size:8px!important;font-weight:900;color:#55745a}.avance-grid.cantidad-only .mini-badge{height:26px!important}.card-menu button{cursor:pointer}.card-menu button:hover{background:#eef8ef!important}.modal-title:has(+ .btn-close){}
</style>

<script>
/* PATCH 253: botón visible para apagar cámara + scanner robusto */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  function closeBtn(){return '<button type="button" class="scanner-close-x" aria-label="Apagar cámara" title="Apagar cámara">×</button>';}
  function ponerX(box){
    if(!box) return;
    box.style.position='relative';
    if(!box.querySelector('.scanner-close-x')) box.insertAdjacentHTML('afterbegin', closeBtn());
    const b=box.querySelector('.scanner-close-x');
    if(b){ b.onclick=function(ev){ev.preventDefault();ev.stopPropagation(); window.cerrarScannerActivo&&window.cerrarScannerActivo();}; }
  }
  async function detener(obj){try{if(obj){await obj.stop?.().catch(()=>{}); await obj.clear?.().catch(()=>{});}}catch(e){}}
  window.cerrarScannerActivo=async function(){
    await detener(window.__scannerActivo253); window.__scannerActivo253=null;
    await detener(window.__qr250); window.__qr250=null;
    document.querySelectorAll('[id^="reader"]').forEach(el=>{el.style.display='none'; el.innerHTML='';});
    try{document.querySelectorAll('video').forEach(v=>{const st=v.srcObject; if(st&&st.getTracks)st.getTracks().forEach(t=>t.stop());});}catch(e){}
  };
  function loadQr(){return new Promise(res=>{if(window.Html5Qrcode)return res(true); const s=document.createElement('script'); s.src='https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js'; s.onload=()=>res(!!window.Html5Qrcode); s.onerror=()=>res(false); document.head.appendChild(s);});}
  window.abrirScanner=async function(readerId,inputId){
    const box=$(readerId), input=$(inputId); if(!box||!input)return;
    await window.cerrarScannerActivo();
    box.style.display='block'; box.innerHTML='<div class="p-2 text-success fw-bold">Abriendo cámara...</div>'; ponerX(box);
    if(location.protocol!=='https:' && location.hostname!=='localhost' && location.hostname!=='127.0.0.1'){
      box.innerHTML='<div class="scan-bad mt-2">La cámara requiere HTTPS. Abre Render con https:// y permite cámara.</div>'; ponerX(box); return;
    }
    if(!(await loadQr())){box.innerHTML='<div class="scan-bad mt-2">No cargó la librería del lector.</div>'; ponerX(box); return;}
    try{
      const scanner=new Html5Qrcode(readerId); window.__scannerActivo253=scanner;
      const cams=await Html5Qrcode.getCameras().catch(()=>[]);
      const camera=(cams&&cams.length)?{deviceId:{exact:cams[cams.length-1].id}}:{facingMode:{ideal:'environment'}};
      await scanner.start(camera,{fps:12,qrbox:{width:260,height:180}},async decoded=>{
        let val=String(decoded||'').trim();
        if(inputId==='dniTrab'||inputId==='dniAvance') val=val.replace(/\D/g,'').slice(-8);
        input.value=val; input.dispatchEvent(new Event('input',{bubbles:true})); input.dispatchEvent(new Event('change',{bubbles:true}));
        await window.cerrarScannerActivo();
      },()=>{});
      ponerX(box);
      const mo=new MutationObserver(()=>ponerX(box)); mo.observe(box,{childList:true,subtree:false}); setTimeout(()=>mo.disconnect(),15000);
    }catch(e){box.innerHTML='<div class="scan-bad mt-2">No se pudo activar cámara. Permite cámara en el candado del navegador.</div>'; ponerX(box);}
  };
})();
</script>


<!-- PATCH FINAL OMAR: Wheel Time Picker táctil estilo iPhone -->
<style>
/* Picker rueda estilo iPhone para campos HH:MM */
.ios-time-wheel-backdrop{position:fixed;inset:0;background:rgba(0,0,0,.30);z-index:200000;display:none;align-items:center;justify-content:center;padding:14px;}
.ios-time-wheel{width:min(330px,92vw);background:rgba(255,255,255,.96);border-radius:18px;box-shadow:0 18px 45px rgba(0,0,0,.28);overflow:hidden;border:1px solid rgba(0,0,0,.08);font-family:Inter,Segoe UI,Arial,sans-serif;}
.ios-wheel-head{height:44px;display:flex;align-items:center;justify-content:space-between;padding:0 12px;border-bottom:1px solid #e5e7eb;font-weight:900;color:#166534;font-size:14px;}
.ios-wheel-btn{border:0;background:transparent;font-weight:900;color:#2f773b;font-size:14px;padding:9px 8px;}
.ios-wheel-cancel{color:#6b7280;}
.ios-wheel-body{position:relative;height:190px;display:grid;grid-template-columns:1fr 1fr;gap:0;background:linear-gradient(#fbfbfb,#fff,#fbfbfb);overflow:hidden;}
.ios-wheel-col{height:190px;overflow-y:auto;scroll-snap-type:y mandatory;-webkit-overflow-scrolling:touch;text-align:center;overscroll-behavior:contain;padding:76px 0;scrollbar-width:none;touch-action:pan-y;}
.ios-wheel-col::-webkit-scrollbar{display:none;}
.ios-wheel-item{height:38px;line-height:38px;font-size:22px;font-weight:800;color:#b9b9b9;scroll-snap-align:center;user-select:none;cursor:pointer;transition:color .12s, font-size .12s, opacity .12s;}
.ios-wheel-item.active{color:#111827;font-size:24px;opacity:1;}
.ios-wheel-center{pointer-events:none;position:absolute;left:14px;right:14px;top:76px;height:38px;border-radius:7px;background:rgba(229,231,235,.82);box-shadow:inset 0 0 0 1px rgba(209,213,219,.55);}
.ios-wheel-fade-top,.ios-wheel-fade-bottom{pointer-events:none;position:absolute;left:0;right:0;height:70px;z-index:2;}
.ios-wheel-fade-top{top:0;background:linear-gradient(to bottom,rgba(255,255,255,.96),rgba(255,255,255,.40),rgba(255,255,255,0));}
.ios-wheel-fade-bottom{bottom:0;background:linear-gradient(to top,rgba(255,255,255,.96),rgba(255,255,255,.40),rgba(255,255,255,0));}
.ios-wheel-sep{pointer-events:none;position:absolute;top:79px;bottom:79px;left:50%;width:1px;background:#e5e7eb;z-index:3;}
input.ios-time-input{cursor:pointer!important;background:#fff!important;caret-color:transparent!important;}
input.ios-time-input:focus{outline:0!important;border-color:#2f773b!important;box-shadow:0 0 0 .25rem rgba(47,119,59,.18)!important;}

/* PATCH 254: minutos 00-59 y horas/títulos centrados */
#modalHora .row.g-2 label.form-label,
#modalEditTareo .row.g-2 label.form-label,
#frmEditTareoStandalone .row.g-2 label.form-label{display:block!important;text-align:center!important;font-weight:900!important;color:#2f773b!important;}
#modalHora .row.g-2 input.form-control,
#modalEditTareo .row.g-2 input.form-control,
#frmEditTareoStandalone .row.g-2 input.form-control,
#horaInicioDefault,#horaFinDefault,#refInicioDefault,#refFinDefault,
#horaInicioTrab,#horaFinTrab,#refInicioTrab,#refFinTrab{text-align:center!important;font-weight:900!important;font-size:16px!important;color:#173322!important;}
.ios-wheel-col{scroll-behavior:smooth;}

</style>

<script>
(function(){
  'use strict';
  const IDS=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault','horaInicioTrab','horaFinTrab','refInicioTrab','refFinTrab','editHi','editHf','editRi','editRf'];
  const $=id=>document.getElementById(id);
  const pad=n=>String(Number(n)||0).padStart(2,'0');
  function parseTime(v){let m=String(v||'00:00').match(/^(\d{1,2}):(\d{1,2})$/);let h=m?parseInt(m[1],10):0, mi=m?parseInt(m[2],10):0;h=Math.max(0,Math.min(23,isNaN(h)?0:h));mi=Math.max(0,Math.min(59,isNaN(mi)?0:mi));return [h, mi];}
  function build(){
    if($('iosTimeWheelBackdrop'))return;
    const html=`<div id="iosTimeWheelBackdrop" class="ios-time-wheel-backdrop">
      <div class="ios-time-wheel" role="dialog" aria-modal="true">
        <div class="ios-wheel-head"><button type="button" id="iosWheelCancel" class="ios-wheel-btn ios-wheel-cancel">Cancelar</button><span>Elegir hora</span><button type="button" id="iosWheelOk" class="ios-wheel-btn">Listo</button></div>
        <div class="ios-wheel-body">
          <div class="ios-wheel-center"></div><div class="ios-wheel-fade-top"></div><div class="ios-wheel-fade-bottom"></div><div class="ios-wheel-sep"></div>
          <div id="iosWheelHours" class="ios-wheel-col" aria-label="Horas"></div>
          <div id="iosWheelMinutes" class="ios-wheel-col" aria-label="Minutos"></div>
        </div>
      </div>
    </div>`;
    document.body.insertAdjacentHTML('beforeend',html);
    const h=$('iosWheelHours'), m=$('iosWheelMinutes');
    h.innerHTML=Array.from({length:24},(_,i)=>`<div class="ios-wheel-item" data-v="${i}">${pad(i)}</div>`).join('');
    m.innerHTML=Array.from({length:60},(_,i)=>`<div class="ios-wheel-item" data-v="${i}">${pad(i)}</div>`).join('');
    h.addEventListener('scroll',()=>requestAnimationFrame(updateActive));
    m.addEventListener('scroll',()=>requestAnimationFrame(updateActive));
    h.addEventListener('click',ev=>{const it=ev.target.closest('.ios-wheel-item'); if(it)scrollToVal(h,Number(it.dataset.v));});
    m.addEventListener('click',ev=>{const it=ev.target.closest('.ios-wheel-item'); if(it)scrollToVal(m,Number(it.dataset.v));});
    $('iosWheelCancel').onclick=close;
    $('iosWheelOk').onclick=apply;
    $('iosTimeWheelBackdrop').addEventListener('click',ev=>{if(ev.target.id==='iosTimeWheelBackdrop')close();});
  }
  let current=null;
  function itemHeight(){return 38;}
  function scrollToVal(col,val){col.scrollTo({top:val*itemHeight(),behavior:'smooth'});setTimeout(updateActive,140);}
  function selected(col){const div=Math.round(col.scrollTop/itemHeight());return col.id==='iosWheelMinutes'?Math.max(0,Math.min(59,div)):Math.max(0,Math.min(23,div));}
  function updateActive(){
    const h=$('iosWheelHours'), m=$('iosWheelMinutes'); if(!h||!m)return;
    const hv=selected(h), mv=selected(m);
    h.querySelectorAll('.ios-wheel-item').forEach((it,i)=>it.classList.toggle('active',i===hv));
    m.querySelectorAll('.ios-wheel-item').forEach((it,i)=>it.classList.toggle('active',i===mv));
    if(current){current.value=pad(hv)+':'+pad(mv); current.dispatchEvent(new Event('input',{bubbles:true})); current.dispatchEvent(new Event('change',{bubbles:true})); syncDisplay();}
  }
  function syncDisplay(){
    try{ if(window.aplicarHorarioRegistro) window.aplicarHorarioRegistro(); }catch(e){}
    try{ if(window.setCampoHorario && current) window.setCampoHorario(current.id); }catch(e){}
    const tv=$('touchClockValue'); if(tv&&current)tv.textContent=current.value;
    const sl=$('timeSlider24'); if(sl&&current){const [h,m]=parseTime(current.value); sl.value=h*60+m;}
  }
  function openFor(input){
    build(); current=input; input.classList.add('ios-time-input');
    try{input.blur();}catch(e){}
    const [h,m]=parseTime(input.value);
    const back=$('iosTimeWheelBackdrop'), hc=$('iosWheelHours'), mc=$('iosWheelMinutes');
    back.style.display='flex';
    hc.scrollTop=h*itemHeight(); mc.scrollTop=m*itemHeight();
    setTimeout(()=>{updateActive();},30);
  }
  function close(){const b=$('iosTimeWheelBackdrop'); if(b)b.style.display='none'; current=null;}
  function apply(){updateActive(); close();}
  function wire(root=document){
    IDS.forEach(id=>{
      const el=$(id); if(!el||el.dataset.iosWheel==='1')return;
      el.dataset.iosWheel='1'; el.classList.add('ios-time-input'); el.setAttribute('autocomplete','off'); el.setAttribute('inputmode','none'); el.readOnly=true;
      ['pointerdown','mousedown','touchstart','focus','click'].forEach(ev=>el.addEventListener(ev,e=>{e.preventDefault();e.stopPropagation();openFor(el);},{capture:true,passive:false}));
    });
  }
  document.addEventListener('DOMContentLoaded',()=>setTimeout(wire,100));
  document.addEventListener('shown.bs.modal',()=>setTimeout(wire,120),true);
  document.addEventListener('click',e=>{const el=e.target&&e.target.closest&&e.target.closest('input'); if(el&&IDS.includes(el.id)){e.preventDefault();e.stopPropagation();openFor(el);}},true);
})();
</script>


<script>
/* PATCH 255: limpiar DNI después del pre-registro + Wheel Time Picker en modificar horario */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  function clearAvanceDni(){
    const d=$('dniAvance');
    if(d && d.value){
      d.value='';
      d.dispatchEvent(new Event('input',{bubbles:true}));
      d.dispatchEvent(new Event('change',{bubbles:true}));
      setTimeout(()=>{try{d.focus({preventScroll:true});}catch(e){d.focus();}},70);
    }
    const st=$('avanceTrabStatus');
    if(st){st.className='field-help mt-2'; st.innerHTML='Listo. Escanee o digite el siguiente DNI.';}
  }
  const obs=new MutationObserver(()=>{
    const msg=$('cantidadDetectada');
    if(msg && /Pre-registro agregado/i.test(msg.textContent||'')) clearAvanceDni();
  });
  document.addEventListener('shown.bs.modal',e=>{
    if(e.target && e.target.id==='modalAvance'){
      const m=$('cantidadDetectada'); if(m) obs.observe(m,{childList:true,subtree:true,characterData:true});
    }
    if(e.target && e.target.id==='modalEditTareo'){
      setTimeout(()=>{
        ['editHi','editHf','editRi','editRf'].forEach(id=>{
          const el=$(id); if(el){el.classList.add('ios-time-input'); el.readOnly=true;}
        });
      },80);
    }
  },true);
})();
</script>

</body></html>
"""

def render_page(body, title="Tareo Móvil", **ctx):
    return render_template_string(BASE_HTML, body=render_template_string(body, **ctx), title=title)


def get_actividades_maestras(limit=50000):
    """Devuelve actividades únicas.
    Antes estaba limitado a 5000, por eso con cargas de 20 mil filas el modal no veía todo.
    Además se usa DISTINCT para que no repita la misma ACTIVIDAD/LABOR/CONSUMIDOR.
    """
    try:
        sql = """SELECT DISTINCT
                    COALESCE(cod_actividad,'') AS cod_actividad,
                    COALESCE(desc_actividad,'') AS desc_actividad,
                    COALESCE(cod_labor,'') AS cod_labor,
                    COALESCE(desc_labor,'') AS desc_labor,
                    COALESCE(cod_consumidor,'') AS cod_consumidor,
                    COALESCE(desc_consumidor,'') AS desc_consumidor
                 FROM actividades_maestras
                 WHERE COALESCE(estado,'ACTIVO')='ACTIVO'
                 ORDER BY desc_actividad, desc_labor, desc_consumidor
                 LIMIT ?"""
        rows = rows_to_dict(execute(sql, (limit,), fetchall=True))
    except Exception:
        rows = []
    return rows

def js_master_options(rows):
    import json
    return json.dumps(rows, ensure_ascii=False)

# ========================= AUTH + HOME =========================

@app.route('/inicio')
def inicio():
    body = """
    <div class="phone-wrap"><div class="splash-card">
      <div class="splash-logo"><i class="bi bi-clipboard2-data"></i></div>
      <div class="splash-title">TAREO MOVIL</div>
      <a class="btn btn-light btn-sm mt-4 fw-bold text-success" href="{{url_for('login')}}">ENTRAR</a>
      <div class="splash-foot">P&A<br>v.1.0</div>
    </div></div>"""
    return render_page(body, title='Tareo Móvil')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario','').strip()
        password = request.form.get('password','')
        row = row_to_dict(execute('SELECT * FROM usuarios WHERE usuario=?', (usuario,), fetchone=True))
        if row and row.get('estado','ACTIVO') == 'ACTIVO' and check_password_hash(row['password_hash'], password):
            if request.form.get('modo') == 'admin' and row.get('rol') != 'admin':
                flash('Este acceso es solo para administradores.', 'danger'); return redirect(url_for('login'))
            session['usuario']=usuario; session['rol']=row.get('rol','operador'); session['nombres']=row.get('nombres') or usuario
            return redirect(url_for('home'))
        flash('Usuario/clave incorrecta o usuario inactivo.', 'danger')
    body = """
    <div class="phone-wrap">
      <div class="green-hero" style="min-height:245px;border-radius:0 0 22px 22px">
        <div class="green-top"><span><i class="bi bi-headset"></i> Soporte</span><span><i class="bi bi-gear"></i> Config.</span></div>
        <div class="splash-logo" style="width:96px;height:96px;font-size:42px;margin:14px auto 6px"><i class="bi bi-clipboard2-data"></i></div>
        <div class="splash-title">TAREO MOVIL</div><div class="login-name">INICIAR SESIÓN</div>
      </div>
      <form method="post" class="login-form">
        <div class="floating-card">
          <div class="role-toggle mb-2"><label><input type="radio" name="modo" value="usuario" checked><span>USUARIO</span></label><label><input type="radio" name="modo" value="admin"><span>ADMINISTRADOR</span></label></div>
          <input class="form-control white-input mb-2" name="usuario" required autofocus placeholder="Usuario">
          <input class="form-control white-input mb-3" name="password" type="password" required placeholder="Clave">
          <button class="btn btn-green w-100"><i class="bi bi-box-arrow-in-right me-1"></i> INGRESAR</button>
          <div class="text-center small mt-2 text-muted">Admin demo: admin / admin123</div>
        </div>
      </form>
      <div class="d-flex justify-content-center gap-3 mt-4">
        <div class="tile"><i class="bi bi-list-check"></i>TAREO</div><div class="tile"><i class="bi bi-file-earmark-bar-graph"></i>REPORTES<br>TAREO</div>
      </div>
      <div class="leaf"></div>
      <div class="bottom-sync"><i class="bi bi-arrow-repeat"></i> Sincronizar Tablas Maestras<br>Actualizado hasta: {{ now }}</div><a class="bottom-out"><i class="bi bi-box-arrow-right"></i></a>
    </div>"""
    return render_page(body, title='Login Tareo Móvil', now=now_str())

@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))

@app.route('/')
def home():
    if not session.get('usuario'):
        return redirect(url_for('inicio'))
    hojas = rows_to_dict(execute('SELECT * FROM hojas_tareo ORDER BY fecha DESC, id DESC LIMIT 8', fetchall=True))
    body = """
    <div class="desktop-grid">
      <div class="phone-wrap">
        <div class="green-hero" style="min-height:220px">
          <div class="green-top"><a class="text-white text-decoration-none" href="{{url_for('soporte')}}"><i class="bi bi-headset"></i> Soporte</a>{% if session.get('rol')=='admin' %}<a class="text-white text-decoration-none" href="{{url_for('configuraciones')}}"><i class="bi bi-gear"></i> Config.</a>{% else %}<span></span>{% endif %}</div>
          <div class="avatar"><i class="bi bi-person-circle"></i></div><div class="login-name">{{ session.get('nombres','USUARIO') }}</div>
          <div class="white-input mt-3"></div>
        </div>
        <div class="top-actions">
          <a class="tile text-decoration-none" href="{{url_for('hojas_tareo')}}"><i class="bi bi-list-check"></i>TAREO</a>
          <a class="tile text-decoration-none" href="{{url_for('asistencia_modulo')}}"><i class="bi bi-fingerprint"></i>ASIST.</a>
          <a class="tile text-decoration-none" href="{{url_for('documentos_firma')}}"><i class="bi bi-pen"></i>FIRMAS</a>
          <a class="tile text-decoration-none" href="{{url_for('transporte')}}"><i class="bi bi-bus-front"></i>TRANSP.</a>
          <a class="tile text-decoration-none" href="{{url_for('reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i>REPORTES<br>TAREO</a>
          <a class="tile text-decoration-none" href="{{url_for('sincronizacion')}}"><i class="bi bi-arrow-repeat"></i>SINC.</a>
        </div>
        <div class="leaf"></div>
        <div class="bottom-sync"><i class="bi bi-arrow-repeat"></i> Sincronizar Tablas Maestras<br>Actualizado hasta: {{ now }}</div><a href="{{url_for('logout')}}" class="bottom-out"><i class="bi bi-box-arrow-right"></i></a>
      </div>
      <div class="desk-panel">
        <h1 class="header-title">TAREO MÓVIL – GRUPO DE COSECHA</h1>
        <div class="card-pro p-4 mb-3"><div class="d-flex justify-content-between align-items-center"><div><h4 class="fw-bold text-success mb-1">Hojas recientes</h4><div class="text-muted small">Crea una hoja y registra labores, trabajadores y avances.</div></div><a class="btn btn-green" href="{{url_for('crear_hoja')}}"><i class="bi bi-plus-lg"></i> Crear hoja</a></div></div>
        <div class="card-pro p-3"><div class="table-responsive"><table class="table list-table"><thead><tr><th>Fecha</th><th>Grupo</th><th>Subgrupo</th><th>Labor</th><th>Responsable</th><th>Estado</th><th></th></tr></thead><tbody>{% for h in hojas %}<tr><td>{{h.fecha}}</td><td>{{h.grupo}}</td><td>{{h.subgrupo}}</td><td>{{h.labor}}</td><td>{{h.responsable}}</td><td><span class="status-pill">{{h.estado}}</span></td><td><a class="btn btn-sm btn-green" href="{{url_for('detalle_hoja', hoja_id=h.id)}}">Abrir</a></td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted py-4">Sin hojas creadas.</td></tr>{% endfor %}</tbody></table></div></div>
      </div>
    </div>"""
    return render_page(body, hojas=hojas, now=now_str())

# ========================= HOJAS TAREO =========================
@app.route('/hojas')
@login_required
def hojas_tareo():
    hojas = rows_to_dict(execute('SELECT * FROM hojas_tareo ORDER BY fecha DESC, id DESC LIMIT 50', fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad tareo-list-page">
      <div class="page-card">
        <div class="green-hero tareo-hero" style="border-radius:0 0 12px 12px">
          <div class="green-top"><span>v.1.0</span><a class="text-white text-decoration-none" href="{{url_for('home')}}"><i class="bi bi-house"></i></a></div>
          <i class="bi bi-list-check" style="font-size:34px;margin-top:14px"></i><div class="login-name mt-1">TAREOS</div>
        </div>
        <div class="toolstrip tareo-toolbar">
          <a class="btn-plus-fab" title="Crear hoja" href="{{url_for('crear_hoja')}}"><i class="bi bi-list-task"></i><i class="bi bi-plus-circle-fill"></i></a>
          <a title="Plantilla Excel" href="{{url_for('plantilla_trabajadores')}}"><i class="bi bi-file-earmark-excel"></i></a>
          <a title="Sincronizar" href="{{url_for('sincronizacion')}}"><i class="bi bi-arrow-clockwise"></i></a>
        </div>
        {% for h in hojas %}
          <div class="swipe-wrap">
            <div class="swipe-actions"><a class="act-edit" href="{{url_for('editar_hoja', hoja_id=h.id)}}">MODIFICAR</a><a class="act-send" href="{{url_for('enviar_hoja', hoja_id=h.id)}}">ENVIAR</a><a class="act-del" href="{{url_for('eliminar_hoja', hoja_id=h.id)}}" onclick="return confirm('¿Eliminar esta hoja?')">ELIMINAR</a></div>
            <a class="text-decoration-none" href="{{url_for('detalle_hoja', hoja_id=h.id, tab='labores')}}"><div class="worker-card">
              <span class="person-dot" style="border-radius:8px"><i class="bi bi-clipboard2-check"></i></span>
              <div class="worker-title"><div>RESPONSABLE<br><b>{{h.responsable}}</b></div><div class="text-end">PRESUPUESTO<br><b>{{h.tipo_tareo or 'JORNAL'}}</b></div></div>
              <div class="worker-grid"><div><label>SUCURSAL</label><div class="small-value">{{h.grupo}}</div></div><div><label>PLANILLA</label><div class="small-value">AGR. PACKING</div></div><div><label>DOCUMENTO</label><div class="small-value">{{h.id}}</div></div></div>
              <div class="small-label mt-2">ZONA CONSUMIDOR</div><div class="small-value">{{h.subgrupo}}</div>
              <div class="worker-grid"><div><label>FECHA</label><div class="small-value">{{h.fecha}}</div></div><div><label>ESTADO</label><div class="mini-badge bg-y">{{h.estado}}</div></div><div class="text-end"><i class="bi bi-chevron-left text-success"></i></div></div>
            </div></a>
          </div>
        {% else %}<div class="worker-card text-center text-muted">No hay hojas. Presiona <b>+</b> para crear.</div>{% endfor %}
        <div class="leaf"></div>
        <div class="bottom-nav"><a href="{{url_for('hojas_tareo')}}"><i class="bi bi-list-check"></i>Listado de Tareos</a><a href="{{url_for('home')}}"><i class="bi bi-file-text"></i>Detalle</a></div>
      </div>
    </div>"""
    return render_page(body, hojas=hojas)

@app.route('/hojas/crear', methods=['GET','POST'])
@login_required
def crear_hoja():
    if request.method == 'POST':
        fecha = request.form.get('fecha') or today_str()
        grupo = limpiar_texto(request.form.get('actividad') or request.form.get('grupo'))
        subgrupo = limpiar_texto(request.form.get('labor') or request.form.get('subgrupo'))
        labor = limpiar_texto(request.form.get('consumidor') or request.form.get('consumidor_desc') or request.form.get('labor_consumidor') or '')
        responsable = limpiar_texto(request.form.get('responsable'))
        turno = limpiar_texto(request.form.get('turno') or 'DIA')
        tipo_tareo = limpiar_texto(request.form.get('tipo_tareo') or 'JORNAL')
        if turno not in ('DIA','NOCHE'): turno = 'DIA'
        if tipo_tareo not in ('JORNAL','RENDIMIENTO'): tipo_tareo = 'JORNAL'
        if not grupo or not subgrupo or not responsable:
            flash('Debe seleccionar Actividad y Labor, además de responsable.', 'danger')
            return redirect(url_for('crear_hoja'))
        execute('INSERT INTO hojas_tareo(fecha,grupo,subgrupo,labor,responsable,turno,tipo_tareo,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?)',
                (fecha,grupo,subgrupo,labor,responsable,turno,tipo_tareo,'ABIERTA',session.get('usuario'),now_str()), commit=True)
        hid = scalar('SELECT MAX(id) AS id FROM hojas_tareo')
        execute('INSERT INTO hoja_labores(hoja_id,grupo,subgrupo,labor,turno,tipo_tareo,responsable,creado_en,creado_por) VALUES(?,?,?,?,?,?,?,?,?)',
                (hid,grupo,subgrupo,labor,turno,tipo_tareo,responsable,now_str(),session.get('usuario')), commit=True)
        return redirect(url_for('hojas_tareo'))
    maestros = get_actividades_maestras()
    body = """
    <div id="createHojaCompact" class="phone-wrap desktop-pad"><h2 class="header-title">CREAR HOJA DE TAREO</h2><div class="page-card">
      <div class="panel-green"><i class="bi bi-clipboard2-plus"></i><h4>NUEVA HOJA – FECHA, ACTIVIDAD, LABOR, CONSUMIDOR, RESPONSABLE, TURNO Y TIPO</h4></div>
      <form method="post" class="floating-card" style="margin:-24px 10px 12px">
        <label class="form-label">FECHA</label><input type="date" name="fecha" class="form-control mb-2 field-required" value="{{today}}" required>
        <label class="form-label">ACTIVIDAD</label><input id="actividadInput" name="actividad" class="form-control mb-2 field-required" list="actividad_list" placeholder="Digite primeras letras de la actividad" required autocomplete="off"><datalist id="actividad_list"></datalist>
        <label class="form-label">LABOR</label><input id="laborInput" name="labor" class="form-control mb-2 field-required" list="labor_list" placeholder="Seleccione labor según actividad" required autocomplete="off"><datalist id="labor_list"></datalist>
        <label class="form-label">CONSUMIDOR <span class="text-muted">(opcional)</span></label><input id="consumidorInput" name="consumidor" class="form-control mb-2" list="consumidor_list" placeholder="Consumidor / zona / campo"><datalist id="consumidor_list"></datalist>
        <label class="form-label">RESPONSABLE</label><input name="responsable" class="form-control mb-2 field-required" placeholder="APELLIDOS Y NOMBRES" required>
        <div class="row g-2 mb-3"><div class="col-6"><label class="form-label">TURNO</label><select name="turno" class="form-select field-required"><option>DIA</option><option>NOCHE</option></select></div><div class="col-6"><label class="form-label">TIPO</label><select name="tipo_tareo" class="form-select field-required"><option>JORNAL</option><option>RENDIMIENTO</option></select></div></div>
        <button class="btn btn-green w-100"><i class="bi bi-check-circle"></i> GUARDAR HOJA</button><a class="btn btn-outline-secondary w-100 mt-2" href="{{url_for('hojas_tareo')}}">VOLVER</a>
      </form></div></div>
      <script>
        const MAESTROS={{ maestros_json|safe }};
        const uniq=a=>[...new Set(a.filter(Boolean))].sort();
        function fillList(id, arr){const dl=document.getElementById(id); dl.innerHTML=''; arr.slice(0,300).forEach(v=>{const o=document.createElement('option'); o.value=v; dl.appendChild(o);});}
        function refreshActividad(){fillList('actividad_list', uniq(MAESTROS.map(x=>x.desc_actividad || x.cod_actividad)));}
        function refreshLabor(){const a=(actividadInput.value||'').toUpperCase(); const rows=MAESTROS.filter(x=>!a || String(x.desc_actividad||'').toUpperCase()===a || String(x.cod_actividad||'').toUpperCase()===a || String(x.desc_actividad||'').toUpperCase().includes(a)); fillList('labor_list', uniq(rows.map(x=>x.desc_labor || x.cod_labor))); refreshConsumidor();}
        function refreshConsumidor(){const a=(actividadInput.value||'').toUpperCase(), l=(laborInput.value||'').toUpperCase(); const rows=MAESTROS.filter(x=>(!a || String(x.desc_actividad||'').toUpperCase().includes(a)||String(x.cod_actividad||'').toUpperCase().includes(a)) && (!l || String(x.desc_labor||'').toUpperCase().includes(l)||String(x.cod_labor||'').toUpperCase().includes(l))); fillList('consumidor_list', uniq(rows.map(x=>x.desc_consumidor || x.cod_consumidor)));}
        actividadInput.addEventListener('input', refreshLabor); laborInput.addEventListener('input', refreshConsumidor); document.addEventListener('DOMContentLoaded',()=>{refreshActividad();refreshLabor();});
      </script>

    """
    return render_page(body, today=today_str(), maestros_json=js_master_options(maestros))

@app.route('/hoja/<int:hoja_id>')
@login_required
def detalle_hoja(hoja_id):
    tab = request.args.get('tab','labores')
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger'); return redirect(url_for('hojas_tareo'))
    labores = rows_to_dict(execute('SELECT * FROM hoja_labores WHERE hoja_id=? ORDER BY id DESC', (hoja_id,), fetchall=True))
    selected_labor_id = limpiar_texto(request.args.get('labor_id') or '', upper=False)
    selected_labor = None
    if selected_labor_id:
        selected_labor = row_to_dict(execute('SELECT * FROM hoja_labores WHERE id=? AND hoja_id=?', (selected_labor_id, hoja_id), fetchone=True))
        if not selected_labor:
            selected_labor_id = ''
    if tab in ('trabajadores','rendimiento') and not selected_labor_id:
        flash('Primero debes elegir una labor: toca/clic en una tarjeta de labor para entrar a Trabajadores o Rend./Avance.', 'danger')
        tab = 'labores'
    if selected_labor_id:
        tareos = rows_to_dict(execute('SELECT * FROM tareos WHERE hoja_id=? AND labor_id=? ORDER BY creado_en DESC LIMIT 100', (hoja_id, selected_labor_id), fetchall=True))
        lecturas = rows_to_dict(execute('SELECT * FROM lecturas_balde WHERE hoja_id=? AND labor_id=? ORDER BY fecha_hora DESC LIMIT 100', (hoja_id, selected_labor_id), fetchall=True))
    else:
        tareos = rows_to_dict(execute('SELECT * FROM tareos WHERE hoja_id=? ORDER BY creado_en DESC LIMIT 100', (hoja_id,), fetchall=True))
        lecturas = rows_to_dict(execute('SELECT * FROM lecturas_balde WHERE hoja_id=? ORDER BY fecha_hora DESC LIMIT 100', (hoja_id,), fetchall=True))
    registros = len(tareos); horas_total = sum(float(x.get('horas') or 0) for x in tareos); rend_total = sum(float(x.get('cantidad') or 0) for x in tareos)
    execute('UPDATE hojas_tareo SET registros=?, horas_total=?, rendimiento_total=? WHERE id=?', (registros, horas_total, rend_total, hoja_id), commit=True)
    body = """
    <div class="phone-wrap desktop-pad"><h2 class="header-title">TAREO MÓVIL – {{ 'DETALLE DE TRABAJADOR POR LABOR' if tab=='trabajadores' else ('DETALLE NÚMERO DE LECTURAS POR BALDE' if tab=='rendimiento' else 'GRUPO DE COSECHA') }}</h2>
      <div class="page-card">
        <div class="tab-main"><a class="{{'active' if tab=='labores' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}">LABORES</a><a class="{{'active' if tab=='trabajadores' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='trabajadores', labor_id=selected_labor_id) if selected_labor_id else url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}" onclick="{% if not selected_labor_id %}alert('Primero toca/clic en una labor.');{% endif %}">TRABAJADORES</a><a class="{{'active' if tab=='rendimiento' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='rendimiento', labor_id=selected_labor_id) if selected_labor_id else url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}" onclick="{% if not selected_labor_id %}alert('Primero toca/clic en una labor.');{% endif %}">REND./AVANCE</a></div>
        <div class="subtabs"><a class="{{'active' if tab=='labores' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}">Labores</a><a class="{{'active' if tab=='trabajadores' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='trabajadores', labor_id=selected_labor_id) if selected_labor_id else url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}" onclick="{% if not selected_labor_id %}alert('Primero toca/clic en una labor.');{% endif %}">Trab.por Labor</a><a class="{{'active' if tab=='rendimiento' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='rendimiento', labor_id=selected_labor_id) if selected_labor_id else url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}" onclick="{% if not selected_labor_id %}alert('Primero toca/clic en una labor.');{% endif %}">Rend/Avance por Labor</a></div>
        <div class="panel-green"><i class="bi {{ 'bi-people-fill' if tab=='trabajadores' else ('bi-person-badge' if tab=='rendimiento' else 'bi-people') }}"></i><h4>{{ 'TRABAJADORES – QR / CÓDIGO BARRAS / DIGITACIÓN' if tab=='trabajadores' else ('PRODUCTIVIDAD POR TRABAJADOR – BALDE / QR / DIGITACIÓN' if tab=='rendimiento' else 'REGISTRO DE ACTIVIDAD, LABOR, CONSUMIDOR, TURNO Y TIPO') }}</h4></div>
        <div class="toolstrip">
          <button title="Crear labor/grupo/subgrupo" data-bs-toggle="modal" data-bs-target="#modalLabor"><i class="bi bi-list-check"></i></button>
          <button title="Buscar trabajador" data-bs-toggle="modal" data-bs-target="#modalBuscar"><i class="bi bi-search"></i></button>
          {% if tab=='trabajadores' %}
            <button title="Elegir horarios" data-bs-toggle="modal" data-bs-target="#modalHora"><i class="bi bi-clock"></i></button>
            <button title="Entrada / salida" data-bs-toggle="modal" data-bs-target="#modalHora"><i class="bi bi-box-arrow-in-right"></i></button>
            <button title="Registrar trabajador" data-bs-toggle="modal" data-bs-target="#modalRegistro"><i class="bi bi-person-plus"></i></button>
          {% elif tab=='rendimiento' %}
            <button title="Registrar avance" data-bs-toggle="modal" data-bs-target="#modalAvance"><i class="bi bi-upc-scan"></i></button>
            <button title="Refrescar" onclick="location.reload()"><i class="bi bi-arrow-clockwise"></i></button>
          {% else %}
            <button title="Copiar labor" data-bs-toggle="modal" data-bs-target="#modalCopiar"><i class="bi bi-files"></i></button>
            <button title="Refrescar" onclick="location.reload()"><i class="bi bi-arrow-clockwise"></i></button>
          {% endif %}
          <a href="{{url_for('hojas_tareo')}}" title="Volver"><i class="bi bi-chevron-left"></i></a>
        </div>
        <div class="info-bar"><div><i class="bi bi-calendar"></i> {{h.fecha}}</div><div><i class="bi bi-list"></i> {{registros}} Reg.</div><div><i class="bi bi-clock"></i> {{'%.2f'|format(horas_total)}} H.</div><div>A.Rend {{'%.2f'|format(rend_total)}}</div><span>⌄</span></div>
        {% if tab=='labores' %}
          {% for l in labores %}<a class="text-decoration-none" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='trabajadores', labor_id=l.id)}}"><div class="worker-card labor-card-compact"><div class="worker-title"><div>ACTIVIDAD<br><b>{{l.grupo}}</b></div><div class="text-end">LABOR<br><b>{{l.subgrupo or 'SIN LABOR'}}</b></div></div><div class="mt-2"><span class="small-label">CONSUMIDOR</span> <b class="labor-main">{{l.labor or 'SIN CONSUMIDOR'}}</b><br><span class="small-label">RESPONSABLE</span> <b class="resp-main">{{l.responsable or h.responsable}}</b></div><div class="worker-grid mt-2"><div><div class="mini-badge {{'bg-y' if l.turno=='NOCHE' else 'bg-g'}}">{{l.turno}}</div></div><div><div class="mini-badge bg-y">{{l.tipo_tareo}}</div></div><div><div class="mini-badge bg-g">ACTIVA</div></div></div></div></a>{% else %}<div class="worker-card text-center text-muted">Presiona <b>+</b> para crear actividad, labor y consumidor.</div>{% endfor %}
        {% elif tab=='trabajadores' %}
          {% for r in tareos %}<div class="worker-card trabajador-card-ref {{'editable-tareo' if h.estado!='ENVIADA' else ''}}"><div class="worker-title"><div>TRABAJADOR<br><b>{{r.trabajador}}</b></div><div>NRO.DOCUMENTO<br><b>{{r.dni}}</b></div></div><div class="trabajador-grid-ref"><div><label>H.INICIO</label><div class="time-box">{{r.hora_inicio or ('22:00' if r.turno=='NOCHE' else '06:30')}}</div></div><div><label>H.FIN</label><div class="time-box">{{r.hora_fin or ('06:00' if r.turno=='NOCHE' else '16:30')}}</div></div><div><label>H.NORMAL</label><div class="metric-box">{{'%.2f'|format((r.horas or 0) - (r.horas_nocturnas or 0))}}</div></div><div><label>REF.INI</label><div class="time-box">{{r.ref_inicio or '12:00'}}</div></div><div><label>REF.FIN</label><div class="time-box">{{r.ref_fin or '13:00'}}</div></div><div><label>H.NOCTURNO</label><div class="metric-box">{{'%.2f'|format(r.horas_nocturnas or 0)}}</div></div><div><label>ESTADO</label><div class="mini-badge bg-g">FIN TOTAL</div></div></div>{% if h.estado!='ENVIADA' %}<div class="card-action-chevron" onclick="event.stopPropagation();toggleCardMenu(this)" title="Acciones"></div><div class="card-menu" onclick="event.stopPropagation()"><a href="{{url_for('editar_horas_tareo_form', tareo_id=r.id)}}" class="btn-edit-tareo" onclick="abrirEditarTareo('{{r.id}}','{{r.hora_inicio or ('22:00' if r.turno=='NOCHE' else '06:30')}}','{{r.hora_fin or ('06:00' if r.turno=='NOCHE' else '16:30')}}','{{r.ref_inicio or '12:00'}}','{{r.ref_fin or '13:00'}}'); return false;" data-id="{{r.id}}" data-hi="{{r.hora_inicio or ('22:00' if r.turno=='NOCHE' else '06:30')}}" data-hf="{{r.hora_fin or ('06:00' if r.turno=='NOCHE' else '16:30')}}" data-ri="{{r.ref_inicio or '12:00'}}" data-rf="{{r.ref_fin or '13:00'}}">Modificar</a><a class="danger" href="{{url_for('eliminar_tareo', tareo_id=r.id)}}" onclick="return confirm('¿Eliminar trabajador del tareo?')">Eliminar</a></div>{% endif %}</div>{% else %}<div class="worker-card text-center text-muted">Presiona el <b>hombresito +</b> para registrar trabajador por QR/código/digitación.</div>{% endfor %}
        {% else %}
          {% for l in lecturas %}<div class="worker-card avance-card-ref"><span class="person-dot"><i class="bi bi-person-circle"></i></span><div class="worker-title"><div>TRABAJADOR<br><b>{{l.trabajador}}</b></div><div>NRO.DOC.<br><b>{{l.dni}}</b></div></div><div class="small-label mt-1">HORA TOMA REGISTRO</div><div class="small-value">{{l.fecha_hora}} · {{l.metodo or 'DIGITACIÓN'}}</div><div class="worker-grid avance-grid cantidad-only"><div><label>CANTIDAD</label><div class="mini-badge bg-y">{{'%.2f'|format(l.a_diurno or 0)}}</div></div><div><label>UNIDAD</label><div class="mini-badge bg-y">BALDE</div></div><div></div></div>{% if h.estado!='ENVIADA' %}<div class="card-action-chevron" onclick="event.stopPropagation();toggleCardMenu(this)" title="Acciones"></div><div class="card-menu" onclick="event.stopPropagation()"><button type="button" onclick="abrirEditarAvance('{{l.id}}','{{l.a_diurno or 0}}')">Modificar</button><a class="danger" href="{{url_for('eliminar_lectura', lectura_id=l.id)}}" onclick="return confirm('¿Eliminar avance?')">Eliminar</a></div>{% else %}<i class="bi bi-lock text-muted"></i>{% endif %}</div>{% else %}<div class="worker-card text-center text-muted">Presiona el icono de escaneo para registrar avance por QR/código/digitación.</div>{% endfor %}
        {% endif %}<div class="leaf"></div>
      </div>
    </div>

    <div class="modal fade" id="modalLabor" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('guardar_labor_hoja', hoja_id=h.id, tab=tab)}}"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-plus-square"></i> Crear nueva labor</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-light border small mb-2">Complete los datos y presione <b>CREAR LABOR</b>. Al cerrar con X no se guarda nada.</div><label class="form-label">ACTIVIDAD</label><input id="modalActividad" name="grupo" class="form-control mb-1" placeholder="Digite primeras letras de la actividad" required autocomplete="off"><datalist id="modal_actividad_list"></datalist><div id="modalMasterStatus" class="master-status">CARGANDO ACTIVIDADES...</div><div id="modalActividadSuggest" class="modal-suggest"></div><label class="form-label">LABOR</label><input id="modalLaborInput" name="subgrupo" class="form-control mb-1" placeholder="Seleccione labor según actividad" required autocomplete="off"><datalist id="modal_labor_list"></datalist><div id="modalLaborSuggest" class="modal-suggest"></div><label class="form-label">CONSUMIDOR (opcional)</label><input id="modalConsumidor" name="labor" class="form-control mb-1" placeholder="Consumidor / zona / campo"><datalist id="modal_consumidor_list"></datalist><div id="modalConsumidorSuggest" class="modal-suggest"></div><label class="form-label">RESPONSABLE</label><input name="responsable" class="form-control mb-2" placeholder="APELLIDOS Y NOMBRES" value="{{h.responsable}}"><div class="row g-2"><div class="col-6"><label class="form-label">TURNO</label><select name="turno" class="form-select"><option>DIA</option><option>NOCHE</option></select></div><div class="col-6"><label class="form-label">TIPO</label><select name="tipo_tareo" class="form-select"><option>JORNAL</option><option>RENDIMIENTO</option></select></div></div></div><div class="modal-footer"><button class="btn btn-green w-100" type="submit">CREAR LABOR</button></div></form></div></div></div>
    <div class="modal fade" id="modalCopiar" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('copiar_labor_hoja', hoja_id=h.id, tab=tab)}}"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-files"></i> Copiar labor existente</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-light border small">Selecciona el documento/labor que deseas copiar. No se copiará nada hasta presionar <b>COPIAR SELECCIONADO</b>.</div><div class="copy-list">{% for l in labores %}<label class="d-block mb-2"><input type="radio" name="labor_id_origen" value="{{l.id}}" required> <b>{{l.labor}}</b><br><span class="small text-muted">{{l.grupo}} / {{l.subgrupo}} / {{l.turno}} / {{l.tipo_tareo}}</span></label>{% endfor %}</div><label class="form-label mt-2">Nuevo nombre de labor (opcional)</label><input name="labor_nueva" class="form-control" placeholder="Dejar vacío para copiar igual"></div><div class="modal-footer"><button class="btn btn-green w-100">COPIAR SELECCIONADO</button></div></form></div></div></div>
    <div class="modal fade" id="modalBuscar" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-search"></i> Buscar trabajador</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><input id="buscarDni" class="form-control mb-2" placeholder="DNI / QR / código barras"><button class="btn btn-green w-100" onclick="buscarTrabajadorLibre()">BUSCAR</button><div id="buscarResultado" class="alert alert-light border mt-2">Esperando búsqueda.</div></div></div></div></div>
    <div class="modal fade" id="modalHora" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('fijar_horario_hoja', hoja_id=h.id, tab=tab)}}"><input type="hidden" name="labor_id" value="{{selected_labor_id}}"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-clock"></i> Fijar horario obligatorio</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-warning small"><b>Obligatorio:</b> fija el horario antes de tarear trabajadores. Formato 24 horas (ej. 06:30, 16:30, 22:00, 06:00).</div><div class="touch-clock-panel"><div class="clock-24-hint">Toque un campo, mueva la manija o digite la hora en formato 24:00.</div><div id="touchClockValue" class="time-display">06:30</div><input id="timeSlider24" class="time-slider" type="range" min="0" max="1435" step="5" value="390"><div id="clockPickFields" class="touch-clock-picks"><button type="button" data-target="horaInicioDefault" class="active">Inicio trabajo</button><button type="button" data-target="horaFinDefault">Fin trabajo</button><button type="button" data-target="refInicioDefault">Inicio refrigerio</button><button type="button" data-target="refFinDefault">Fin refrigerio</button></div></div><div class="row g-2"><div class="col-6"><label class="form-label">Inicio trabajo</label><input name="hora_inicio_default" id="horaInicioDefault" type="text" class="form-control locked-input" value="{{h.hora_inicio_default or '06:30'}}" required></div><div class="col-6"><label class="form-label">Fin trabajo</label><input name="hora_fin_default" id="horaFinDefault" type="text" class="form-control locked-input" value="{{h.hora_fin_default or '16:30'}}" required></div><div class="col-6"><label class="form-label">Inicio refrigerio</label><input name="ref_inicio_default" id="refInicioDefault" type="text" class="form-control locked-input" value="{{h.ref_inicio_default or '12:00'}}" required></div><div class="col-6"><label class="form-label">Fin refrigerio</label><input name="ref_fin_default" id="refFinDefault" type="text" class="form-control locked-input" value="{{h.ref_fin_default or '13:00'}}" required></div></div><div id="horarioCoherencia" class="field-help mt-2">El refrigerio debe quedar dentro de la jornada.</div><button class="btn btn-green w-100 mt-3" type="submit">FIJAR HORARIO</button></div></form></div></div></div>
    <div class="modal fade" id="modalEditTareo" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form id="frmEditTareo" method="post"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-pencil-square"></i> Editar horas del trabajador</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-light border small">Solo se puede editar si la hoja aún no fue enviada.</div><div class="row g-2"><div class="col-6"><label class="form-label">Hora inicio</label><input id="editHi" name="hora_inicio" class="form-control" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div><div class="col-6"><label class="form-label">Hora fin</label><input id="editHf" name="hora_fin" class="form-control" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div><div class="col-6"><label class="form-label">Ref. ini</label><input id="editRi" name="ref_inicio" class="form-control" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div><div class="col-6"><label class="form-label">Ref. fin</label><input id="editRf" name="ref_fin" class="form-control" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div></div></div><div class="modal-footer"><button class="btn btn-green w-100">GUARDAR CAMBIOS</button></div></form></div></div></div>

    <div class="modal fade" id="modalRegistro" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('guardar_registro_hoja', hoja_id=h.id, tab='trabajadores')}}" id="frmTrab"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-person-plus"></i> Registrar trabajador</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="scan-box mb-2"><label class="form-label">DNI / QR / CÓDIGO BARRAS</label><div class="input-group"><input name="dni" id="dniTrab" class="form-control" placeholder="Escanee o digite DNI" autocomplete="off" inputmode="numeric" maxlength="30" oninput="autoDetectarDniInline(this)" onkeyup="autoDetectarDniInline(this)" onchange="autoDetectarDniInline(this)"><button type="button" class="btn btn-green" onclick="abrirScanner('readerTrab','dniTrab')"><i class="bi bi-upc-scan"></i></button></div><div id="readerTrab" style="display:none;margin-top:8px"></div><div id="dniStatus" class="mt-2 field-help">Escanee o digite DNI: al completar 8 dígitos se agregará al pre-registro con sonido.</div><input type="hidden" name="dnis_masivos" id="dnisMasivos"><div class="queue-title">PRE-REGISTRO DE TRABAJADORES</div><div id="workerQueue" class="worker-queue"><div class="text-muted small text-center">Aún no hay trabajadores detectados.</div></div></div><label class="form-label">LABOR SELECCIONADA</label><input type="hidden" name="labor_id" value="{{selected_labor_id}}"><div class="form-control mb-2" style="height:auto;min-height:37px;background:#f8fff9;color:#166534;font-weight:900">{% if selected_labor %}{{selected_labor.grupo}} / {{selected_labor.subgrupo}} / {{selected_labor.labor}} / {{selected_labor.turno}} / {{selected_labor.tipo_tareo}}{% else %}PRIMERO SELECCIONA UNA LABOR{% endif %}</div><input name="turno" id="turnoTrab" type="hidden" value="DIA"><input name="tipo_tareo" type="hidden" value="JORNAL"><input name="hora_inicio" id="horaInicioTrab" type="hidden" value="{{h.hora_inicio_default or '06:30'}}"><input name="hora_fin" id="horaFinTrab" type="hidden" value="{{h.hora_fin_default or '16:30'}}"><input name="ref_inicio" id="refInicioTrab" type="hidden" value="{{h.ref_inicio_default or '12:00'}}"><input name="ref_fin" id="refFinTrab" type="hidden" value="{{h.ref_fin_default or '13:00'}}"><input name="horas" id="horasTrab" type="hidden" value="0"><input name="cantidad" type="hidden" value="0.00"><div id="horarioActivoTxt" class="alert {{'alert-success' if h.horario_fijado else 'alert-warning'}} small mt-2 mb-0"><b>Horario activo:</b> {{h.hora_inicio_default or 'NO FIJADO'}} - {{h.hora_fin_default or 'NO FIJADO'}} / Refrigerio {{h.ref_inicio_default or '--:--'}} - {{h.ref_fin_default or '--:--'}}. {% if not h.horario_fijado %}<b>Primero fija el horario desde el icono de reloj.</b>{% endif %}</div></div><div class="modal-footer"><button class="btn btn-green w-100">GUARDAR TRABAJADORES</button></div></form></div></div></div>
    <div class="modal fade" id="modalAvance" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('guardar_registro_hoja', hoja_id=h.id, tab='rendimiento')}}" id="frmAvance"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-upc-scan"></i> Registrar avance / lectura</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="scan-box mb-2"><label class="form-label">DNI / QR / CÓDIGO BARRAS DEL TRABAJADOR</label><div class="input-group"><input name="dni" id="dniAvance" class="form-control" placeholder="Escanee o digite DNI" required autocomplete="off" inputmode="numeric"><button type="button" class="btn btn-green" onclick="abrirScanner('readerAvance','dniAvance')"><i class="bi bi-upc-scan"></i></button></div><div id="avanceTrabStatus" class="field-help mt-2">Debe estar registrado en Trabajadores de esta labor.</div><div id="readerAvance" style="display:none;margin-top:8px"></div><label class="form-label mt-2">LECTOR AVANCE / QR / CÓDIGO BARRAS</label><div class="input-group"><input name="codigo_avance" id="codigoAvance" class="form-control" placeholder="OPCIONAL: CÓDIGO DE AVANCE / BALDE / ETIQUETA"><button type="button" class="btn btn-green" onclick="abrirScanner('readerCodigoAvance','codigoAvance')"><i class="bi bi-upc-scan"></i></button></div><div id="readerCodigoAvance" style="display:none;margin-top:8px"></div><div id="cantidadDetectada" class="scan-ok mt-2" style="display:none">Cantidad detectada: <b>1.00</b></div></div><label class="form-label">LABOR SELECCIONADA</label><input type="hidden" name="labor_id" value="{{selected_labor_id}}" id="avanceLaborId"><div class="form-control mb-2" style="height:auto;min-height:37px;background:#f8fff9;color:#166534;font-weight:900">{% if selected_labor %}{{selected_labor.labor}} / {{selected_labor.turno}} / {{selected_labor.tipo_tareo}}{% else %}PRIMERO SELECCIONA UNA LABOR{% endif %}</div><input type="hidden" name="avance_pre_json" id="avancePreJson"><div class="row g-2"><div class="col-12"><label class="form-label">CANTIDAD</label><input name="cantidad" id="cantidadAvance" type="number" step="0.01" class="form-control" value="1.00"></div><input name="a_noct" id="aNoctAvance" type="hidden" value="0.00"><div class="col-6"><label class="form-label">UNIDAD</label><select name="unidad" class="form-select"><option>BALDE</option><option>KG</option><option>JABA</option><option>UNIDAD</option></select></div><div class="col-6"><label class="form-label">MÉTODO</label><select name="metodo" class="form-select"><option>QR/CÓDIGO</option><option>DIGITACIÓN</option><option>LECTOR USB</option></select></div></div><div class="queue-title">PRE-REGISTRO DE AVANCES</div><div id="avanceQueue" class="worker-queue"><div class="text-muted small text-center">Aún no hay avances pre-registrados.</div></div></div><div class="modal-footer"><button class="btn btn-green w-100" id="btnGuardarAvance">GUARDAR AVANCE FINAL</button></div></form></div></div></div>
    <div class="modal fade" id="modalEditAvance" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form id="frmEditAvance" method="post"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-pencil-square"></i> Editar avance</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-light border small">Solo se puede editar si la hoja aún no fue enviada.</div><div class="row g-2"><div class="col-12"><label class="form-label">CANTIDAD</label><input id="editCantAvance" name="cantidad" type="number" step="0.01" class="form-control" required></div><input id="editNoctAvance" name="a_noct" type="hidden" value="0.00"></div></div><div class="modal-footer"><button class="btn btn-green w-100">GUARDAR CAMBIOS</button></div></form></div></div></div>

    <script>
(function(){
  'use strict';
  const MAESTROS_DET={{ maestros_json|safe }};
  const $=(id)=>document.getElementById(id);
  const pad=(n)=>String(Number(n)||0).padStart(2,'0');
  const sleep=(ms)=>new Promise(r=>setTimeout(r,ms));
  function playOk(){try{beep();}catch(e){}}
  function onlyDni(v){
    const raw=String(v||'');
    const m=raw.match(/(?:^|\D)(\d{8})(?:\D|$)/);
    const d=m?m[1]:raw.replace(/\D/g,'');
    return d.length>=8?d.slice(-8):d;
  }
  window.limpiarDni=onlyDni;

  // ================== BUSCADOR Y QR/CÓDIGO ==================
  window.buscarTrabajadorLibre=async function(){
    const inp=$('buscarDni'), box=$('buscarResultado');
    const dni=onlyDni(inp?inp.value:'');
    if(inp)inp.value=dni;
    if(!box)return;
    if(dni.length!==8){box.className='alert alert-warning mt-2';box.textContent='Ingrese DNI válido de 8 dígitos.';return;}
    box.className='alert alert-light border mt-2'; box.innerHTML='Buscando <b>'+dni+'</b>...';
    try{
      const r=await fetch('/api/trabajador/'+encodeURIComponent(dni),{cache:'no-store',credentials:'same-origin'});
      const j=await r.json();
      if(!j.ok){box.className='alert alert-danger mt-2'; box.textContent=j.msg||'DNI no encontrado.'; return;}
      const t=j.trabajador||{};
      box.className='alert alert-success mt-2';
      box.innerHTML='<b>'+(t.trabajador||'TRABAJADOR')+'</b><br>'+dni+' · '+(t.cargo||'')+' · '+(t.area||'');
      playOk();
    }catch(e){box.className='alert alert-danger mt-2';box.textContent='No se pudo consultar la base de trabajadores.';}
  };

  let scanner=null;
  function playBad(){try{playOk();}catch(e){}}
  function loadQrLib(){
    return new Promise(resolve=>{
      if(window.Html5Qrcode) return resolve(true);
      const old=document.querySelector('script[data-qr-loader="1"]');
      if(old){old.addEventListener('load',()=>resolve(true)); old.addEventListener('error',()=>resolve(false)); return;}
      const sc=document.createElement('script');
      sc.dataset.qrLoader='1';
      sc.src='https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js';
      sc.onload=()=>resolve(!!window.Html5Qrcode);
      sc.onerror=()=>resolve(false);
      document.head.appendChild(sc);
    });
  }
  window.cerrarScannerActivo=async function(){
    try{ if(scanner){ await scanner.stop().catch(()=>{}); await scanner.clear().catch(()=>{}); scanner=null; } }catch(e){}
    try{ if(window.__qr250){ await window.__qr250.stop().catch(()=>{}); await window.__qr250.clear().catch(()=>{}); window.__qr250=null; } }catch(e){}
    document.querySelectorAll('[id^="reader"]').forEach(el=>{el.style.display='none'; el.innerHTML='';});
  };
  function scannerCloseHtml(){
    return '<button type="button" class="scanner-close-x" title="Apagar cámara" onclick="cerrarScannerActivo()">×</button>';
  }
  window.abrirScanner=async function(readerId,inputId){
    const el=$(readerId), input=$(inputId); if(!el||!input)return;
    await window.cerrarScannerActivo();
    el.style.display='block';
    el.style.position='relative';
    el.innerHTML=scannerCloseHtml()+'<div class="p-2 text-success fw-bold">Abriendo cámara...</div>';
    if(location.protocol!=='https:' && location.hostname!=='localhost' && location.hostname!=='127.0.0.1'){
      el.innerHTML=scannerCloseHtml()+'<div class="scan-bad mt-2">La cámara requiere HTTPS. Abre la app desde Render con https://</div>';
      playBad(); return;
    }
    if(!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia){
      el.innerHTML=scannerCloseHtml()+'<div class="scan-bad mt-2">Este navegador no permite cámara. Usa Chrome actualizado.</div>';
      playBad(); return;
    }
    const libOk=await loadQrLib();
    if(!libOk){
      el.innerHTML=scannerCloseHtml()+'<div class="scan-bad mt-2">No cargó la librería del lector. Revisa internet o CDN.</div>';
      playBad(); return;
    }
    try{
      let cfg={};
      if(window.Html5QrcodeSupportedFormats){
        cfg={formatsToSupport:[
          Html5QrcodeSupportedFormats.QR_CODE,
          Html5QrcodeSupportedFormats.CODE_128,
          Html5QrcodeSupportedFormats.CODE_39,
          Html5QrcodeSupportedFormats.EAN_13,
          Html5QrcodeSupportedFormats.EAN_8,
          Html5QrcodeSupportedFormats.UPC_A,
          Html5QrcodeSupportedFormats.UPC_E
        ]};
      }
      scanner=new Html5Qrcode(readerId,cfg);
      await scanner.start({facingMode:{ideal:'environment'}},{fps:12,qrbox:{width:240,height:160},aspectRatio:1.333},async decoded=>{
        let val=String(decoded||'').trim();
        if(inputId==='dniTrab'||inputId==='dniAvance') val=onlyDni(val);
        input.value=val;
        input.dispatchEvent(new Event('input',{bubbles:true}));
        input.dispatchEvent(new Event('change',{bubbles:true}));
        if(inputId==='dniTrab') await procesarDni(val,true);
        else playOk();
        await window.cerrarScannerActivo();
      },()=>{});
      if(!el.querySelector('.scanner-close-x')) el.insertAdjacentHTML('afterbegin', scannerCloseHtml());
    }catch(err){
      el.innerHTML=scannerCloseHtml()+'<div class="scan-bad mt-2">No se pudo activar cámara. Permite cámara en el candado del navegador y vuelve a intentar.</div>';
      playBad();
    }
  };

  // ================== DNI AUTOMÁTICO REAL ==================
  const workerMap=new Map();
  window.workerMap=workerMap;
  window.renderQueue=function(){
    const q=$('workerQueue'), h=$('dnisMasivos'); if(!q||!h)return;
    h.value=[...workerMap.keys()].join(',');
    if(workerMap.size===0){q.innerHTML='<div class="text-muted small text-center">Aún no hay trabajadores detectados.</div>';return;}
    q.innerHTML=[...workerMap.entries()].map(([dni,n])=>
      '<div class="queue-item"><div><b>'+dni+'</b><br><span>'+n+'</span></div><button type="button" class="btn btn-sm btn-outline-danger" onclick="workerMap.delete(\''+dni+'\');renderQueue();">×</button></div>'
    ).join('');
  };
  function dniStatus(kind,html){
    const st=$('dniStatus'); if(!st)return;
    st.className=(kind==='ok'?'scan-ok mt-2 flash':kind==='bad'?'scan-bad mt-2 flash':'mt-2 field-help');
    st.innerHTML=html;
  }
  let dniTimer=null, dniBusy=false, dniLast='';
  async function procesarDni(valor, forzar=false){
    const inp=$('dniTrab'); if(!inp)return;
    const dni=onlyDni(valor || inp.value);
    if(dni.length<8){ if(forzar) dniStatus('help','Escanee o digite DNI: al completar 8 dígitos se agregará al pre-registro con sonido.'); return; }
    inp.value=dni;
    if(!forzar && dni===dniLast)return;
    if(dniBusy){ clearTimeout(dniTimer); dniTimer=setTimeout(()=>procesarDni(dni,true),120); return; }
    dniBusy=true; dniLast=dni;
    dniStatus('ok','Buscando DNI <b>'+dni+'</b> en base trabajadores...');
    try{
      const r=await fetch('/api/trabajador/'+encodeURIComponent(dni),{cache:'no-store',credentials:'same-origin'});
      let j={ok:false,msg:'Respuesta inválida'}; try{j=await r.json();}catch(e){}
      if(!j.ok){
        dniStatus('bad','✕ '+(j.msg||'DNI no encontrado en base trabajadores')+' <b>'+dni+'</b>');
        playOk(); inp.select(); return;
      }
      const t=j.trabajador||{}; const nombre=t.trabajador||t.nombres||t.nombre||'TRABAJADOR';
      if(!workerMap.has(dni)){workerMap.set(dni,nombre); window.renderQueue();}
      dniStatus('ok','✓ Reconocido automáticamente: <b>'+nombre+'</b> · '+dni);
      playOk();
      await sleep(180); inp.value=''; dniLast=''; inp.focus();
    }catch(e){
      dniStatus('bad','Error consultando trabajador. Revisa conexión o sesión.');
    }finally{dniBusy=false;}
  }
  window.autoDetectarDniInline=function(el){
    const inp=el||$('dniTrab'); if(!inp)return;
    clearTimeout(dniTimer);
    dniTimer=setTimeout(()=>procesarDni(inp.value,false),40);
  };
  function instalarDniAuto(){
    const inp=$('dniTrab'); if(!inp || inp.dataset.finalAuto==='1')return; inp.dataset.finalAuto='1';
    ['input','keyup','change','paste','blur'].forEach(ev=>inp.addEventListener(ev,()=>setTimeout(()=>procesarDni(inp.value,ev!=='input'), ev==='paste'?80:5), true));
    inp.addEventListener('keydown',e=>{ if(e.key==='Enter'||e.key==='Tab'){procesarDni(inp.value,true); if(e.key==='Enter')e.preventDefault();} }, true);
    setInterval(()=>{const x=$('dniTrab'); if(x && x.value && onlyDni(x.value).length>=8) procesarDni(x.value,false);},250);
  }

  // ================== HORARIO TÁCTIL / CURSOR PC POR DESLIZADOR 24H ==================
  const IDS=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'];
  let campoActivo='horaInicioDefault';
  function toMin(v){const p=String(v||'00:00').split(':');let h=parseInt(p[0]||0,10),m=parseInt(p[1]||0,10); if(isNaN(h))h=0;if(isNaN(m))m=0;return Math.max(0,Math.min(1435,h*60+m));}
  function minToTime(m){m=((parseInt(m||0,10)%1440)+1440)%1440;return pad(Math.floor(m/60))+':'+pad(m%60);}
  function horasNetas(hi,hf,ri,rf){let a=toMin(hi), b=toMin(hf); if(b<=a)b+=1440; let total=b-a; if(ri&&rf){let c=toMin(ri), d=toMin(rf); if(d<=c)d+=1440; if(b>1440&&c<a){c+=1440;d+=1440;} total-=Math.max(0,Math.min(b,d)-Math.max(a,c));} return (Math.max(0,total)/60).toFixed(2);}
  function activeInput(){return $(campoActivo)||$('horaInicioDefault');}
  function sincHorario(){const hi=$('horaInicioDefault')?.value||'06:30', hf=$('horaFinDefault')?.value||'16:30', ri=$('refInicioDefault')?.value||'12:00', rf=$('refFinDefault')?.value||'13:00'; [['horaInicioTrab',hi],['horaFinTrab',hf],['refInicioTrab',ri],['refFinTrab',rf]].forEach(([id,v])=>{const e=$(id); if(e)e.value=v;}); const h=horasNetas(hi,hf,ri,rf); const ht=$('horasTrab'); if(ht)ht.value=h; const box=$('horarioActivoTxt'); if(box)box.innerHTML='<b>Horario activo:</b> '+hi+' - '+hf+' / Refrigerio '+ri+' - '+rf+' / H.Normal '+h+'.';}
  function pintarReloj(){const input=activeInput(), tv=$('touchClockValue'), sl=$('timeSlider24'), pills=$('clockPickFields'); if(tv&&input)tv.textContent=input.value; if(sl&&input)sl.value=toMin(input.value); if(pills)pills.querySelectorAll('button').forEach(b=>b.classList.toggle('active',b.dataset.target===campoActivo));}
  function setCampo(id){campoActivo=id;pintarReloj();const el=$(id);if(el){try{el.focus({preventScroll:true});}catch(_){}}}
  function setDesdeMinutos(m){const i=activeInput(); if(!i)return; i.value=minToTime(Math.round(Number(m||0))); sincHorario(); pintarReloj();}
  function instalarReloj(){
    const sl=$('timeSlider24'), pills=$('clockPickFields');
    if(pills){pills.querySelectorAll('button').forEach(b=>{b.onclick=(ev)=>{ev.preventDefault(); setCampo(b.dataset.target);}; b.onpointerdown=(ev)=>{ev.preventDefault(); setCampo(b.dataset.target);};});}
    IDS.forEach(id=>{const e=$(id); if(e){e.readOnly=true; e.tabIndex=0; e.onclick=()=>setCampo(id); e.onpointerdown=()=>setCampo(id); e.onfocus=()=>setCampo(id);}});
    if(sl){
      sl.style.pointerEvents='auto'; sl.style.touchAction='pan-x';
      const handler=(ev)=>{setDesdeMinutos(sl.value);};
      ['input','change','mousemove','touchmove'].forEach(ev=>{sl.addEventListener(ev,handler,{passive:true});});
      sl.onpointerdown=(ev)=>{setDesdeMinutos(sl.value);};
      sl.onpointerup=(ev)=>{setDesdeMinutos(sl.value); playOk();};
      sl.onclick=(ev)=>{setDesdeMinutos(sl.value);};
    }
    sincHorario(); pintarReloj();
  }
  window.aplicarHorarioRegistro=function(){sincHorario();pintarReloj();playOk();};
  window.tocarHora=function(delta){setDesdeMinutos(toMin(activeInput()?.value)+delta*60);};
  window.tocarMin=function(delta){setDesdeMinutos(toMin(activeInput()?.value)+delta);};
  window.setCampoHorario=setCampo;
  
  window.abrirEditarTareo=function(id,hi,hf,ri,rf){
    const f=$('frmEditTareo'); if(!f){alert('No se pudo abrir editor de horario.');return;}
    document.querySelectorAll('.card-menu.show').forEach(m=>m.classList.remove('show'));
    f.action='/tareo/'+id+'/editar-horas';
    $('editHi').value=hi||'06:30'; $('editHf').value=hf||'16:30'; $('editRi').value=ri||'12:00'; $('editRf').value=rf||'13:00';
    const modalEl=$('modalEditTareo');
    if(!modalEl){ window.location.href='/tareo/'+id+'/editar-horas-form'; return; }
    modalEl.style.zIndex='20000';
    if(window.bootstrap&&bootstrap.Modal){bootstrap.Modal.getOrCreateInstance(modalEl).show();}
    else {modalEl.style.display='block'; modalEl.classList.add('show'); modalEl.removeAttribute('aria-hidden');}
  };

  document.addEventListener('click',e=>{
    const b=e.target.closest && e.target.closest('.btn-edit-tareo');
    if(!b)return;
    e.preventDefault(); e.stopPropagation();
    window.abrirEditarTareo(b.dataset.id,b.dataset.hi,b.dataset.hf,b.dataset.ri,b.dataset.rf);
  },true);


  // PATCH 252: valida coherencia antes de guardar horario general o editar horario del trabajador
  function normHHMM(v){
    v=String(v||'').trim();
    if(/^\d{1,2}:\d{2}$/.test(v)){let [h,m]=v.split(':'); return String(Number(h)).padStart(2,'0')+':'+m;}
    const d=v.replace(/\D/g,'');
    if(d.length===3) return '0'+d[0]+':'+d.slice(1);
    if(d.length>=4) return d.slice(0,2)+':'+d.slice(2,4);
    return v;
  }
  function minHH(v){v=normHHMM(v); const p=v.split(':'); return (Number(p[0]||0)*60)+Number(p[1]||0);}
  function horarioOkClient(hi,hf,ri,rf){
    hi=normHHMM(hi); hf=normHHMM(hf); ri=normHHMM(ri); rf=normHHMM(rf);
    const re=/^([01][0-9]|2[0-3]):[0-5][0-9]$/;
    if(![hi,hf,ri,rf].every(x=>re.test(x))) return {ok:false,msg:'Usa formato 24 horas HH:MM. Ejemplo: 06:30, 16:30, 22:00.'};
    let a=minHH(hi), b=minHH(hf), c=minHH(ri), d=minHH(rf);
    if(b<=a) b+=1440;
    if(d<=c) d+=1440;
    if(b>1440 && c<a){c+=1440; d+=1440;}
    if(!(a<=c && c<d && d<=b)) return {ok:false,msg:'El refrigerio debe quedar dentro del horario de inicio y fin de trabajo.'};
    if((d-c)>(b-a)) return {ok:false,msg:'El refrigerio no puede ser mayor que la jornada.'};
    // Permite jornadas en formato 24 horas e incluso cruce de medianoche.
    // La única regla bloqueante es que el refrigerio quede dentro del inicio y fin de trabajo.
    return {ok:true,msg:'Horario coherente.'};
  }
  function validarFormHorario(form, ids){
    const vals=ids.map(id=>$(id)); if(vals.some(x=>!x)) return true;
    vals.forEach(x=>x.value=normHHMM(x.value));
    const r=horarioOkClient(vals[0].value, vals[1].value, vals[2].value, vals[3].value);
    const box=$('horarioCoherencia');
    if(box){box.className=r.ok?'scan-ok mt-2':'scan-bad mt-2'; box.innerHTML=r.msg;}
    if(!r.ok){alert(r.msg); return false;}
    return true;
  }
  document.addEventListener('submit',e=>{
    const f=e.target;
    if(f && f.action && f.action.includes('/fijar-horario/')){
      if(!validarFormHorario(f,['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'])){e.preventDefault(); e.stopPropagation(); return false;}
    }
    if(f && f.id==='frmEditTareo'){
      const vals=['editHi','editHf','editRi','editRf'].map(id=>$(id)); vals.forEach(x=>{if(x)x.value=normHHMM(x.value);});
      const r=horarioOkClient(vals[0]?.value, vals[1]?.value, vals[2]?.value, vals[3]?.value);
      if(!r.ok){e.preventDefault(); e.stopPropagation(); alert(r.msg); return false;}
    }
  },true);


  // ================== MAYÚSCULAS Y MAESTROS ACTIVIDAD/LABOR/CONSUMIDOR ==================
  document.addEventListener('input',e=>{const el=e.target;if(el && (el.tagName==='INPUT'||el.tagName==='TEXTAREA') && el.type!=='password' && el.type!=='date' && el.type!=='number' && el.id!=='dniTrab' && el.id!=='dniAvance' && el.id!=='buscarDni'){const p=el.selectionStart; el.value=String(el.value||'').toUpperCase(); try{el.setSelectionRange(p,p)}catch(_){}}},true);
  let MAESTROS_CACHE = (typeof MAESTROS_DET !== 'undefined' && Array.isArray(MAESTROS_DET)) ? MAESTROS_DET.slice() : [];
  const DEMO_MAESTROS=[{desc_actividad:'ADMINISTRACION',desc_labor:'LABOR ADMINISTRATIVA',desc_consumidor:'OFICINA CENTRAL'},{desc_actividad:'ADMISION',desc_labor:'CONTROL DOCUMENTARIO',desc_consumidor:'OFICINA 01'},{desc_actividad:'COSECHA',desc_labor:'COSECHA MANUAL',desc_consumidor:'CAMPO 01'},{desc_actividad:'COSECHA',desc_labor:'COSECHA SELECTIVA',desc_consumidor:'CAMPO 02'},{desc_actividad:'PODA',desc_labor:'PODA SANITARIA',desc_consumidor:'LOTE 01'}];
  const uniq=(a)=>[...new Set((a||[]).map(x=>String(x||'').trim()).filter(Boolean))].sort((a,b)=>a.localeCompare(b));
  function norm(v){return String(v||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toUpperCase().trim();}
  function firstVal(x,names){for(const n of names){if(x && x[n]!==undefined && x[n]!==null && String(x[n]).trim()!=='')return String(x[n]).trim().toUpperCase();}return '';}
  function normalizaFila(x){return {desc_actividad:firstVal(x,['desc_actividad','actividad','grupo','ACTIVIDAD','DESCRIPCION_ACTIVIDAD','DESCRIPCION ACTIVIDAD','DESC ACTIVIDAD','cod_actividad','COD_ACTIVIDAD']),cod_actividad:firstVal(x,['cod_actividad','codigo_actividad','COD ACTIVIDAD','COD_ACTIVIDAD']),desc_labor:firstVal(x,['desc_labor','labor','subgrupo','LABOR','DESCRIPCION_LABOR','DESCRIPCION LABOR','DESC LABOR','cod_labor','COD_LABOR']),cod_labor:firstVal(x,['cod_labor','codigo_labor','COD LABOR','COD_LABOR']),desc_consumidor:firstVal(x,['desc_consumidor','consumidor','CONSUMIDOR','zona','campo','ZONA','CAMPO','DESCRIPCION_CONSUMIDOR','DESCRIPCION CONSUMIDOR','cod_consumidor','COD_CONSUMIDOR']),cod_consumidor:firstVal(x,['cod_consumidor','codigo_consumidor','COD CONSUMIDOR','COD_CONSUMIDOR'])};}
  function setMasterStatus(msg,bad=false){const st=$('modalMasterStatus'); if(st){st.className='master-status'+(bad?' bad':'');st.textContent=msg;}}
  async function cargarMaestrosSiHaceFalta(){
    try{
      const r=await fetch('/api/actividades-maestras?ts='+Date.now(),{cache:'no-store',credentials:'same-origin'});
      const j=await r.json();
      if(j && j.ok && Array.isArray(j.data)){MAESTROS_CACHE=j.data.map(normalizaFila).filter(x=>x.desc_actividad||x.desc_labor||x.desc_consumidor);}
    }catch(e){setMasterStatus('NO SE PUDO CONECTAR A LA API DE ACTIVIDADES',true);}
    if(!MAESTROS_CACHE.length){MAESTROS_CACHE=DEMO_MAESTROS.slice(); setMasterStatus('SIN DATA REAL EN API: USANDO DEMO TEMPORAL',true);} else {setMasterStatus('ACTIVIDADES CARGADAS: '+MAESTROS_CACHE.length+' REGISTROS');}
  }
  function fillDL(id,arr){const dl=$(id); if(!dl)return; dl.innerHTML=''; arr.slice(0,500).forEach(v=>{const o=document.createElement('option');o.value=v;dl.appendChild(o);});}
  function showSuggest(id,input,arr,cb){const box=$(id); if(!box||!input)return; const q=norm(input.value); const vals=(q?arr.filter(v=>norm(v).includes(q)):arr).slice(0,20); if(!vals.length){box.style.display='none';box.innerHTML='';return;} box.innerHTML=vals.map(v=>'<div>'+String(v).replace(/</g,'&lt;')+'</div>').join(''); box.style.display='block'; [...box.children].forEach(div=>{div.onmousedown=(ev)=>ev.preventDefault(); div.onclick=()=>{input.value=div.textContent.toUpperCase();box.style.display='none';cb&&cb();};});}
  async function instalarMaestros(){
    await cargarMaestrosSiHaceFalta();
    const a=$('modalActividad'), l=$('modalLaborInput'), c=$('modalConsumidor'); if(!a||!l)return;
    const acts=uniq(MAESTROS_CACHE.map(x=>x.desc_actividad||x.cod_actividad)); fillDL('modal_actividad_list',acts);
    const rowsA=()=>{const q=norm(a.value);return MAESTROS_CACHE.filter(x=>!q||norm(x.desc_actividad).includes(q)||norm(x.cod_actividad).includes(q));};
    const laborVals=()=>uniq(rowsA().map(x=>x.desc_labor||x.cod_labor));
    const rowsL=()=>{const q=norm(l.value);return rowsA().filter(x=>!q||norm(x.desc_labor).includes(q)||norm(x.cod_labor).includes(q));};
    const consVals=()=>uniq(rowsL().map(x=>x.desc_consumidor||x.cod_consumidor));
    const refreshC=(show=false)=>{const vals=consVals(); fillDL('modal_consumidor_list',vals); if(c&&show)showSuggest('modalConsumidorSuggest',c,vals);};
    const refreshL=(show=false)=>{const vals=laborVals(); fillDL('modal_labor_list',vals); if(show)showSuggest('modalLaborSuggest',l,vals,()=>{refreshC(true); c&&c.focus();}); refreshC(false);};
    a.oninput=()=>{showSuggest('modalActividadSuggest',a,acts,()=>{l.value=''; if(c)c.value=''; refreshL(true);l.focus();});refreshL(true);};
    a.onfocus=()=>showSuggest('modalActividadSuggest',a,acts,()=>{l.value=''; refreshL(true);});
    l.oninput=()=>{showSuggest('modalLaborSuggest',l,laborVals(),()=>{if(c)c.value=''; refreshC(true);c&&c.focus();});refreshC(false);};
    l.onfocus=()=>showSuggest('modalLaborSuggest',l,laborVals(),()=>refreshC(true));
    if(c){c.oninput=()=>refreshC(true); c.onfocus=()=>refreshC(true);}
    document.addEventListener('click',ev=>{if(!ev.target.closest('#modalLabor'))['modalActividadSuggest','modalLaborSuggest','modalConsumidorSuggest'].forEach(id=>{const b=$(id);if(b)b.style.display='none';});},true);
    refreshL(false);
  }

  document.addEventListener('shown.bs.modal',e=>{
    if(e.target&&e.target.id==='modalRegistro'){instalarDniAuto();sincHorario();const i=$('dniTrab');if(i){setTimeout(()=>i.focus(),80);}}
    if(e.target&&e.target.id==='modalHora'){setTimeout(instalarReloj,60);}
    if(e.target&&e.target.id==='modalLabor'){instalarMaestros();}
  });
  document.addEventListener('DOMContentLoaded',()=>{instalarDniAuto();instalarReloj();instalarMaestros();});
  document.addEventListener('submit',e=>{if(e.target&&e.target.id==='frmTrab'){sincHorario(); window.renderQueue();}});
})();
</script>

<script>
/* === PARCHE REAL 247: maestros + horario deslizable robusto === */
(function(){
  'use strict';
  const $ = (id)=>document.getElementById(id);
  const norm = (v)=>String(v||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toUpperCase().trim();
  const pad = (n)=>String(Number(n)||0).padStart(2,'0');
  function minToTime(m){m=Math.max(0,Math.min(1435,parseInt(m||0,10)));return pad(Math.floor(m/60))+':'+pad(m%60);}
  function toMin(v){let p=String(v||'00:00').split(':'),h=parseInt(p[0]||0,10),m=parseInt(p[1]||0,10); if(isNaN(h))h=0;if(isNaN(m))m=0;return Math.max(0,Math.min(1435,h*60+m));}
  function safeText(v){return String(v||'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));}

  // ---------- ACTIVIDAD / LABOR / CONSUMIDOR DESDE API ----------
  let maestros=[];
  const demo=[
    {desc_actividad:'ADMINISTRACION',desc_labor:'LABOR ADMINISTRATIVA',desc_consumidor:'OFICINA CENTRAL'},
    {desc_actividad:'ADMISION',desc_labor:'CONTROL DOCUMENTARIO',desc_consumidor:'OFICINA 01'},
    {desc_actividad:'COSECHA',desc_labor:'COSECHA MANUAL',desc_consumidor:'CAMPO 01'},
    {desc_actividad:'COSECHA',desc_labor:'COSECHA SELECTIVA',desc_consumidor:'CAMPO 02'},
    {desc_actividad:'PODA',desc_labor:'PODA SANITARIA',desc_consumidor:'LOTE 01'}
  ];
  function fila(x){
    const pick=(names)=>{for(const n of names){if(x&&x[n]!=null&&String(x[n]).trim())return norm(x[n]);}return '';};
    return {
      desc_actividad:pick(['desc_actividad','actividad','grupo','cod_actividad','ACTIVIDAD','DESCRIPCION_ACTIVIDAD','DESCRIPCION ACTIVIDAD']),
      desc_labor:pick(['desc_labor','labor','subgrupo','cod_labor','LABOR','DESCRIPCION_LABOR','DESCRIPCION LABOR']),
      desc_consumidor:pick(['desc_consumidor','consumidor','zona','campo','cod_consumidor','CONSUMIDOR','ZONA','CAMPO'])
    };
  }
  function unique(a){return [...new Set((a||[]).filter(Boolean))].sort((x,y)=>x.localeCompare(y));}
  function status(msg,bad){let s=$('modalMasterStatus'); if(s){s.className='master-status'+(bad?' bad':''); s.textContent=msg;}}
  function fillDatalist(id, arr){let dl=$(id); if(!dl)return; dl.innerHTML=''; arr.slice(0,500).forEach(v=>{let o=document.createElement('option');o.value=v;dl.appendChild(o);});}
  function showBox(id,input,vals,onpick){
    let b=$(id); if(!b||!input)return; let q=norm(input.value); let list=(q?vals.filter(v=>norm(v).includes(q)):vals).slice(0,30);
    if(!list.length){b.style.display='none'; b.innerHTML=''; return;}
    b.innerHTML=list.map(v=>'<div>'+safeText(v)+'</div>').join(''); b.style.display='block';
    [...b.children].forEach(d=>{d.onmousedown=e=>e.preventDefault(); d.onclick=()=>{input.value=d.textContent; b.style.display='none'; if(onpick)onpick();};});
  }
  async function cargarMaestros(){
    try{
      let r=await fetch('/api/actividades-maestras?ts='+Date.now(),{cache:'no-store',credentials:'same-origin'});
      let j=await r.json();
      maestros=(j&&j.ok&&Array.isArray(j.data)?j.data:[]).map(fila).filter(x=>x.desc_actividad||x.desc_labor||x.desc_consumidor);
    }catch(e){maestros=[];}
    if(!maestros.length){maestros=demo.map(fila); status('SIN DATA REAL EN API: USANDO DEMO TEMPORAL',true);} else status('ACTIVIDADES CARGADAS: '+maestros.length+' REGISTROS',false);
  }
  async function instalarMaestrosReal(){
    let a=$('modalActividad'), l=$('modalLaborInput'), c=$('modalConsumidor'); if(!a||!l)return;
    await cargarMaestros();
    const getActs=()=>unique(maestros.map(x=>x.desc_actividad));
    const rowsA=()=>{let q=norm(a.value);return maestros.filter(x=>!q||norm(x.desc_actividad).includes(q)||q.includes(norm(x.desc_actividad)));};
    const getLabs=()=>unique(rowsA().map(x=>x.desc_labor));
    const rowsL=()=>{let q=norm(l.value);return rowsA().filter(x=>!q||norm(x.desc_labor).includes(q)||q.includes(norm(x.desc_labor)));};
    const getCons=()=>unique(rowsL().map(x=>x.desc_consumidor));
    function refreshActividad(show=true){let vals=getActs();fillDatalist('modal_actividad_list',vals); if(show)showBox('modalActividadSuggest',a,vals,()=>{l.value=''; if(c)c.value=''; refreshLabor(true); setTimeout(()=>l.focus(),20);});}
    function refreshLabor(show=true){let vals=getLabs();fillDatalist('modal_labor_list',vals); if(show)showBox('modalLaborSuggest',l,vals,()=>{if(c)c.value=''; refreshConsumidor(true); setTimeout(()=>c&&c.focus(),20);}); refreshConsumidor(false);}
    function refreshConsumidor(show=true){let vals=getCons();fillDatalist('modal_consumidor_list',vals); if(c&&show)showBox('modalConsumidorSuggest',c,vals,()=>{});}
    a.oninput=()=>{a.value=norm(a.value); refreshActividad(true); refreshLabor(true);};
    a.onfocus=()=>refreshActividad(true);
    l.oninput=()=>{l.value=norm(l.value); refreshLabor(true);};
    l.onfocus=()=>refreshLabor(true);
    if(c){c.oninput=()=>{c.value=norm(c.value); refreshConsumidor(true);}; c.onfocus=()=>refreshConsumidor(true);}
    refreshActividad(false); refreshLabor(false);
  }

  // ---------- HORARIO POR DESLIZADOR: PC + TÁCTIL ----------
  const ids=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'];
  let active='horaInicioDefault';
  function inputActive(){return $(active)||$('horaInicioDefault');}
  function paint(){let inp=inputActive(), sl=$('timeSlider24'), val=$('touchClockValue'); if(inp&&sl)sl.value=toMin(inp.value); if(inp&&val)val.textContent=inp.value; let box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(b=>b.classList.toggle('active',b.dataset.target===active));}
  function setActive(id){active=id; paint(); let e=$(id); if(e)try{e.focus({preventScroll:true});}catch(_){} }
  function applySlider(){let inp=inputActive(); if(!inp)return; let sl=$('timeSlider24'); inp.value=minToTime(sl?sl.value:0); paint(); syncHidden();}
  function syncHidden(){
    const hi=$('horaInicioDefault')?.value||'06:30', hf=$('horaFinDefault')?.value||'16:30', ri=$('refInicioDefault')?.value||'12:00', rf=$('refFinDefault')?.value||'13:00';
    [['horaInicioTrab',hi],['horaFinTrab',hf],['refInicioTrab',ri],['refFinTrab',rf]].forEach(([id,v])=>{let e=$(id);if(e)e.value=v;});
    let a=toMin(hi),b=toMin(hf); if(b<=a)b+=1440; let c=toMin(ri),d=toMin(rf); if(d<=c)d+=1440; if(b>1440&&c<a){c+=1440;d+=1440;} let total=Math.max(0,(b-a)-Math.max(0,Math.min(b,d)-Math.max(a,c)))/60;
    let h=$('horasTrab'); if(h)h.value=total.toFixed(2);
    let txt=$('horarioActivoTxt'); if(txt)txt.innerHTML='<b>Horario activo:</b> '+hi+' - '+hf+' / Refrigerio '+ri+' - '+rf+' / H.Normal '+total.toFixed(2)+'.';
  }
  function instalarRelojReal(){
    let sl=$('timeSlider24');
    let box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(b=>{b.onclick=e=>{e.preventDefault();setActive(b.dataset.target);}; b.onpointerdown=e=>{setActive(b.dataset.target);};});
    ids.forEach(id=>{let e=$(id); if(e){e.readOnly=true; e.style.cursor='pointer'; e.onclick=()=>setActive(id); e.onpointerdown=()=>setActive(id); e.onfocus=()=>setActive(id);}});
    if(sl && sl.dataset.boundReal!=='1'){
      sl.dataset.boundReal='1'; sl.style.pointerEvents='auto'; sl.style.touchAction='none';
      ['input','change','pointermove','mousemove','touchmove','click','pointerup','touchend'].forEach(ev=>sl.addEventListener(ev,(e)=>{applySlider();},{passive:false}));
    }
    paint(); syncHidden();
  }
  window.setCampoHorario=setActive; window.aplicarHorarioRegistro=syncHidden;
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalLabor')setTimeout(instalarMaestrosReal,50); if(e.target&&e.target.id==='modalHora')setTimeout(instalarRelojReal,50);});
  document.addEventListener('DOMContentLoaded',()=>{setTimeout(()=>{instalarMaestrosReal(); instalarRelojReal();},100);});
})();
</script>

    """
    return render_page(body, h=h, tab=tab, tareos=tareos, lecturas=lecturas, labores=labores, registros=registros, horas_total=horas_total, rend_total=rend_total, maestros_json=js_master_options(get_actividades_maestras()), selected_labor_id=selected_labor_id, selected_labor=selected_labor)


@app.route('/hoja/<int:hoja_id>/copiar-labor/<tab>', methods=['POST'])
@login_required
def copiar_labor_hoja(hoja_id, tab):
    origen_id = request.form.get('labor_id_origen')
    lab = row_to_dict(execute('SELECT * FROM hoja_labores WHERE id=? AND hoja_id=?', (origen_id, hoja_id), fetchone=True))
    if not lab:
        flash('Selecciona una labor válida para copiar.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))
    labor_nueva = limpiar_texto(request.form.get('labor_nueva') or lab.get('labor'))
    execute('INSERT INTO hoja_labores(hoja_id,grupo,subgrupo,labor,turno,tipo_tareo,responsable,creado_en,creado_por) VALUES(?,?,?,?,?,?,?,?,?)',
            (hoja_id, lab.get('grupo'), lab.get('subgrupo'), labor_nueva, lab.get('turno'), lab.get('tipo_tareo'), lab.get('responsable'), now_str(), session.get('usuario')), commit=True)
    flash('Labor copiada correctamente.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))

@app.route('/hoja/<int:hoja_id>/labor/<tab>', methods=['POST'])
@login_required
def guardar_labor_hoja(hoja_id, tab):
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    if not h: flash('Hoja no encontrada.', 'danger'); return redirect(url_for('hojas_tareo'))
    grupo = limpiar_texto(request.form.get('grupo') or h.get('grupo'))
    subgrupo = limpiar_texto(request.form.get('subgrupo') or h.get('subgrupo'))
    labor = limpiar_texto(request.form.get('labor') or h.get('labor'))
    responsable = limpiar_texto(request.form.get('responsable') or h.get('responsable'))
    turno = limpiar_texto(request.form.get('turno') or 'DIA')
    tipo_tareo = limpiar_texto(request.form.get('tipo_tareo') or 'JORNAL')
    if not grupo or not subgrupo:
        flash('Debe seleccionar Actividad y Labor.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))
    execute('INSERT INTO hoja_labores(hoja_id,grupo,subgrupo,labor,turno,tipo_tareo,responsable,creado_en,creado_por) VALUES(?,?,?,?,?,?,?,?,?)',
            (hoja_id,grupo,subgrupo,labor,turno,tipo_tareo,responsable,now_str(),session.get('usuario')), commit=True)
    flash('Nueva labor creada dentro de la hoja.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))


@app.route('/hoja/<int:hoja_id>/fijar-horario/<tab>', methods=['POST'])
@login_required
def fijar_horario_hoja(hoja_id, tab):
    labor_id = request.form.get('labor_id') or request.args.get('labor_id') or ''
    def okhora(v):
        return bool(re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', str(v or '').strip()))
    hi = (request.form.get('hora_inicio_default') or '').strip().zfill(5)
    hf = (request.form.get('hora_fin_default') or '').strip().zfill(5)
    ri = (request.form.get('ref_inicio_default') or '').strip().zfill(5)
    rf = (request.form.get('ref_fin_default') or '').strip().zfill(5)
    if not all(okhora(x) for x in [hi,hf,ri,rf]):
        flash('Horario inválido. Usa formato 24 horas: HH:MM, por ejemplo 06:30 o 22:00.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id))
    coherente, msg_coh = horario_coherente(hi, hf, ri, rf)
    if not coherente:
        flash(msg_coh, 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id))
    execute('UPDATE hojas_tareo SET horario_fijado=1, hora_inicio_default=?, hora_fin_default=?, ref_inicio_default=?, ref_fin_default=? WHERE id=?',
            (hi,hf,ri,rf,hoja_id), commit=True)
    flash('Horario fijado correctamente para esta hoja de tareo.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id))

@app.route('/hoja/<int:hoja_id>/registro/<tab>', methods=['POST'])
@login_required
def guardar_registro_hoja(hoja_id, tab):
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger')
        return redirect(url_for('hojas_tareo'))

    avance_pre_items = []
    if tab == 'rendimiento' and request.form.get('avance_pre_json'):
        try:
            raw_items = json.loads(request.form.get('avance_pre_json') or '[]')
            if isinstance(raw_items, list):
                for it in raw_items:
                    d = limpiar_dni((it or {}).get('dni'))
                    try:
                        cant_it = float((it or {}).get('cantidad') or 0)
                    except Exception:
                        cant_it = 0
                    metodo_it = limpiar_texto((it or {}).get('metodo') or request.form.get('metodo') or 'QR/CÓDIGO')
                    if len(d) == 8 and cant_it > 0:
                        avance_pre_items.append({'dni': d, 'cantidad': cant_it, 'a_noct': 0.0, 'metodo': metodo_it})
        except Exception:
            avance_pre_items = []

    dnis_raw = request.form.get('dnis_masivos') or request.form.get('dni') or ''
    dnis = []
    for part in re.split(r'[,;\s]+', dnis_raw):
        d = limpiar_dni(part)
        if len(d) == 8 and d not in dnis:
            dnis.append(d)
    if avance_pre_items:
        dnis = []
        for it in avance_pre_items:
            if it['dni'] not in dnis:
                dnis.append(it['dni'])
    if not dnis:
        flash('Debe digitar o escanear al menos un DNI válido.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab))
    if tab == 'trabajadores' and not int(h.get('horario_fijado') or 0):
        flash('Antes de tarear trabajadores debes fijar el horario de la hoja desde el icono de reloj.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab))

    labor_id = request.form.get('labor_id') or None
    lab = row_to_dict(execute('SELECT * FROM hoja_labores WHERE id=? AND hoja_id=?', (labor_id, hoja_id), fetchone=True)) if labor_id else None
    if str(h.get('estado') or '').upper() == 'ENVIADA':
        flash('No se puede registrar: la hoja ya fue enviada.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id or ''))
    if not lab:
        flash('Primero debes seleccionar una labor válida.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))
    labor = (lab or h).get('labor')
    grupo = (lab or h).get('grupo')
    turno = limpiar_texto(request.form.get('turno') or (lab or h).get('turno') or 'DIA')
    tipo_tareo = limpiar_texto(request.form.get('tipo_tareo') or (lab or h).get('tipo_tareo') or 'JORNAL')
    try:
        horas = float(request.form.get('horas') or 0)
        cantidad = float(request.form.get('cantidad') or 0)
        a_noct = 0.0
    except Exception:
        flash('Horas / avance inválido.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab))
    hora_inicio = request.form.get('hora_inicio') or h.get('hora_inicio_default') or ('22:00' if turno == 'NOCHE' else '06:30')
    hora_fin = request.form.get('hora_fin') or h.get('hora_fin_default') or ('06:00' if turno == 'NOCHE' else '16:30')
    ref_inicio = request.form.get('ref_inicio') or h.get('ref_inicio_default') or '12:00'
    ref_fin = request.form.get('ref_fin') or h.get('ref_fin_default') or '13:00'
    horas_noct = 0
    if tab == 'trabajadores':
        coherente, msg_coh = horario_coherente(hora_inicio, hora_fin, ref_inicio, ref_fin)
        if not coherente:
            flash(msg_coh, 'danger')
            return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id or ''))
        horas = calcular_horas_laborales(hora_inicio, hora_fin, ref_inicio, ref_fin)
        horas_noct = calcular_horas_nocturnas(hora_inicio, hora_fin, ref_inicio, ref_fin)
    unidad = limpiar_texto(request.form.get('unidad') or ('BALDE' if tab == 'rendimiento' else tipo_tareo))
    metodo = limpiar_texto(request.form.get('metodo') or 'DIGITACIÓN')

    ok = 0; no_encontrados = []
    for dni in dnis:
        t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
        if not t:
            no_encontrados.append(dni)
            continue
        if tab == 'trabajadores' and scalar('SELECT COUNT(*) AS c FROM tareos WHERE hoja_id=? AND labor_id=? AND dni=?', (hoja_id, labor_id, dni)):
            no_encontrados.append(dni + ' duplicado')
            continue
        if tab == 'rendimiento' and not scalar('SELECT COUNT(*) AS c FROM tareos WHERE hoja_id=? AND labor_id=? AND dni=?', (hoja_id, labor_id, dni)):
            no_encontrados.append(dni + ' no registrado en Trabajadores de esta labor')
            continue
        h_reg = horas
        if tab == 'rendimiento':
            item = next((x for x in avance_pre_items if x.get('dni') == dni), None) if avance_pre_items else None
            cant_reg = float(item.get('cantidad') if item else cantidad)
            metodo_reg = item.get('metodo') if item else metodo
            execute('INSERT INTO lecturas_balde(hoja_id,labor_id,dni,trabajador,fecha_hora,a_diurno,a_noct,metodo,registrado_por) VALUES(?,?,?,?,?,?,?,?,?)',
                    (hoja_id,labor_id,dni,t.get('trabajador',''),now_str(),cant_reg,0.0,metodo_reg,session.get('usuario')), commit=True)
            ok += 1
            continue
        execute('''INSERT INTO tareos(hoja_id,labor_id,dni,trabajador,empresa,area,cargo,fecha,labor,lote,fundo,horas,cantidad,unidad,observacion,registrado_por,creado_en,hora_inicio,hora_fin,ref_inicio,ref_fin,turno,tipo_tareo,horas_nocturnas)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (hoja_id,labor_id,dni,t.get('trabajador',''),t.get('empresa',''),t.get('area',''),t.get('cargo',''),h.get('fecha'),labor,limpiar_texto(request.form.get('lote') or grupo),grupo,h_reg,cantidad,unidad,'',session.get('usuario'),now_str(),hora_inicio,hora_fin,ref_inicio,ref_fin,turno,tipo_tareo,horas_noct), commit=True)
        ok += 1
    msg = f'Registro guardado correctamente. Registros guardados: {ok}.'
    if no_encontrados:
        msg += ' No encontrados: ' + ', '.join(no_encontrados)
    flash(msg, 'success' if ok else 'danger')
    return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id or ''))




@app.route('/tareo/<int:tareo_id>/editar-horas-form', methods=['GET'])
@login_required
def editar_horas_tareo_form(tareo_id):
    r = row_to_dict(execute('SELECT * FROM tareos WHERE id=?', (tareo_id,), fetchone=True))
    if not r:
        flash('Registro de trabajador no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (r.get('hoja_id'),), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger')
        return redirect(url_for('hojas_tareo'))
    hi = r.get('hora_inicio') or ('22:00' if r.get('turno') == 'NOCHE' else '06:30')
    hf = r.get('hora_fin') or ('06:00' if r.get('turno') == 'NOCHE' else '16:30')
    ri = r.get('ref_inicio') or '12:00'
    rf = r.get('ref_fin') or '13:00'
    body = f'''
    <div class="phone-wrap">
      <div class="page-card p-3">
        <div class="d-flex align-items-center justify-content-between mb-2">
          <h5 class="fw-bold text-success m-0"><i class="bi bi-pencil-square"></i> Modificar horario</h5>
          <a class="back-mini" href="{url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or '')}">×</a>
        </div>
        <div class="alert alert-light border small mb-2">
          Trabajador: <b>{r.get('trabajador') or ''}</b><br>DNI: <b>{r.get('dni') or ''}</b>
        </div>
        <form method="post" action="{url_for('editar_horas_tareo', tareo_id=tareo_id)}" id="frmEditTareoStandalone">
          <div class="row g-2">
            <div class="col-6"><label class="form-label">Hora inicio</label><input id="editHi" name="hora_inicio" class="form-control" value="{hi}" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div>
            <div class="col-6"><label class="form-label">Hora fin</label><input id="editHf" name="hora_fin" class="form-control" value="{hf}" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div>
            <div class="col-6"><label class="form-label">Ref. inicio</label><input id="editRi" name="ref_inicio" class="form-control" value="{ri}" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div>
            <div class="col-6"><label class="form-label">Ref. fin</label><input id="editRf" name="ref_fin" class="form-control" value="{rf}" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div>
          </div>
          <div class="field-help mt-2">Se permiten horarios 24 horas y cruce de medianoche. El refrigerio debe estar dentro de la jornada.</div>
          <button class="btn btn-green w-100 mt-3" type="submit">GUARDAR CAMBIOS</button>
        </form>
      </div>
    </div>'''
    return render_template_string(BASE_HTML, title='Modificar horario', body=body)

@app.route('/tareo/<int:tareo_id>/editar-horas', methods=['POST'])
@login_required
def editar_horas_tareo(tareo_id):
    r = row_to_dict(execute('SELECT * FROM tareos WHERE id=?', (tareo_id,), fetchone=True))
    if not r:
        flash('Registro de trabajador no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (r.get('hoja_id'),), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if str(h.get('estado') or '').upper() == 'ENVIADA':
        flash('No se puede editar: la hoja de tareo ya fue enviada.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))
    def okhora(v):
        return bool(re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', str(v or '').strip()))
    hi=(request.form.get('hora_inicio') or '').strip().zfill(5)
    hf=(request.form.get('hora_fin') or '').strip().zfill(5)
    ri=(request.form.get('ref_inicio') or '').strip().zfill(5)
    rf=(request.form.get('ref_fin') or '').strip().zfill(5)
    if not all(okhora(x) for x in [hi,hf,ri,rf]):
        flash('Horario inválido. Usa formato 24 horas HH:MM.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))
    coherente, msg_coh = horario_coherente(hi, hf, ri, rf)
    if not coherente:
        flash(msg_coh, 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))
    horas = calcular_horas_laborales(hi,hf,ri,rf)
    noct = calcular_horas_nocturnas(hi,hf,ri,rf)
    execute('UPDATE tareos SET hora_inicio=?, hora_fin=?, ref_inicio=?, ref_fin=?, horas=?, horas_nocturnas=? WHERE id=?', (hi,hf,ri,rf,horas,noct,tareo_id), commit=True)
    flash('Horas del trabajador actualizadas correctamente.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))


@app.route('/tareo/<int:tareo_id>/eliminar')
@login_required
def eliminar_tareo(tareo_id):
    r = row_to_dict(execute('SELECT * FROM tareos WHERE id=?', (tareo_id,), fetchone=True))
    if not r:
        flash('Registro no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if hoja_enviada(r.get('hoja_id')):
        flash('No se puede eliminar: la hoja ya fue enviada.', 'danger')
    else:
        execute('DELETE FROM tareos WHERE id=?', (tareo_id,), commit=True)
        flash('Trabajador eliminado del tareo.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))

@app.route('/lectura/<int:lectura_id>/eliminar')
@login_required
def eliminar_lectura(lectura_id):
    l = row_to_dict(execute('SELECT * FROM lecturas_balde WHERE id=?', (lectura_id,), fetchone=True))
    if not l:
        flash('Avance no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if hoja_enviada(l.get('hoja_id')):
        flash('No se puede eliminar: la hoja ya fue enviada.', 'danger')
    else:
        execute('DELETE FROM lecturas_balde WHERE id=?', (lectura_id,), commit=True)
        flash('Avance eliminado correctamente.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=l.get('hoja_id'), tab='rendimiento', labor_id=l.get('labor_id') or ''))

@app.route('/lectura/<int:lectura_id>/editar', methods=['POST'])
@login_required
def editar_lectura(lectura_id):
    l = row_to_dict(execute('SELECT * FROM lecturas_balde WHERE id=?', (lectura_id,), fetchone=True))
    if not l:
        flash('Avance no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if hoja_enviada(l.get('hoja_id')):
        flash('No se puede editar: la hoja ya fue enviada.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=l.get('hoja_id'), tab='rendimiento', labor_id=l.get('labor_id') or ''))
    try:
        cant = float(request.form.get('cantidad') or 0)
        noct = float(request.form.get('a_noct') or 0)
    except Exception:
        flash('Cantidad inválida.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=l.get('hoja_id'), tab='rendimiento', labor_id=l.get('labor_id') or ''))
    execute('UPDATE lecturas_balde SET a_diurno=?, a_noct=? WHERE id=?', (cant, noct, lectura_id), commit=True)
    flash('Avance actualizado correctamente.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=l.get('hoja_id'), tab='rendimiento', labor_id=l.get('labor_id') or ''))

@app.route('/hoja/<int:hoja_id>/editar', methods=['GET','POST'])
@login_required
def editar_hoja(hoja_id):
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if request.method == 'POST':
        fecha = request.form.get('fecha') or h.get('fecha')
        grupo = limpiar_texto(request.form.get('actividad'))
        subgrupo = limpiar_texto(request.form.get('labor'))
        labor = limpiar_texto(request.form.get('consumidor'))
        responsable = limpiar_texto(request.form.get('responsable'))
        turno = limpiar_texto(request.form.get('turno') or 'DIA')
        tipo_tareo = limpiar_texto(request.form.get('tipo_tareo') or 'JORNAL')
        if not grupo or not subgrupo or not responsable:
            flash('Actividad, labor y responsable son obligatorios.', 'danger')
            return redirect(url_for('editar_hoja', hoja_id=hoja_id))
        execute('UPDATE hojas_tareo SET fecha=?,grupo=?,subgrupo=?,labor=?,responsable=?,turno=?,tipo_tareo=? WHERE id=?', (fecha,grupo,subgrupo,labor,responsable,turno,tipo_tareo,hoja_id), commit=True)
        execute('UPDATE hoja_labores SET grupo=?,subgrupo=?,labor=?,responsable=?,turno=?,tipo_tareo=? WHERE hoja_id=?', (grupo,subgrupo,labor,responsable,turno,tipo_tareo,hoja_id), commit=True)
        flash('Hoja modificada correctamente.', 'success')
        return redirect(url_for('hojas_tareo'))
    body = """
    <div class="phone-wrap desktop-pad"><h2 class="header-title">MODIFICAR HOJA</h2><div class="page-card"><form method="post" class="floating-card m-2">
      <a class="back-mini" href="{{url_for('hojas_tareo')}}"><i class="bi bi-chevron-left"></i></a>
      <label class="form-label mt-2">FECHA</label><input type="date" name="fecha" class="form-control mb-2" value="{{h.fecha}}" required>
      <label class="form-label">ACTIVIDAD</label><input name="actividad" class="form-control mb-2" value="{{h.grupo}}" required>
      <label class="form-label">LABOR</label><input name="labor" class="form-control mb-2" value="{{h.subgrupo}}" required>
      <label class="form-label">CONSUMIDOR</label><input name="consumidor" class="form-control mb-2" value="{{h.labor}}">
      <label class="form-label">RESPONSABLE</label><input name="responsable" class="form-control mb-2" value="{{h.responsable}}" required>
      <div class="row g-2 mb-3"><div class="col-6"><label class="form-label">TURNO</label><select name="turno" class="form-select"><option {{'selected' if h.turno=='DIA' else ''}}>DIA</option><option {{'selected' if h.turno=='NOCHE' else ''}}>NOCHE</option></select></div><div class="col-6"><label class="form-label">TIPO</label><select name="tipo_tareo" class="form-select"><option {{'selected' if h.tipo_tareo=='JORNAL' else ''}}>JORNAL</option><option {{'selected' if h.tipo_tareo=='RENDIMIENTO' else ''}}>RENDIMIENTO</option></select></div></div>
      <button class="btn btn-green w-100">GUARDAR CAMBIOS</button>
    </form></div></div>"""
    return render_page(body, h=h)

@app.route('/hoja/<int:hoja_id>/eliminar')
@login_required
def eliminar_hoja(hoja_id):
    execute('DELETE FROM lecturas_balde WHERE hoja_id=?', (hoja_id,), commit=True)
    execute('DELETE FROM tareos WHERE hoja_id=?', (hoja_id,), commit=True)
    execute('DELETE FROM hoja_labores WHERE hoja_id=?', (hoja_id,), commit=True)
    execute('DELETE FROM hojas_tareo WHERE id=?', (hoja_id,), commit=True)
    flash('Hoja eliminada correctamente.', 'success')
    return redirect(url_for('hojas_tareo'))

@app.route('/hoja/<int:hoja_id>/enviar')
@login_required
def enviar_hoja(hoja_id):
    execute("UPDATE hojas_tareo SET estado='ENVIADA' WHERE id=?", (hoja_id,), commit=True)
    flash('Hoja marcada como ENVIADA.', 'success')
    return redirect(url_for('hojas_tareo'))


# ========================= ASISTENCIA + FIRMA DOCUMENTOS =========================
def guardar_data_url(data_url, carpeta, prefijo):
    try:
        if not data_url or ',' not in data_url:
            return ''
        head, data = data_url.split(',', 1)
        if 'base64' not in head:
            return ''
        raw = base64.b64decode(data)
        nombre = f"{prefijo}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.png"
        path = os.path.join(carpeta, nombre)
        with open(path, 'wb') as f:
            f.write(raw)
        return path
    except Exception as e:
        print('No se pudo guardar imagen:', e)
        return ''

@app.route('/asistencia', methods=['GET','POST'])
@login_required
def asistencia_modulo():
    if request.method == 'POST':
        dni = limpiar_dni(request.form.get('dni'))
        tipo = limpiar_texto(request.form.get('tipo') or 'ENTRADA')
        metodo = limpiar_texto(request.form.get('metodo') or 'QR/BARRAS')
        obs = limpiar_texto(request.form.get('observacion'), upper=False)
        if len(dni) != 8:
            flash('DNI inválido. Escanea QR, código de barras, huella/lector o digita 8 dígitos.', 'danger')
            return redirect(url_for('asistencia_modulo'))
        t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
        if not t:
            flash('DNI no existe en la base de trabajadores. No se registró asistencia.', 'danger')
            return redirect(url_for('asistencia_modulo'))
        ahora = datetime.now(); fecha = ahora.strftime('%Y-%m-%d'); hora = ahora.strftime('%H:%M:%S')
        foto_path = guardar_data_url(request.form.get('foto_data'), FOTO_DIR, f"asistencia_{dni}")
        sql = 'INSERT INTO asistencia(dni,trabajador,empresa,area,cargo,tipo,fecha,hora,fecha_hora,metodo,foto_path,latitud,longitud,registrado_por,observacion) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'
        execute(sql, (dni,t.get('trabajador'),t.get('empresa'),t.get('area'),t.get('cargo'),tipo,fecha,hora,ahora.strftime('%Y-%m-%d %H:%M:%S'),metodo,foto_path,request.form.get('latitud',''),request.form.get('longitud',''),session.get('usuario'),obs), commit=True)
        flash(f"Asistencia registrada: {tipo} - {dni} - {t.get('trabajador')}", 'success')
        return redirect(url_for('asistencia_modulo'))
    registros = rows_to_dict(execute('SELECT * FROM asistencia ORDER BY fecha_hora DESC LIMIT 80', fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad"><div class="page-card">
      <div class="panel-green"><a class="text-white text-decoration-none float-start" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><i class="bi bi-fingerprint"></i><h4>REGISTRO DE ASISTENCIA</h4></div>
      <form class="floating-card" style="margin:-24px 9px 10px" method="post" id="frmAsistencia">
        <div class="alert alert-light border small mb-2">Lee <b>QR/código de barras</b> con cámara, digita DNI o usa lector de <b>huella</b> compatible que envíe el DNI/código al campo.</div>
        <label class="form-label">DNI / QR / Código / Huella</label>
        <div class="input-group mb-2"><input name="dni" id="dniAsistencia" class="form-control" maxlength="20" placeholder="ESCANEAR O DIGITAR" required autofocus><button type="button" class="btn btn-green" onclick="abrirScanner('readerAsistencia','dniAsistencia')"><i class="bi bi-camera"></i></button></div>
        <div id="readerAsistencia" class="scan-box mb-2" style="display:none"></div><div id="asisStatus" class="field-help mb-2">Al completar 8 dígitos se validará contra trabajadores.</div>
        <div class="row g-2"><div class="col-6"><label class="form-label">Tipo</label><select name="tipo" class="form-select"><option>ENTRADA</option><option>SALIDA</option><option>REFRIGERIO INICIO</option><option>REFRIGERIO FIN</option></select></div><div class="col-6"><label class="form-label">Método</label><select name="metodo" class="form-select"><option>QR</option><option>CODIGO DE BARRAS</option><option>HUELLA</option><option>DNI DIGITADO</option></select></div></div>
        <input type="hidden" name="latitud" id="latitud"><input type="hidden" name="longitud" id="longitud"><input type="hidden" name="foto_data" id="fotoAsistencia">
        <label class="form-label mt-2">Observación</label><input name="observacion" class="form-control" placeholder="Opcional"><button class="btn btn-green w-100 mt-3"><i class="bi bi-check2-circle"></i> REGISTRAR ASISTENCIA</button>
      </form><div class="mx-2 mb-2 d-grid"><a class="btn btn-outline-success" href="{{url_for('exportar_asistencia')}}"><i class="bi bi-file-earmark-excel"></i> Exportar asistencia</a></div>
      {% for r in registros %}<div class="worker-card"><div class="worker-title"><div>{{r.fecha}} {{r.hora}}<br><b>{{r.trabajador}}</b></div><div class="text-end">{{r.dni}}<br><b>{{r.tipo}}</b></div></div><div class="worker-grid"><div><label>MÉTODO</label><div class="small-value">{{r.metodo}}</div></div><div><label>ÁREA</label><div class="small-value">{{r.area}}</div></div><div><label>CARGO</label><div class="small-value">{{r.cargo}}</div></div></div></div>{% else %}<div class="worker-card text-center text-muted">Sin registros de asistencia.</div>{% endfor %}
    </div></div><script>(function(){const dni=v=>String(v||'').replace(/\D/g,'').slice(-8), i=document.getElementById('dniAsistencia'), st=document.getElementById('asisStatus'); async function val(){const d=dni(i.value); if(d.length<8){st.className='field-help mb-2'; st.innerHTML='Esperando 8 dígitos...'; return;} i.value=d; st.className='scan-ok mb-2'; st.innerHTML='Validando DNI '+d+'...'; try{const r=await fetch('/api/trabajador/'+d); const j=await r.json(); if(j.ok){st.className='scan-ok mb-2'; st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR')+' encontrado'; beep();}else{st.className='scan-bad mb-2'; st.innerHTML='✕ '+j.msg;}}catch(e){st.className='scan-bad mb-2'; st.innerHTML='Error validando DNI';}} if(i){i.addEventListener('input',val); i.addEventListener('paste',()=>setTimeout(val,80));} if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{document.getElementById('latitud').value=p.coords.latitude;document.getElementById('longitud').value=p.coords.longitude;},()=>{});}})();</script>"""
    return render_page(body, registros=registros)

@app.route('/exportar/asistencia')
@login_required
def exportar_asistencia():
    rows = rows_to_dict(execute('SELECT fecha,hora,dni,trabajador,empresa,area,cargo,tipo,metodo,latitud,longitud,registrado_por,observacion,fecha_hora FROM asistencia ORDER BY fecha_hora DESC', fetchall=True))
    headers=['FECHA','HORA','DNI','TRABAJADOR','EMPRESA','AREA','CARGO','TIPO','METODO','LATITUD','LONGITUD','REGISTRADO_POR','OBSERVACION','FECHA_HORA']
    return excel_response(headers, rows, 'asistencia.xlsx', 'ASISTENCIA')

@app.route('/documentos-firma', methods=['GET','POST'])
@login_required
def documentos_firma():
    if request.method == 'POST':
        dni = limpiar_dni(request.form.get('dni'))
        documento = limpiar_texto(request.form.get('documento') or 'DOCUMENTO')
        metodo = limpiar_texto(request.form.get('metodo') or 'FIRMA DIGITAL')
        obs = limpiar_texto(request.form.get('observacion'), upper=False)
        if len(dni) != 8:
            flash('DNI inválido. No se registró la firma.', 'danger'); return redirect(url_for('documentos_firma'))
        t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
        if not t:
            flash('DNI no existe en trabajadores. No se registró la firma.', 'danger'); return redirect(url_for('documentos_firma'))
        firma_path = guardar_data_url(request.form.get('firma_data'), FIRMA_DIR, f"firma_{dni}")
        if not firma_path:
            flash('Debe registrar la firma en el recuadro antes de guardar.', 'danger'); return redirect(url_for('documentos_firma'))
        ahora = datetime.now(); fecha = ahora.strftime('%Y-%m-%d'); hora = ahora.strftime('%H:%M:%S')
        sql = 'INSERT INTO documentos_firma(dni,trabajador,empresa,area,cargo,documento,estado,fecha,hora,fecha_hora,metodo,firma_path,foto_path,registrado_por,observacion) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'
        execute(sql, (dni,t.get('trabajador'),t.get('empresa'),t.get('area'),t.get('cargo'),documento,'FIRMADO',fecha,hora,ahora.strftime('%Y-%m-%d %H:%M:%S'),metodo,firma_path,'',session.get('usuario'),obs), commit=True)
        flash(f'Documento firmado: {documento} - {t.get("trabajador")}', 'success')
        return redirect(url_for('documentos_firma'))
    firmas = rows_to_dict(execute('SELECT * FROM documentos_firma ORDER BY fecha_hora DESC LIMIT 80', fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad"><div class="page-card"><div class="panel-green"><a class="text-white text-decoration-none float-start" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><i class="bi bi-pen"></i><h4>FIRMA DE DOCUMENTOS</h4></div>
      <form class="floating-card" style="margin:-24px 9px 10px" method="post" id="frmFirma" onsubmit="return prepararFirma()"><label class="form-label">DNI trabajador</label><div class="input-group mb-2"><input name="dni" id="dniFirma" class="form-control" maxlength="20" required placeholder="ESCANEAR O DIGITAR"><button type="button" class="btn btn-green" onclick="abrirScanner('readerFirma','dniFirma')"><i class="bi bi-qr-code-scan"></i></button></div><div id="readerFirma" class="scan-box mb-2" style="display:none"></div><div id="firmaStatus" class="field-help mb-2">Escanee o digite DNI.</div>
        <label class="form-label">Documento</label><select name="documento" class="form-select mb-2"><option>CONTRATO DE TRABAJO</option><option>ENTREGA DE DOCUMENTOS</option><option>DECLARACIÓN JURADA</option><option>REGLAMENTO INTERNO</option><option>OTRO DOCUMENTO</option></select><label class="form-label">Método</label><select name="metodo" class="form-select mb-2"><option>FIRMA DIGITAL EN PANTALLA</option><option>HUELLA / BIOMETRÍA</option><option>QR / CÓDIGO VALIDADO</option></select>
        <label class="form-label">Firma en pantalla</label><canvas id="canvasFirma" width="340" height="150" style="width:100%;height:150px;border:2px dashed #8dbf93;border-radius:12px;background:#fff;touch-action:none"></canvas><input type="hidden" name="firma_data" id="firmaData"><div class="d-flex gap-2 mt-2"><button type="button" class="btn btn-outline-secondary w-50" onclick="limpiarFirma()">Limpiar</button><button class="btn btn-green w-50">Guardar firma</button></div><label class="form-label mt-2">Observación</label><input name="observacion" class="form-control" placeholder="Opcional"></form>
      <div class="mx-2 mb-2 d-grid"><a class="btn btn-outline-success" href="{{url_for('exportar_firmas')}}"><i class="bi bi-file-earmark-excel"></i> Exportar firmas</a></div>{% for r in firmas %}<div class="worker-card"><div class="worker-title"><div>{{r.fecha}} {{r.hora}}<br><b>{{r.trabajador}}</b></div><div class="text-end">{{r.dni}}<br><b>{{r.estado}}</b></div></div><div class="worker-grid"><div><label>DOCUMENTO</label><div class="small-value">{{r.documento}}</div></div><div><label>MÉTODO</label><div class="small-value">{{r.metodo}}</div></div><div><label>ÁREA</label><div class="small-value">{{r.area}}</div></div></div></div>{% else %}<div class="worker-card text-center text-muted">Sin documentos firmados.</div>{% endfor %}</div></div>
    <script>let firmaDibujada=false;(function(){const c=document.getElementById('canvasFirma'),ctx=c.getContext('2d');ctx.lineWidth=3;ctx.lineCap='round';let down=false,last=null;function pos(e){const r=c.getBoundingClientRect(),p=e.touches?e.touches[0]:e;return{x:(p.clientX-r.left)*(c.width/r.width),y:(p.clientY-r.top)*(c.height/r.height)}}function draw(e){if(!down)return;e.preventDefault();const p=pos(e);ctx.beginPath();ctx.moveTo(last.x,last.y);ctx.lineTo(p.x,p.y);ctx.stroke();last=p;firmaDibujada=true;}['mousedown','touchstart','pointerdown'].forEach(ev=>c.addEventListener(ev,e=>{down=true;last=pos(e);e.preventDefault();},{passive:false}));['mousemove','touchmove','pointermove'].forEach(ev=>c.addEventListener(ev,draw,{passive:false}));['mouseup','mouseleave','touchend','pointerup'].forEach(ev=>c.addEventListener(ev,()=>down=false));})();function limpiarFirma(){const c=document.getElementById('canvasFirma');c.getContext('2d').clearRect(0,0,c.width,c.height);firmaDibujada=false;}function prepararFirma(){if(!firmaDibujada){alert('Primero firme en el recuadro.');return false;}document.getElementById('firmaData').value=document.getElementById('canvasFirma').toDataURL('image/png');return true;}(function(){const i=document.getElementById('dniFirma'),st=document.getElementById('firmaStatus'),dni=v=>String(v||'').replace(/\D/g,'').slice(-8); async function val(){const d=dni(i.value); if(d.length<8){st.className='field-help mb-2';st.innerHTML='Esperando DNI...';return;} i.value=d; try{const r=await fetch('/api/trabajador/'+d); const j=await r.json(); if(j.ok){st.className='scan-ok mb-2';st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR');beep();}else{st.className='scan-bad mb-2';st.innerHTML='✕ '+j.msg;}}catch(e){st.className='scan-bad mb-2';st.innerHTML='Error validando DNI';}} if(i){i.addEventListener('input',val);i.addEventListener('paste',()=>setTimeout(val,80));}})();</script>"""
    return render_page(body, firmas=firmas)

@app.route('/exportar/firmas')
@login_required
def exportar_firmas():
    rows = rows_to_dict(execute('SELECT fecha,hora,dni,trabajador,empresa,area,cargo,documento,estado,metodo,registrado_por,observacion,fecha_hora FROM documentos_firma ORDER BY fecha_hora DESC', fetchall=True))
    headers=['FECHA','HORA','DNI','TRABAJADOR','EMPRESA','AREA','CARGO','DOCUMENTO','ESTADO','METODO','REGISTRADO_POR','OBSERVACION','FECHA_HORA']
    return excel_response(headers, rows, 'firmas_documentos.xlsx', 'FIRMAS')

# ========================= TRANSPORTE =========================
def transporte_estado_requisito(vencimiento):
    if not vencimiento:
        return 'SIN FECHA'
    try:
        dias = (datetime.strptime(vencimiento, '%Y-%m-%d').date() - date.today()).days
        if dias < 0: return 'VENCIDO'
        if dias <= 30: return 'POR VENCER'
        return 'VIGENTE'
    except Exception:
        return 'SIN FECHA'


def _valor(row, posibles):
    """Lee una columna Excel usando nombres flexibles."""
    for k in posibles:
        if k in row and row.get(k) not in (None, ''):
            return str(row.get(k)).strip()
    return ''

def _iter_excel_upload(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    headers = [normalizar_columna(c.value) for c in ws[1]]
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        if any(v not in (None, '') for v in values):
            yield row

def _pin_conductor(dni):
    d = limpiar_dni(dni)
    return d[-4:] if len(d) >= 4 else '1234'


@app.route('/transporte')
@login_required
def transporte():
    """Panel principal del Módulo Transporte.
    Vista compacta tipo app móvil: solo módulos principales aprobados por Omar.
    No modifica flujo de Tareo.
    """
    body = """
    <style>
      /* ===== UI TRANSPORTE COMPACTA - SOLO MÓDULO TRANSPORTE ===== */
      .shell{max-width:100%!important;padding:0!important;background:#f6f8f6!important;}
      .phone-wrap.trans-phone{width:100%!important;max-width:430px!important;margin:0 auto!important;padding:0 8px 18px!important;}
      .trans-app{max-width:390px;margin:8px auto 16px;background:#fff;border-radius:18px;overflow:hidden;box-shadow:0 10px 28px rgba(0,0,0,.12);border:1px solid #e8eee8;}
      .trans-hero{height:156px;background:linear-gradient(135deg,#075d2a,#137a37);color:white;position:relative;text-align:center;padding-top:16px;border-radius:18px 18px 0 0;}
      .trans-hero .back{position:absolute;left:18px;top:30px;color:white;text-decoration:none;font-size:42px;line-height:1;font-weight:300;}
      .trans-hero .config{position:absolute;right:14px;top:20px;border:1px solid rgba(255,255,255,.65);border-radius:14px;color:white;text-decoration:none;padding:8px 10px;font-weight:900;font-size:13px;display:inline-flex;align-items:center;gap:5px;background:rgba(255,255,255,.08)}
      .trans-hero .bus-icon{font-size:52px;line-height:1;margin-top:2px;display:block;}
      .trans-hero h1{font-size:19px;letter-spacing:.5px;font-weight:950;margin:8px 0 0;text-transform:uppercase;color:#fff;}
      .trans-card{background:#fff;margin-top:-16px;border-radius:20px 20px 0 0;padding:28px 18px 22px;position:relative;z-index:2;}
      .trans-section-title{color:#07642d;font-size:20px;font-weight:950;letter-spacing:.4px;text-transform:uppercase;margin:4px 0 17px;}
      .trans-section-title.oper{font-size:18px;margin-top:28px;margin-bottom:14px;}
      .trans-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;}
      .trans-grid.oper{grid-template-columns:repeat(2,1fr);gap:18px;margin-bottom:20px;}
      .trans-tile{background:linear-gradient(135deg,#07642d,#0b7a38);color:white;text-decoration:none;border-radius:13px;min-height:110px;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 9px 16px rgba(0,0,0,.15);padding:10px 7px;}
      .trans-grid.oper .trans-tile{min-height:118px;}
      .trans-tile i{font-size:38px;line-height:1;margin-bottom:12px;color:#fff;}
      .trans-grid.oper .trans-tile i{font-size:40px;margin-bottom:12px;}
      .trans-tile .label{font-size:16px;font-weight:950;line-height:1.08;color:#fff;text-shadow:0 1px 1px rgba(0,0,0,.12);}
      .trans-grid.oper .trans-tile .label{font-size:17px;}
      .trans-tile .sub{font-size:11px;font-weight:900;margin-top:5px;color:#fff;opacity:.98;line-height:1.1;}
      .trans-info{border:1px solid #b9d8ff;background:#eef6ff;border-radius:12px;color:#073b8e;font-size:14px;line-height:1.45;font-weight:900;padding:14px 14px;margin-top:18px;display:grid;grid-template-columns:26px 1fr;gap:9px;align-items:start;}
      .trans-info i{font-size:21px;color:#0b46a0;margin-top:1px;}
      @media(max-width:390px){
        .phone-wrap.trans-phone{max-width:390px!important;padding:0 6px 14px!important;}
        .trans-app{max-width:365px;margin-top:6px;border-radius:16px;}
        .trans-hero{height:146px;}
        .trans-hero h1{font-size:17px;}
        .trans-hero .bus-icon{font-size:47px;}
        .trans-hero .back{font-size:38px;left:14px;}
        .trans-hero .config{font-size:12px;right:10px;padding:7px 9px;}
        .trans-card{padding:24px 14px 20px;}
        .trans-section-title{font-size:18px;margin-bottom:14px;}
        .trans-section-title.oper{font-size:16px;margin-top:25px;}
        .trans-grid{gap:10px;}
        .trans-grid.oper{gap:14px;}
        .trans-tile{min-height:102px;border-radius:12px;}
        .trans-grid.oper .trans-tile{min-height:108px;}
        .trans-tile i{font-size:32px;margin-bottom:10px;}
        .trans-grid.oper .trans-tile i{font-size:35px;}
        .trans-tile .label{font-size:14px;}
        .trans-grid.oper .trans-tile .label{font-size:15px;}
        .trans-tile .sub{font-size:10px;}
        .trans-info{font-size:12.5px;padding:12px 11px;}
      }
    </style>
    <div class="phone-wrap trans-phone desktop-pad">
      <div class="trans-app">
        <div class="trans-hero">
          <a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>
          <a class="config" href="{{url_for('transporte_config')}}"><i class="bi bi-gear"></i> Config.</a>
          <i class="bi bi-bus-front bus-icon"></i>
          <h1>Módulo Transporte</h1>
        </div>
        <div class="trans-card">
          <div class="trans-section-title">Módulos</div>
          <div class="trans-grid">
            <a class="trans-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span class="label">Conductores</span></a>
            <a class="trans-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span class="label">Buses</span></a>
            <a class="trans-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i><span class="label">Rutas</span></a>
          </div>

          <div class="trans-section-title oper">Operación</div>
          <div class="trans-grid oper">
            <a class="trans-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-geo"></i><span class="label">GPS / Seguimiento</span><span class="sub">Ver ubicación</span></a>
            <a class="trans-tile" href="{{url_for('conductor_movil_login')}}"><i class="bi bi-phone"></i><span class="label">Móvil conductor</span><span class="sub">Abordaje y GPS</span></a>
          </div>

          <div class="trans-info"><i class="bi bi-info-circle-fill"></i><div>Así queda conectado: primero cargas conductores y buses, luego creas/cargas rutas asignando bus + conductor. El conductor entra a Móvil conductor, registra abordajes y envía GPS.</div></div>
        </div>
      </div>
    </div>
    """
    return render_page(body)


@app.route('/transporte/config')
@login_required
def transporte_config():
    """Configuración limpia. Datos maestros queda adentro, no en la portada."""
    body = _tm_css()+"""
    <style>
      .cfg-head{height:58px;background:linear-gradient(135deg,#075d2a,#2f773b);color:#fff;border-radius:12px 12px 0 0;display:flex;align-items:center;gap:10px;padding:0 14px;font-weight:900;font-size:17px}.cfg-head a{color:#fff;text-decoration:none;font-size:27px}.cfg-list{background:#fff;border:1px solid #e5e7eb;border-top:0;border-radius:0 0 12px 12px;box-shadow:0 6px 18px rgba(0,0,0,.08);overflow:hidden}.cfg-item{display:flex;align-items:center;gap:13px;padding:15px 16px;border-bottom:1px solid #edf0ed;color:#111827;text-decoration:none;font-weight:850;font-size:13px}.cfg-item:last-child{border-bottom:0}.cfg-item i{font-size:22px;color:#08713b;width:24px;text-align:center}.cfg-item .chev{margin-left:auto;color:#111;font-size:17px;width:auto}.cfg-item.highlight{background:#f0fbf2;color:#065f2a}
    </style>
    <div class="phone-wrap desktop-pad"><div class="page-card">
      <div class="cfg-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><i class="bi bi-gear"></i> CONFIGURACIONES</div>
      <div class="cfg-list">
        <a class="cfg-item" href="{{url_for('transporte')}}"><i class="bi bi-sliders"></i>Parámetros del módulo <i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg-item" href="{{url_for('transporte_pin_conductor')}}"><i class="bi bi-phone"></i>Relación conductor - móvil <i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg-item" href="{{url_for('transporte')}}"><i class="bi bi-clipboard2-check"></i>Causas de no abordaje <i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg-item" href="{{url_for('transporte')}}"><i class="bi bi-geo-alt"></i>Zonas / paraderos <i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg-item" href="{{url_for('transporte')}}"><i class="bi bi-bell"></i>Notificaciones <i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg-item" href="{{url_for('transporte')}}"><i class="bi bi-cloud-arrow-up"></i>Respaldos <i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg-item" href="{{url_for('transporte')}}"><i class="bi bi-info-circle"></i>Acerca del módulo <i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg-item highlight" href="{{url_for('transporte_datos_maestros')}}"><i class="bi bi-database"></i>Datos maestros <i class="bi bi-chevron-right chev"></i></a>
      </div>
    </div></div>"""
    return render_page(body)

@app.route('/transporte/conductores', methods=['GET','POST'])
@login_required
def transporte_conductores():
    if request.method == 'POST':
        dni=limpiar_dni(request.form.get('dni')); nombres=limpiar_texto(request.form.get('nombres'))
        if len(dni)!=8 or not nombres:
            flash('Ingrese DNI de 8 dígitos y nombres del conductor.', 'danger'); return redirect(url_for('transporte_conductores'))
        movil_pin=(request.form.get('movil_pin') or _pin_conductor(dni)).strip()
        movil_estado=limpiar_texto(request.form.get('movil_estado') or 'ACTIVO')
        params=(dni,nombres,request.form.get('telefono',''),limpiar_texto(request.form.get('licencia')),limpiar_texto(request.form.get('categoria')),request.form.get('venc_licencia',''),request.form.get('venc_cert_medico',''),request.form.get('venc_sctr',''),limpiar_texto(request.form.get('estado') or 'APTO'),limpiar_texto(request.form.get('observacion'), upper=False),dni,movil_pin,movil_estado,now_str())
        try:
            execute("""INSERT INTO transporte_conductores(dni,nombres,telefono,licencia,categoria,venc_licencia,venc_cert_medico,venc_sctr,estado,observacion,movil_usuario,movil_pin,movil_estado,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", params, commit=True)
            flash(f'Conductor registrado. Acceso móvil: DNI {dni} + PIN {movil_pin}.', 'success')
        except Exception:
            execute("""UPDATE transporte_conductores SET nombres=?,telefono=?,licencia=?,categoria=?,venc_licencia=?,venc_cert_medico=?,venc_sctr=?,estado=?,observacion=?,movil_usuario=?,movil_pin=?,movil_estado=? WHERE dni=?""", (nombres,params[2],params[3],params[4],params[5],params[6],params[7],params[8],params[9],dni,movil_pin,movil_estado,dni), commit=True)
            flash('Conductor actualizado correctamente.', 'success')
        return redirect(url_for('transporte_conductores'))
    rows=rows_to_dict(execute('SELECT * FROM transporte_conductores ORDER BY id DESC LIMIT 100', fetchall=True))
    for r in rows:
        r['lic_estado']=transporte_estado_requisito(r.get('venc_licencia'))
        r['med_estado']=transporte_estado_requisito(r.get('venc_cert_medico'))
        r['sctr_estado']=transporte_estado_requisito(r.get('venc_sctr'))
    body="""
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">CONDUCTORES</h2></div>
      <form class="floating-card mb-2" method="post"><div class="alert alert-light border small mb-2">Aquí también se crea el acceso al <b>Móvil conductor</b>. El conductor ingresa con su <b>DNI + PIN</b>, no con usuario ADMIN.</div><div class="row g-2"><div class="col-5"><label class="form-label">DNI</label><input name="dni" class="form-control" maxlength="8" required></div><div class="col-7"><label class="form-label">Conductor</label><input name="nombres" class="form-control" required></div><div class="col-6"><label class="form-label">Licencia</label><input name="licencia" class="form-control"></div><div class="col-6"><label class="form-label">Categoría</label><input name="categoria" class="form-control"></div><div class="col-4"><label class="form-label">Venc. Lic.</label><input name="venc_licencia" type="date" class="form-control"></div><div class="col-4"><label class="form-label">Cert. Médico</label><input name="venc_cert_medico" type="date" class="form-control"></div><div class="col-4"><label class="form-label">SCTR</label><input name="venc_sctr" type="date" class="form-control"></div><div class="col-6"><label class="form-label">Teléfono</label><input name="telefono" class="form-control"></div><div class="col-6"><label class="form-label">Estado</label><select name="estado" class="form-select"><option>APTO</option><option>OBSERVADO</option><option>VENCIDO</option></select></div><div class="col-6"><label class="form-label">PIN móvil</label><input name="movil_pin" class="form-control" placeholder="Por defecto últimos 4 DNI"></div><div class="col-6"><label class="form-label">Acceso móvil</label><select name="movil_estado" class="form-select"><option>ACTIVO</option><option>BLOQUEADO</option></select></div><div class="col-12"><input name="observacion" class="form-control" placeholder="Observación"></div></div><button class="btn btn-green w-100 mt-2">Guardar conductor / PIN</button></form>
      {% for c in rows %}<div class="worker-card"><div class="worker-title"><div>{{c.dni}}<br><b>{{c.nombres}}</b></div><div class="text-end">{{c.categoria or ''}}<br><b>{{c.estado}}</b></div></div><div class="worker-grid"><div><label>LICENCIA</label><div class="small-value">{{c.lic_estado}}</div></div><div><label>MÉDICO</label><div class="small-value">{{c.med_estado}}</div></div><div><label>SCTR</label><div class="small-value">{{c.sctr_estado}}</div></div></div><div class="worker-grid mt-2"><div><label>USUARIO MÓVIL</label><div class="small-value">{{c.movil_usuario or c.dni}}</div></div><div><label>PIN</label><div class="small-value">{{c.movil_pin or 'SIN PIN'}}</div></div><div><label>MÓVIL</label><div class="small-value">{{c.movil_estado or 'ACTIVO'}}</div></div></div><a class="btn btn-outline-success btn-sm w-100 mt-2" href="{{url_for('transporte_reset_pin', conductor_id=c.id)}}" onclick="return confirm('¿Resetear PIN a los últimos 4 dígitos del DNI?')"><i class="bi bi-key"></i> Resetear PIN</a></div>{% else %}<div class="worker-card text-center text-muted">Sin conductores.</div>{% endfor %}
    </div>"""
    return render_page(body, rows=rows)

@app.route('/transporte/conductor/<int:conductor_id>/reset-pin')
@login_required
def transporte_reset_pin(conductor_id):
    c = row_to_dict(execute('SELECT * FROM transporte_conductores WHERE id=?', (conductor_id,), fetchone=True))
    if not c:
        flash('Conductor no encontrado.', 'danger'); return redirect(url_for('transporte_conductores'))
    pin = _pin_conductor(c.get('dni'))
    execute("UPDATE transporte_conductores SET movil_usuario=?, movil_pin=?, movil_estado='ACTIVO' WHERE id=?", (c.get('dni'), pin, conductor_id), commit=True)
    flash(f'PIN móvil reseteado para {c.get("nombres")}: {pin}', 'success')
    return redirect(url_for('transporte_conductores'))

@app.route('/transporte/vehiculos', methods=['GET','POST'])
@login_required
def transporte_vehiculos():
    if request.method == 'POST':
        placa=limpiar_texto(request.form.get('placa'))
        if not placa:
            flash('Ingrese placa del vehículo.', 'danger'); return redirect(url_for('transporte_vehiculos'))
        try: capacidad=int(request.form.get('capacidad') or 0)
        except Exception: capacidad=0
        params=(placa,limpiar_texto(request.form.get('tipo') or 'BUS'),capacidad,limpiar_texto(request.form.get('empresa_transportista')),request.form.get('soat_venc',''),request.form.get('revision_tecnica_venc',''),limpiar_texto(request.form.get('gps_codigo')),limpiar_texto(request.form.get('estado') or 'ACTIVO'),limpiar_texto(request.form.get('observacion'), upper=False),now_str())
        try:
            execute("""INSERT INTO transporte_vehiculos(placa,tipo,capacidad,empresa_transportista,soat_venc,revision_tecnica_venc,gps_codigo,estado,observacion,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?)""", params, commit=True)
            flash('Vehículo registrado correctamente.', 'success')
        except Exception:
            execute("""UPDATE transporte_vehiculos SET tipo=?,capacidad=?,empresa_transportista=?,soat_venc=?,revision_tecnica_venc=?,gps_codigo=?,estado=?,observacion=? WHERE placa=?""", (params[1],params[2],params[3],params[4],params[5],params[6],params[7],params[8],placa), commit=True)
            flash('Vehículo actualizado correctamente.', 'success')
        return redirect(url_for('transporte_vehiculos'))
    rows=rows_to_dict(execute('SELECT * FROM transporte_vehiculos ORDER BY id DESC LIMIT 100', fetchall=True))
    for r in rows:
        r['soat_estado']=transporte_estado_requisito(r.get('soat_venc'))
        r['rev_estado']=transporte_estado_requisito(r.get('revision_tecnica_venc'))
    body="""
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">BUSES / MINIBUSES</h2></div>
      <form class="floating-card mb-2" method="post"><div class="row g-2"><div class="col-5"><label class="form-label">Placa</label><input name="placa" class="form-control" required></div><div class="col-7"><label class="form-label">Tipo</label><select name="tipo" class="form-select"><option>BUS</option><option>MINIBUS</option><option>VAN</option><option>CAMIONETA</option></select></div><div class="col-6"><label class="form-label">Capacidad</label><input name="capacidad" type="number" class="form-control" value="30"></div><div class="col-6"><label class="form-label">Transportista</label><input name="empresa_transportista" class="form-control"></div><div class="col-6"><label class="form-label">SOAT vence</label><input name="soat_venc" type="date" class="form-control"></div><div class="col-6"><label class="form-label">Rev. técnica</label><input name="revision_tecnica_venc" type="date" class="form-control"></div><div class="col-6"><label class="form-label">Código GPS</label><input name="gps_codigo" class="form-control"></div><div class="col-6"><label class="form-label">Estado</label><select name="estado" class="form-select"><option>ACTIVO</option><option>OBSERVADO</option><option>INACTIVO</option></select></div><div class="col-12"><input name="observacion" class="form-control" placeholder="Observación"></div></div><button class="btn btn-green w-100 mt-2">Guardar vehículo</button></form>
      {% for v in rows %}<div class="worker-card"><div class="worker-title"><div>{{v.placa}}<br><b>{{v.tipo}}</b></div><div class="text-end">CAP. {{v.capacidad}}<br><b>{{v.estado}}</b></div></div><div class="worker-grid"><div><label>SOAT</label><div class="small-value">{{v.soat_estado}}</div></div><div><label>REV. TÉCNICA</label><div class="small-value">{{v.rev_estado}}</div></div><div><label>GPS</label><div class="small-value">{{v.gps_codigo or '-'}}</div></div></div></div>{% else %}<div class="worker-card text-center text-muted">Sin vehículos.</div>{% endfor %}
    </div>"""
    return render_page(body, rows=rows)

@app.route('/transporte/rutas', methods=['GET','POST'])
@login_required
def transporte_rutas():
    if request.method == 'POST':
        fecha=request.form.get('fecha') or today_str(); nombre=limpiar_texto(request.form.get('nombre') or 'RUTA')
        vehiculo_id=request.form.get('vehiculo_id') or None; conductor_id=request.form.get('conductor_id') or None
        execute("""INSERT INTO transporte_rutas(fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,vehiculo_id,conductor_id,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", (fecha,nombre,limpiar_texto(request.form.get('origen')),limpiar_texto(request.form.get('destino')),limpiar_texto(request.form.get('sede')),request.form.get('hora_salida',''),request.form.get('hora_retorno',''),vehiculo_id,conductor_id,limpiar_texto(request.form.get('estado') or 'PROGRAMADA'),session.get('usuario'),now_str()), commit=True)
        flash('Ruta programada correctamente.', 'success'); return redirect(url_for('transporte_rutas'))
    vehiculos=rows_to_dict(execute("SELECT id, placa, tipo, capacidad FROM transporte_vehiculos WHERE COALESCE(estado,'ACTIVO')<>'INACTIVO' ORDER BY placa", fetchall=True))
    conductores=rows_to_dict(execute("SELECT id, dni, nombres, estado FROM transporte_conductores WHERE COALESCE(estado,'APTO')<>'INACTIVO' ORDER BY nombres", fetchall=True))
    rutas=rows_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id ORDER BY r.fecha DESC, r.id DESC LIMIT 80""", fetchall=True))
    body="""
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">RUTAS</h2></div>
      <form class="floating-card mb-2" method="post"><div class="row g-2"><div class="col-6"><label class="form-label">Fecha</label><input name="fecha" type="date" value="{{hoy}}" class="form-control"></div><div class="col-6"><label class="form-label">Ruta</label><input name="nombre" class="form-control" placeholder="Ej. RUTA 01"></div><div class="col-6"><label class="form-label">Origen</label><input name="origen" class="form-control"></div><div class="col-6"><label class="form-label">Destino</label><input name="destino" class="form-control"></div><div class="col-6"><label class="form-label">Sede/Fundo</label><input name="sede" class="form-control"></div><div class="col-3"><label class="form-label">Salida</label><input name="hora_salida" type="time" class="form-control"></div><div class="col-3"><label class="form-label">Retorno</label><input name="hora_retorno" type="time" class="form-control"></div><div class="col-6"><label class="form-label">Vehículo</label><select name="vehiculo_id" class="form-select"><option value="">Seleccionar</option>{% for v in vehiculos %}<option value="{{v.id}}">{{v.placa}} - {{v.tipo}} ({{v.capacidad}})</option>{% endfor %}</select></div><div class="col-6"><label class="form-label">Conductor</label><select name="conductor_id" class="form-select"><option value="">Seleccionar</option>{% for c in conductores %}<option value="{{c.id}}">{{c.nombres}}</option>{% endfor %}</select></div><div class="col-12"><select name="estado" class="form-select"><option>PROGRAMADA</option><option>EN RUTA</option><option>LLEGÓ</option><option>RETORNO</option><option>CERRADA</option></select></div></div><button class="btn btn-green w-100 mt-2">Crear ruta</button></form>
      {% for r in rutas %}<div class="worker-card"><div class="worker-title"><div>{{r.fecha}} {{r.hora_salida or ''}}<br><b>{{r.nombre}}</b></div><div class="text-end">{{r.placa or 'SIN BUS'}}<br><b>{{r.estado}}</b></div></div><div class="worker-grid"><div><label>CONDUCTOR</label><div class="small-value">{{r.conductor or '-'}}</div></div><div><label>ORIGEN</label><div class="small-value">{{r.origen or '-'}}</div></div><div><label>DESTINO</label><div class="small-value">{{r.destino or '-'}}</div></div></div><a class="btn btn-green btn-sm w-100 mt-2" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}">Abrir ruta</a></div>{% else %}<div class="worker-card text-center text-muted">Sin rutas.</div>{% endfor %}
    </div>"""
    return render_page(body, vehiculos=vehiculos, conductores=conductores, rutas=rutas, hoy=today_str())

@app.route('/transporte/ruta/<int:ruta_id>')
@login_required
def transporte_ruta_detalle(ruta_id):
    ruta=row_to_dict(execute("""SELECT r.*, v.placa, v.tipo, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id WHERE r.id=?""", (ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.', 'danger'); return redirect(url_for('transporte'))
    pasajeros=rows_to_dict(execute('SELECT * FROM transporte_pasajeros WHERE ruta_id=? ORDER BY fecha_hora DESC', (ruta_id,), fetchall=True))
    ocupados=len(pasajeros); capacidad=int(ruta.get('capacidad') or 0); libres=max(0, capacidad-ocupados) if capacidad else 0
    body="""
    <div class="phone-wrap desktop-pad"><div class="page-card"><div class="panel-green"><a class="text-white text-decoration-none float-start" href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><i class="bi bi-bus-front"></i><h4>{{ruta.nombre}} - {{ruta.placa or 'SIN BUS'}}</h4></div>
      <form class="floating-card" style="margin:-24px 9px 10px" method="post" action="{{url_for('transporte_abordar', ruta_id=ruta.id)}}" id="frmAbordar"><div class="alert alert-light border small mb-2"><b>{{ruta.origen}}</b> → <b>{{ruta.destino}}</b><br>Conductor: {{ruta.conductor or '-'}} | Capacidad: {{capacidad or 'SIN DEFINIR'}} | Ocupados: {{ocupados}} | Libres: {{libres}}</div>
        <label class="form-label">DNI / QR / Código de barras / Huella</label><div class="input-group mb-2"><input name="dni" id="dniTransporte" class="form-control" maxlength="20" placeholder="ESCANEAR O DIGITAR" required autofocus><button type="button" class="btn btn-green" onclick="abrirScanner('readerTransporte','dniTransporte')"><i class="bi bi-camera"></i></button></div><div id="readerTransporte" class="scan-box mb-2" style="display:none"></div><div id="transpStatus" class="field-help mb-2">Al completar 8 dígitos se validará y registrará el abordaje.</div>
        <div class="row g-2"><div class="col-6"><label class="form-label">Método</label><select name="metodo" class="form-select"><option>QR</option><option>CODIGO DE BARRAS</option><option>HUELLA</option><option>DNI DIGITADO</option></select></div><div class="col-6"><label class="form-label">Observación</label><input name="observacion" class="form-control"></div></div><input type="hidden" name="latitud" id="latitudTrans"><input type="hidden" name="longitud" id="longitudTrans"><button class="btn btn-green w-100 mt-2"><i class="bi bi-person-check"></i> REGISTRAR SUBIDA</button>
      </form><div class="mx-2 mb-2 d-grid gap-2"><button class="btn btn-outline-success" onclick="enviarGpsTransporte({{ruta.id}})"><i class="bi bi-geo-alt"></i> Actualizar GPS ruta</button><a class="btn btn-outline-success" href="{{url_for('exportar_transporte_pasajeros', ruta_id=ruta.id)}}"><i class="bi bi-file-earmark-excel"></i> Exportar ruta</a></div>
      {% for p in pasajeros %}<div class="worker-card"><div class="worker-title"><div>{{p.hora}}<br><b>{{p.trabajador}}</b></div><div class="text-end">{{p.dni}}<br><b>{{p.metodo}}</b></div></div><div class="worker-grid"><div><label>EMPRESA</label><div class="small-value">{{p.empresa or '-'}}</div></div><div><label>ÁREA</label><div class="small-value">{{p.area or '-'}}</div></div><div><label>CARGO</label><div class="small-value">{{p.cargo or '-'}}</div></div></div></div>{% else %}<div class="worker-card text-center text-muted">Aún no hay trabajadores abordados.</div>{% endfor %}
    </div></div><script>(function(){const dni=v=>String(v||'').replace(/\D/g,'').slice(-8), i=document.getElementById('dniTransporte'), st=document.getElementById('transpStatus'); async function val(){const d=dni(i.value); if(d.length<8){st.className='field-help mb-2'; st.innerHTML='Esperando 8 dígitos...'; return;} i.value=d; st.className='scan-ok mb-2'; st.innerHTML='Validando DNI '+d+'...'; try{const r=await fetch('/api/trabajador/'+d); const j=await r.json(); if(j.ok){st.className='scan-ok mb-2'; st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR')+' encontrado'; beep();}else{st.className='scan-bad mb-2'; st.innerHTML='✕ '+j.msg;}}catch(e){st.className='scan-bad mb-2'; st.innerHTML='Error validando DNI';}} if(i){i.addEventListener('input',val); i.addEventListener('paste',()=>setTimeout(val,80));} if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{document.getElementById('latitudTrans').value=p.coords.latitude;document.getElementById('longitudTrans').value=p.coords.longitude;},()=>{});}})(); async function enviarGpsTransporte(rid){if(!navigator.geolocation){alert('GPS no disponible');return;} navigator.geolocation.getCurrentPosition(async p=>{let fd=new FormData();fd.append('latitud',p.coords.latitude);fd.append('longitud',p.coords.longitude);let r=await fetch('/transporte/ruta/'+rid+'/gps',{method:'POST',body:fd});let j=await r.json();alert(j.msg||'GPS actualizado');location.reload();},()=>alert('Permite ubicación/GPS en el navegador'));}</script>"""
    return render_page(body, ruta=ruta, pasajeros=pasajeros, ocupados=ocupados, capacidad=capacidad, libres=libres)

@app.route('/transporte/ruta/<int:ruta_id>/abordar', methods=['POST'])
@login_required
def transporte_abordar(ruta_id):
    ruta=row_to_dict(execute('SELECT r.*, v.capacidad FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id WHERE r.id=?', (ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.', 'danger'); return redirect(url_for('transporte'))
    dni=limpiar_dni(request.form.get('dni'))
    if len(dni)!=8:
        flash('DNI inválido. No se registró subida.', 'danger'); return redirect(url_for('transporte_ruta_detalle', ruta_id=ruta_id))
    t=row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    if not t:
        flash('DNI no existe en base trabajadores. No se registró subida.', 'danger'); return redirect(url_for('transporte_ruta_detalle', ruta_id=ruta_id))
    dup=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE ruta_id=? AND dni=?', (ruta_id,dni))
    if dup:
        flash('Este trabajador ya fue registrado en esta ruta.', 'danger'); return redirect(url_for('transporte_ruta_detalle', ruta_id=ruta_id))
    capacidad=int(ruta.get('capacidad') or 0); ocupados=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE ruta_id=?', (ruta_id,))
    if capacidad and ocupados >= capacidad:
        flash('Capacidad completa del vehículo. No se registró subida.', 'danger'); return redirect(url_for('transporte_ruta_detalle', ruta_id=ruta_id))
    ahora=datetime.now(); fecha=ahora.strftime('%Y-%m-%d'); hora=ahora.strftime('%H:%M:%S')
    execute("""INSERT INTO transporte_pasajeros(ruta_id,dni,trabajador,empresa,area,cargo,fecha,hora,fecha_hora,metodo,latitud,longitud,registrado_por,observacion) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (ruta_id,dni,t.get('trabajador'),t.get('empresa'),t.get('area'),t.get('cargo'),fecha,hora,ahora.strftime('%Y-%m-%d %H:%M:%S'),limpiar_texto(request.form.get('metodo') or 'QR'),request.form.get('latitud',''),request.form.get('longitud',''),session.get('usuario'),limpiar_texto(request.form.get('observacion'), upper=False)), commit=True)
    flash(f'Subida registrada: {dni} - {t.get("trabajador")}', 'success')
    return redirect(url_for('transporte_ruta_detalle', ruta_id=ruta_id))

@app.route('/transporte/ruta/<int:ruta_id>/gps', methods=['POST'])
@login_required
def transporte_gps_actualizar(ruta_id):
    lat=request.form.get('latitud',''); lon=request.form.get('longitud',''); fh=now_str()
    ruta = row_to_dict(execute('SELECT r.*, v.placa FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id WHERE r.id=?', (ruta_id,), fetchone=True)) or {}
    conductor_id = ruta.get('conductor_id') or session.get('conductor_id')
    execute('INSERT INTO transporte_gps(ruta_id,latitud,longitud,fecha_hora,registrado_por,conductor_id,placa,ruta_nombre) VALUES(?,?,?,?,?,?,?,?)', (ruta_id,lat,lon,fh,session.get('usuario') or session.get('conductor_nombre') or 'CONDUCTOR MOVIL',conductor_id,ruta.get('placa'),ruta.get('nombre')), commit=True)
    execute('UPDATE transporte_rutas SET latitud=?, longitud=?, ultima_ubicacion=?, estado=? WHERE id=?', (lat,lon,fh,'EN RUTA',ruta_id), commit=True)
    if conductor_id:
        execute('UPDATE transporte_conductores SET ultima_latitud=?, ultima_longitud=?, ultimo_gps=? WHERE id=?', (lat,lon,fh,conductor_id), commit=True)
    return jsonify(ok=True, msg='GPS actualizado correctamente')

@app.route('/exportar/transporte/pasajeros')
@login_required
def exportar_transporte_pasajeros():
    ruta_id=request.args.get('ruta_id')
    params=[]; where=''
    if ruta_id:
        where='WHERE p.ruta_id=?'; params=[ruta_id]
    rows=rows_to_dict(execute(f"""SELECT r.fecha AS fecha_ruta, r.nombre AS ruta, r.origen, r.destino, v.placa, c.nombres AS conductor, p.fecha, p.hora, p.dni, p.trabajador, p.empresa, p.area, p.cargo, p.metodo, p.latitud, p.longitud, p.registrado_por, p.observacion, p.fecha_hora FROM transporte_pasajeros p LEFT JOIN transporte_rutas r ON r.id=p.ruta_id LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id {where} ORDER BY p.fecha_hora DESC""", params, fetchall=True))
    headers=['FECHA_RUTA','RUTA','ORIGEN','DESTINO','PLACA','CONDUCTOR','FECHA','HORA','DNI','TRABAJADOR','EMPRESA','AREA','CARGO','METODO','LATITUD','LONGITUD','REGISTRADO_POR','OBSERVACION','FECHA_HORA']
    return excel_response(headers, rows, 'transporte_abordajes.xlsx', 'ABORDAJES')


@app.route('/transporte/carga-masiva', methods=['GET','POST'])
@login_required
def transporte_carga_masiva():
    resumen = None
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        f = request.files.get('archivo')
        if not f or not f.filename.lower().endswith(('.xlsx','.xlsm')):
            flash('Adjunta un Excel .xlsx para cargar.', 'danger')
            return redirect(url_for('transporte_carga_masiva'))
        ok = err = 0
        errores = []
        for n, row in enumerate(_iter_excel_upload(f), start=2):
            try:
                if tipo == 'conductores':
                    dni = limpiar_dni(_valor(row, ['DNI','DOCUMENTO','DOC']))
                    nombres = limpiar_texto(_valor(row, ['NOMBRES','CONDUCTOR','APELLIDOS Y NOMBRES','TRABAJADOR']))
                    if len(dni) != 8 or not nombres:
                        raise ValueError('DNI/NOMBRES obligatorio')
                    telefono = _valor(row, ['TELEFONO','CELULAR','MOVIL'])
                    licencia = limpiar_texto(_valor(row, ['LICENCIA','BREVETE']))
                    categoria = limpiar_texto(_valor(row, ['CATEGORIA','CLASE']))
                    venc_lic = _valor(row, ['VENC_LICENCIA','VENCIMIENTO LICENCIA','VENCE LICENCIA'])
                    venc_med = _valor(row, ['VENC_CERT_MEDICO','CERTIFICADO MEDICO','VENCE CERT MEDICO'])
                    venc_sctr = _valor(row, ['VENC_SCTR','SCTR','VENCE SCTR'])
                    estado = limpiar_texto(_valor(row, ['ESTADO']) or 'APTO')
                    pin = _valor(row, ['PIN','CLAVE','MOVIL_PIN']) or _pin_conductor(dni)
                    try:
                        execute("""INSERT INTO transporte_conductores(dni,nombres,telefono,licencia,categoria,venc_licencia,venc_cert_medico,venc_sctr,estado,observacion,movil_usuario,movil_pin,movil_estado,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (dni,nombres,telefono,licencia,categoria,venc_lic,venc_med,venc_sctr,estado,'CARGA MASIVA',dni,pin,'ACTIVO',now_str()), commit=True)
                    except Exception:
                        execute("""UPDATE transporte_conductores SET nombres=?,telefono=?,licencia=?,categoria=?,venc_licencia=?,venc_cert_medico=?,venc_sctr=?,estado=?,movil_usuario=?,movil_pin=?,movil_estado=? WHERE dni=?""", (nombres,telefono,licencia,categoria,venc_lic,venc_med,venc_sctr,estado,dni,pin,'ACTIVO',dni), commit=True)
                    ok += 1
                elif tipo == 'vehiculos':
                    placa = limpiar_texto(_valor(row, ['PLACA','BUS','VEHICULO']))
                    if not placa:
                        raise ValueError('PLACA obligatoria')
                    tipo_v = limpiar_texto(_valor(row, ['TIPO','CLASE']) or 'BUS')
                    capacidad = int(float(_valor(row, ['CAPACIDAD','ASIENTOS']) or 0))
                    empresa = limpiar_texto(_valor(row, ['EMPRESA','TRANSPORTISTA','EMPRESA TRANSPORTISTA']))
                    soat = _valor(row, ['SOAT_VENC','VENCE SOAT','SOAT'])
                    rev = _valor(row, ['REVISION_TECNICA_VENC','REVISION TECNICA','VENCE REVISION'])
                    gps = limpiar_texto(_valor(row, ['GPS','GPS_CODIGO','CODIGO GPS']))
                    estado = limpiar_texto(_valor(row, ['ESTADO']) or 'ACTIVO')
                    try:
                        execute("""INSERT INTO transporte_vehiculos(placa,tipo,capacidad,empresa_transportista,soat_venc,revision_tecnica_venc,gps_codigo,estado,observacion,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?)""", (placa,tipo_v,capacidad,empresa,soat,rev,gps,estado,'CARGA MASIVA',now_str()), commit=True)
                    except Exception:
                        execute("""UPDATE transporte_vehiculos SET tipo=?,capacidad=?,empresa_transportista=?,soat_venc=?,revision_tecnica_venc=?,gps_codigo=?,estado=? WHERE placa=?""", (tipo_v,capacidad,empresa,soat,rev,gps,estado,placa), commit=True)
                    ok += 1
                else:
                    raise ValueError('Tipo de carga inválido')
            except Exception as e:
                err += 1
                if len(errores) < 8:
                    errores.append(f'Fila {n}: {e}')
        resumen = {'ok': ok, 'err': err, 'errores': errores}
        flash(f'Carga finalizada: {ok} registros OK, {err} observados.', 'success' if err == 0 else 'warning')
    body = """
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">CARGA MASIVA TRANSPORTE</h2></div>
      <form class="floating-card" method="post" enctype="multipart/form-data">
        <label class="form-label">Tipo de carga</label><select name="tipo" class="form-select mb-2"><option value="conductores">Conductores</option><option value="vehiculos">Buses / minibuses / vans</option></select>
        <label class="form-label">Archivo Excel .xlsx</label><input type="file" name="archivo" class="form-control mb-2" accept=".xlsx,.xlsm" required>
        <div class="field-help mb-2"><b>Conductores:</b> DNI, NOMBRES, TELEFONO, LICENCIA, CATEGORIA, VENC_LICENCIA, VENC_CERT_MEDICO, VENC_SCTR, PIN.<br><b>Buses:</b> PLACA, TIPO, CAPACIDAD, EMPRESA, SOAT_VENC, REVISION_TECNICA_VENC, GPS.</div>
        <button class="btn btn-green w-100"><i class="bi bi-upload"></i> CARGAR EXCEL</button>
      </form>
      {% if resumen %}<div class="worker-card"><b>Resultado</b><br>OK: {{resumen.ok}} | Observados: {{resumen.err}}{% for e in resumen.errores %}<div class="scan-bad mt-1">{{e}}</div>{% endfor %}</div>{% endif %}
    </div>"""
    return render_page(body, resumen=resumen)

@app.route('/movil/conductor', methods=['GET','POST'])
def conductor_movil_login():
    if request.method == 'POST':
        dni = limpiar_dni(request.form.get('dni'))
        pin = (request.form.get('pin') or '').strip()
        c = row_to_dict(execute("SELECT * FROM transporte_conductores WHERE dni=? AND movil_pin=? AND COALESCE(movil_estado,'ACTIVO')='ACTIVO'", (dni,pin), fetchone=True))
        if not c:
            flash('DNI o PIN móvil incorrecto.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        session['conductor_id'] = c['id']; session['conductor_nombre'] = c.get('nombres')
        return redirect(url_for('conductor_movil_panel'))
    body = """
    <div class="phone-wrap"><div class="page-card"><div class="panel-green"><a class="text-white text-decoration-none float-start" href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><i class="bi bi-phone"></i><h4>ACCESO MÓVIL CONDUCTOR</h4></div>
      <form class="floating-card" style="margin:-24px 9px 10px" method="post"><div class="alert alert-light border small mb-2"><b>Solo conductores:</b> ingrese DNI del conductor + PIN móvil creado en Transporte &gt; Conductores. No usar ADMIN.</div><label class="form-label">DNI conductor</label><input name="dni" class="form-control mb-2" maxlength="8" inputmode="numeric" required autofocus placeholder="Ej. 12345678"><label class="form-label">PIN móvil</label><input name="pin" class="form-control mb-2" type="password" required placeholder="Últimos 4 DNI o PIN asignado"><button class="btn btn-green w-100">INGRESAR</button><a class="btn btn-outline-success w-100 mt-2" href="{{url_for('transporte_conductores')}}">Crear / resetear PIN</a><a class="btn btn-outline-secondary w-100 mt-2" href="{{url_for('transporte')}}">Volver</a></form>
    </div></div>"""
    return render_page(body)

@app.route('/movil/conductor/panel')
def conductor_movil_panel():
    cid = session.get('conductor_id')
    if not cid:
        return redirect(url_for('conductor_movil_login'))
    rutas = rows_to_dict(execute("""SELECT r.*, v.placa, v.capacidad FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id WHERE r.conductor_id=? AND r.fecha>=? ORDER BY r.fecha, r.hora_salida LIMIT 20""", (cid,today_str()), fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad"><h2 class="header-title">MIS RUTAS MÓVIL</h2>
      {% for r in rutas %}<div class="worker-card"><div class="worker-title"><div>{{r.fecha}} {{r.hora_salida or ''}}<br><b>{{r.nombre}}</b></div><div class="text-end">{{r.placa or 'SIN BUS'}}<br><b>{{r.estado}}</b></div></div><a class="btn btn-green btn-sm w-100 mt-2" href="{{url_for('conductor_movil_ruta', ruta_id=r.id)}}"><i class="bi bi-qr-code-scan"></i> Abrir / registrar subida</a></div>{% else %}<div class="worker-card text-center text-muted">No tienes rutas asignadas desde hoy.</div>{% endfor %}
      <a class="btn btn-outline-danger w-100 mt-2" href="{{url_for('conductor_movil_logout')}}">Cerrar móvil</a>
    </div>"""
    return render_page(body, rutas=rutas)

@app.route('/movil/conductor/ruta/<int:ruta_id>')
def conductor_movil_ruta(ruta_id):
    cid = session.get('conductor_id')
    if not cid:
        return redirect(url_for('conductor_movil_login'))
    ruta=row_to_dict(execute("""SELECT r.*, v.placa, v.tipo, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id WHERE r.id=? AND r.conductor_id=?""", (ruta_id,cid), fetchone=True))
    if not ruta:
        flash('Ruta no asignada a este conductor.', 'danger'); return redirect(url_for('conductor_movil_panel'))
    pasajeros=rows_to_dict(execute('SELECT * FROM transporte_pasajeros WHERE ruta_id=? ORDER BY fecha_hora DESC', (ruta_id,), fetchall=True))
    ocupados=len(pasajeros); capacidad=int(ruta.get('capacidad') or 0); libres=max(0, capacidad-ocupados) if capacidad else 0
    body="""
    <div class="phone-wrap desktop-pad"><div class="page-card"><div class="panel-green"><a class="text-white text-decoration-none float-start" href="{{url_for('conductor_movil_panel')}}"><i class="bi bi-chevron-left"></i></a><i class="bi bi-bus-front"></i><h4>{{ruta.nombre}} - {{ruta.placa or 'SIN BUS'}}</h4></div>
      <form class="floating-card" style="margin:-24px 9px 10px" method="post" action="{{url_for('transporte_abordar', ruta_id=ruta.id)}}" id="frmAbordarMovil"><div class="alert alert-light border small mb-2"><b>{{ruta.origen}}</b> → <b>{{ruta.destino}}</b><br>Capacidad: {{capacidad or 'SIN DEFINIR'}} | Ocupados: {{ocupados}} | Libres: {{libres}}</div>
        <label class="form-label">DNI / QR / Código de barras</label><div class="input-group mb-2"><input name="dni" id="dniTransporteMovil" class="form-control" maxlength="20" placeholder="ESCANEAR O DIGITAR" required autofocus><button type="button" class="btn btn-green" onclick="abrirScanner('readerTransporteMovil','dniTransporteMovil')"><i class="bi bi-camera"></i></button></div><div id="readerTransporteMovil" class="scan-box mb-2" style="display:none"></div><div id="transpStatusMovil" class="field-help mb-2">Completa 8 dígitos y presiona registrar.</div>
        <select name="metodo" class="form-select mb-2"><option>QR</option><option>CODIGO DE BARRAS</option><option>DNI DIGITADO</option></select><input type="hidden" name="latitud" id="latitudTransMovil"><input type="hidden" name="longitud" id="longitudTransMovil"><button class="btn btn-green w-100"><i class="bi bi-person-check"></i> REGISTRAR SUBIDA</button>
      </form><div class="mx-2 mb-2 d-grid gap-2"><button class="btn btn-outline-success" onclick="enviarGpsTransporteMovil({{ruta.id}})"><i class="bi bi-geo-alt"></i> Enviar GPS conductor</button></div>
      {% for p in pasajeros %}<div class="worker-card"><div class="worker-title"><div>{{p.hora}}<br><b>{{p.trabajador}}</b></div><div class="text-end">{{p.dni}}<br><b>{{p.metodo}}</b></div></div></div>{% else %}<div class="worker-card text-center text-muted">Sin abordajes.</div>{% endfor %}
    </div></div><script>(function(){const dni=v=>String(v||'').replace(/\D/g,'').slice(-8), i=document.getElementById('dniTransporteMovil'), st=document.getElementById('transpStatusMovil'); async function val(){const d=dni(i.value); if(d.length<8){st.className='field-help mb-2'; st.innerHTML='Esperando 8 dígitos...'; return;} i.value=d; try{const r=await fetch('/api/trabajador/'+d); const j=await r.json(); if(j.ok){st.className='scan-ok mb-2'; st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR'); beep();}else{st.className='scan-bad mb-2'; st.innerHTML='✕ '+j.msg;}}catch(e){st.className='scan-bad mb-2'; st.innerHTML='Error validando DNI';}} if(i){i.addEventListener('input',val); i.addEventListener('paste',()=>setTimeout(val,80)); i.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();document.getElementById('frmAbordarMovil').requestSubmit();}});} if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{document.getElementById('latitudTransMovil').value=p.coords.latitude;document.getElementById('longitudTransMovil').value=p.coords.longitude;},()=>{});}})(); async function enviarGpsTransporteMovil(rid){if(!navigator.geolocation){alert('GPS no disponible');return;} navigator.geolocation.getCurrentPosition(async p=>{document.getElementById('latitudTransMovil').value=p.coords.latitude;document.getElementById('longitudTransMovil').value=p.coords.longitude;let fd=new FormData();fd.append('latitud',p.coords.latitude);fd.append('longitud',p.coords.longitude);let r=await fetch('/transporte/ruta/'+rid+'/gps',{method:'POST',body:fd});let j=await r.json();alert(j.msg||'GPS actualizado');location.reload();},()=>alert('Permite ubicación/GPS en el navegador'));}</script>"""
    return render_page(body, ruta=ruta, pasajeros=pasajeros, ocupados=ocupados, capacidad=capacidad, libres=libres)

@app.route('/movil/conductor/logout')
def conductor_movil_logout():
    session.pop('conductor_id', None); session.pop('conductor_nombre', None)
    return redirect(url_for('conductor_movil_login'))


@app.route('/transporte/ruta/<int:ruta_id>/esperados', methods=['GET','POST'])
@login_required
def transporte_esperados(ruta_id):
    ruta = row_to_dict(execute('SELECT * FROM transporte_rutas WHERE id=?', (ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.', 'danger'); return redirect(url_for('transporte'))
    resumen = None
    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f or not f.filename.lower().endswith(('.xlsx','.xlsm')):
            flash('Adjunta un Excel .xlsx con DNI de trabajadores esperados.', 'danger')
            return redirect(url_for('transporte_esperados', ruta_id=ruta_id))
        ok = err = 0; errores = []
        for n, row in enumerate(_iter_excel_upload(f), start=2):
            try:
                dni = limpiar_dni(_valor(row, ['DNI','DOCUMENTO','DOC','CODIGO']))
                if len(dni) != 8: raise ValueError('DNI inválido')
                t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
                if not t: raise ValueError('DNI no existe en trabajadores')
                existe = scalar('SELECT COUNT(*) AS c FROM transporte_ruta_esperados WHERE ruta_id=? AND dni=?', (ruta_id,dni))
                if existe:
                    execute('UPDATE transporte_ruta_esperados SET trabajador=?,empresa=?,area=?,cargo=? WHERE ruta_id=? AND dni=?', (t.get('trabajador'),t.get('empresa'),t.get('area'),t.get('cargo'),ruta_id,dni), commit=True)
                else:
                    execute('INSERT INTO transporte_ruta_esperados(ruta_id,dni,trabajador,empresa,area,cargo,estado,creado_en,creado_por) VALUES(?,?,?,?,?,?,?,?,?)', (ruta_id,dni,t.get('trabajador'),t.get('empresa'),t.get('area'),t.get('cargo'),'ESPERADO',now_str(),session.get('usuario')), commit=True)
                ok += 1
            except Exception as e:
                err += 1
                if len(errores) < 8: errores.append(f'Fila {n}: {e}')
        resumen = {'ok':ok,'err':err,'errores':errores}
        flash(f'Lista esperada cargada: {ok} OK, {err} observados.', 'success' if err==0 else 'warning')
    esperados = rows_to_dict(execute("""SELECT e.*, CASE WHEN p.id IS NULL THEN 'NO SUBIO' ELSE 'SUBIO' END AS abordaje, p.hora
                                         FROM transporte_ruta_esperados e
                                         LEFT JOIN transporte_pasajeros p ON p.ruta_id=e.ruta_id AND p.dni=e.dni
                                         WHERE e.ruta_id=? ORDER BY e.trabajador""", (ruta_id,), fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('transporte_ruta_detalle', ruta_id=ruta.id)}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">TRABAJADORES ESPERADOS</h2></div>
      <form class="floating-card" method="post" enctype="multipart/form-data"><div class="field-help mb-2">Excel con columna: DNI. Se valida contra la base de trabajadores.</div><input type="file" name="archivo" class="form-control mb-2" accept=".xlsx,.xlsm" required><button class="btn btn-green w-100"><i class="bi bi-upload"></i> CARGAR LISTA</button></form>
      <div class="mx-2 my-2 d-grid"><a class="btn btn-outline-success" href="{{url_for('exportar_transporte_no_subieron', ruta_id=ruta.id)}}"><i class="bi bi-file-earmark-excel"></i> Exportar no subieron</a></div>
      {% for e in esperados %}<div class="worker-card"><div class="worker-title"><div>{{e.dni}}<br><b>{{e.trabajador}}</b></div><div class="text-end">{{e.hora or ''}}<br><b class="{{'text-success' if e.abordaje=='SUBIO' else 'text-danger'}}">{{e.abordaje}}</b></div></div><div class="small-value">{{e.area or ''}} · {{e.cargo or ''}}</div></div>{% else %}<div class="worker-card text-center text-muted">Aún no hay lista esperada cargada.</div>{% endfor %}
    </div>"""
    return render_page(body, ruta=ruta, esperados=esperados, resumen=resumen)

@app.route('/exportar/transporte/no-subieron')
@login_required
def exportar_transporte_no_subieron():
    ruta_id = request.args.get('ruta_id')
    if not ruta_id:
        flash('Selecciona una ruta.', 'danger'); return redirect(url_for('transporte'))
    rows = rows_to_dict(execute("""SELECT r.fecha AS fecha_ruta, r.nombre AS ruta, r.origen, r.destino, e.dni, e.trabajador, e.empresa, e.area, e.cargo, 'NO SUBIO' AS estado
                                   FROM transporte_ruta_esperados e
                                   LEFT JOIN transporte_rutas r ON r.id=e.ruta_id
                                   LEFT JOIN transporte_pasajeros p ON p.ruta_id=e.ruta_id AND p.dni=e.dni
                                   WHERE e.ruta_id=? AND p.id IS NULL ORDER BY e.trabajador""", (ruta_id,), fetchall=True))
    headers=['FECHA_RUTA','RUTA','ORIGEN','DESTINO','DNI','TRABAJADOR','EMPRESA','AREA','CARGO','ESTADO']
    return excel_response(headers, rows, 'transporte_no_subieron.xlsx', 'NO_SUBIERON')

@app.route('/transporte/ruta/<int:ruta_id>/cerrar', methods=['POST'])
@login_required
def transporte_cerrar_ruta(ruta_id):
    hora_llegada = request.form.get('hora_llegada') or datetime.now().strftime('%H:%M')
    km_fin = request.form.get('km_fin') or 0
    execute("UPDATE transporte_rutas SET estado='CERRADA', hora_llegada=?, km_fin=?, cerrado_por=?, cerrado_en=? WHERE id=?", (hora_llegada, km_fin, session.get('usuario'), now_str(), ruta_id), commit=True)
    flash('Ruta cerrada correctamente.', 'success')
    return redirect(url_for('transporte_ruta_detalle', ruta_id=ruta_id))

@app.route('/transporte/ruta/<int:ruta_id>/mapa')
@login_required
def transporte_mapa(ruta_id):
    ruta = row_to_dict(execute('SELECT r.*, v.placa, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id WHERE r.id=?', (ruta_id,), fetchone=True))
    puntos = rows_to_dict(execute('SELECT * FROM transporte_gps WHERE ruta_id=? ORDER BY fecha_hora DESC LIMIT 20', (ruta_id,), fetchall=True))
    lat = ruta.get('latitud') if ruta else ''; lon = ruta.get('longitud') if ruta else ''
    map_url = ''
    if lat and lon: map_url = f'https://maps.google.com/maps?q={lat},{lon}&z=16&output=embed'
    body = """
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('transporte_ruta_detalle', ruta_id=ruta.id)}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">MAPA GPS RUTA</h2></div>
      <div class="worker-card"><b>{{ruta.nombre}}</b><br>{{ruta.placa or 'SIN PLACA'}} · {{ruta.conductor or 'SIN CONDUCTOR'}}<br><span class="small-value">Última ubicación: {{ruta.ultima_ubicacion or 'Sin GPS'}}</span></div>
      {% if map_url %}<iframe src="{{map_url}}" style="width:100%;height:320px;border:0;border-radius:12px" loading="lazy"></iframe>{% else %}<div class="scan-bad m-2">Aún no hay ubicación GPS para esta ruta.</div>{% endif %}
      {% for p in puntos %}<div class="worker-card"><div class="worker-title"><div>{{p.fecha_hora}}<br><b>{{p.latitud}}, {{p.longitud}}</b></div><div class="text-end"><b>{{p.registrado_por}}</b></div></div></div>{% endfor %}
    </div>"""
    return render_page(body, ruta=ruta, puntos=puntos, map_url=map_url)


@app.route('/transporte/requisitos')
@login_required
def transporte_requisitos():
    hoy=today_str()
    conductores=rows_to_dict(execute("""SELECT dni,nombres,'CONDUCTOR' AS tipo,
        venc_licencia, venc_cert_medico, venc_sctr, estado
        FROM transporte_conductores
        WHERE (venc_licencia IS NOT NULL AND venc_licencia<>'' AND venc_licencia<=?)
           OR (venc_cert_medico IS NOT NULL AND venc_cert_medico<>'' AND venc_cert_medico<=?)
           OR (venc_sctr IS NOT NULL AND venc_sctr<>'' AND venc_sctr<=?)
        ORDER BY nombres""", (hoy,hoy,hoy), fetchall=True))
    vehiculos=rows_to_dict(execute("""SELECT placa,tipo,empresa_transportista,soat_venc,revision_tecnica_venc,estado
        FROM transporte_vehiculos
        WHERE (soat_venc IS NOT NULL AND soat_venc<>'' AND soat_venc<=?)
           OR (revision_tecnica_venc IS NOT NULL AND revision_tecnica_venc<>'' AND revision_tecnica_venc<=?)
        ORDER BY placa""", (hoy,hoy), fetchall=True))
    body="""
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">REQUISITOS VENCIDOS</h2></div>
      <div class="worker-card text-center"><b>Conductores:</b> {{conductores|length}} &nbsp; <b>Vehículos:</b> {{vehiculos|length}}</div>
      <div class="trans-section mx-2">Conductores</div>
      {% for c in conductores %}<div class="worker-card"><div class="worker-title"><div>{{c.dni}}<br><b>{{c.nombres}}</b></div><div class="text-end"><b>{{c.estado}}</b></div></div><div class="worker-grid"><div><label>LICENCIA</label><div class="small-value">{{c.venc_licencia or '-'}}</div></div><div><label>MÉDICO</label><div class="small-value">{{c.venc_cert_medico or '-'}}</div></div><div><label>SCTR</label><div class="small-value">{{c.venc_sctr or '-'}}</div></div></div></div>{% else %}<div class="worker-card text-center text-muted">Sin conductores vencidos.</div>{% endfor %}
      <div class="trans-section mx-2">Vehículos</div>
      {% for v in vehiculos %}<div class="worker-card"><div class="worker-title"><div>{{v.placa}}<br><b>{{v.tipo or 'VEHÍCULO'}}</b></div><div class="text-end"><b>{{v.estado}}</b></div></div><div class="worker-grid"><div><label>SOAT</label><div class="small-value">{{v.soat_venc or '-'}}</div></div><div><label>REV. TÉCNICA</label><div class="small-value">{{v.revision_tecnica_venc or '-'}}</div></div><div><label>EMPRESA</label><div class="small-value">{{v.empresa_transportista or '-'}}</div></div></div></div>{% else %}<div class="worker-card text-center text-muted">Sin vehículos vencidos.</div>{% endfor %}
    </div>"""
    return render_page(body, conductores=conductores, vehiculos=vehiculos)

@app.route('/transporte/mapa-general')
@login_required
def transporte_mapa_general():
    rutas=rows_to_dict(execute("""SELECT r.*, v.placa, c.nombres AS conductor FROM transporte_rutas r
                               LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                               LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                               WHERE COALESCE(r.latitud,'')<>'' AND COALESCE(r.longitud,'')<>''
                               ORDER BY r.ultima_ubicacion DESC, r.fecha DESC LIMIT 50""", fetchall=True))
    body="""
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">GPS / SEGUIMIENTO</h2></div>
      {% for r in rutas %}<div class="worker-card"><div class="worker-title"><div>{{r.fecha}}<br><b>{{r.nombre}}</b></div><div class="text-end">{{r.placa or 'SIN BUS'}}<br><b>{{r.estado}}</b></div></div><div class="small-value">{{r.conductor or '-'}} · {{r.ultima_ubicacion or ''}}</div><a class="btn btn-outline-success btn-sm w-100 mt-2" href="{{url_for('transporte_mapa', ruta_id=r.id)}}"><i class="bi bi-map"></i> Ver mapa</a></div>{% else %}<div class="worker-card text-center text-muted">Aún no hay rutas con GPS registrado.</div>{% endfor %}
    </div>"""
    return render_page(body, rutas=rutas)

# ========================= SOPORTE / CONFIG / SINC =========================
@app.route('/soporte')
@login_required
def soporte():
    return render_page('<div class="phone-wrap"><h2 class="header-title">SOPORTE</h2><div class="floating-card"><b>Canal de soporte</b><p class="small text-muted mb-2">Registra incidencias de sincronización, acceso o lectura de fotocheck.</p><textarea class="form-control" rows="4" placeholder="Describe el problema..."></textarea><a class="btn btn-green w-100 mt-3" href="{{url_for(\'home\')}}">ENVIAR / VOLVER</a></div></div>')

@app.route('/configuraciones')
@admin_required
def configuraciones():
    body = '<div class="phone-wrap"><a class="back-mini" href="{{url_for(\'home\')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title">CONFIGURACIÓN</h2><div class="floating-card"><a class="btn btn-green w-100 mb-2" href="{{url_for(\'cargar_base\')}}"><i class="bi bi-people"></i> Cargar trabajadores</a><a class="btn btn-green w-100 mb-2" href="{{url_for(\'cargar_actividades\')}}"><i class="bi bi-diagram-3"></i> Actividades / Labores / Consumidores</a><a class="btn btn-outline-success w-100 mb-2" href="{{url_for(\'plantilla_trabajadores\')}}">Plantilla trabajadores</a><a class="btn btn-outline-success w-100 mb-2" href="{{url_for(\'plantilla_actividades\')}}">Plantilla actividades</a><a class="btn btn-outline-secondary w-100 mb-2" href="{{url_for(\'usuarios\')}}"><i class="bi bi-people"></i> Usuarios</a><a class="btn btn-outline-secondary w-100" href="{{url_for(\'home\')}}">Volver</a></div></div>'
    return render_page(body)

@app.route('/sincronizacion')
@login_required
def sincronizacion():
    total = scalar('SELECT COUNT(*) AS c FROM trabajadores')
    body = '<div class="phone-wrap"><h2 class="header-title">SINCRONIZACIÓN</h2><div class="panel-green"><i class="bi bi-arrow-repeat"></i><h4>TABLAS MAESTRAS</h4></div><div class="floating-card" style="margin:-20px 10px 0"><p><b>Trabajadores:</b> {{total}}</p><p><b>Última actualización:</b> {{now}}</p><a class="btn btn-green w-100" href="{{url_for(\'home\')}}">SINCRONIZADO</a></div></div>'
    return render_page(body, total=total, now=now_str())

# ========================= REPORTES / CARGA / USUARIOS =========================
@app.route('/reportes')
@login_required
def reportes():
    desde = request.args.get('desde') or today_str(); hasta = request.args.get('hasta') or today_str(); q=request.args.get('q','').strip()
    params=[desde,hasta]; where='WHERE fecha>=? AND fecha<=?'
    if q:
        like=f"%{q.upper()}%"; where += ' AND (dni LIKE ? OR UPPER(trabajador) LIKE ? OR UPPER(labor) LIKE ?)'; params += [like,like,like]
    tareos = rows_to_dict(execute(f'SELECT * FROM tareos {where} ORDER BY creado_en DESC LIMIT 500', params, fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad report-wrap"><div class="config-header"><a class="back-mini" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">REPORTES TAREO</h2></div><form class="floating-card mb-2"><div class="row g-2"><div class="col-6"><input class="form-control" type="date" name="desde" value="{{desde}}"></div><div class="col-6"><input class="form-control" type="date" name="hasta" value="{{hasta}}"></div><div class="col-9"><input class="form-control" name="q" value="{{q}}" placeholder="DNI / trabajador / labor"></div><div class="col-3 d-grid"><button class="btn btn-green"><i class="bi bi-search"></i></button></div></div><a class="btn btn-outline-success w-100 mt-2" href="{{url_for('exportar_tareos',desde=desde,hasta=hasta,q=q)}}">EXPORTAR EXCEL</a></form>{% for r in tareos %}<div class="worker-card"><div class="worker-title"><div>{{r.fecha}}<br><b>{{r.trabajador}}</b></div><div class="text-end">{{r.dni}}<br><b>{{r.labor}}</b></div></div><div class="worker-grid"><div><label>HORAS</label><div class="mini-input">{{r.horas}}</div></div><div><label>CANT.</label><div class="mini-input">{{r.cantidad}}</div></div><div><label>UNIDAD</label><div class="mini-input">{{r.unidad}}</div></div></div></div>{% else %}<div class="alert alert-light border text-center">Sin datos.</div>{% endfor %}</div>"""
    return render_page(body, desde=desde, hasta=hasta, q=q, tareos=tareos)

@app.route('/exportar/tareos')
@login_required
def exportar_tareos():
    desde=request.args.get('desde') or today_str(); hasta=request.args.get('hasta') or today_str(); q=request.args.get('q','').strip()
    params=[desde,hasta]; where='WHERE fecha>=? AND fecha<=?'
    if q:
        like=f"%{q.upper()}%"; where += ' AND (dni LIKE ? OR UPPER(trabajador) LIKE ? OR UPPER(labor) LIKE ?)'; params += [like,like,like]
    rows=rows_to_dict(execute(f'SELECT fecha,dni,trabajador,empresa,area,cargo,labor,fundo,lote,horas,cantidad,unidad,turno,tipo_tareo,hora_inicio,hora_fin,ref_inicio,ref_fin,observacion,registrado_por,creado_en FROM tareos {where} ORDER BY creado_en DESC', params, fetchall=True))
    headers=['FECHA','DNI','TRABAJADOR','EMPRESA','AREA','CARGO','LABOR','FUNDO','LOTE','HORAS','CANTIDAD','UNIDAD','TURNO','TIPO_TAREO','HORA_INICIO','HORA_FIN','REF_INICIO','REF_FIN','OBSERVACION','REGISTRADO_POR','CREADO_EN']
    return excel_response(headers, rows, f'tareos_{desde}_a_{hasta}.xlsx', 'TAREOS')

@app.route('/cargar-base', methods=['GET','POST'])
@admin_required
def cargar_base():
    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f or not f.filename.lower().endswith('.xlsx'):
            flash('Suba un Excel .xlsx válido.', 'danger'); return redirect(url_for('cargar_base'))
        wb = load_workbook(f, data_only=True, read_only=True); ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: flash('Excel vacío.', 'danger'); return redirect(url_for('cargar_base'))
        headers = [normalizar_columna(c) for c in rows[0]]
        if 'DNI' not in headers:
            flash('La plantilla debe tener la columna DNI.', 'danger'); return redirect(url_for('cargar_base'))
        colmap = {'TRABAJADOR':'trabajador','NOMBRE':'trabajador','APELLIDOS Y NOMBRES':'trabajador','EMPRESA':'empresa','AREA':'area','ÁREA':'area','CARGO':'cargo','ACTIVIDAD':'actividad','PLANILLA':'planilla','ESTADO':'estado'}
        ins=upd=omi=0; conn=get_conn(); cur=conn.cursor(); ahora=now_str()
        for row in rows[1:]:
            item={headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
            dni=limpiar_dni(item.get('DNI'))
            if len(dni)!=8: omi+=1; continue
            data={'trabajador':'','empresa':'','area':'','cargo':'','actividad':'','planilla':'','estado':'ACTIVO'}
            for col,key in colmap.items():
                if col in headers and item.get(col) not in (None,''): data[key]=limpiar_texto(item.get(col))
            cur.execute(qmark('SELECT id FROM trabajadores WHERE dni=?'), (dni,))
            if cur.fetchone():
                cur.execute(qmark('UPDATE trabajadores SET trabajador=?,empresa=?,area=?,cargo=?,actividad=?,planilla=?,estado=?,fecha_carga=? WHERE dni=?'), (data['trabajador'],data['empresa'],data['area'],data['cargo'],data['actividad'],data['planilla'],data['estado'],ahora,dni)); upd+=1
            else:
                cur.execute(qmark('INSERT INTO trabajadores(dni,trabajador,empresa,area,cargo,actividad,planilla,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?,?)'), (dni,data['trabajador'],data['empresa'],data['area'],data['cargo'],data['actividad'],data['planilla'],data['estado'],ahora)); ins+=1
        conn.commit(); cur.close(); conn.close(); flash(f'Carga completa. Insertados: {ins} | Actualizados: {upd} | Omitidos: {omi}', 'success')
        return redirect(url_for('cargar_base'))
    trabajadores = rows_to_dict(execute('SELECT * FROM trabajadores ORDER BY fecha_carga DESC, trabajador LIMIT 100', fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('configuraciones')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title">CONFIGURACIÓN – TRABAJADORES</h2></div><form method="post" enctype="multipart/form-data" class="floating-card mb-2"><label class="form-label">Archivo Excel .xlsx</label><input class="form-control mb-2" type="file" name="archivo" accept=".xlsx" required><button class="btn btn-green w-100">CARGAR BASE</button><a class="btn btn-outline-success w-100 mt-2" href="{{url_for('plantilla_trabajadores')}}">PLANTILLA EXCEL</a></form>{% for r in trabajadores %}<div class="worker-card"><div class="worker-title"><div>{{r.dni}}<br><b>{{r.trabajador}}</b></div><div class="text-end">{{r.empresa}}<br><b>{{r.cargo}}</b></div></div></div>{% endfor %}</div>"""
    return render_page(body, trabajadores=trabajadores)

@app.route('/plantilla-trabajadores')
@admin_required
def plantilla_trabajadores():
    headers=['DNI','TRABAJADOR','EMPRESA','CARGO','ESTADO','FECHA']
    rows=[{'DNI':'12345678','TRABAJADOR':'APELLIDOS Y NOMBRES','EMPRESA':'AQUANQA I','CARGO':'OPERARIO','ESTADO':'ACTIVO','FECHA':today_str()}]
    return excel_response(headers, rows, 'plantilla_trabajadores.xlsx', 'TRABAJADORES')


@app.route('/cargar-actividades', methods=['GET','POST'])
@admin_required
def cargar_actividades():
    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f or not f.filename.lower().endswith('.xlsx'):
            flash('Suba un Excel .xlsx válido.', 'danger'); return redirect(url_for('cargar_actividades'))
        wb = load_workbook(f, data_only=True, read_only=True); ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash('Excel vacío.', 'danger'); return redirect(url_for('cargar_actividades'))
        headers = [normalizar_columna(c) for c in rows[0]]
        aliases = {
            'COD. ACTIVIDAD':'cod_actividad','COD ACTIVIDAD':'cod_actividad','COD_ACTIVIDAD':'cod_actividad',
            'DESCRIPCION ACTIVIDAD':'desc_actividad','DESCRIPCIÓN ACTIVIDAD':'desc_actividad','DESC ACTIVIDAD':'desc_actividad','DESC. ACTIVIDAD':'desc_actividad','ACTIVIDAD':'desc_actividad',
            'COD. LABOR':'cod_labor','COD LABOR':'cod_labor','COD_LABOR':'cod_labor',
            'DESCRIPCION LABOR':'desc_labor','DESCRIPCIÓN LABOR':'desc_labor','DESC LABOR':'desc_labor','DESC. LABOR':'desc_labor','LABOR':'desc_labor',
            'COD. CONSUMIDOR':'cod_consumidor','COD CONSUMIDOR':'cod_consumidor','COD_CONSUMIDOR':'cod_consumidor',
            'DESCRIPCION CONSUMIDOR':'desc_consumidor','DESCRIPCIÓN CONSUMIDOR':'desc_consumidor','DESC CONSUMIDOR':'desc_consumidor','DESC. CONSUMIDOR':'desc_consumidor','CONSUMIDOR':'desc_consumidor','ZONA':'desc_consumidor','CAMPO':'desc_consumidor'
        }
        required = ['desc_actividad','desc_labor']
        mapped = {i: aliases.get(h) for i,h in enumerate(headers)}
        if not any(v=='desc_actividad' for v in mapped.values()) or not any(v=='desc_labor' for v in mapped.values()):
            flash('La plantilla debe tener Descripción Actividad y Descripción Labor.', 'danger'); return redirect(url_for('cargar_actividades'))
        conn=get_conn(); cur=conn.cursor(); ins=0; omi=0; ahora=now_str()
        # Carga tipo reemplazo: evita que al subir el mismo Excel se dupliquen miles de filas.
        cur.execute(qmark('DELETE FROM actividades_maestras'))
        vistos_import=set()
        for row in rows[1:]:
            data={'cod_actividad':'','desc_actividad':'','cod_labor':'','desc_labor':'','cod_consumidor':'','desc_consumidor':''}
            for i,val in enumerate(row):
                k=mapped.get(i)
                if k: data[k]=limpiar_texto(val)
            if not data['desc_actividad'] or not data['desc_labor']:
                omi += 1; continue
            key=(data['desc_actividad'],data['desc_labor'],data['desc_consumidor'])
            if key in vistos_import:
                omi += 1; continue
            vistos_import.add(key)
            cur.execute(qmark('INSERT INTO actividades_maestras(cod_actividad,desc_actividad,cod_labor,desc_labor,cod_consumidor,desc_consumidor,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?)'),
                        (data['cod_actividad'],data['desc_actividad'],data['cod_labor'],data['desc_labor'],data['cod_consumidor'],data['desc_consumidor'],'ACTIVO',ahora)); ins += 1
        conn.commit(); cur.close(); conn.close()
        flash(f'Actividades cargadas. Insertados: {ins} | Omitidos: {omi}', 'success')
        return redirect(url_for('cargar_actividades'))
    datos = get_actividades_maestras(2000)
    total_reg = scalar('SELECT COUNT(*) AS total FROM actividades_maestras')
    body = """
    <div class="phone-wrap desktop-pad"><a class="back-mini" href="{{url_for('configuraciones')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title">ACTIVIDADES / LABORES / CONSUMIDORES</h2>
    <form method="post" enctype="multipart/form-data" class="floating-card mb-2"><label class="form-label">Archivo Excel .xlsx</label><input class="form-control mb-2" type="file" name="archivo" accept=".xlsx" required><button class="btn btn-green w-100">CARGAR ACTIVIDADES</button><a class="btn btn-outline-success w-100 mt-2" href="{{url_for('plantilla_actividades')}}">PLANTILLA ACTIVIDADES</a></form>
    <div class="alert alert-success border">Base cargada: {{total_reg}} filas. Mostrando combinaciones únicas ACTIVIDAD/LABOR/CONSUMIDOR para evitar repetidos.</div>
    {% for r in datos %}<div class="worker-card"><div class="worker-title"><div>ACTIVIDAD<br><b>{{r.desc_actividad}}</b></div><div class="text-end">LABOR<br><b>{{r.desc_labor}}</b></div></div><div class="small-label mt-2">CONSUMIDOR</div><div class="small-value">{{r.desc_consumidor or 'NO OBLIGATORIO'}}</div></div>{% else %}<div class="alert alert-light border text-center">Sin actividades cargadas.</div>{% endfor %}</div>"""
    return render_page(body, datos=datos, total_reg=total_reg)

@app.route('/plantilla-actividades')
@admin_required
def plantilla_actividades():
    headers=['Cod. Actividad','Descripción Actividad','Cod. Labor','Descripción Labor','Cod. Consumidor','Descripción Consumidor']
    rows=[{'Cod. Actividad':'ACT001','Descripción Actividad':'COSECHA','Cod. Labor':'LAB001','Descripción Labor':'COSECHA MANUAL','Cod. Consumidor':'CON001','Descripción Consumidor':'CAMPO 01'},
          {'Cod. Actividad':'ACT001','Descripción Actividad':'COSECHA','Cod. Labor':'LAB002','Descripción Labor':'COSECHA SELECTIVA','Cod. Consumidor':'','Descripción Consumidor':''}]
    return excel_response(headers, rows, 'plantilla_actividades_labores_consumidores.xlsx', 'ACTIVIDADES')

@app.route('/api/actividades-maestras')
@login_required
def api_actividades_maestras():
    """Devuelve el árbol Actividad/Labor/Consumidor de forma tolerante.
    Primero usa actividades_maestras; si no hay data, arma opciones con lo ya registrado en hojas/labores
    y con actividad/cargo de trabajadores. Así el modal nunca queda vacío.
    """
    data = []
    try:
        data.extend(get_actividades_maestras())
    except Exception:
        data = []
    try:
        for r in rows_to_dict(execute("SELECT DISTINCT grupo AS desc_actividad, subgrupo AS desc_labor, labor AS desc_consumidor FROM hoja_labores WHERE COALESCE(grupo,'')<>'' OR COALESCE(subgrupo,'')<>'' OR COALESCE(labor,'')<>'' LIMIT 1000", fetchall=True)):
            data.append(r)
    except Exception:
        pass
    try:
        for r in rows_to_dict(execute("SELECT DISTINCT grupo AS desc_actividad, subgrupo AS desc_labor, labor AS desc_consumidor FROM hojas_tareo WHERE COALESCE(grupo,'')<>'' OR COALESCE(subgrupo,'')<>'' OR COALESCE(labor,'')<>'' LIMIT 1000", fetchall=True)):
            data.append(r)
    except Exception:
        pass
    try:
        for r in rows_to_dict(execute("SELECT DISTINCT actividad AS desc_actividad, cargo AS desc_labor, area AS desc_consumidor FROM trabajadores WHERE COALESCE(actividad,'')<>'' OR COALESCE(cargo,'')<>'' OR COALESCE(area,'')<>'' LIMIT 1000", fetchall=True)):
            data.append(r)
    except Exception:
        pass
    clean=[]; seen=set()
    for x in data:
        da=limpiar_texto(x.get('desc_actividad') or x.get('actividad') or x.get('grupo') or x.get('cod_actividad') or '')
        dl=limpiar_texto(x.get('desc_labor') or x.get('labor') or x.get('subgrupo') or x.get('cod_labor') or '')
        dc=limpiar_texto(x.get('desc_consumidor') or x.get('consumidor') or x.get('zona') or x.get('campo') or x.get('cod_consumidor') or '')
        key=(da,dl,dc)
        if (da or dl or dc) and key not in seen:
            seen.add(key); clean.append({'desc_actividad':da,'desc_labor':dl,'desc_consumidor':dc,'cod_actividad':x.get('cod_actividad',''),'cod_labor':x.get('cod_labor',''),'cod_consumidor':x.get('cod_consumidor','')})
    if not clean:
        clean = [
            {'desc_actividad':'ADMINISTRACION','desc_labor':'LABOR ADMINISTRATIVA','desc_consumidor':'OFICINA CENTRAL','cod_actividad':'ACT000','cod_labor':'LAB000','cod_consumidor':'CON000'},
            {'desc_actividad':'ADMISION','desc_labor':'CONTROL DOCUMENTARIO','desc_consumidor':'OFICINA 01','cod_actividad':'ACT003','cod_labor':'LAB004','cod_consumidor':'CON004'},
            {'desc_actividad':'COSECHA','desc_labor':'COSECHA MANUAL','desc_consumidor':'CAMPO 01','cod_actividad':'ACT001','cod_labor':'LAB001','cod_consumidor':'CON001'},
            {'desc_actividad':'COSECHA','desc_labor':'COSECHA SELECTIVA','desc_consumidor':'CAMPO 02','cod_actividad':'ACT001','cod_labor':'LAB002','cod_consumidor':'CON002'},
            {'desc_actividad':'PODA','desc_labor':'PODA SANITARIA','desc_consumidor':'LOTE 01','cod_actividad':'ACT002','cod_labor':'LAB003','cod_consumidor':'CON003'},
        ]
    return jsonify(ok=True, data=clean, total=len(clean))

@app.route('/usuarios', methods=['GET','POST'])
@admin_required
def usuarios():
    if request.method == 'POST':
        usuario=request.form.get('usuario','').strip(); nombres=limpiar_texto(request.form.get('nombres')); clave=request.form.get('clave',''); rol=request.form.get('rol','operador'); estado=request.form.get('estado','ACTIVO')
        if not usuario or not clave: flash('Usuario y clave son obligatorios.', 'danger')
        else:
            try:
                execute('INSERT INTO usuarios(usuario,password_hash,nombres,rol,estado,creado_en) VALUES(?,?,?,?,?,?)', (usuario,generate_password_hash(clave),nombres,rol,estado,now_str()), commit=True); flash('Usuario creado.', 'success')
            except Exception as e: flash(f'No se pudo crear usuario: {e}', 'danger')
        return redirect(url_for('usuarios'))
    users=rows_to_dict(execute('SELECT usuario,nombres,rol,estado,creado_en FROM usuarios ORDER BY usuario', fetchall=True))
    body='<div class="phone-wrap"><a class="back-mini" href="{{url_for(\'configuraciones\')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title">USUARIOS</h2><form method="post" class="floating-card mb-2"><input class="form-control mb-2" name="usuario" placeholder="Usuario" required><input class="form-control mb-2" name="nombres" placeholder="Nombres"><input class="form-control mb-2" type="password" name="clave" placeholder="Clave" required><select class="form-select mb-2" name="rol"><option value="operador">operador</option><option value="admin">admin</option></select><select class="form-select mb-2" name="estado"><option>ACTIVO</option><option>INACTIVO</option></select><button class="btn btn-green w-100">Guardar</button></form>{% for u in users %}<div class="worker-card"><b>{{u.usuario}}</b> · {{u.rol}}<br><span class="small text-muted">{{u.nombres}} · {{u.estado}}</span></div>{% endfor %}</div>'
    return render_page(body, users=users)

# ========================= COMPAT API MARCACIÓN BÁSICA =========================

@app.route('/api/trabajador-labor/<int:hoja_id>/<int:labor_id>/<dni>')
@login_required
def api_trabajador_labor(hoja_id, labor_id, dni):
    dni = limpiar_dni(dni)
    if len(dni) != 8:
        return jsonify(ok=False, msg='DNI inválido.')
    r = row_to_dict(execute('SELECT dni, trabajador FROM tareos WHERE hoja_id=? AND labor_id=? AND dni=? ORDER BY id DESC LIMIT 1', (hoja_id, labor_id, dni), fetchone=True))
    if not r:
        return jsonify(ok=False, msg='Debe registrar primero al trabajador en el módulo Trabajadores de esta labor.')
    return jsonify(ok=True, trabajador=r)

@app.route('/api/trabajador/<dni>')
@login_required
def api_trabajador(dni):
    dni = limpiar_dni(dni)
    if len(dni) != 8: return jsonify(ok=False, msg='DNI inválido.')
    t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    if not t: return jsonify(ok=False, msg='DNI no encontrado en la base de trabajadores.')
    return jsonify(ok=True, trabajador=t, sugerido='ENTRADA')

# ========================= PWA =========================
@app.route('/manifest.webmanifest')
@app.route('/manifest.json')
def manifest():
    return jsonify({"name":"Tareo Móvil PRIZE","short_name":"Tareo","start_url":"/","display":"standalone","background_color":"#ffffff","theme_color":"#2f773b","icons":[]})

@app.route('/sw.js')
def sw():
    return Response("self.addEventListener('install',e=>self.skipWaiting()); self.addEventListener('fetch',e=>{});", mimetype='application/javascript')

try:
    init_db()
except Exception as e:
    print('ERROR inicializando base de datos:', e)



# ========================= PATCH UI TRANSPORTE MÓVIL SUBMÓDULOS =========================
# Implementación compacta solicitada: portada delgada + submódulos internos.

def _tm_css():
    return """
    <style>
      .shell{max-width:430px!important;margin:0 auto!important;padding:4px 6px!important;background:#fff!important;overflow-x:hidden!important}
      .phone-wrap,.page-card{max-width:390px!important;width:100%!important;margin:0 auto!important;border-radius:14px!important;overflow:hidden!important}
      .tm-head{height:132px;background:#28783b;color:white;border-radius:12px 12px 0 0;position:relative;text-align:center;padding:16px 12px 0;box-shadow:0 4px 12px rgba(0,0,0,.12)}
      .tm-back{position:absolute;left:12px;top:24px;color:#fff;text-decoration:none;font-size:35px;line-height:1}.tm-bus{font-size:34px;line-height:1;margin-top:4px}.tm-title{font-weight:900;font-size:13px;margin-top:8px;letter-spacing:.2px}.tm-config{position:absolute;right:12px;top:12px;color:#fff;text-decoration:none;border:1px solid rgba(255,255,255,.75);border-radius:9px;padding:7px 9px;font-weight:900;font-size:12px;display:flex;gap:5px;align-items:center}.tm-card{background:#fff;border:1px solid #e5e7eb;border-radius:12px;box-shadow:0 6px 16px rgba(0,0,0,.10);padding:12px;margin:-38px 9px 10px;position:relative;z-index:4}.tm-tabs{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}.tm-tab{height:74px;border-radius:10px;background:#2f773b;color:#fff;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;font-weight:900;font-size:13px;box-shadow:0 7px 14px rgba(0,0,0,.12)}.tm-tab i{font-size:23px;margin-bottom:5px}.tm-section{font-size:12px;color:#06622d;font-weight:900;margin:13px 2px 7px;text-transform:uppercase}.tm-list{border:1px solid #e5e7eb;border-radius:9px;overflow:hidden;background:#fff}.tm-item{display:flex;align-items:center;gap:11px;padding:12px 13px;border-bottom:1px solid #edf0ed;color:#1f2937;text-decoration:none;font-weight:800;font-size:13px}.tm-item:last-child{border-bottom:0}.tm-item i{color:#08713b;font-size:20px;width:23px;text-align:center}.tm-item .chev{margin-left:auto;color:#111;font-size:17px;width:auto}.tm-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin-top:10px}.tm-kpi{border:1px solid #9fcea8;background:#f7fff8;border-radius:8px;text-align:center;color:#06622d;padding:8px 3px;font-weight:900}.tm-kpi label{display:block;font-size:9px;line-height:1.1}.tm-kpi strong{font-size:19px;line-height:1.15}.tm-page{background:#fff;border:1px solid #e5e7eb;border-radius:0 0 12px 12px;box-shadow:0 6px 18px rgba(0,0,0,.08);margin-top:-38px;position:relative;z-index:4;padding:12px}.tm-topbar{height:56px;background:linear-gradient(135deg,#075d2a,#2f773b);color:#fff;display:flex;align-items:center;gap:10px;padding:0 12px;border-radius:12px 12px 0 0;font-weight:900}.tm-topbar a{color:#fff;text-decoration:none;font-size:28px}.tm-topbar .ttl{flex:1;text-align:center;margin-right:26px}.tm-search-row{display:flex;gap:8px;margin:10px 0}.tm-search{flex:1;border:1px solid #e5e7eb;border-radius:8px;height:37px;padding:0 10px;font-weight:700;font-size:12px}.tm-btn{background:#08713b;color:#fff!important;border:0;border-radius:8px;font-weight:900;text-decoration:none;padding:8px 11px;font-size:12px}.tm-btn-outline{background:#fff;color:#08713b!important;border:1px solid #08713b;border-radius:8px;font-weight:900;text-decoration:none;padding:8px 11px;font-size:12px}.tm-table{width:100%;border-collapse:collapse;font-size:11px}.tm-table th{background:#fafafa;color:#374151;font-weight:900;border-bottom:1px solid #e5e7eb;padding:8px 6px}.tm-table td{border-bottom:1px solid #edf0ed;padding:8px 6px;vertical-align:middle}.tm-table tr:last-child td{border-bottom:0}.tm-scroll{overflow-x:auto}.tm-scroll .tm-table{min-width:430px}.tm-badge{display:inline-block;padding:3px 6px;border-radius:5px;background:#dcfce7;color:#166534;font-weight:900;font-size:9px}.tm-warn{background:#fff7ed!important;color:#ea580c!important}.tm-bad{background:#fee2e2!important;color:#dc2626!important}.tm-info{background:#eff6ff;border:1px solid #bfdbfe;color:#1e3a8a;border-radius:9px;padding:10px;font-size:12px;font-weight:700}.tm-mini-cards{display:grid;grid-template-columns:1fr;gap:8px;margin-top:10px}.tm-mini{border:1px solid #e5e7eb;border-radius:10px;background:#fff;padding:10px;color:#065f2a;font-weight:900}.tm-mini small{display:block;color:#6b7280;font-weight:700}.tm-action-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:10px}.tm-scanner-box{height:150px;border:1px solid #e5e7eb;border-radius:10px;display:grid;place-items:center;color:#08713b;background:#fbfffb;text-align:center;font-weight:900}.tm-map{height:220px;background:linear-gradient(135deg,#f3f4f6,#e5e7eb);border-radius:10px;display:grid;place-items:center;color:#08713b;font-weight:900}.tm-form-card{border:1px solid #e5e7eb;border-radius:10px;padding:10px;margin:8px 0;background:#fbfffb}.tm-form-card label{font-size:10px;font-weight:900;color:#2f773b;margin-bottom:3px}.tm-form-card .form-control,.tm-form-card .form-select{height:34px;font-size:12px;border-radius:8px}.tm-hide{display:none}
    </style>
    """

def _tm_top(title, back='transporte', icon='bi-bus-front'):
    return f"""<div class='phone-wrap desktop-pad'><div class='tm-topbar'><a href='{{{{url_for('{back}')}}}}'><i class='bi bi-chevron-left'></i></a><div class='ttl'><i class='bi {icon}'></i> {title}</div></div><div class='tm-page'>"""

def _estado_badge(estado):
    e=(estado or 'ACTIVO').upper()
    cls='tm-badge'
    if 'VENC' in e or 'INACT' in e: cls+=' tm-bad'
    elif 'PEND' in e or 'POR' in e: cls+=' tm-warn'
    return f"<span class='{cls}'>{e}</span>"

@login_required
def transporte_mobile_home():
    """Portada compacta solicitada: solo tiles verdes. Sin indicadores ni Maestros visibles."""
    body=_tm_css()+"""
    <style>
      .tm-card.clean-home{padding:12px 11px 14px!important;margin:-38px 9px 10px!important}
      .tm-tile-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin:8px 0 12px}
      .tm-tile-grid.two{grid-template-columns:repeat(2,1fr)}
      .tm-tile-green{height:78px;border-radius:9px;background:linear-gradient(135deg,#075d2a,#2f773b);color:#fff!important;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;font-weight:900;font-size:11px;line-height:1.12;box-shadow:0 6px 12px rgba(0,0,0,.15)}
      .tm-tile-green i{font-size:24px;margin-bottom:6px;color:#fff!important}
      .tm-section{font-size:12px!important;margin:12px 1px 7px!important;color:#065f2a!important;font-weight:900!important;text-transform:uppercase!important}
      .tm-head{height:132px!important}.tm-title{font-size:13px!important}.tm-config{top:12px!important;right:12px!important}
    </style>
    <div class="phone-wrap desktop-pad"><div class="page-card">
      <div class="tm-head">
        <a class="tm-back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>
        <a class="tm-config" href="{{url_for('transporte_config')}}"><i class="bi bi-gear"></i> Config.</a>
        <div class="tm-bus"><i class="bi bi-bus-front"></i></div><div class="tm-title">MÓDULO TRANSPORTE</div>
      </div>
      <div class="tm-card clean-home">
        <div class="tm-tile-grid">
          <a class="tm-tile-green" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i>Conductores</a>
          <a class="tm-tile-green" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i>Buses</a>
          <a class="tm-tile-green" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i>Rutas</a>
        </div>
        <div class="tm-section">Operación</div>
        <div class="tm-tile-grid">
          <a class="tm-tile-green" href="{{url_for('transporte_abordaje_home')}}"><i class="bi bi-people"></i>Abordaje<br>trabajadores</a>
          <a class="tm-tile-green" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-geo-alt"></i>GPS /<br>seguimiento</a>
          <a class="tm-tile-green" href="{{url_for('transporte_rutas')}}"><i class="bi bi-calendar2-check"></i>Rutas de hoy</a>
          <a class="tm-tile-green" href="{{url_for('conductor_movil_login')}}"><i class="bi bi-phone"></i>Móvil<br>conductor</a>
        </div>
        <div class="tm-section">Control</div>
        <div class="tm-tile-grid two">
          <a class="tm-tile-green" href="{{url_for('transporte_reporte_abordajes')}}"><i class="bi bi-file-earmark-text"></i>Reporte<br>abordajes</a>
          <a class="tm-tile-green" href="{{url_for('transporte_requisitos')}}"><i class="bi bi-exclamation-triangle"></i>Requisitos<br>vencidos</a>
        </div>
      </div>
    </div></div>"""
    return render_page(body)

@login_required
def transporte_conductores_mobile():
    if request.method=='POST':
        dni=limpiar_dni(request.form.get('dni')); nombres=limpiar_texto(request.form.get('nombres'))
        if len(dni)!=8 or not nombres:
            flash('Ingrese DNI de 8 dígitos y nombres del conductor.','danger'); return redirect(url_for('transporte_conductores'))
        pin=(request.form.get('movil_pin') or _pin_conductor(dni)).strip()
        vals=(dni,nombres,request.form.get('telefono',''),limpiar_texto(request.form.get('licencia')),limpiar_texto(request.form.get('categoria')),request.form.get('venc_licencia',''),request.form.get('venc_cert_medico',''),request.form.get('venc_sctr',''),limpiar_texto(request.form.get('estado') or 'ACTIVO'),dni,pin,now_str())
        try:
            execute('INSERT INTO transporte_conductores(dni,nombres,telefono,licencia,categoria,venc_licencia,venc_cert_medico,venc_sctr,estado,movil_usuario,movil_pin,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)', vals, commit=True)
            flash(f'Conductor registrado. PIN: {pin}','success')
        except Exception:
            execute('UPDATE transporte_conductores SET nombres=?,telefono=?,licencia=?,categoria=?,venc_licencia=?,venc_cert_medico=?,venc_sctr=?,estado=?,movil_usuario=?,movil_pin=? WHERE dni=?', (nombres,vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],vals[8],dni,pin,dni), commit=True)
            flash('Conductor actualizado.','success')
        return redirect(url_for('transporte_conductores'))
    rows=rows_to_dict(execute('SELECT * FROM transporte_conductores ORDER BY id DESC LIMIT 80', fetchall=True))
    total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO','APTO',''))
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<today_str() for c in ('venc_licencia','venc_cert_medico','venc_sctr')))
    body=_tm_css()+_tm_top('Conductores')+"""
      <form method="post" class="tm-form-card"><div class="row g-1"><div class="col-5"><label>DNI</label><input name="dni" class="form-control" maxlength="8" required></div><div class="col-7"><label>Nombre</label><input name="nombres" class="form-control" required></div><div class="col-6"><label>Teléfono</label><input name="telefono" class="form-control"></div><div class="col-6"><label>Licencia</label><input name="licencia" class="form-control"></div><div class="col-6"><label>Categoría</label><input name="categoria" class="form-control"></div><div class="col-6"><label>PIN móvil</label><input name="movil_pin" class="form-control" placeholder="Auto"></div><div class="col-6"><label>Venc. licencia</label><input name="venc_licencia" type="date" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>ACTIVO</option><option>APTO</option><option>POR VENCER</option><option>VENCIDA</option><option>INACTIVO</option></select></div></div><button class="tm-btn w-100 mt-2">+ Guardar conductor</button></form>
      <div class="tm-search-row"><input class="tm-search" placeholder="Buscar conductor..."><a class="tm-btn" href="#">+ Nuevo</a></div>
      <div class="tm-scroll"><table class="tm-table"><thead><tr><th>DNI</th><th>Nombre</th><th>Licencia</th><th>Estado</th><th>PIN</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.nombres}}</td><td>{{r.categoria or r.licencia or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td>{{r.movil_pin or '-'}}</td></tr>{% else %}<tr><td colspan="5" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}</tbody></table></div>
      <div class="tm-mini-cards"><div class="tm-mini"><small>Total conductores</small>{{total}}</div><div class="tm-mini"><small>Activos</small>{{activos}}</div><div class="tm-mini"><small>Vencidos</small>{{venc}}</div></div>
    </div></div>"""
    return render_page(body, rows=rows, total=total, activos=activos, venc=venc, badge=_estado_badge)

@login_required
def transporte_vehiculos_mobile():
    if request.method=='POST':
        placa=limpiar_texto(request.form.get('placa'))
        if not placa:
            flash('Ingrese placa.','danger'); return redirect(url_for('transporte_vehiculos'))
        vals=(placa,limpiar_texto(request.form.get('tipo') or 'BUS'),int(request.form.get('capacidad') or 0),limpiar_texto(request.form.get('empresa_transportista')),request.form.get('soat_venc',''),request.form.get('revision_tecnica_venc',''),limpiar_texto(request.form.get('gps_codigo')),limpiar_texto(request.form.get('estado') or 'ACTIVO'),now_str())
        try:
            execute('INSERT INTO transporte_vehiculos(placa,tipo,capacidad,empresa_transportista,soat_venc,revision_tecnica_venc,gps_codigo,estado,creado_en) VALUES(?,?,?,?,?,?,?,?,?)', vals, commit=True)
            flash('Vehículo registrado.','success')
        except Exception:
            execute('UPDATE transporte_vehiculos SET tipo=?,capacidad=?,empresa_transportista=?,soat_venc=?,revision_tecnica_venc=?,gps_codigo=?,estado=? WHERE placa=?', (vals[1],vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],placa), commit=True)
            flash('Vehículo actualizado.','success')
        return redirect(url_for('transporte_vehiculos'))
    rows=rows_to_dict(execute('SELECT * FROM transporte_vehiculos ORDER BY id DESC LIMIT 80', fetchall=True))
    total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO',''))
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<today_str() for c in ('soat_venc','revision_tecnica_venc')))
    body=_tm_css()+_tm_top('Buses / Vehículos')+"""
      <form method="post" class="tm-form-card"><div class="row g-1"><div class="col-5"><label>Placa</label><input name="placa" class="form-control" required></div><div class="col-7"><label>Tipo</label><select name="tipo" class="form-select"><option>BUS</option><option>MINIBUS</option><option>VAN</option><option>CAMIONETA</option></select></div><div class="col-6"><label>Capacidad</label><input name="capacidad" type="number" class="form-control"></div><div class="col-6"><label>Empresa</label><input name="empresa_transportista" class="form-control"></div><div class="col-6"><label>SOAT venc.</label><input name="soat_venc" type="date" class="form-control"></div><div class="col-6"><label>Rev. técnica</label><input name="revision_tecnica_venc" type="date" class="form-control"></div></div><button class="tm-btn w-100 mt-2">+ Guardar vehículo</button></form>
      <div class="tm-search-row"><input class="tm-search" placeholder="Buscar vehículo..."><a class="tm-btn" href="#">+ Nuevo</a></div><div class="tm-scroll"><table class="tm-table"><thead><tr><th>Placa</th><th>Tipo</th><th>Capacidad</th><th>Estado</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.placa}}</td><td>{{r.tipo}}</td><td>{{r.capacidad}}</td><td>{{badge(r.estado)|safe}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin vehículos.</td></tr>{% endfor %}</tbody></table></div>
      <div class="tm-mini-cards"><div class="tm-mini"><small>Total vehículos</small>{{total}}</div><div class="tm-mini"><small>Activos</small>{{activos}}</div><div class="tm-mini"><small>Vencidos</small>{{venc}}</div></div>
    </div></div>"""
    return render_page(body, rows=rows, total=total, activos=activos, venc=venc, badge=_estado_badge)

@login_required
def transporte_rutas_mobile():
    if request.method=='POST':
        fecha=request.form.get('fecha') or today_str(); nombre=limpiar_texto(request.form.get('nombre') or 'RUTA')
        execute('INSERT INTO transporte_rutas(fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,vehiculo_id,conductor_id,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)', (fecha,nombre,limpiar_texto(request.form.get('origen')),limpiar_texto(request.form.get('destino')),limpiar_texto(request.form.get('sede')),request.form.get('hora_salida',''),request.form.get('hora_retorno',''),request.form.get('vehiculo_id') or None,request.form.get('conductor_id') or None,limpiar_texto(request.form.get('estado') or 'PROGRAMADA'),session.get('usuario'),now_str()), commit=True)
        flash('Ruta programada.','success'); return redirect(url_for('transporte_rutas'))
    rutas=rows_to_dict(execute('SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id ORDER BY r.fecha DESC, r.id DESC LIMIT 80', fetchall=True))
    vehiculos=rows_to_dict(execute('SELECT id, placa, tipo, capacidad FROM transporte_vehiculos ORDER BY placa', fetchall=True))
    conductores=rows_to_dict(execute('SELECT id, dni, nombres FROM transporte_conductores ORDER BY nombres', fetchall=True))
    hoy=today_str(); prog=len(rutas); en=sum(1 for r in rutas if (r.get('estado') or '').upper()=='EN RUTA'); pend=sum(1 for r in rutas if (r.get('estado') or '').upper() in ('PROGRAMADA','PENDIENTE')); fin=sum(1 for r in rutas if (r.get('estado') or '').upper() in ('FINALIZADA','CERRADA'))
    body=_tm_css()+_tm_top('Rutas')+"""
      <form method="post" class="tm-form-card"><div class="row g-1"><div class="col-6"><label>Fecha</label><input name="fecha" type="date" value="{{hoy}}" class="form-control"></div><div class="col-6"><label>Ruta</label><input name="nombre" class="form-control" placeholder="R-001"></div><div class="col-6"><label>Origen</label><input name="origen" class="form-control"></div><div class="col-6"><label>Destino</label><input name="destino" class="form-control"></div><div class="col-6"><label>Salida</label><input name="hora_salida" type="time" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>PROGRAMADA</option><option>EN RUTA</option><option>CERRADA</option></select></div><div class="col-6"><label>Vehículo</label><select name="vehiculo_id" class="form-select"><option value="">-</option>{% for v in vehiculos %}<option value="{{v.id}}">{{v.placa}}</option>{% endfor %}</select></div><div class="col-6"><label>Conductor</label><select name="conductor_id" class="form-select"><option value="">-</option>{% for c in conductores %}<option value="{{c.id}}">{{c.nombres}}</option>{% endfor %}</select></div></div><button class="tm-btn w-100 mt-2">+ Nueva ruta</button></form>
      <div class="tm-search-row"><input class="tm-search" placeholder="Buscar ruta..."><a class="tm-btn" href="#">+ Nueva ruta</a></div><div class="tm-scroll"><table class="tm-table"><thead><tr><th>Código</th><th>Ruta</th><th>Salida</th><th>Estado</th><th></th></tr></thead><tbody>{% for r in rutas %}<tr><td>R-{{'%03d'%r.id}}</td><td>{{r.nombre}}<br><small>{{r.origen}} → {{r.destino}}</small></td><td>{{r.hora_salida or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td><a class="btn-icon" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}">›</a></td></tr>{% else %}<tr><td colspan="5" class="text-center text-muted">Sin rutas.</td></tr>{% endfor %}</tbody></table></div>
      <div class="tm-kpis"><div class="tm-kpi"><label>Programadas</label><strong>{{prog}}</strong></div><div class="tm-kpi"><label>En ruta</label><strong>{{en}}</strong></div><div class="tm-kpi"><label>Pendientes</label><strong>{{pend}}</strong></div><div class="tm-kpi"><label>Finalizadas</label><strong>{{fin}}</strong></div></div>
    </div></div>"""
    return render_page(body, rutas=rutas, vehiculos=vehiculos, conductores=conductores, hoy=hoy, prog=prog, en=en, pend=pend, fin=fin, badge=_estado_badge)

@login_required
def transporte_abordaje_home():
    hoy=today_str()
    rutas=rows_to_dict(execute('SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id ORDER BY CASE WHEN r.fecha=? THEN 0 ELSE 1 END, r.fecha DESC LIMIT 30',(hoy,), fetchall=True))
    total=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=?',(hoy,))
    body=_tm_css()+_tm_top('Abordaje trabajadores')+"""
      <div class="tm-info">Escanee QR / Código de barras desde la ruta o ingrese DNI manualmente.</div>
      <div class="tm-scanner-box mt-2"><div><i class="bi bi-qr-code-scan" style="font-size:42px"></i><br>Seleccione una ruta para escanear</div></div>
      <div class="tm-section">Rutas disponibles</div><div class="tm-list">{% for r in rutas %}<a class="tm-item" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}"><i class="bi bi-bus-front"></i>{{r.nombre}} · {{r.placa or 'SIN BUS'}}<i class="bi bi-chevron-right chev"></i></a>{% else %}<div class="tm-item text-muted">No hay rutas programadas.</div>{% endfor %}</div>
      <div class="tm-mini-cards"><div class="tm-mini"><small>Total abordados hoy</small>{{total}}</div></div>
    </div></div>"""
    return render_page(body, rutas=rutas, total=total)

@app.route('/transporte/abordaje')
@login_required
def transporte_abordaje_home_route():
    return transporte_abordaje_home()

@app.route('/transporte/pin')
@login_required
def transporte_pin_conductor():
    rows=rows_to_dict(execute('SELECT id,dni,nombres,movil_pin,movil_estado FROM transporte_conductores ORDER BY nombres', fetchall=True))
    body=_tm_css()+_tm_top('PIN conductor')+"""
      <div class="tm-info"><b>El PIN</b> es utilizado por el conductor para acceder al app móvil. Debe tener 4 dígitos.</div>
      <div class="tm-search-row"><input class="tm-search" placeholder="Buscar conductor..."></div><div class="tm-scroll"><table class="tm-table"><thead><tr><th>Conductor</th><th>PIN móvil</th><th>Acción</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.nombres}}<br><small>{{r.dni}}</small></td><td>{{r.movil_pin or '-'}}</td><td><a class="btn-icon" href="{{url_for('transporte_reset_pin', conductor_id=r.id)}}" onclick="return confirm('¿Resetear PIN?')"><i class="bi bi-arrow-clockwise"></i></a></td></tr>{% else %}<tr><td colspan="3" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}</tbody></table></div><a class="tm-btn w-100 mt-3 d-block text-center" href="{{url_for('transporte_conductores')}}">Gestionar conductores</a>
    </div></div>"""
    return render_page(body, rows=rows)


@app.route('/transporte/datos-maestros')
@login_required
def transporte_datos_maestros():
    body=_tm_css()+"""
    <style>
      .dm-list{border:1px solid #e5e7eb;border-radius:10px;background:#fff;overflow:hidden}.dm-item{display:flex;align-items:center;gap:13px;padding:14px 15px;border-bottom:1px solid #edf0ed;color:#111827;text-decoration:none;font-weight:850;font-size:13px}.dm-item:last-child{border-bottom:0}.dm-item i{font-size:21px;color:#08713b;width:23px;text-align:center}.dm-item .chev{margin-left:auto;color:#111;width:auto}.dm-note{background:#f2fbf4;border:1px solid #bde5c6;border-radius:9px;padding:9px 10px;color:#065f2a;font-size:11px;font-weight:800;margin-bottom:10px}
    </style>
    <div class="phone-wrap desktop-pad"><div class="page-card">
      <div class="tm-topbar"><a href="{{url_for('transporte_config')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Datos maestros</div></div>
      <div class="tm-page">
        <div class="dm-note">Aquí se administran las bases del módulo Transporte.</div>
        <div class="dm-list">
          <a class="dm-item" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i>Conductores <i class="bi bi-chevron-right chev"></i></a>
          <a class="dm-item" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i>Buses / Vehículos <i class="bi bi-chevron-right chev"></i></a>
          <a class="dm-item" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i>Rutas <i class="bi bi-chevron-right chev"></i></a>
          <a class="dm-item" href="{{url_for('transporte_carga_masiva')}}"><i class="bi bi-upload"></i>Carga masiva <i class="bi bi-chevron-right chev"></i></a>
          <a class="dm-item" href="{{url_for('transporte_pin_conductor')}}"><i class="bi bi-key"></i>PIN conductor <i class="bi bi-chevron-right chev"></i></a>
          <a class="dm-item" href="{{url_for('transporte')}}"><i class="bi bi-building"></i>Empresas transportistas <i class="bi bi-chevron-right chev"></i></a>
          <a class="dm-item" href="{{url_for('transporte')}}"><i class="bi bi-card-checklist"></i>Tipos de licencia <i class="bi bi-chevron-right chev"></i></a>
          <a class="dm-item" href="{{url_for('transporte')}}"><i class="bi bi-truck-front"></i>Tipos de vehículo <i class="bi bi-chevron-right chev"></i></a>
        </div>
      </div>
    </div></div>"""
    return render_page(body)

@app.route('/transporte/reporte-abordajes')
@login_required
def transporte_reporte_abordajes():
    hoy=today_str()
    rutas_hoy=scalar('SELECT COUNT(*) AS c FROM transporte_rutas WHERE fecha=?',(hoy,))
    abordaron=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=?',(hoy,))
    no_subieron=scalar("""SELECT COUNT(*) AS c FROM transporte_ruta_esperados e
        LEFT JOIN transporte_pasajeros p ON p.ruta_id=e.ruta_id AND p.dni=e.dni
        LEFT JOIN transporte_rutas r ON r.id=e.ruta_id
        WHERE r.fecha=? AND p.id IS NULL""",(hoy,))
    conductores=scalar('SELECT COUNT(*) AS c FROM transporte_conductores')
    buses=scalar('SELECT COUNT(*) AS c FROM transporte_vehiculos')
    vencidos=scalar("""SELECT COUNT(*) AS c FROM transporte_conductores WHERE (venc_licencia IS NOT NULL AND venc_licencia<>'' AND venc_licencia<=?) OR (venc_cert_medico IS NOT NULL AND venc_cert_medico<>'' AND venc_cert_medico<=?) OR (venc_sctr IS NOT NULL AND venc_sctr<>'' AND venc_sctr<=?)""",(hoy,hoy,hoy)) + scalar("""SELECT COUNT(*) AS c FROM transporte_vehiculos WHERE (soat_venc IS NOT NULL AND soat_venc<>'' AND soat_venc<=?) OR (revision_tecnica_venc IS NOT NULL AND revision_tecnica_venc<>'' AND revision_tecnica_venc<=?)""",(hoy,hoy))
    body=_tm_css()+"""
    <div class="phone-wrap desktop-pad"><div class="page-card">
      <div class="tm-topbar"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Reporte abordajes</div></div>
      <div class="tm-page">
        <form class="tm-form-card" method="get">
          <div class="row g-1">
            <div class="col-6"><label>Fecha inicio</label><input type="date" name="fi" value="{{hoy}}" class="form-control"></div>
            <div class="col-6"><label>Fecha fin</label><input type="date" name="ff" value="{{hoy}}" class="form-control"></div>
            <div class="col-12"><label>Estado</label><select class="form-select"><option>Todos</option><option>Abordó</option><option>No subió</option></select></div>
          </div>
          <button class="tm-btn w-100 mt-2">Generar reporte</button>
        </form>
        <div class="tm-kpis"><div class="tm-kpi"><label>Rutas hoy</label><strong>{{rutas_hoy}}</strong></div><div class="tm-kpi"><label>Abordaron hoy</label><strong>{{abordaron}}</strong></div><div class="tm-kpi"><label>No subieron</label><strong>{{no_subieron}}</strong></div><div class="tm-kpi"><label>Conductores</label><strong>{{conductores}}</strong></div><div class="tm-kpi"><label>Buses</label><strong>{{buses}}</strong></div><div class="tm-kpi"><label>Vencidos</label><strong>{{vencidos}}</strong></div></div>
        <a class="tm-btn w-100 mt-3 d-block text-center" href="{{url_for('exportar_transporte_pasajeros')}}"><i class="bi bi-file-earmark-excel"></i> Exportar Excel</a>
      </div>
    </div></div>"""
    return render_page(body, hoy=hoy, rutas_hoy=rutas_hoy, abordaron=abordaron, no_subieron=no_subieron, conductores=conductores, buses=buses, vencidos=vencidos)

# Reasignación de vistas existentes sin duplicar URL rules.
app.view_functions['transporte'] = transporte_mobile_home
app.view_functions['transporte_conductores'] = transporte_conductores_mobile
app.view_functions['transporte_vehiculos'] = transporte_vehiculos_mobile
app.view_functions['transporte_rutas'] = transporte_rutas_mobile
# endpoint lógico para url_for desde el menú principal
app.add_url_rule('/transporte/abordaje-home', 'transporte_abordaje_home', transporte_abordaje_home)

# ========================= PATCH 273 FINAL TRANSPORTE =========================
# Correcciones solicitadas:
# - Todas las pantallas transporte en formato móvil delgado con flecha volver.
# - Búsquedas activas por DNI/placa/ruta.
# - Botones +Nuevo funcionales.
# - Abordaje por QR/código de barras/digitación manual.
# - GPS explicado y funcional desde navegador/celular.
# - Configuración y datos maestros fuera de portada principal.

def _tf_css():
    return """
    <style>
      html,body{max-width:100%!important;overflow-x:hidden!important;background:#fff!important}.shell{max-width:430px!important;width:100%!important;margin:0 auto!important;padding:4px 6px!important;background:#fff!important}.phone-wrap,.page-card{max-width:390px!important;width:100%!important;margin:0 auto!important;border-radius:13px!important;overflow:hidden!important}.page-card{border:1px solid #e5e7eb!important;background:#fff!important;box-shadow:0 7px 18px rgba(0,0,0,.08)!important}.tf-hero{height:132px;background:linear-gradient(135deg,#075d2a,#2f773b);color:#fff;border-radius:13px 13px 0 0;text-align:center;position:relative;padding-top:15px}.tf-hero .back{position:absolute;left:11px;top:22px;color:#fff;text-decoration:none;font-size:34px;line-height:1}.tf-hero .cfg{position:absolute;right:11px;top:13px;color:#fff;text-decoration:none;border:1px solid rgba(255,255,255,.78);border-radius:9px;padding:7px 9px;font-weight:900;font-size:12px;display:flex;gap:5px;align-items:center}.tf-hero .bus{font-size:35px;line-height:1}.tf-hero .title{font-size:13px;font-weight:900;margin-top:8px;letter-spacing:.2px}.tf-card{background:#fff;border:1px solid #e5e7eb;border-radius:12px;box-shadow:0 7px 17px rgba(0,0,0,.10);padding:12px;margin:-38px 9px 10px;position:relative;z-index:5}.tf-section{font-size:12px;color:#065f2a;font-weight:900;margin:13px 2px 7px;text-transform:uppercase}.tf-grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}.tf-grid2{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}.tf-tile{height:78px;border-radius:9px;background:linear-gradient(135deg,#075d2a,#2f773b);color:#fff!important;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;font-weight:900;font-size:11px;line-height:1.12;box-shadow:0 7px 14px rgba(0,0,0,.16)}.tf-tile i{font-size:24px;margin-bottom:6px;color:#fff!important}.tf-top{height:56px;background:linear-gradient(135deg,#075d2a,#2f773b);color:#fff;display:flex;align-items:center;padding:0 12px;border-radius:13px 13px 0 0}.tf-top a{color:#fff;text-decoration:none;font-size:28px;width:31px}.tf-top .ttl{flex:1;text-align:center;font-weight:900;font-size:15px;margin-right:31px}.tf-page{background:#fff;padding:12px;border-radius:0 0 13px 13px}.tf-form{border:1px solid #dce8df;border-radius:11px;background:#fbfffb;padding:10px;margin-bottom:10px}.tf-form label{font-size:10px;color:#247037;font-weight:900;margin:0 0 3px}.tf-form .form-control,.tf-form .form-select{height:36px;border-radius:9px;border:1px solid #dfe7df;font-size:12px;font-weight:700}.tf-btn{background:#08713b;color:#fff!important;border:0;border-radius:9px;text-decoration:none;font-size:12px;font-weight:900;padding:9px 12px;display:inline-flex;align-items:center;justify-content:center;gap:5px}.tf-btn.w100{width:100%;display:flex}.tf-searchrow{display:flex;gap:8px;margin:10px 0}.tf-search{flex:1;height:38px;border:1px solid #dfe7df;border-radius:9px;padding:0 12px;font-size:12px;font-weight:800;text-transform:uppercase}.tf-tablewrap{overflow-x:auto;border-bottom:1px solid #edf0ed}.tf-table{width:100%;min-width:430px;border-collapse:collapse;font-size:11px}.tf-table th{background:#fafafa;color:#374151;font-weight:900;border-bottom:1px solid #e5e7eb;padding:8px 7px;text-align:left}.tf-table td{border-bottom:1px solid #edf0ed;padding:8px 7px;vertical-align:middle}.tf-table tr:last-child td{border-bottom:0}.tf-badge{display:inline-block;border-radius:6px;padding:3px 6px;font-size:9px;font-weight:900;background:#dcfce7;color:#166534}.tf-warn{background:#fff7ed;color:#ea580c}.tf-bad{background:#fee2e2;color:#dc2626}.tf-cards{display:grid;grid-template-columns:1fr;gap:8px;margin-top:10px}.tf-stat{border:1px solid #e5e7eb;border-radius:10px;padding:10px;color:#065f2a;font-weight:900;background:#fff}.tf-stat small{display:block;color:#6b7280;font-weight:800;font-size:12px}.tf-list{border:1px solid #e5e7eb;border-radius:10px;background:#fff;overflow:hidden}.tf-item{display:flex;align-items:center;gap:12px;text-decoration:none;color:#111827;border-bottom:1px solid #edf0ed;padding:13px;font-size:13px;font-weight:850}.tf-item:last-child{border-bottom:0}.tf-item i{color:#08713b;font-size:20px;width:24px;text-align:center}.tf-item .chev{margin-left:auto;color:#111}.tf-info{background:#eff6ff;border:1px solid #bfdbfe;color:#1e3a8a;border-radius:9px;padding:10px;font-size:12px;font-weight:800;margin-bottom:10px}.tf-scanner{height:142px;border:1px solid #dce8df;border-radius:10px;background:#fbfffb;display:grid;place-items:center;text-align:center;color:#08713b;font-weight:900;margin:10px 0}.tf-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin:10px 0}.tf-kpi{border:1px solid #9fcea8;border-radius:8px;background:#f7fff8;color:#065f2a;text-align:center;font-weight:900;padding:8px 3px}.tf-kpi label{display:block;font-size:9px;line-height:1.1}.tf-kpi strong{font-size:19px}.tf-map{height:190px;border:1px solid #dce8df;border-radius:10px;background:linear-gradient(135deg,#f3f4f6,#e5e7eb);display:grid;place-items:center;color:#08713b;text-align:center;font-weight:900}.tf-hide{display:none!important}.scan-ok{background:#dcfce7!important;border:1px solid #86efac!important;color:#166534!important;border-radius:9px!important;padding:8px!important;font-size:12px!important;font-weight:800!important}.scan-bad{background:#fee2e2!important;border:1px solid #fecaca!important;color:#991b1b!important;border-radius:9px!important;padding:8px!important;font-size:12px!important;font-weight:800!important}
    </style>
    """

def _tf_top(title, back='transporte'):
    return f"""<div class='phone-wrap desktop-pad'><div class='page-card'><div class='tf-top'><a href='{{{{url_for('{back}')}}}}'><i class='bi bi-chevron-left'></i></a><div class='ttl'>{title}</div></div><div class='tf-page'>"""

def _tf_end():
    return "</div></div></div>"

def _tf_badge(estado):
    e=(estado or 'ACTIVO').upper()
    cls='tf-badge'
    if 'VENC' in e or 'INACT' in e or 'NO APTO' in e: cls+=' tf-bad'
    elif 'PEND' in e or 'POR' in e: cls+=' tf-warn'
    return f"<span class='{cls}'>{e}</span>"

def _tf_filter_script(input_id='q', table_id='tbl'):
    return f"""
    <script>
    (function(){{
      const q=document.getElementById('{input_id}'), tbl=document.getElementById('{table_id}');
      if(!q||!tbl)return;
      q.addEventListener('input',()=>{{const v=q.value.toUpperCase().trim(); tbl.querySelectorAll('tbody tr').forEach(tr=>{{tr.style.display=tr.innerText.toUpperCase().includes(v)?'':'none';}});}});
    }})();
    function tfNuevo(){{const f=document.querySelector('.tf-form input, .tf-form select'); if(f){{f.scrollIntoView({{behavior:'smooth',block:'center'}}); setTimeout(()=>f.focus(),250);}}}}
    </script>
    """

@login_required
def tf_transporte_home():
    body=_tf_css()+"""
    <div class="phone-wrap desktop-pad"><div class="page-card">
      <div class="tf-hero">
        <a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>
        <a class="cfg" href="{{url_for('transporte_config')}}"><i class="bi bi-gear"></i> Config.</a>
        <div class="bus"><i class="bi bi-bus-front"></i></div><div class="title">MÓDULO TRANSPORTE</div>
      </div>
      <div class="tf-card">
        <div class="tf-grid3">
          <a class="tf-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i>Conductores</a>
          <a class="tf-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i>Buses</a>
          <a class="tf-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i>Rutas</a>
        </div>
        <div class="tf-section">Operación</div>
        <div class="tf-grid3">
          <a class="tf-tile" href="{{url_for('transporte_abordaje_home')}}"><i class="bi bi-people"></i>Abordaje<br>trabajadores</a>
          <a class="tf-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-geo-alt"></i>GPS /<br>seguimiento</a>
          <a class="tf-tile" href="{{url_for('transporte_rutas')}}?hoy=1"><i class="bi bi-calendar2-check"></i>Rutas de hoy</a>
          <a class="tf-tile" href="{{url_for('conductor_movil_login')}}"><i class="bi bi-phone"></i>Móvil<br>conductor</a>
        </div>
        <div class="tf-section">Control</div>
        <div class="tf-grid2">
          <a class="tf-tile" href="{{url_for('transporte_reporte_abordajes')}}"><i class="bi bi-file-earmark-text"></i>Reporte<br>abordajes</a>
          <a class="tf-tile" href="{{url_for('transporte_requisitos')}}"><i class="bi bi-exclamation-triangle"></i>Requisitos<br>vencidos</a>
        </div>
      </div>
    </div></div>"""
    return render_page(body)

@login_required
def tf_transporte_config():
    body=_tf_css()+_tf_top('Configuraciones')+"""
      <div class="tf-list">
        <a class="tf-item" href="{{url_for('transporte')}}"><i class="bi bi-sliders"></i>Parámetros del módulo <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte_pin_conductor')}}"><i class="bi bi-phone"></i>Relación conductor - móvil <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte')}}"><i class="bi bi-clipboard-check"></i>Causas de no abordaje <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte')}}"><i class="bi bi-geo-alt"></i>Zonas / paraderos <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte')}}"><i class="bi bi-bell"></i>Notificaciones <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte_carga_masiva')}}"><i class="bi bi-cloud-arrow-up"></i>Respaldos / carga masiva <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte_datos_maestros')}}"><i class="bi bi-database"></i>Datos maestros <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte')}}"><i class="bi bi-info-circle"></i>Acerca del módulo <i class="bi bi-chevron-right chev"></i></a>
      </div>
    """+_tf_end()
    return render_page(body)

@login_required
def tf_transporte_datos_maestros():
    body=_tf_css()+_tf_top('Datos maestros','transporte_config')+"""
      <div class="tf-list">
        <a class="tf-item" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i>Conductores <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i>Buses / vehículos <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i>Rutas <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte_carga_masiva')}}"><i class="bi bi-upload"></i>Carga masiva <i class="bi bi-chevron-right chev"></i></a>
        <a class="tf-item" href="{{url_for('transporte_pin_conductor')}}"><i class="bi bi-key"></i>PIN conductor <i class="bi bi-chevron-right chev"></i></a>
      </div>
    """+_tf_end()
    return render_page(body)

@login_required
def tf_transporte_conductores():
    if request.method=='POST':
        dni=limpiar_dni(request.form.get('dni')); nombres=limpiar_texto(request.form.get('nombres'))
        if len(dni)!=8 or not nombres:
            flash('Ingrese DNI de 8 dígitos y nombres del conductor.','danger'); return redirect(url_for('transporte_conductores'))
        pin=(request.form.get('movil_pin') or _pin_conductor(dni)).strip()[:8]
        vals=(dni,nombres,request.form.get('telefono',''),limpiar_texto(request.form.get('licencia')),limpiar_texto(request.form.get('categoria')),request.form.get('venc_licencia',''),request.form.get('venc_cert_medico',''),request.form.get('venc_sctr',''),limpiar_texto(request.form.get('estado') or 'ACTIVO'),dni,pin,'ACTIVO',now_str())
        try:
            execute('INSERT INTO transporte_conductores(dni,nombres,telefono,licencia,categoria,venc_licencia,venc_cert_medico,venc_sctr,estado,movil_usuario,movil_pin,movil_estado,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)', vals, commit=True)
            flash(f'Conductor registrado. PIN: {pin}','success')
        except Exception:
            execute('UPDATE transporte_conductores SET nombres=?,telefono=?,licencia=?,categoria=?,venc_licencia=?,venc_cert_medico=?,venc_sctr=?,estado=?,movil_usuario=?,movil_pin=?,movil_estado=? WHERE dni=?', (nombres,vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],vals[8],dni,pin,'ACTIVO',dni), commit=True)
            flash('Conductor actualizado.','success')
        return redirect(url_for('transporte_conductores'))
    rows=rows_to_dict(execute('SELECT * FROM transporte_conductores ORDER BY id DESC LIMIT 200', fetchall=True))
    total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO','APTO',''))
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<today_str() for c in ('venc_licencia','venc_cert_medico','venc_sctr')))
    body=_tf_css()+_tf_top('Conductores')+"""
      <form method="post" class="tf-form"><div class="row g-2">
        <div class="col-5"><label>DNI</label><input name="dni" class="form-control" maxlength="8" inputmode="numeric" required></div><div class="col-7"><label>Nombre</label><input name="nombres" class="form-control" required></div>
        <div class="col-6"><label>Teléfono</label><input name="telefono" class="form-control"></div><div class="col-6"><label>Licencia</label><input name="licencia" class="form-control"></div>
        <div class="col-6"><label>Categoría</label><input name="categoria" class="form-control"></div><div class="col-6"><label>PIN móvil</label><input name="movil_pin" class="form-control" placeholder="AUTO"></div>
        <div class="col-6"><label>Venc. licencia</label><input name="venc_licencia" type="date" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>ACTIVO</option><option>APTO</option><option>POR VENCER</option><option>VENCIDA</option><option>INACTIVO</option></select></div>
      </div><button class="tf-btn w100 mt-2">+ Guardar conductor</button></form>
      <div class="tf-searchrow"><input id="qcond" class="tf-search" placeholder="Buscar DNI o conductor..."><button type="button" onclick="tfNuevo()" class="tf-btn">+ Nuevo</button></div>
      <div class="tf-tablewrap"><table id="tblcond" class="tf-table"><thead><tr><th>DNI</th><th>Nombre</th><th>Licencia</th><th>Estado</th><th>PIN</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.nombres}}</td><td>{{r.categoria or r.licencia or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td>{{r.movil_pin or '-'}}</td></tr>{% else %}<tr><td colspan="5" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}</tbody></table></div>
      <div class="tf-cards"><div class="tf-stat"><small>Total conductores</small>{{total}}</div><div class="tf-stat"><small>Activos</small>{{activos}}</div><div class="tf-stat"><small>Vencidos</small>{{venc}}</div></div>
    """+_tf_end()+_tf_filter_script('qcond','tblcond')
    return render_page(body, rows=rows, total=total, activos=activos, venc=venc, badge=_tf_badge)

@login_required
def tf_transporte_vehiculos():
    if request.method=='POST':
        placa=limpiar_texto(request.form.get('placa'))
        if not placa:
            flash('Ingrese placa del vehículo.','danger'); return redirect(url_for('transporte_vehiculos'))
        vals=(placa,limpiar_texto(request.form.get('tipo') or 'BUS'),int(request.form.get('capacidad') or 0),limpiar_texto(request.form.get('empresa_transportista')),request.form.get('soat_venc',''),request.form.get('revision_tecnica_venc',''),limpiar_texto(request.form.get('gps_codigo')),limpiar_texto(request.form.get('estado') or 'ACTIVO'),now_str())
        try:
            execute('INSERT INTO transporte_vehiculos(placa,tipo,capacidad,empresa_transportista,soat_venc,revision_tecnica_venc,gps_codigo,estado,creado_en) VALUES(?,?,?,?,?,?,?,?,?)', vals, commit=True)
            flash('Vehículo registrado.','success')
        except Exception:
            execute('UPDATE transporte_vehiculos SET tipo=?,capacidad=?,empresa_transportista=?,soat_venc=?,revision_tecnica_venc=?,gps_codigo=?,estado=? WHERE placa=?', (vals[1],vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],placa), commit=True)
            flash('Vehículo actualizado.','success')
        return redirect(url_for('transporte_vehiculos'))
    rows=rows_to_dict(execute('SELECT * FROM transporte_vehiculos ORDER BY id DESC LIMIT 200', fetchall=True))
    total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO',''))
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<today_str() for c in ('soat_venc','revision_tecnica_venc')))
    body=_tf_css()+_tf_top('Buses / Vehículos')+"""
      <form method="post" class="tf-form"><div class="row g-2"><div class="col-5"><label>Placa</label><input name="placa" class="form-control" required></div><div class="col-7"><label>Tipo</label><select name="tipo" class="form-select"><option>BUS</option><option>MINIBUS</option><option>VAN</option><option>CAMIONETA</option></select></div><div class="col-6"><label>Capacidad</label><input name="capacidad" type="number" class="form-control"></div><div class="col-6"><label>Empresa</label><input name="empresa_transportista" class="form-control"></div><div class="col-6"><label>SOAT venc.</label><input name="soat_venc" type="date" class="form-control"></div><div class="col-6"><label>Rev. técnica</label><input name="revision_tecnica_venc" type="date" class="form-control"></div><div class="col-12"><label>Código GPS / placa GPS</label><input name="gps_codigo" class="form-control" placeholder="Opcional"></div></div><button class="tf-btn w100 mt-2">+ Guardar vehículo</button></form>
      <div class="tf-searchrow"><input id="qveh" class="tf-search" placeholder="Buscar placa o vehículo..."><button type="button" onclick="tfNuevo()" class="tf-btn">+ Nuevo</button></div><div class="tf-tablewrap"><table id="tblveh" class="tf-table"><thead><tr><th>Placa</th><th>Tipo</th><th>Capacidad</th><th>Estado</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.placa}}</td><td>{{r.tipo}}</td><td>{{r.capacidad}}</td><td>{{badge(r.estado)|safe}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin vehículos.</td></tr>{% endfor %}</tbody></table></div>
      <div class="tf-cards"><div class="tf-stat"><small>Total vehículos</small>{{total}}</div><div class="tf-stat"><small>Activos</small>{{activos}}</div><div class="tf-stat"><small>Vencidos</small>{{venc}}</div></div>
    """+_tf_end()+_tf_filter_script('qveh','tblveh')
    return render_page(body, rows=rows, total=total, activos=activos, venc=venc, badge=_tf_badge)

@login_required
def tf_transporte_rutas():
    if request.method=='POST':
        fecha=request.form.get('fecha') or today_str(); nombre=limpiar_texto(request.form.get('nombre') or 'RUTA')
        execute('INSERT INTO transporte_rutas(fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,vehiculo_id,conductor_id,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)', (fecha,nombre,limpiar_texto(request.form.get('origen')),limpiar_texto(request.form.get('destino')),limpiar_texto(request.form.get('sede')),request.form.get('hora_salida',''),request.form.get('hora_retorno',''),request.form.get('vehiculo_id') or None,request.form.get('conductor_id') or None,limpiar_texto(request.form.get('estado') or 'PROGRAMADA'),session.get('usuario'),now_str()), commit=True)
        flash('Ruta programada.','success'); return redirect(url_for('transporte_rutas'))
    hoy=today_str(); solo_hoy=request.args.get('hoy')=='1'
    where='WHERE r.fecha=?' if solo_hoy else ''
    params=(hoy,) if solo_hoy else ()
    rutas=rows_to_dict(execute(f'SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id {where} ORDER BY r.fecha DESC, r.id DESC LIMIT 120', params, fetchall=True))
    vehiculos=rows_to_dict(execute('SELECT id, placa, tipo, capacidad FROM transporte_vehiculos ORDER BY placa', fetchall=True))
    conductores=rows_to_dict(execute('SELECT id, dni, nombres FROM transporte_conductores ORDER BY nombres', fetchall=True))
    body=_tf_css()+_tf_top('Rutas de hoy' if solo_hoy else 'Rutas')+"""
      <form method="post" class="tf-form"><div class="row g-2"><div class="col-6"><label>Fecha</label><input name="fecha" type="date" value="{{hoy}}" class="form-control"></div><div class="col-6"><label>Ruta</label><input name="nombre" class="form-control" placeholder="R-001"></div><div class="col-6"><label>Origen</label><input name="origen" class="form-control"></div><div class="col-6"><label>Destino</label><input name="destino" class="form-control"></div><div class="col-6"><label>Salida</label><input name="hora_salida" type="time" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>PROGRAMADA</option><option>EN RUTA</option><option>CERRADA</option></select></div><div class="col-6"><label>Vehículo</label><select name="vehiculo_id" class="form-select"><option value="">-</option>{% for v in vehiculos %}<option value="{{v.id}}">{{v.placa}}</option>{% endfor %}</select></div><div class="col-6"><label>Conductor</label><select name="conductor_id" class="form-select"><option value="">-</option>{% for c in conductores %}<option value="{{c.id}}">{{c.nombres}}</option>{% endfor %}</select></div></div><button class="tf-btn w100 mt-2">+ Nueva ruta</button></form>
      <div class="tf-searchrow"><input id="qruta" class="tf-search" placeholder="Buscar ruta..."><button type="button" onclick="tfNuevo()" class="tf-btn">+ Nueva</button></div><div class="tf-tablewrap"><table id="tblruta" class="tf-table"><thead><tr><th>Código</th><th>Ruta</th><th>Salida</th><th>Estado</th><th></th></tr></thead><tbody>{% for r in rutas %}<tr><td>R-{{'%03d'%r.id}}</td><td>{{r.nombre}}<br><small>{{r.origen}} → {{r.destino}}</small></td><td>{{r.hora_salida or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td><a class="tf-btn" style="padding:4px 7px" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}">›</a></td></tr>{% else %}<tr><td colspan="5" class="text-center text-muted">Sin rutas.</td></tr>{% endfor %}</tbody></table></div>
    """+_tf_end()+_tf_filter_script('qruta','tblruta')
    return render_page(body, rutas=rutas, vehiculos=vehiculos, conductores=conductores, hoy=hoy, badge=_tf_badge)

@login_required
def tf_transporte_abordaje_home():
    hoy=today_str()
    rutas=rows_to_dict(execute('SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id ORDER BY CASE WHEN r.fecha=? THEN 0 ELSE 1 END, r.fecha DESC, r.hora_salida LIMIT 40',(hoy,), fetchall=True))
    total=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=?',(hoy,))
    body=_tf_css()+_tf_top('Abordaje trabajadores')+"""
      <div class="tf-info">Seleccione una ruta y luego escanee QR / código de barras o digite DNI manualmente.</div>
      <div class="tf-scanner"><div><i class="bi bi-qr-code-scan" style="font-size:42px"></i><br>Seleccione una ruta para escanear</div></div>
      <div class="tf-section">Rutas disponibles</div><div class="tf-list">{% for r in rutas %}<a class="tf-item" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}"><i class="bi bi-bus-front"></i><span>{{r.nombre}}<br><small>{{r.placa or 'SIN BUS'}} · {{r.hora_salida or '-'}}</small></span><i class="bi bi-chevron-right chev"></i></a>{% else %}<div class="tf-item text-muted">No hay rutas programadas.</div>{% endfor %}</div>
      <div class="tf-cards"><div class="tf-stat"><small>Total abordados hoy</small>{{total}}</div></div>
    """+_tf_end()
    return render_page(body, rutas=rutas, total=total)

@login_required
def tf_transporte_ruta_detalle(ruta_id):
    ruta=row_to_dict(execute('SELECT r.*, v.placa, v.tipo, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id WHERE r.id=?',(ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.','danger'); return redirect(url_for('transporte_abordaje_home'))
    pasajeros=rows_to_dict(execute('SELECT * FROM transporte_pasajeros WHERE ruta_id=? ORDER BY fecha_hora DESC',(ruta_id,), fetchall=True))
    ocupados=len(pasajeros); capacidad=int(ruta.get('capacidad') or 0); libres=max(0,capacidad-ocupados) if capacidad else 0
    body=_tf_css()+_tf_top('Abordaje ruta','transporte_abordaje_home')+"""
      <div class="tf-info"><b>{{ruta.nombre}}</b><br>{{ruta.origen or '-'}} → {{ruta.destino or '-'}}<br>Bus: {{ruta.placa or 'SIN BUS'}} | Conductor: {{ruta.conductor or '-'}} | Capacidad: {{capacidad or 'SIN DEFINIR'}} | Ocupados: {{ocupados}}</div>
      <form method="post" action="{{url_for('transporte_abordar', ruta_id=ruta.id)}}" id="frmAbordarTF" class="tf-form">
        <label>DNI / QR / Código de barras</label><div class="input-group"><input name="dni" id="dniTransporteTF" class="form-control" maxlength="20" placeholder="ESCANEAR O DIGITAR DNI" required autofocus><button type="button" class="tf-btn" onclick="abrirScanner('readerTransporteTF','dniTransporteTF')"><i class="bi bi-camera"></i></button></div>
        <div id="readerTransporteTF" class="scan-box mt-2" style="display:none"></div><div id="statusTransporteTF" class="field-help mt-2">Digita 8 dígitos o usa la cámara.</div>
        <select name="metodo" id="metodoTransporteTF" class="form-select mt-2"><option>DNI DIGITADO</option><option>QR</option><option>CODIGO DE BARRAS</option></select><input type="hidden" name="latitud" id="latitudTransTF"><input type="hidden" name="longitud" id="longitudTransTF"><button class="tf-btn w100 mt-2"><i class="bi bi-person-check"></i> Registrar subida</button>
      </form>
      <button class="tf-btn w100 mb-2" onclick="tfEnviarGps({{ruta.id}})"><i class="bi bi-geo-alt"></i> Enviar GPS de esta ruta</button>
      <div class="tf-kpis"><div class="tf-kpi"><label>Abordados</label><strong>{{ocupados}}</strong></div><div class="tf-kpi"><label>Libres</label><strong>{{libres}}</strong></div><div class="tf-kpi"><label>Capacidad</label><strong>{{capacidad or 0}}</strong></div></div>
      <div class="tf-tablewrap"><table class="tf-table"><thead><tr><th>Hora</th><th>DNI</th><th>Trabajador</th><th>Método</th></tr></thead><tbody>{% for p in pasajeros %}<tr><td>{{p.hora}}</td><td>{{p.dni}}</td><td>{{p.trabajador}}</td><td>{{p.metodo}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin abordajes.</td></tr>{% endfor %}</tbody></table></div>
    """+_tf_end()+"""
    <script>
    (function(){
      const input=document.getElementById('dniTransporteTF'), st=document.getElementById('statusTransporteTF'), lat=document.getElementById('latitudTransTF'), lon=document.getElementById('longitudTransTF');
      const dni=v=>{const m=String(v||'').match(/(?:^|\D)(\d{8})(?:\D|$)/); return m?m[1]:String(v||'').replace(/\D/g,'').slice(-8)};
      async function validar(){const d=dni(input.value); if(d.length<8){st.className='field-help mt-2';st.innerHTML='Esperando 8 dígitos...';return;} input.value=d; try{let r=await fetch('/api/trabajador/'+d,{cache:'no-store'});let j=await r.json(); if(j.ok){st.className='scan-ok mt-2';st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR'); if(typeof beep==='function')beep();} else {st.className='scan-bad mt-2';st.innerHTML='✕ '+(j.msg||'DNI no encontrado');}}catch(e){st.className='scan-bad mt-2';st.innerHTML='Error validando DNI';}}
      input.addEventListener('input',validar); input.addEventListener('paste',()=>setTimeout(validar,80)); input.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();document.getElementById('frmAbordarTF').requestSubmit();}});
      if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{lat.value=p.coords.latitude;lon.value=p.coords.longitude;},()=>{});}
    })();
    async function tfEnviarGps(rid){if(!navigator.geolocation){alert('GPS no disponible en este navegador');return;} navigator.geolocation.getCurrentPosition(async p=>{let fd=new FormData();fd.append('latitud',p.coords.latitude);fd.append('longitud',p.coords.longitude);let r=await fetch('/transporte/ruta/'+rid+'/gps',{method:'POST',body:fd});let j=await r.json();alert(j.msg||'GPS actualizado');location.reload();},()=>alert('Permite ubicación/GPS en el navegador. Debe estar en HTTPS.'));}
    </script>
    """
    return render_page(body, ruta=ruta, pasajeros=pasajeros, ocupados=ocupados, capacidad=capacidad, libres=libres)

@login_required
def tf_transporte_mapa_general():
    rutas=rows_to_dict(execute("SELECT r.*, v.placa, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id WHERE COALESCE(r.latitud,'')<>'' AND COALESCE(r.longitud,'')<>'' ORDER BY r.ultima_ubicacion DESC, r.id DESC LIMIT 30", fetchall=True))
    body=_tf_css()+_tf_top('GPS / Seguimiento')+"""
      <div class="tf-info"><b>¿Cómo configurar GPS?</b><br>1) Crea conductor y PIN. 2) Asigna el conductor a una ruta. 3) El conductor entra a <b>Móvil conductor</b> desde su celular. 4) Debe permitir ubicación/GPS y tocar <b>Enviar GPS</b>. Funciona mejor en HTTPS/Render.</div>
      {% for r in rutas %}<div class="tf-stat"><small>{{r.ultima_ubicacion or 'GPS registrado'}}</small><b>{{r.nombre}}</b><br>{{r.placa or 'SIN BUS'}} · {{r.conductor or '-' }}<br><span class="text-muted">Lat: {{r.latitud}} | Lon: {{r.longitud}}</span><br><a class="tf-btn mt-2" target="_blank" href="https://www.google.com/maps?q={{r.latitud}},{{r.longitud}}">Ver en mapa</a></div>{% else %}<div class="tf-map">Aún no hay rutas con GPS registrado.<br><small>Ingresa por Móvil conductor y envía ubicación.</small></div>{% endfor %}
    """+_tf_end()
    return render_page(body, rutas=rutas)

@login_required
def tf_transporte_pin_conductor():
    rows=rows_to_dict(execute('SELECT id,dni,nombres,movil_pin,movil_estado FROM transporte_conductores ORDER BY nombres', fetchall=True))
    body=_tf_css()+_tf_top('PIN conductor','transporte_config')+"""
      <div class="tf-info">El PIN es utilizado por el conductor para acceder al app móvil. Debe tener 4 dígitos. El acceso es con DNI conductor + PIN.</div>
      <div class="tf-searchrow"><input id="qpin" class="tf-search" placeholder="Buscar conductor..."></div><div class="tf-tablewrap"><table id="tblpin" class="tf-table"><thead><tr><th>Conductor</th><th>PIN móvil</th><th>Acción</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.nombres}}<br><small>{{r.dni}}</small></td><td>{{r.movil_pin or '-'}}</td><td><a class="tf-btn" style="padding:5px 8px" href="{{url_for('transporte_reset_pin', conductor_id=r.id)}}" onclick="return confirm('¿Resetear PIN?')"><i class="bi bi-arrow-clockwise"></i></a></td></tr>{% else %}<tr><td colspan="3" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}</tbody></table></div><a class="tf-btn w100 mt-3" href="{{url_for('transporte_conductores')}}">Gestionar conductores</a>
    """+_tf_end()+_tf_filter_script('qpin','tblpin')
    return render_page(body, rows=rows)

@login_required
def tf_transporte_reporte_abordajes():
    hoy=today_str()
    rutas_hoy=scalar('SELECT COUNT(*) AS c FROM transporte_rutas WHERE fecha=?',(hoy,)); abordaron=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=?',(hoy,))
    no_subieron=scalar("""SELECT COUNT(*) AS c FROM transporte_ruta_esperados e LEFT JOIN transporte_pasajeros p ON p.ruta_id=e.ruta_id AND p.dni=e.dni LEFT JOIN transporte_rutas r ON r.id=e.ruta_id WHERE r.fecha=? AND p.id IS NULL""",(hoy,))
    conductores=scalar('SELECT COUNT(*) AS c FROM transporte_conductores'); buses=scalar('SELECT COUNT(*) AS c FROM transporte_vehiculos')
    vencidos=scalar("""SELECT COUNT(*) AS c FROM transporte_conductores WHERE (venc_licencia IS NOT NULL AND venc_licencia<>'' AND venc_licencia<=?) OR (venc_cert_medico IS NOT NULL AND venc_cert_medico<>'' AND venc_cert_medico<=?) OR (venc_sctr IS NOT NULL AND venc_sctr<>'' AND venc_sctr<=?)""",(hoy,hoy,hoy)) + scalar("""SELECT COUNT(*) AS c FROM transporte_vehiculos WHERE (soat_venc IS NOT NULL AND soat_venc<>'' AND soat_venc<=?) OR (revision_tecnica_venc IS NOT NULL AND revision_tecnica_venc<>'' AND revision_tecnica_venc<=?)""",(hoy,hoy))
    body=_tf_css()+_tf_top('Reporte abordajes')+"""
      <form class="tf-form" method="get"><div class="row g-2"><div class="col-6"><label>Fecha inicio</label><input type="date" name="fi" value="{{hoy}}" class="form-control"></div><div class="col-6"><label>Fecha fin</label><input type="date" name="ff" value="{{hoy}}" class="form-control"></div><div class="col-12"><label>Estado</label><select class="form-select"><option>Todos</option><option>Abordó</option><option>No subió</option></select></div></div><button class="tf-btn w100 mt-2">Generar reporte</button></form>
      <div class="tf-kpis"><div class="tf-kpi"><label>Rutas hoy</label><strong>{{rutas_hoy}}</strong></div><div class="tf-kpi"><label>Abordaron hoy</label><strong>{{abordaron}}</strong></div><div class="tf-kpi"><label>No subieron</label><strong>{{no_subieron}}</strong></div><div class="tf-kpi"><label>Conductores</label><strong>{{conductores}}</strong></div><div class="tf-kpi"><label>Buses</label><strong>{{buses}}</strong></div><div class="tf-kpi"><label>Vencidos</label><strong>{{vencidos}}</strong></div></div>
      <a class="tf-btn w100 mt-2" href="{{url_for('exportar_transporte_pasajeros')}}"><i class="bi bi-file-earmark-excel"></i> Exportar Excel</a>
    """+_tf_end()
    return render_page(body, hoy=hoy, rutas_hoy=rutas_hoy, abordaron=abordaron, no_subieron=no_subieron, conductores=conductores, buses=buses, vencidos=vencidos)

@login_required
def tf_transporte_requisitos():
    hoy=today_str(); venc=[]
    for c in rows_to_dict(execute('SELECT * FROM transporte_conductores ORDER BY nombres', fetchall=True)):
        for campo,tipo in [('venc_licencia','Licencia'),('venc_cert_medico','Cert. médico'),('venc_sctr','SCTR')]:
            v=c.get(campo)
            if v and str(v)<=hoy: venc.append({'tipo':tipo,'persona':c.get('nombres'),'detalle':c.get('dni'),'vencimiento':v})
    for b in rows_to_dict(execute('SELECT * FROM transporte_vehiculos ORDER BY placa', fetchall=True)):
        for campo,tipo in [('soat_venc','SOAT'),('revision_tecnica_venc','Rev. técnica')]:
            v=b.get(campo)
            if v and str(v)<=hoy: venc.append({'tipo':tipo,'persona':b.get('placa'),'detalle':b.get('tipo'),'vencimiento':v})
    body=_tf_css()+_tf_top('Requisitos vencidos')+"""
      <div class="tf-list">{% for v in venc %}<div class="tf-item"><i class="bi bi-exclamation-triangle"></i><span><b>{{v.tipo}}</b><br>{{v.persona}} · {{v.detalle}}<br><small>Vence: {{v.vencimiento}}</small></span></div>{% else %}<div class="tf-item text-muted">No hay requisitos vencidos.</div>{% endfor %}</div>
    """+_tf_end()
    return render_page(body, venc=venc)

# Reasignación final de endpoints para corregir UI y navegación.
app.view_functions['transporte'] = tf_transporte_home
app.view_functions['transporte_config'] = tf_transporte_config
app.view_functions['transporte_datos_maestros'] = tf_transporte_datos_maestros
app.view_functions['transporte_conductores'] = tf_transporte_conductores
app.view_functions['transporte_vehiculos'] = tf_transporte_vehiculos
app.view_functions['transporte_rutas'] = tf_transporte_rutas
app.view_functions['transporte_ruta_detalle'] = tf_transporte_ruta_detalle
app.view_functions['transporte_mapa_general'] = tf_transporte_mapa_general
app.view_functions['transporte_pin_conductor'] = tf_transporte_pin_conductor
app.view_functions['transporte_reporte_abordajes'] = tf_transporte_reporte_abordajes
app.view_functions['transporte_requisitos'] = tf_transporte_requisitos
app.view_functions['transporte_abordaje_home'] = tf_transporte_abordaje_home


# ========================= PATCH TRANSPORTE OMAR 274 =========================
# Mejoras: carga masiva por módulo, KPIs verdes en 3 columnas, listas legales de licencia/categoría,
# y abordaje con QR / código de barras / digitación manual.
LICENCIAS_PERU = [
    'A-I','A-IIa','A-IIb','A-IIIa','A-IIIb','A-IIIc','B-I','B-IIa','B-IIb','B-IIc'
]
CATEGORIAS_CONDUCTOR_PERU = [
    'PARTICULAR / AUTO', 'TAXI / COLECTIVO', 'BUS / TRANSPORTE DE PERSONAS',
    'MINIBUS / VAN', 'CAMIÓN', 'REMOLQUE / SEMIRREMOLQUE', 'MOTO / MOTOTAXI'
]

def _option_list(items, selected=''):
    selected = (selected or '').upper()
    return ''.join([f"<option value='{x}' {'selected' if x.upper()==selected else ''}>{x}</option>" for x in items])

def _tf_extra_transport_css():
    return """
    <style>
    .tf-kpi-green{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:10px 0 12px}
    .tf-kpi-green .tf-kpi{background:linear-gradient(135deg,#08713b,#19a35b)!important;color:#fff!important;border:0!important;box-shadow:0 6px 14px rgba(8,113,59,.23)}
    .tf-kpi-green .tf-kpi label{color:#eaffef!important;font-size:10px!important;font-weight:900!important}.tf-kpi-green .tf-kpi strong{font-size:22px!important;color:#fff!important}
    .tf-mass{border:1px dashed #21a15a;background:#f0fff4;border-radius:10px;padding:9px;margin:8px 0 10px}.tf-mass label{font-size:10px;color:#08713b;font-weight:900}.tf-mass input{font-size:11px}.tf-help-mini{font-size:10px;color:#166534;font-weight:800;margin-top:4px;line-height:1.25}.tf-searchrow .tf-btn{white-space:nowrap}.tf-scanner.live{height:auto;min-height:120px;padding:8px;display:block}.tf-manual{display:grid;grid-template-columns:1fr auto;gap:8px;align-items:end}.tf-camera-btn{min-width:44px}.tf-mini-note{background:#ecfdf5;border:1px solid #bbf7d0;color:#166534;border-radius:9px;padding:8px;font-size:11px;font-weight:800;margin-bottom:8px}.tf-tablewrap{scrollbar-color:#08713b #e5e7eb}.tf-tablewrap::-webkit-scrollbar{height:9px}.tf-tablewrap::-webkit-scrollbar-thumb{background:#08713b;border-radius:999px}.tf-tablewrap::-webkit-scrollbar-track{background:#e5e7eb}
    </style>
    """

def _excel_date(v):
    if v is None: return ''
    if isinstance(v, (datetime, date)): return v.strftime('%Y-%m-%d')
    txt=str(v).strip()
    if not txt: return ''
    for fmt in ('%Y-%m-%d','%d/%m/%Y','%d-%m-%Y'):
        try: return datetime.strptime(txt[:10], fmt).strftime('%Y-%m-%d')
        except Exception: pass
    return txt[:10]

def _find_id_by_text(table, id_col, fields, value):
    value=limpiar_texto(value)
    if not value: return None
    for f in fields:
        r=row_to_dict(execute(f"SELECT {id_col} FROM {table} WHERE UPPER(COALESCE({f},''))=? LIMIT 1", (value,), fetchone=True))
        if r: return r.get(id_col)
    return None

def _importar_conductores_excel(file_storage):
    ok=upd=bad=0
    for row in _iter_excel_upload(file_storage):
        dni=limpiar_dni(_valor(row,['DNI','DOCUMENTO','DOC','NRO DOCUMENTO']))
        nombres=limpiar_texto(_valor(row,['NOMBRE','NOMBRES','CONDUCTOR','APELLIDOS Y NOMBRES','TRABAJADOR']))
        if len(dni)!=8 or not nombres:
            bad+=1; continue
        licencia=limpiar_texto(_valor(row,['LICENCIA','CLASE','TIPO LICENCIA'])) or 'A-IIIc'
        categoria=limpiar_texto(_valor(row,['CATEGORIA','CATEGORÍA','CAT'])) or 'BUS / TRANSPORTE DE PERSONAS'
        pin=limpiar_texto(_valor(row,['PIN','PIN MOVIL','PIN MÓVIL']), upper=False) or _pin_conductor(dni)
        vals=(dni,nombres,_valor(row,['TELEFONO','TELÉFONO','CELULAR']),licencia,categoria,_excel_date(_valor(row,['VENC LICENCIA','VENC. LICENCIA','VENCIMIENTO LICENCIA'])),_excel_date(_valor(row,['CERT MEDICO','CERT. MEDICO','CERTIFICADO MEDICO'])),_excel_date(_valor(row,['SCTR','VENC SCTR'])),limpiar_texto(_valor(row,['ESTADO']) or 'ACTIVO'),_valor(row,['OBSERVACION','OBSERVACIÓN']),dni,pin,'ACTIVO',now_str())
        try:
            execute('INSERT INTO transporte_conductores(dni,nombres,telefono,licencia,categoria,venc_licencia,venc_cert_medico,venc_sctr,estado,observacion,movil_usuario,movil_pin,movil_estado,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)', vals, commit=True); ok+=1
        except Exception:
            execute('UPDATE transporte_conductores SET nombres=?,telefono=?,licencia=?,categoria=?,venc_licencia=?,venc_cert_medico=?,venc_sctr=?,estado=?,observacion=?,movil_usuario=?,movil_pin=?,movil_estado=? WHERE dni=?', (vals[1],vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],vals[8],vals[9],dni,pin,'ACTIVO',dni), commit=True); upd+=1
    return ok,upd,bad

def _importar_vehiculos_excel(file_storage):
    ok=upd=bad=0
    for row in _iter_excel_upload(file_storage):
        placa=limpiar_texto(_valor(row,['PLACA','BUS','VEHICULO','VEHÍCULO']))
        if not placa: bad+=1; continue
        try: capacidad=int(float(_valor(row,['CAPACIDAD','ASIENTOS']) or 0))
        except Exception: capacidad=0
        vals=(placa,limpiar_texto(_valor(row,['TIPO','TIPO VEHICULO','TIPO VEHÍCULO']) or 'BUS'),capacidad,limpiar_texto(_valor(row,['EMPRESA','TRANSPORTISTA','EMPRESA TRANSPORTISTA'])),_excel_date(_valor(row,['SOAT','SOAT VENCE','VENC SOAT'])),_excel_date(_valor(row,['REV TECNICA','REV. TECNICA','REVISIÓN TÉCNICA','REVISION TECNICA'])),limpiar_texto(_valor(row,['GPS','CODIGO GPS','CÓDIGO GPS'])),limpiar_texto(_valor(row,['ESTADO']) or 'ACTIVO'),_valor(row,['OBSERVACION','OBSERVACIÓN']),now_str())
        try:
            execute('INSERT INTO transporte_vehiculos(placa,tipo,capacidad,empresa_transportista,soat_venc,revision_tecnica_venc,gps_codigo,estado,observacion,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?)', vals, commit=True); ok+=1
        except Exception:
            execute('UPDATE transporte_vehiculos SET tipo=?,capacidad=?,empresa_transportista=?,soat_venc=?,revision_tecnica_venc=?,gps_codigo=?,estado=?,observacion=? WHERE placa=?', (vals[1],vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],vals[8],placa), commit=True); upd+=1
    return ok,upd,bad

def _importar_rutas_excel(file_storage):
    ok=bad=0
    for row in _iter_excel_upload(file_storage):
        fecha=_excel_date(_valor(row,['FECHA','FECHA RUTA'])) or today_str()
        nombre=limpiar_texto(_valor(row,['RUTA','NOMBRE','NOMBRE RUTA']) or 'RUTA')
        if not nombre: bad+=1; continue
        placa=_valor(row,['PLACA','BUS','VEHICULO','VEHÍCULO'])
        dni_cond=limpiar_dni(_valor(row,['DNI CONDUCTOR','DOCUMENTO CONDUCTOR']))
        conductor_txt=_valor(row,['CONDUCTOR','NOMBRE CONDUCTOR'])
        vehiculo_id=_find_id_by_text('transporte_vehiculos','id',['placa'],placa)
        conductor_id=None
        if dni_cond:
            r=row_to_dict(execute('SELECT id FROM transporte_conductores WHERE dni=? LIMIT 1',(dni_cond,), fetchone=True)); conductor_id=r.get('id') if r else None
        if not conductor_id and conductor_txt:
            conductor_id=_find_id_by_text('transporte_conductores','id',['nombres'],conductor_txt)
        execute('INSERT INTO transporte_rutas(fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,vehiculo_id,conductor_id,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)', (fecha,nombre,limpiar_texto(_valor(row,['ORIGEN','PARADERO ORIGEN'])),limpiar_texto(_valor(row,['DESTINO','PARADERO DESTINO'])),limpiar_texto(_valor(row,['SEDE','FUNDO'])),str(_valor(row,['SALIDA','HORA SALIDA']))[:5],str(_valor(row,['RETORNO','HORA RETORNO']))[:5],vehiculo_id,conductor_id,limpiar_texto(_valor(row,['ESTADO']) or 'PROGRAMADA'),session.get('usuario'),now_str()), commit=True); ok+=1
    return ok,0,bad

@login_required
def transporte_upload_conductores():
    f=request.files.get('archivo')
    if not f:
        flash('Seleccione un Excel para carga masiva de conductores.', 'danger'); return redirect(url_for('transporte_conductores'))
    ok,upd,bad=_importar_conductores_excel(f)
    flash(f'Carga masiva conductores: nuevos {ok}, actualizados {upd}, omitidos {bad}.', 'success' if ok or upd else 'warning')
    return redirect(url_for('transporte_conductores'))

@login_required
def transporte_upload_vehiculos():
    f=request.files.get('archivo')
    if not f:
        flash('Seleccione un Excel para carga masiva de buses/vehículos.', 'danger'); return redirect(url_for('transporte_vehiculos'))
    ok,upd,bad=_importar_vehiculos_excel(f)
    flash(f'Carga masiva buses: nuevos {ok}, actualizados {upd}, omitidos {bad}.', 'success' if ok or upd else 'warning')
    return redirect(url_for('transporte_vehiculos'))

@login_required
def transporte_upload_rutas():
    f=request.files.get('archivo')
    if not f:
        flash('Seleccione un Excel para carga masiva de rutas.', 'danger'); return redirect(url_for('transporte_rutas'))
    ok,upd,bad=_importar_rutas_excel(f)
    flash(f'Carga masiva rutas: creadas {ok}, omitidas {bad}.', 'success' if ok else 'warning')
    return redirect(url_for('transporte_rutas'))

# Registrar endpoints de carga masiva si no existen.
try:
    app.add_url_rule('/transporte/conductores/carga-masiva', 'transporte_upload_conductores', transporte_upload_conductores, methods=['POST'])
    app.add_url_rule('/transporte/vehiculos/carga-masiva', 'transporte_upload_vehiculos', transporte_upload_vehiculos, methods=['POST'])
    app.add_url_rule('/transporte/rutas/carga-masiva', 'transporte_upload_rutas', transporte_upload_rutas, methods=['POST'])
except Exception:
    pass

@login_required
def tf_transporte_conductores_omar():
    if request.method=='POST':
        dni=limpiar_dni(request.form.get('dni')); nombres=limpiar_texto(request.form.get('nombres'))
        if len(dni)!=8 or not nombres:
            flash('Ingrese DNI de 8 dígitos y nombres del conductor.','danger'); return redirect(url_for('transporte_conductores'))
        pin=(request.form.get('movil_pin') or _pin_conductor(dni)).strip()
        vals=(dni,nombres,request.form.get('telefono',''),limpiar_texto(request.form.get('licencia') or 'A-IIIc'),limpiar_texto(request.form.get('categoria') or 'BUS / TRANSPORTE DE PERSONAS'),request.form.get('venc_licencia',''),request.form.get('venc_cert_medico',''),request.form.get('venc_sctr',''),limpiar_texto(request.form.get('estado') or 'ACTIVO'),limpiar_texto(request.form.get('observacion'), upper=False),dni,pin,'ACTIVO',now_str())
        try:
            execute('INSERT INTO transporte_conductores(dni,nombres,telefono,licencia,categoria,venc_licencia,venc_cert_medico,venc_sctr,estado,observacion,movil_usuario,movil_pin,movil_estado,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)', vals, commit=True); flash('Conductor registrado correctamente.','success')
        except Exception:
            execute('UPDATE transporte_conductores SET nombres=?,telefono=?,licencia=?,categoria=?,venc_licencia=?,venc_cert_medico=?,venc_sctr=?,estado=?,observacion=?,movil_usuario=?,movil_pin=?,movil_estado=? WHERE dni=?', (vals[1],vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],vals[8],vals[9],dni,pin,'ACTIVO',dni), commit=True); flash('Conductor actualizado correctamente.','success')
        return redirect(url_for('transporte_conductores'))
    rows=rows_to_dict(execute('SELECT * FROM transporte_conductores ORDER BY id DESC LIMIT 300', fetchall=True))
    hoy=today_str(); total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO','APTO',''))
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<hoy for c in ('venc_licencia','venc_cert_medico','venc_sctr')))
    body=_tf_css()+_tf_extra_transport_css()+_tf_top('Conductores')+"""
      <form method="post" class="tf-form"><div class="row g-2">
        <div class="col-5"><label>DNI</label><input name="dni" class="form-control" maxlength="8" inputmode="numeric" required></div><div class="col-7"><label>Nombre</label><input name="nombres" class="form-control" required></div>
        <div class="col-6"><label>Teléfono</label><input name="telefono" class="form-control"></div><div class="col-6"><label>Licencia</label><select name="licencia" class="form-select">{{lic_opts|safe}}</select></div>
        <div class="col-6"><label>Categoría</label><select name="categoria" class="form-select">{{cat_opts|safe}}</select></div><div class="col-6"><label>PIN móvil</label><input name="movil_pin" class="form-control" placeholder="AUTO"></div>
        <div class="col-6"><label>Venc. licencia</label><input name="venc_licencia" type="date" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>ACTIVO</option><option>APTO</option><option>POR VENCER</option><option>VENCIDA</option><option>INACTIVO</option></select></div>
      </div><button class="tf-btn w100 mt-2">+ Guardar conductor</button></form>
      <form method="post" action="{{url_for('transporte_upload_conductores')}}" enctype="multipart/form-data" class="tf-mass"><label><i class="bi bi-cloud-arrow-up"></i> Carga masiva conductores Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control mt-1" required><div class="tf-help-mini">Columnas sugeridas: DNI, NOMBRE, TELEFONO, LICENCIA, CATEGORIA, VENC LICENCIA, ESTADO, PIN.</div><button class="tf-btn w100 mt-2" type="submit"><i class="bi bi-upload"></i> Cargar conductores</button></form>
      <div class="tf-searchrow"><input id="qcond" class="tf-search" placeholder="Buscar DNI o conductor..."><button type="button" onclick="document.querySelector('.tf-mass input').click()" class="tf-btn"><i class="bi bi-upload"></i> Carga masiva</button></div>
      <div class="tf-tablewrap"><table id="tblcond" class="tf-table"><thead><tr><th>DNI</th><th>Nombre</th><th>Licencia</th><th>Estado</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.nombres}}</td><td>{{r.licencia or '-'}}<br><small>{{r.categoria or ''}}</small></td><td>{{badge(r.estado)|safe}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}</tbody></table></div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total conductores</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activos</label><strong>{{activos}}</strong></div><div class="tf-kpi"><label>Vencidos</label><strong>{{venc}}</strong></div></div>
    """+_tf_end()+_tf_filter_script('qcond','tblcond')
    return render_page(body, rows=rows,total=total,activos=activos,venc=venc,badge=_tf_badge,lic_opts=_option_list(LICENCIAS_PERU,'A-IIIc'),cat_opts=_option_list(CATEGORIAS_CONDUCTOR_PERU,'BUS / TRANSPORTE DE PERSONAS'))

@login_required
def tf_transporte_vehiculos_omar():
    if request.method=='POST':
        placa=limpiar_texto(request.form.get('placa'))
        if not placa:
            flash('Ingrese placa del vehículo.','danger'); return redirect(url_for('transporte_vehiculos'))
        try: cap=int(request.form.get('capacidad') or 0)
        except Exception: cap=0
        vals=(placa,limpiar_texto(request.form.get('tipo') or 'BUS'),cap,limpiar_texto(request.form.get('empresa_transportista')),request.form.get('soat_venc',''),request.form.get('revision_tecnica_venc',''),limpiar_texto(request.form.get('gps_codigo')),limpiar_texto(request.form.get('estado') or 'ACTIVO'),limpiar_texto(request.form.get('observacion'),upper=False),now_str())
        try: execute('INSERT INTO transporte_vehiculos(placa,tipo,capacidad,empresa_transportista,soat_venc,revision_tecnica_venc,gps_codigo,estado,observacion,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?)', vals, commit=True); flash('Bus/vehículo registrado.','success')
        except Exception: execute('UPDATE transporte_vehiculos SET tipo=?,capacidad=?,empresa_transportista=?,soat_venc=?,revision_tecnica_venc=?,gps_codigo=?,estado=?,observacion=? WHERE placa=?', (vals[1],vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],vals[8],placa), commit=True); flash('Bus/vehículo actualizado.','success')
        return redirect(url_for('transporte_vehiculos'))
    rows=rows_to_dict(execute('SELECT * FROM transporte_vehiculos ORDER BY id DESC LIMIT 300', fetchall=True)); hoy=today_str(); total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO',''))
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<hoy for c in ('soat_venc','revision_tecnica_venc')))
    body=_tf_css()+_tf_extra_transport_css()+_tf_top('Buses / Vehículos')+"""
      <form method="post" class="tf-form"><div class="row g-2"><div class="col-5"><label>Placa</label><input name="placa" class="form-control" required></div><div class="col-7"><label>Tipo</label><select name="tipo" class="form-select"><option>BUS</option><option>MINIBUS</option><option>VAN</option><option>CAMIONETA</option></select></div><div class="col-6"><label>Capacidad</label><input name="capacidad" type="number" class="form-control"></div><div class="col-6"><label>Empresa</label><input name="empresa_transportista" class="form-control"></div><div class="col-6"><label>SOAT venc.</label><input name="soat_venc" type="date" class="form-control"></div><div class="col-6"><label>Rev. técnica</label><input name="revision_tecnica_venc" type="date" class="form-control"></div><div class="col-12"><label>Código GPS / placa GPS</label><input name="gps_codigo" class="form-control"></div></div><button class="tf-btn w100 mt-2">+ Guardar bus</button></form>
      <form method="post" action="{{url_for('transporte_upload_vehiculos')}}" enctype="multipart/form-data" class="tf-mass"><label><i class="bi bi-cloud-arrow-up"></i> Carga masiva buses/vehículos Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control mt-1" required><div class="tf-help-mini">Columnas sugeridas: PLACA, TIPO, CAPACIDAD, EMPRESA, SOAT, REV TECNICA, GPS, ESTADO.</div><button class="tf-btn w100 mt-2" type="submit"><i class="bi bi-upload"></i> Cargar buses</button></form>
      <div class="tf-searchrow"><input id="qveh" class="tf-search" placeholder="Buscar placa o bus..."><button type="button" onclick="document.querySelector('.tf-mass input').click()" class="tf-btn"><i class="bi bi-upload"></i> Carga masiva</button></div><div class="tf-tablewrap"><table id="tblveh" class="tf-table"><thead><tr><th>Placa</th><th>Tipo</th><th>Capacidad</th><th>Estado</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.placa}}</td><td>{{r.tipo}}</td><td>{{r.capacidad}}</td><td>{{badge(r.estado)|safe}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin buses.</td></tr>{% endfor %}</tbody></table></div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total buses</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activos</label><strong>{{activos}}</strong></div><div class="tf-kpi"><label>Vencidos</label><strong>{{venc}}</strong></div></div>
    """+_tf_end()+_tf_filter_script('qveh','tblveh')
    return render_page(body, rows=rows,total=total,activos=activos,venc=venc,badge=_tf_badge)

@login_required
def tf_transporte_rutas_omar():
    if request.method=='POST':
        fecha=request.form.get('fecha') or today_str(); nombre=limpiar_texto(request.form.get('nombre') or 'RUTA')
        execute('INSERT INTO transporte_rutas(fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,vehiculo_id,conductor_id,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)', (fecha,nombre,limpiar_texto(request.form.get('origen')),limpiar_texto(request.form.get('destino')),limpiar_texto(request.form.get('sede')),request.form.get('hora_salida',''),request.form.get('hora_retorno',''),request.form.get('vehiculo_id') or None,request.form.get('conductor_id') or None,limpiar_texto(request.form.get('estado') or 'PROGRAMADA'),session.get('usuario'),now_str()), commit=True)
        flash('Ruta programada.','success'); return redirect(url_for('transporte_rutas'))
    hoy=today_str(); solo_hoy=request.args.get('hoy')=='1'; where='WHERE r.fecha=?' if solo_hoy else ''; params=(hoy,) if solo_hoy else ()
    rutas=rows_to_dict(execute(f'SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id {where} ORDER BY r.fecha DESC, r.id DESC LIMIT 200', params, fetchall=True))
    vehiculos=rows_to_dict(execute('SELECT id, placa, tipo, capacidad FROM transporte_vehiculos ORDER BY placa', fetchall=True)); conductores=rows_to_dict(execute('SELECT id, dni, nombres FROM transporte_conductores ORDER BY nombres', fetchall=True))
    prog=len(rutas); en=sum(1 for r in rutas if (r.get('estado') or '').upper()=='EN RUTA'); pend=sum(1 for r in rutas if (r.get('estado') or '').upper() in ('PROGRAMADA','PENDIENTE')); fin=sum(1 for r in rutas if (r.get('estado') or '').upper() in ('FINALIZADA','CERRADA'))
    body=_tf_css()+_tf_extra_transport_css()+_tf_top('Rutas de hoy' if solo_hoy else 'Rutas')+"""
      <form method="post" class="tf-form"><div class="row g-2"><div class="col-6"><label>Fecha</label><input name="fecha" type="date" value="{{hoy}}" class="form-control"></div><div class="col-6"><label>Ruta</label><input name="nombre" class="form-control" placeholder="R-001"></div><div class="col-6"><label>Origen</label><input name="origen" class="form-control"></div><div class="col-6"><label>Destino</label><input name="destino" class="form-control"></div><div class="col-6"><label>Salida</label><input name="hora_salida" type="time" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>PROGRAMADA</option><option>EN RUTA</option><option>CERRADA</option></select></div><div class="col-6"><label>Bus</label><select name="vehiculo_id" class="form-select"><option value="">-</option>{% for v in vehiculos %}<option value="{{v.id}}">{{v.placa}}</option>{% endfor %}</select></div><div class="col-6"><label>Conductor</label><select name="conductor_id" class="form-select"><option value="">-</option>{% for c in conductores %}<option value="{{c.id}}">{{c.nombres}}</option>{% endfor %}</select></div></div><button class="tf-btn w100 mt-2">+ Nueva ruta</button></form>
      <form method="post" action="{{url_for('transporte_upload_rutas')}}" enctype="multipart/form-data" class="tf-mass"><label><i class="bi bi-cloud-arrow-up"></i> Carga masiva rutas Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control mt-1" required><div class="tf-help-mini">Columnas sugeridas: FECHA, RUTA, ORIGEN, DESTINO, SALIDA, PLACA, DNI CONDUCTOR, ESTADO.</div><button class="tf-btn w100 mt-2" type="submit"><i class="bi bi-upload"></i> Cargar rutas</button></form>
      <div class="tf-searchrow"><input id="qruta" class="tf-search" placeholder="Buscar ruta..."><button type="button" onclick="document.querySelector('.tf-mass input').click()" class="tf-btn"><i class="bi bi-upload"></i> Carga masiva</button></div><div class="tf-tablewrap"><table id="tblruta" class="tf-table"><thead><tr><th>Código</th><th>Ruta</th><th>Salida</th><th>Estado</th><th></th></tr></thead><tbody>{% for r in rutas %}<tr><td>R-{{'%03d'%r.id}}</td><td>{{r.nombre}}<br><small>{{r.origen}} → {{r.destino}}</small></td><td>{{r.hora_salida or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td><a class="tf-btn" style="padding:4px 7px" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}">›</a></td></tr>{% else %}<tr><td colspan="5" class="text-center text-muted">Sin rutas.</td></tr>{% endfor %}</tbody></table></div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Programadas</label><strong>{{prog}}</strong></div><div class="tf-kpi"><label>En ruta</label><strong>{{en}}</strong></div><div class="tf-kpi"><label>Finalizadas</label><strong>{{fin}}</strong></div></div>
    """+_tf_end()+_tf_filter_script('qruta','tblruta')
    return render_page(body,rutas=rutas,vehiculos=vehiculos,conductores=conductores,hoy=hoy,prog=prog,en=en,pend=pend,fin=fin,badge=_tf_badge)

@login_required
def tf_transporte_abordaje_home_omar():
    hoy=today_str()
    rutas=rows_to_dict(execute("SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id WHERE r.fecha=? OR UPPER(COALESCE(r.estado,'')) IN ('PROGRAMADA','EN RUTA') ORDER BY CASE WHEN r.fecha=? THEN 0 ELSE 1 END, r.fecha DESC, r.hora_salida LIMIT 40",(hoy,hoy), fetchall=True))
    total=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=?',(hoy,)); rutas_cnt=len(rutas)
    body=_tf_css()+_tf_extra_transport_css()+_tf_top('Abordaje trabajadores')+"""
      <div class="tf-info">Seleccione una ruta y luego escanee QR / código de barras o digite DNI manualmente.</div>
      <div class="tf-mini-note"><i class="bi bi-check-circle"></i> Lectura habilitada: QR, código de barras y DNI manual. La cámara se abre dentro de cada ruta.</div>
      <div class="tf-section">Rutas disponibles</div><div class="tf-list">{% for r in rutas %}<a class="tf-item" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}"><i class="bi bi-bus-front"></i><span>{{r.nombre}}<br><small>{{r.placa or 'SIN BUS'}} · {{r.hora_salida or '-'}} · {{r.conductor or '-'}}</small></span><i class="bi bi-chevron-right chev"></i></a>{% else %}<div class="tf-item text-muted">No hay rutas programadas.</div>{% endfor %}</div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Rutas visibles</label><strong>{{rutas_cnt}}</strong></div><div class="tf-kpi"><label>Abordados hoy</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Pendientes</label><strong>{{rutas_cnt}}</strong></div></div>
    """+_tf_end()
    return render_page(body,rutas=rutas,total=total,rutas_cnt=rutas_cnt)

@login_required
def tf_transporte_ruta_detalle_omar(ruta_id):
    ruta=row_to_dict(execute('SELECT r.*, v.placa, v.tipo, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id WHERE r.id=?',(ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.','danger'); return redirect(url_for('transporte_abordaje_home'))
    pasajeros=rows_to_dict(execute('SELECT * FROM transporte_pasajeros WHERE ruta_id=? ORDER BY fecha_hora DESC',(ruta_id,), fetchall=True)); ocupados=len(pasajeros); capacidad=int(ruta.get('capacidad') or 0); libres=max(0,capacidad-ocupados) if capacidad else 0
    body=_tf_css()+_tf_extra_transport_css()+_tf_top('Abordaje ruta','transporte_abordaje_home')+"""
      <div class="tf-info"><b>{{ruta.nombre}}</b><br>{{ruta.origen or '-'}} → {{ruta.destino or '-'}}<br>Bus: {{ruta.placa or 'SIN BUS'}} | Conductor: {{ruta.conductor or '-'}} | Capacidad: {{capacidad or 'SIN DEFINIR'}} | Ocupados: {{ocupados}}</div>
      <form method="post" action="{{url_for('transporte_abordar', ruta_id=ruta.id)}}" id="frmAbordarTF" class="tf-form">
        <label>DNI / QR / Código de barras</label><div class="tf-manual"><input name="dni" id="dniTransporteTF" class="form-control" maxlength="20" placeholder="ESCANEAR O DIGITAR DNI" required autofocus><button type="button" class="tf-btn tf-camera-btn" onclick="abrirScanner('readerTransporteTF','dniTransporteTF')"><i class="bi bi-camera"></i></button></div>
        <div id="readerTransporteTF" class="scan-box mt-2 live" style="display:none"></div><div id="statusTransporteTF" class="field-help mt-2">Digita 8 dígitos o usa la cámara.</div>
        <select name="metodo" id="metodoTransporteTF" class="form-select mt-2"><option>DNI DIGITADO</option><option>QR</option><option>CODIGO DE BARRAS</option></select><input type="hidden" name="latitud" id="latitudTransTF"><input type="hidden" name="longitud" id="longitudTransTF"><button class="tf-btn w100 mt-2"><i class="bi bi-person-check"></i> Registrar subida</button>
      </form>
      <button class="tf-btn w100 mb-2" onclick="tfEnviarGps({{ruta.id}})"><i class="bi bi-geo-alt"></i> Enviar GPS de esta ruta</button>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Abordados</label><strong>{{ocupados}}</strong></div><div class="tf-kpi"><label>Libres</label><strong>{{libres}}</strong></div><div class="tf-kpi"><label>Capacidad</label><strong>{{capacidad or 0}}</strong></div></div>
      <div class="tf-tablewrap"><table class="tf-table"><thead><tr><th>Hora</th><th>DNI</th><th>Trabajador</th><th>Método</th></tr></thead><tbody>{% for p in pasajeros %}<tr><td>{{p.hora}}</td><td>{{p.dni}}</td><td>{{p.trabajador}}</td><td>{{p.metodo}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin abordajes.</td></tr>{% endfor %}</tbody></table></div>
    """+_tf_end()+"""
    <script>
    (function(){
      const input=document.getElementById('dniTransporteTF'), st=document.getElementById('statusTransporteTF'), metodo=document.getElementById('metodoTransporteTF'), lat=document.getElementById('latitudTransTF'), lon=document.getElementById('longitudTransTF');
      const dni=v=>{const m=String(v||'').match(/(?:^|\D)(\d{8})(?:\D|$)/); return m?m[1]:String(v||'').replace(/\D/g,'').slice(-8)};
      async function validar(){const d=dni(input.value); if(d.length<8){st.className='field-help mt-2';st.innerHTML='Esperando 8 dígitos...';return;} input.value=d; try{let r=await fetch('/api/trabajador/'+d,{cache:'no-store'});let j=await r.json(); if(j.ok){st.className='scan-ok mt-2';st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR')+' encontrado. Puede registrar.'; if(typeof beep==='function')beep();} else {st.className='scan-bad mt-2';st.innerHTML='✕ '+(j.msg||'DNI no encontrado');}}catch(e){st.className='scan-bad mt-2';st.innerHTML='Error validando DNI';}}
      input.addEventListener('input',()=>{metodo.value='DNI DIGITADO'; validar();}); input.addEventListener('paste',()=>setTimeout(validar,80)); input.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();document.getElementById('frmAbordarTF').requestSubmit();}});
      const old=window.abrirScanner; window.abrirScanner=function(readerId,inputId){metodo.value='QR'; return old(readerId,inputId)};
      if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{lat.value=p.coords.latitude;lon.value=p.coords.longitude;},()=>{});}
    })();
    async function tfEnviarGps(rid){if(!navigator.geolocation){alert('GPS no disponible en este navegador');return;} navigator.geolocation.getCurrentPosition(async p=>{let fd=new FormData();fd.append('latitud',p.coords.latitude);fd.append('longitud',p.coords.longitude);let r=await fetch('/transporte/ruta/'+rid+'/gps',{method:'POST',body:fd});let j=await r.json();alert(j.msg||'GPS actualizado');location.reload();},()=>alert('Permite ubicación/GPS en el navegador. Debe estar en HTTPS.'));}
    </script>
    """
    return render_page(body,ruta=ruta,pasajeros=pasajeros,ocupados=ocupados,capacidad=capacidad,libres=libres)

# Reasignación final del patch 274.
app.view_functions['transporte_conductores'] = tf_transporte_conductores_omar
app.view_functions['transporte_vehiculos'] = tf_transporte_vehiculos_omar
app.view_functions['transporte_rutas'] = tf_transporte_rutas_omar
app.view_functions['transporte_abordaje_home'] = tf_transporte_abordaje_home_omar
app.view_functions['transporte_ruta_detalle'] = tf_transporte_ruta_detalle_omar

# ========================= PATCH TRANSPORTE OMAR 275 =========================
def _tpl_xlsx(headers, example, filename, sheet='PLANTILLA'):
    return excel_response(headers, [dict(zip([h.lower() for h in headers], example))], filename, sheet)

@app.route('/transporte/conductores/plantilla')
@login_required
def transporte_plantilla_conductores():
    headers=['DNI','NOMBRE','TELEFONO','LICENCIA','CATEGORIA','VENC LICENCIA','CERT MEDICO','SCTR','ESTADO','PIN','OBSERVACION']
    example=['74324033','JOSE GARCIA','999999999','A-IIIc','BUS / TRANSPORTE DE PERSONAS','2026-12-31','2026-12-31','2026-12-31','ACTIVO','AUTO','']
    return _tpl_xlsx(headers, example, 'plantilla_conductores.xlsx', 'CONDUCTORES')

@app.route('/transporte/buses/plantilla')
@login_required
def transporte_plantilla_vehiculos():
    headers=['PLACA','TIPO','CAPACIDAD','EMPRESA','SOAT','REV TECNICA','GPS','ESTADO','OBSERVACION']
    example=['ABC-123','BUS','50','TRANSPORTISTA SAC','2026-12-31','2026-12-31','GPS-001','ACTIVO','']
    return _tpl_xlsx(headers, example, 'plantilla_buses.xlsx', 'BUSES')

@app.route('/transporte/rutas/plantilla')
@login_required
def transporte_plantilla_rutas():
    headers=['FECHA','RUTA','ORIGEN','DESTINO','SEDE','SALIDA','RETORNO','PLACA','DNI CONDUCTOR','ESTADO']
    example=[today_str(),'RUTA NORTE','PARADERO 1','FUNDO','SEDE','06:00','15:00','ABC-123','74324033','PROGRAMADA']
    return _tpl_xlsx(headers, example, 'plantilla_rutas.xlsx', 'RUTAS')

@app.route('/transporte/conductores/<int:item_id>/toggle')
@login_required
def transporte_toggle_conductor(item_id):
    r=row_to_dict(execute('SELECT estado FROM transporte_conductores WHERE id=?',(item_id,), fetchone=True))
    if r:
        nuevo='INACTIVO' if (r.get('estado') or '').upper() in ('ACTIVO','APTO','') else 'ACTIVO'
        execute('UPDATE transporte_conductores SET estado=?, movil_estado=? WHERE id=?',(nuevo,nuevo,item_id), commit=True)
        flash('Conductor '+('desactivado.' if nuevo=='INACTIVO' else 'activado.'),'success')
    return redirect(url_for('transporte_conductores'))

@app.route('/transporte/buses/<int:item_id>/toggle')
@login_required
def transporte_toggle_vehiculo(item_id):
    r=row_to_dict(execute('SELECT estado FROM transporte_vehiculos WHERE id=?',(item_id,), fetchone=True))
    if r:
        nuevo='INACTIVO' if (r.get('estado') or '').upper()=='ACTIVO' else 'ACTIVO'
        execute('UPDATE transporte_vehiculos SET estado=? WHERE id=?',(nuevo,item_id), commit=True)
        flash('Bus '+('desactivado.' if nuevo=='INACTIVO' else 'activado.'),'success')
    return redirect(url_for('transporte_vehiculos'))

@app.route('/transporte/rutas/<int:item_id>/toggle')
@login_required
def transporte_toggle_ruta(item_id):
    r=row_to_dict(execute('SELECT estado FROM transporte_rutas WHERE id=?',(item_id,), fetchone=True))
    if r:
        nuevo='INACTIVA' if (r.get('estado') or '').upper() not in ('INACTIVA','CERRADA','FINALIZADA') else 'PROGRAMADA'
        execute('UPDATE transporte_rutas SET estado=? WHERE id=?',(nuevo,item_id), commit=True)
        flash('Ruta '+('desactivada.' if nuevo=='INACTIVA' else 'activada.'),'success')
    return redirect(url_for('transporte_rutas'))

def _tf_switch(active, url):
    checked='checked' if active else ''
    return f"<a class='tf-switch {checked}' href='{url}' title='Activar / desactivar'><span></span></a>"

def _tf_transport_css_275():
    return _tf_extra_transport_css()+'''
    <style>
    .tf-mass{border:1.5px dashed #14a34a!important;background:#ecfdf3!important;border-radius:12px!important;padding:11px!important;margin:10px 0 14px!important}
    .tf-mass label{font-size:12px!important;color:#086b32!important;font-weight:900!important;margin-bottom:6px!important;display:block}.tf-mass .form-control{height:38px!important;background:white!important}.tf-help-mini{font-size:11px!important;color:#08713b!important;font-weight:900!important;margin:7px 0!important;line-height:1.25!important}.tf-template-btn{display:flex;align-items:center;justify-content:center;gap:6px;background:#f0fff4!important;border:1px solid #b7e4c7!important;color:#08713b!important;border-radius:9px!important;font-weight:900!important;text-decoration:none!important;padding:10px 8px!important;margin-top:8px!important}.tf-searchrow{display:grid!important;grid-template-columns:1fr auto!important;gap:8px!important;align-items:center!important}.tf-kpi-green{grid-template-columns:repeat(3,1fr)!important}.tf-kpi-green .tf-kpi{text-align:center!important;border-radius:8px!important;padding:10px 6px!important;background:#12964f!important}.tf-kpi-green .tf-kpi strong{font-size:26px!important;line-height:1!important}.tf-kpi-green .tf-kpi label{font-size:11px!important;line-height:1.1!important;display:block;margin-bottom:5px}.tf-switch{width:42px;height:23px;border-radius:999px;background:#cbd5e1;display:inline-flex;align-items:center;padding:2px;text-decoration:none}.tf-switch span{width:19px;height:19px;border-radius:999px;background:white;display:block;box-shadow:0 1px 4px rgba(0,0,0,.25)}.tf-switch.checked{background:#07823f}.tf-switch.checked span{margin-left:19px}.tf-base-title{font-size:11px;color:#08713b;font-weight:900;margin:8px 0 5px;text-transform:uppercase}.tf-gps-card{background:#fff;border:1px solid #dfe7df;border-radius:14px;padding:12px;margin:10px 0;box-shadow:0 5px 16px rgba(0,0,0,.06)}.tf-gps-driver{display:flex;gap:10px;align-items:center;border-bottom:1px solid #edf2ed;padding-bottom:10px;margin-bottom:10px}.tf-gps-avatar{width:48px;height:48px;border-radius:999px;background:#dff6df;color:#08713b;display:grid;place-items:center;font-size:25px}.tf-gps-driver b{font-size:14px;color:#102a43}.tf-gps-meta{font-size:11px;color:#475569;font-weight:800}.tf-gps-row{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin:8px 0}.tf-gps-box{background:#f8fff9;border:1px solid #dbeadb;border-radius:10px;padding:8px}.tf-gps-box small{display:block;color:#64748b;font-size:10px;font-weight:900}.tf-gps-box strong{font-size:12px;color:#065f2a}.tf-table td,.tf-table th{white-space:nowrap}.tf-home-note{background:#eff6ff;border:1px solid #bfdbfe;color:#0b2e83;border-radius:10px;padding:10px;font-size:12px;font-weight:800;margin-top:10px}.tf-mobile-sub{border:1px solid #d8eadc;background:#f7fff8;border-radius:12px;padding:10px;margin:9px 0}.tf-mobile-sub .tf-item{margin-bottom:7px}
    </style>
    '''

@login_required
def tf_transporte_conductores_275():
    if request.method=='POST':
        return tf_transporte_conductores_omar()
    rows=rows_to_dict(execute('SELECT * FROM transporte_conductores ORDER BY id DESC LIMIT 300', fetchall=True))
    hoy=today_str(); total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO','APTO',''))
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<hoy for c in ('venc_licencia','venc_cert_medico','venc_sctr')))
    body=_tf_css()+_tf_transport_css_275()+_tf_top('Conductores')+'''
      <form method="post" enctype="multipart/form-data" action="{{url_for('transporte_upload_conductores')}}" class="tf-mass">
        <label><i class="bi bi-cloud-arrow-up"></i> Carga masiva conductores Excel</label>
        <input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required>
        <div class="tf-help-mini">Columnas sugeridas: DNI, NOMBRE, TELEFONO, LICENCIA, CATEGORIA, VENC LICENCIA, ESTADO, PIN.</div>
        <button class="tf-btn w100" type="submit"><i class="bi bi-upload"></i> Cargar conductores</button>
        <a class="tf-template-btn" href="{{url_for('transporte_plantilla_conductores')}}"><i class="bi bi-file-earmark-excel"></i> Descargar plantilla</a>
      </form>
      <div class="tf-searchrow"><input id="qcond" class="tf-search" placeholder="Buscar DNI o conductor..."><a class="tf-template-btn" style="margin:0;padding:9px 11px" href="{{url_for('transporte_plantilla_conductores')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total conductores</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activos</label><strong>{{activos}}</strong></div><div class="tf-kpi"><label>Vencidos</label><strong>{{venc}}</strong></div></div>
      <div class="tf-base-title">Base de conductores</div>
      <div class="tf-tablewrap"><table id="tblcond" class="tf-table"><thead><tr><th>DNI</th><th>Nombre</th><th>Licencia</th><th>Categoría</th><th>Estado</th><th>Activo</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.nombres}}</td><td>{{r.licencia or '-'}}</td><td>{{r.categoria or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td>{{switch((r.estado or '').upper() in ['ACTIVO','APTO',''], url_for('transporte_toggle_conductor', item_id=r.id))|safe}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}</tbody></table></div>
      <form method="post" class="tf-form mt-2"><div class="row g-2"><div class="col-5"><label>DNI</label><input name="dni" class="form-control" maxlength="8" inputmode="numeric" required></div><div class="col-7"><label>Nombre</label><input name="nombres" class="form-control" required></div><div class="col-6"><label>Teléfono</label><input name="telefono" class="form-control"></div><div class="col-6"><label>Licencia</label><select name="licencia" class="form-select">{{lic_opts|safe}}</select></div><div class="col-6"><label>Categoría</label><select name="categoria" class="form-select">{{cat_opts|safe}}</select></div><div class="col-6"><label>PIN móvil</label><input name="movil_pin" class="form-control" placeholder="AUTO"></div><div class="col-6"><label>Venc. licencia</label><input name="venc_licencia" type="date" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>ACTIVO</option><option>APTO</option><option>POR VENCER</option><option>VENCIDA</option><option>INACTIVO</option></select></div></div><button class="tf-btn w100 mt-2">+ Guardar conductor</button></form>
    '''+_tf_end()+_tf_filter_script('qcond','tblcond')
    return render_page(body, rows=rows,total=total,activos=activos,venc=venc,badge=_tf_badge,switch=_tf_switch,lic_opts=_option_list(LICENCIAS_PERU,'A-IIIc'),cat_opts=_option_list(CATEGORIAS_CONDUCTOR_PERU,'BUS / TRANSPORTE DE PERSONAS'))

@login_required
def tf_transporte_vehiculos_275():
    if request.method=='POST': return tf_transporte_vehiculos_omar()
    rows=rows_to_dict(execute('SELECT * FROM transporte_vehiculos ORDER BY id DESC LIMIT 300', fetchall=True)); hoy=today_str(); total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper()=='ACTIVO')
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<hoy for c in ('soat_venc','revision_tecnica_venc')))
    body=_tf_css()+_tf_transport_css_275()+_tf_top('Buses')+'''
      <form method="post" enctype="multipart/form-data" action="{{url_for('transporte_upload_vehiculos')}}" class="tf-mass"><label><i class="bi bi-cloud-arrow-up"></i> Carga masiva buses Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="tf-help-mini">Columnas sugeridas: PLACA, TIPO, CAPACIDAD, EMPRESA, SOAT, REV TECNICA, GPS, ESTADO.</div><button class="tf-btn w100" type="submit"><i class="bi bi-upload"></i> Cargar buses</button><a class="tf-template-btn" href="{{url_for('transporte_plantilla_vehiculos')}}"><i class="bi bi-file-earmark-excel"></i> Descargar plantilla</a></form>
      <div class="tf-searchrow"><input id="qveh" class="tf-search" placeholder="Buscar placa o bus..."><a class="tf-template-btn" style="margin:0;padding:9px 11px" href="{{url_for('transporte_plantilla_vehiculos')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total buses</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activos</label><strong>{{activos}}</strong></div><div class="tf-kpi"><label>Vencidos</label><strong>{{venc}}</strong></div></div>
      <div class="tf-base-title">Base de buses</div><div class="tf-tablewrap"><table id="tblveh" class="tf-table"><thead><tr><th>Placa</th><th>Tipo</th><th>Capacidad</th><th>GPS</th><th>Estado</th><th>Activo</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.placa}}</td><td>{{r.tipo}}</td><td>{{r.capacidad}}</td><td>{{r.gps_codigo or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td>{{switch((r.estado or '').upper()=='ACTIVO', url_for('transporte_toggle_vehiculo', item_id=r.id))|safe}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin buses.</td></tr>{% endfor %}</tbody></table></div>
      <form method="post" class="tf-form mt-2"><div class="row g-2"><div class="col-5"><label>Placa</label><input name="placa" class="form-control" required></div><div class="col-7"><label>Tipo</label><select name="tipo" class="form-select"><option>BUS</option><option>MINIBUS</option><option>VAN</option><option>CAMIONETA</option></select></div><div class="col-6"><label>Capacidad</label><input name="capacidad" type="number" class="form-control"></div><div class="col-6"><label>Empresa</label><input name="empresa_transportista" class="form-control"></div><div class="col-6"><label>SOAT venc.</label><input name="soat_venc" type="date" class="form-control"></div><div class="col-6"><label>Rev. técnica</label><input name="revision_tecnica_venc" type="date" class="form-control"></div><div class="col-12"><label>Código GPS</label><input name="gps_codigo" class="form-control"></div></div><button class="tf-btn w100 mt-2">+ Guardar bus</button></form>
    '''+_tf_end()+_tf_filter_script('qveh','tblveh')
    return render_page(body, rows=rows,total=total,activos=activos,venc=venc,badge=_tf_badge,switch=_tf_switch)

@login_required
def tf_transporte_rutas_275():
    if request.method=='POST': return tf_transporte_rutas_omar()
    hoy=today_str(); rutas=rows_to_dict(execute('SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id ORDER BY r.fecha DESC, r.id DESC LIMIT 250', fetchall=True))
    vehiculos=rows_to_dict(execute("SELECT id, placa, tipo, capacidad FROM transporte_vehiculos WHERE COALESCE(estado,'ACTIVO')='ACTIVO' ORDER BY placa", fetchall=True)); conductores=rows_to_dict(execute("SELECT id, dni, nombres FROM transporte_conductores WHERE COALESCE(estado,'ACTIVO') IN ('ACTIVO','APTO') ORDER BY nombres", fetchall=True))
    total=len(rutas); activas=sum(1 for r in rutas if (r.get('estado') or '').upper() in ('PROGRAMADA','EN RUTA','ACTIVA')); cerradas=sum(1 for r in rutas if (r.get('estado') or '').upper() in ('CERRADA','FINALIZADA','INACTIVA'))
    body=_tf_css()+_tf_transport_css_275()+_tf_top('Rutas')+'''
      <form method="post" enctype="multipart/form-data" action="{{url_for('transporte_upload_rutas')}}" class="tf-mass"><label><i class="bi bi-cloud-arrow-up"></i> Carga masiva rutas Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="tf-help-mini">Columnas sugeridas: FECHA, RUTA, ORIGEN, DESTINO, SALIDA, PLACA, DNI CONDUCTOR, ESTADO.</div><button class="tf-btn w100" type="submit"><i class="bi bi-upload"></i> Cargar rutas</button><a class="tf-template-btn" href="{{url_for('transporte_plantilla_rutas')}}"><i class="bi bi-file-earmark-excel"></i> Descargar plantilla</a></form>
      <div class="tf-searchrow"><input id="qruta" class="tf-search" placeholder="Buscar ruta..."><a class="tf-template-btn" style="margin:0;padding:9px 11px" href="{{url_for('transporte_plantilla_rutas')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total rutas</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activas</label><strong>{{activas}}</strong></div><div class="tf-kpi"><label>Cerradas / inactivas</label><strong>{{cerradas}}</strong></div></div>
      <div class="tf-base-title">Base de rutas</div><div class="tf-tablewrap"><table id="tblruta" class="tf-table"><thead><tr><th>Código</th><th>Ruta</th><th>Bus</th><th>Conductor</th><th>Estado</th><th>Activo</th><th></th></tr></thead><tbody>{% for r in rutas %}<tr><td>R-{{'%03d'%r.id}}</td><td>{{r.nombre}}<br><small>{{r.origen}} → {{r.destino}}</small></td><td>{{r.placa or '-'}}</td><td>{{r.conductor or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td>{{switch((r.estado or '').upper() not in ['INACTIVA','CERRADA','FINALIZADA'], url_for('transporte_toggle_ruta', item_id=r.id))|safe}}</td><td><a class="tf-btn" style="padding:4px 7px" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}">›</a></td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted">Sin rutas.</td></tr>{% endfor %}</tbody></table></div>
      <form method="post" class="tf-form mt-2"><div class="row g-2"><div class="col-6"><label>Fecha</label><input name="fecha" type="date" value="{{hoy}}" class="form-control"></div><div class="col-6"><label>Ruta</label><input name="nombre" class="form-control" placeholder="R-001"></div><div class="col-6"><label>Origen</label><input name="origen" class="form-control"></div><div class="col-6"><label>Destino</label><input name="destino" class="form-control"></div><div class="col-6"><label>Salida</label><input name="hora_salida" type="time" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>PROGRAMADA</option><option>EN RUTA</option><option>CERRADA</option></select></div><div class="col-6"><label>Bus</label><select name="vehiculo_id" class="form-select"><option value="">-</option>{% for v in vehiculos %}<option value="{{v.id}}">{{v.placa}}</option>{% endfor %}</select></div><div class="col-6"><label>Conductor</label><select name="conductor_id" class="form-select"><option value="">-</option>{% for c in conductores %}<option value="{{c.id}}">{{c.nombres}}</option>{% endfor %}</select></div></div><button class="tf-btn w100 mt-2">+ Nueva ruta</button></form>
    '''+_tf_end()+_tf_filter_script('qruta','tblruta')
    return render_page(body,rutas=rutas,vehiculos=vehiculos,conductores=conductores,hoy=hoy,total=total,activas=activas,cerradas=cerradas,badge=_tf_badge,switch=_tf_switch)

@login_required
def tf_transporte_mobile_home_275():
    body=_tf_css()+_tf_transport_css_275()+_tf_top('Móvil conductor')+'''
      <div class="tf-info"><b>Este módulo es del conductor.</b><br>Desde aquí debe ver su ruta, registrar abordaje de trabajadores, enviar GPS y revisar su estado.</div>
      <div class="tf-mobile-sub"><a class="tf-item" href="{{url_for('transporte_abordaje_home')}}"><i class="bi bi-people"></i><span><b>Abordaje trabajadores</b><br><small>Escanear QR / código de barras / digitar DNI</small></span><i class="bi bi-chevron-right chev"></i></a><a class="tf-item" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-geo-alt"></i><span><b>GPS / Seguimiento</b><br><small>Enviar o revisar ubicación</small></span><i class="bi bi-chevron-right chev"></i></a><a class="tf-item" href="{{url_for('conductor_movil_login')}}"><i class="bi bi-phone"></i><span><b>Ingreso con PIN conductor</b><br><small>Acceso real desde celular</small></span><i class="bi bi-chevron-right chev"></i></a></div>
      <div class="tf-home-note">Recomendación: el administrador solo debe ver bases y seguimiento. El conductor debe usar <b>Móvil conductor</b> para abordaje y GPS desde su celular.</div>
    '''+_tf_end()
    return render_page(body)

@login_required
def tf_transporte_home_275():
    body=_tf_css()+_tf_transport_css_275()+'''
    <div class="tf-wrap"><div class="tf-head"><a class="tf-back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><div class="tf-title">Módulo Transporte</div></div><div class="tf-card">
      <div class="tf-section">Bases maestras</div>
      <div class="tf-list"><a class="tf-item" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span><b>Conductores</b><br><small>Base, plantilla, carga masiva y activo/inactivo</small></span><i class="bi bi-chevron-right chev"></i></a><a class="tf-item" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span><b>Buses</b><br><small>Base, plantilla, carga masiva y activo/inactivo</small></span><i class="bi bi-chevron-right chev"></i></a><a class="tf-item" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i><span><b>Rutas</b><br><small>Programación, bus, conductor y estado</small></span><i class="bi bi-chevron-right chev"></i></a></div>
      <div class="tf-section">Operación</div><div class="tf-list"><a class="tf-item" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span><b>GPS / Seguimiento</b><br><small>Ver conductor, ruta, bus y última ubicación</small></span><i class="bi bi-chevron-right chev"></i></a><a class="tf-item" href="{{url_for('transporte_mobile_home')}}"><i class="bi bi-phone"></i><span><b>Móvil conductor</b><br><small>Abordaje trabajadores y envío GPS</small></span><i class="bi bi-chevron-right chev"></i></a></div>
      <div class="tf-home-note">Así queda conectado: primero cargas <b>conductores</b> y <b>buses</b>, luego creas/cargas <b>rutas</b> asignando bus + conductor. El conductor entra a <b>Móvil conductor</b>, registra abordajes y envía GPS.</div>
    </div></div>'''
    return render_page(body)

@login_required
def tf_transporte_mapa_general_275():
    sql = """SELECT r.*, v.placa, v.tipo, c.dni AS conductor_dni, c.nombres AS conductor, c.estado AS conductor_estado, c.ultima_latitud, c.ultima_longitud, c.ultimo_gps
             FROM transporte_rutas r
             LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
             LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
             ORDER BY COALESCE(r.ultima_ubicacion,r.creado_en) DESC, r.id DESC LIMIT 40"""
    rutas=rows_to_dict(execute(sql, fetchall=True))
    total=len(rutas); con_gps=sum(1 for r in rutas if r.get('latitud') or r.get('ultima_latitud')); sin_bus=sum(1 for r in rutas if not r.get('placa'))
    body=_tf_css()+_tf_transport_css_275()+_tf_top('GPS / Seguimiento')+'''
      <div class="tf-info"><b>¿Cómo funciona?</b><br>1) Crea conductor y PIN. 2) Asigna conductor y bus a una ruta. 3) El conductor entra a Móvil conductor desde su celular. 4) Permite ubicación/GPS y toca Enviar GPS.</div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Rutas</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Con GPS</label><strong>{{con_gps}}</strong></div><div class="tf-kpi"><label>Sin bus</label><strong>{{sin_bus}}</strong></div></div>
      {% for r in rutas %}<div class="tf-gps-card"><div class="tf-gps-driver"><div class="tf-gps-avatar"><i class="bi bi-person-fill"></i></div><div><b>{{r.conductor or 'SIN CONDUCTOR'}}</b><div class="tf-gps-meta">DNI: {{r.conductor_dni or '-'}} · {{badge(r.conductor_estado or 'SIN ASIGNAR')|safe}}</div></div></div><div class="tf-gps-row"><div class="tf-gps-box"><small>Ruta</small><strong>{{r.nombre}}</strong></div><div class="tf-gps-box"><small>Bus</small><strong>{{r.placa or 'SIN BUS'}}</strong></div><div class="tf-gps-box"><small>Última actualización</small><strong>{{r.ultima_ubicacion or r.ultimo_gps or r.creado_en or '-'}}</strong></div><div class="tf-gps-box"><small>Coordenadas</small><strong>Lat: {{r.latitud or r.ultima_latitud or '-'}}<br>Lon: {{r.longitud or r.ultima_longitud or '-'}}</strong></div></div>{% if r.latitud or r.ultima_latitud %}<a class="tf-btn" href="https://www.google.com/maps?q={{r.latitud or r.ultima_latitud}},{{r.longitud or r.ultima_longitud}}" target="_blank"><i class="bi bi-map"></i> Ver en mapa</a>{% endif %}</div>{% else %}<div class="tf-card text-center text-muted">Sin rutas para seguimiento.</div>{% endfor %}
    '''+_tf_end()
    return render_page(body,rutas=rutas,total=total,con_gps=con_gps,sin_bus=sin_bus,badge=_tf_badge)

app.view_functions['transporte'] = tf_transporte_home_275
app.view_functions['transporte_mobile_home'] = tf_transporte_mobile_home_275
app.view_functions['transporte_conductores'] = tf_transporte_conductores_275
app.view_functions['transporte_vehiculos'] = tf_transporte_vehiculos_275
app.view_functions['transporte_rutas'] = tf_transporte_rutas_275
app.view_functions['transporte_mapa_general'] = tf_transporte_mapa_general_275

# PATCH URGENTE 276: registrar rutas faltantes para evitar Internal Server Error en /transporte.
# El home usa url_for('transporte_mobile_home'), pero antes solo se reemplazaba
# view_functions y no existía una regla URL asociada a ese endpoint.
try:
    if 'transporte_mobile_home' not in {r.endpoint for r in app.url_map.iter_rules()}:
        app.add_url_rule('/transporte/movil-conductor', endpoint='transporte_mobile_home', view_func=tf_transporte_mobile_home_275, methods=['GET'])
except Exception as e:
    print('No se pudo registrar /transporte/movil-conductor:', e)

# Alias amigable solicitado: buses debe abrir la misma base de vehículos.
try:
    if 'transporte_buses_alias' not in {r.endpoint for r in app.url_map.iter_rules()}:
        app.add_url_rule('/transporte/buses', endpoint='transporte_buses_alias', view_func=tf_transporte_vehiculos_275, methods=['GET','POST'])
except Exception as e:
    print('No se pudo registrar /transporte/buses:', e)



# ========================= PATCH 277 OMAR: MÓDULOS EN COLUMNAS Y ORDEN KPI/BUSCADOR/DETALLE =========================
def _tf_transport_css_277():
    return _tf_transport_css_275()+'''<style>
    .tf-home-columns{display:grid!important;grid-template-columns:repeat(3,1fr)!important;gap:12px!important;margin:12px 0 14px!important}
    .tf-home-columns.op{grid-template-columns:repeat(2,1fr)!important}
    .tf-home-tile{height:112px!important;background:#0f7334!important;color:white!important;border-radius:10px!important;text-decoration:none!important;display:flex!important;flex-direction:column!important;align-items:center!important;justify-content:center!important;box-shadow:0 7px 15px rgba(0,0,0,.18)!important;font-weight:900!important;text-align:center!important;line-height:1.12!important}
    .tf-home-tile i{font-size:34px!important;margin-bottom:9px!important;color:white!important}.tf-home-tile span{font-size:16px!important;color:white!important}.tf-home-tile small{display:block!important;font-size:10px!important;color:#eaffef!important;margin-top:3px!important;font-weight:800!important}
    .tf-card.home-flat{box-shadow:none!important;border:0!important;padding:0!important;background:transparent!important}.tf-section{font-size:13px!important;color:#08713b!important;font-weight:900!important;margin:12px 0 7px!important;text-transform:uppercase!important}
    .tf-kpi-green{display:grid!important;grid-template-columns:repeat(3,1fr)!important;gap:10px!important;margin:8px 0 12px!important}.tf-searchrow{display:grid!important;grid-template-columns:minmax(0,1fr) auto!important;gap:8px!important;align-items:center!important;margin:8px 0 10px!important;max-width:100%!important}.tf-search{height:42px!important;border-radius:9px!important;border:1px solid #dce8dc!important;padding:0 12px!important;font-weight:800!important;min-width:0!important}.tf-template-btn{white-space:nowrap!important;margin-top:0!important}.tf-mass{margin:12px 0!important}
    @media(max-width:520px){.tf-home-columns{grid-template-columns:repeat(3,1fr)!important;gap:10px!important}.tf-home-columns.op{grid-template-columns:repeat(2,1fr)!important}.tf-home-tile{height:96px!important;border-radius:9px!important}.tf-home-tile i{font-size:30px!important;margin-bottom:7px!important}.tf-home-tile span{font-size:14px!important}.tf-home-tile small{display:none!important}.tf-searchrow{grid-template-columns:1fr!important}.tf-template-btn{width:100%!important}}
    </style>'''

@login_required
def tf_transporte_home_277():
    body=_tf_css()+_tf_transport_css_277()+'''<div class="tf-wrap"><div class="tf-head"><a class="tf-back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><div class="tf-title">Módulo Transporte</div></div><div class="tf-card home-flat">
      <div class="tf-section">Módulos</div><div class="tf-home-columns"><a class="tf-home-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span>Conductores</span></a><a class="tf-home-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span>Buses</span></a><a class="tf-home-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i><span>Rutas</span></a></div>
      <div class="tf-section">Operación</div><div class="tf-home-columns op"><a class="tf-home-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span>GPS / Seguimiento</span><small>Ver ubicación</small></a><a class="tf-home-tile" href="{{url_for('transporte_mobile_home')}}"><i class="bi bi-phone"></i><span>Móvil conductor</span><small>Abordaje y GPS</small></a></div>
      <div class="tf-home-note">Así queda conectado: primero cargas <b>conductores</b> y <b>buses</b>, luego creas/cargas <b>rutas</b> asignando bus + conductor. El conductor entra a <b>Móvil conductor</b>, registra abordajes y envía GPS.</div></div></div>'''
    return render_page(body)

@login_required
def tf_transporte_conductores_277():
    if request.method=='POST': return tf_transporte_conductores_omar()
    rows=rows_to_dict(execute('SELECT * FROM transporte_conductores ORDER BY id DESC LIMIT 300', fetchall=True))
    hoy=today_str(); total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO','APTO',''))
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<hoy for c in ('venc_licencia','venc_cert_medico','venc_sctr')))
    body=_tf_css()+_tf_transport_css_277()+_tf_top('Conductores')+'''
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total conductores</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activos</label><strong>{{activos}}</strong></div><div class="tf-kpi"><label>Vencidos</label><strong>{{venc}}</strong></div></div><div class="tf-searchrow"><input id="qcond" class="tf-search" placeholder="Buscar DNI o conductor..."><a class="tf-template-btn" href="{{url_for('transporte_plantilla_conductores')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div><div class="tf-base-title">Base de conductores</div><div class="tf-tablewrap"><table id="tblcond" class="tf-table"><thead><tr><th>DNI</th><th>Nombre</th><th>Licencia</th><th>Categoría</th><th>Estado</th><th>Activo</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.nombres}}</td><td>{{r.licencia or '-'}}</td><td>{{r.categoria or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td>{{switch((r.estado or '').upper() in ['ACTIVO','APTO',''], url_for('transporte_toggle_conductor', item_id=r.id))|safe}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}</tbody></table></div>
      <form method="post" enctype="multipart/form-data" action="{{url_for('transporte_upload_conductores')}}" class="tf-mass"><label><i class="bi bi-cloud-arrow-up"></i> Carga masiva conductores Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="tf-help-mini">Columnas sugeridas: DNI, NOMBRE, TELEFONO, LICENCIA, CATEGORIA, VENC LICENCIA, ESTADO, PIN.</div><button class="tf-btn w100" type="submit"><i class="bi bi-upload"></i> Cargar conductores</button></form>
      <form method="post" class="tf-form mt-2"><div class="row g-2"><div class="col-5"><label>DNI</label><input name="dni" class="form-control" maxlength="8" inputmode="numeric" required></div><div class="col-7"><label>Nombre</label><input name="nombres" class="form-control" required></div><div class="col-6"><label>Teléfono</label><input name="telefono" class="form-control"></div><div class="col-6"><label>Licencia</label><select name="licencia" class="form-select">{{lic_opts|safe}}</select></div><div class="col-6"><label>Categoría</label><select name="categoria" class="form-select">{{cat_opts|safe}}</select></div><div class="col-6"><label>PIN móvil</label><input name="movil_pin" class="form-control" placeholder="AUTO"></div><div class="col-6"><label>Venc. licencia</label><input name="venc_licencia" type="date" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>ACTIVO</option><option>APTO</option><option>POR VENCER</option><option>VENCIDA</option><option>INACTIVO</option></select></div></div><button class="tf-btn w100 mt-2">+ Guardar conductor</button></form>'''+_tf_end()+_tf_filter_script('qcond','tblcond')
    return render_page(body, rows=rows,total=total,activos=activos,venc=venc,badge=_tf_badge,switch=_tf_switch,lic_opts=_option_list(LICENCIAS_PERU,'A-IIIc'),cat_opts=_option_list(CATEGORIAS_CONDUCTOR_PERU,'BUS / TRANSPORTE DE PERSONAS'))

@login_required
def tf_transporte_vehiculos_277():
    if request.method=='POST': return tf_transporte_vehiculos_omar()
    rows=rows_to_dict(execute('SELECT * FROM transporte_vehiculos ORDER BY id DESC LIMIT 300', fetchall=True)); hoy=today_str(); total=len(rows); activos=sum(1 for r in rows if (r.get('estado') or '').upper()=='ACTIVO')
    venc=sum(1 for r in rows if any(r.get(c) and str(r.get(c))<hoy for c in ('soat_venc','revision_tecnica_venc')))
    body=_tf_css()+_tf_transport_css_277()+_tf_top('Buses')+'''
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total buses</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activos</label><strong>{{activos}}</strong></div><div class="tf-kpi"><label>Vencidos</label><strong>{{venc}}</strong></div></div><div class="tf-searchrow"><input id="qveh" class="tf-search" placeholder="Buscar placa o bus..."><a class="tf-template-btn" href="{{url_for('transporte_plantilla_vehiculos')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div><div class="tf-base-title">Base de buses</div><div class="tf-tablewrap"><table id="tblveh" class="tf-table"><thead><tr><th>Placa</th><th>Tipo</th><th>Capacidad</th><th>GPS</th><th>Estado</th><th>Activo</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.placa}}</td><td>{{r.tipo}}</td><td>{{r.capacidad}}</td><td>{{r.gps_codigo or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td>{{switch((r.estado or '').upper()=='ACTIVO', url_for('transporte_toggle_vehiculo', item_id=r.id))|safe}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin buses.</td></tr>{% endfor %}</tbody></table></div>
      <form method="post" enctype="multipart/form-data" action="{{url_for('transporte_upload_vehiculos')}}" class="tf-mass"><label><i class="bi bi-cloud-arrow-up"></i> Carga masiva buses Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="tf-help-mini">Columnas sugeridas: PLACA, TIPO, CAPACIDAD, EMPRESA, SOAT, REV TECNICA, GPS, ESTADO.</div><button class="tf-btn w100" type="submit"><i class="bi bi-upload"></i> Cargar buses</button></form>
      <form method="post" class="tf-form mt-2"><div class="row g-2"><div class="col-5"><label>Placa</label><input name="placa" class="form-control" required></div><div class="col-7"><label>Tipo</label><select name="tipo" class="form-select"><option>BUS</option><option>MINIBUS</option><option>VAN</option><option>CAMIONETA</option></select></div><div class="col-6"><label>Capacidad</label><input name="capacidad" type="number" class="form-control"></div><div class="col-6"><label>Empresa</label><input name="empresa_transportista" class="form-control"></div><div class="col-6"><label>SOAT venc.</label><input name="soat_venc" type="date" class="form-control"></div><div class="col-6"><label>Rev. técnica</label><input name="revision_tecnica_venc" type="date" class="form-control"></div><div class="col-12"><label>Código GPS</label><input name="gps_codigo" class="form-control"></div></div><button class="tf-btn w100 mt-2">+ Guardar bus</button></form>'''+_tf_end()+_tf_filter_script('qveh','tblveh')
    return render_page(body, rows=rows,total=total,activos=activos,venc=venc,badge=_tf_badge,switch=_tf_switch)

@login_required
def tf_transporte_rutas_277():
    if request.method=='POST': return tf_transporte_rutas_omar()
    hoy=today_str(); rutas=rows_to_dict(execute('SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id LEFT JOIN transporte_conductores c ON c.id=r.conductor_id ORDER BY r.fecha DESC, r.id DESC LIMIT 250', fetchall=True))
    vehiculos=rows_to_dict(execute("SELECT id, placa, tipo, capacidad FROM transporte_vehiculos WHERE COALESCE(estado,'ACTIVO')='ACTIVO' ORDER BY placa", fetchall=True)); conductores=rows_to_dict(execute("SELECT id, dni, nombres FROM transporte_conductores WHERE COALESCE(estado,'ACTIVO') IN ('ACTIVO','APTO') ORDER BY nombres", fetchall=True))
    total=len(rutas); activas=sum(1 for r in rutas if (r.get('estado') or '').upper() in ('PROGRAMADA','EN RUTA','ACTIVA')); cerradas=sum(1 for r in rutas if (r.get('estado') or '').upper() in ('CERRADA','FINALIZADA','INACTIVA'))
    body=_tf_css()+_tf_transport_css_277()+_tf_top('Rutas')+'''
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total rutas</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activas</label><strong>{{activas}}</strong></div><div class="tf-kpi"><label>Cerradas</label><strong>{{cerradas}}</strong></div></div><div class="tf-searchrow"><input id="qruta" class="tf-search" placeholder="Buscar ruta..."><a class="tf-template-btn" href="{{url_for('transporte_plantilla_rutas')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div><div class="tf-base-title">Base de rutas</div><div class="tf-tablewrap"><table id="tblruta" class="tf-table"><thead><tr><th>Código</th><th>Ruta</th><th>Bus</th><th>Conductor</th><th>Estado</th><th>Activo</th><th></th></tr></thead><tbody>{% for r in rutas %}<tr><td>R-{{'%03d'%r.id}}</td><td>{{r.nombre}}<br><small>{{r.origen}} → {{r.destino}}</small></td><td>{{r.placa or '-'}}</td><td>{{r.conductor or '-'}}</td><td>{{badge(r.estado)|safe}}</td><td>{{switch((r.estado or '').upper() not in ['INACTIVA','CERRADA','FINALIZADA'], url_for('transporte_toggle_ruta', item_id=r.id))|safe}}</td><td><a class="tf-btn" style="padding:4px 7px" href="{{url_for('transporte_ruta_detalle', ruta_id=r.id)}}">›</a></td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted">Sin rutas.</td></tr>{% endfor %}</tbody></table></div>
      <form method="post" enctype="multipart/form-data" action="{{url_for('transporte_upload_rutas')}}" class="tf-mass"><label><i class="bi bi-cloud-arrow-up"></i> Carga masiva rutas Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="tf-help-mini">Columnas sugeridas: FECHA, RUTA, ORIGEN, DESTINO, SALIDA, PLACA, DNI CONDUCTOR, ESTADO.</div><button class="tf-btn w100" type="submit"><i class="bi bi-upload"></i> Cargar rutas</button></form>
      <form method="post" class="tf-form mt-2"><div class="row g-2"><div class="col-6"><label>Fecha</label><input name="fecha" type="date" value="{{hoy}}" class="form-control"></div><div class="col-6"><label>Ruta</label><input name="nombre" class="form-control" placeholder="R-001"></div><div class="col-6"><label>Origen</label><input name="origen" class="form-control"></div><div class="col-6"><label>Destino</label><input name="destino" class="form-control"></div><div class="col-6"><label>Salida</label><input name="hora_salida" type="time" class="form-control"></div><div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>PROGRAMADA</option><option>EN RUTA</option><option>CERRADA</option></select></div><div class="col-6"><label>Bus</label><select name="vehiculo_id" class="form-select"><option value="">-</option>{% for v in vehiculos %}<option value="{{v.id}}">{{v.placa}}</option>{% endfor %}</select></div><div class="col-6"><label>Conductor</label><select name="conductor_id" class="form-select"><option value="">-</option>{% for c in conductores %}<option value="{{c.id}}">{{c.nombres}}</option>{% endfor %}</select></div></div><button class="tf-btn w100 mt-2">+ Nueva ruta</button></form>'''+_tf_end()+_tf_filter_script('qruta','tblruta')
    return render_page(body,rutas=rutas,vehiculos=vehiculos,conductores=conductores,hoy=hoy,total=total,activas=activas,cerradas=cerradas,badge=_tf_badge,switch=_tf_switch)

app.view_functions['transporte'] = tf_transporte_home_277
app.view_functions['transporte_conductores'] = tf_transporte_conductores_277
app.view_functions['transporte_vehiculos'] = tf_transporte_vehiculos_277
app.view_functions['transporte_rutas'] = tf_transporte_rutas_277

# ========================= PATCH 279 OMAR: PORTADA TRANSPORTE EXACTA / DELGADA =========================
def _transporte_ui_279_css():
    return """
    <style>
      /* SOLO /transporte: vista móvil delgada, limpia y en cuadros */
      html,body{background:#fff!important;max-width:100%!important;overflow-x:hidden!important;}
      .shell{width:100%!important;max-width:470px!important;margin:0 auto!important;padding:0 8px 26px!important;background:#fff!important;}
      .tr279-phone{width:100%!important;max-width:430px!important;margin:8px auto 26px!important;}
      .tr279-app{width:100%!important;background:#fff;border:1px solid #e5e7eb;border-radius:18px;overflow:hidden;box-shadow:0 12px 28px rgba(0,0,0,.10);}
      .tr279-hero{height:128px;background:linear-gradient(135deg,#075d2a 0%,#0b7837 100%);color:#fff;position:relative;text-align:center;padding-top:13px;}
      .tr279-back{position:absolute;left:16px;top:35px;color:#fff!important;text-decoration:none;font-size:44px;line-height:1;font-weight:300;display:grid;place-items:center;width:44px;height:44px;}
      .tr279-config{position:absolute;right:14px;top:23px;color:#fff!important;text-decoration:none;border:1px solid rgba(255,255,255,.70);border-radius:13px;padding:8px 11px;font-weight:950;font-size:15px;display:inline-flex;align-items:center;gap:6px;background:rgba(255,255,255,.06);}
      .tr279-config i{font-size:18px;}
      .tr279-bus{font-size:52px;line-height:1;color:#fff;text-shadow:0 2px 5px rgba(0,0,0,.12);}
      .tr279-title{margin-top:5px;font-size:18px;font-weight:950;letter-spacing:.35px;text-transform:uppercase;color:#fff;}
      .tr279-body{background:#fff;border-radius:16px 16px 0 0;margin-top:-1px;padding:29px 20px 28px;min-height:515px;}
      .tr279-section{font-size:21px;line-height:1;color:#08713b;font-weight:950;text-transform:uppercase;letter-spacing:.35px;margin:0 0 18px;}
      .tr279-section.op{margin-top:30px;margin-bottom:17px;}
      .tr279-grid3{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:13px;}
      .tr279-grid2{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:18px;}
      .tr279-tile{min-height:118px;background:linear-gradient(135deg,#07632d,#0b7837);border-radius:13px;color:#fff!important;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 10px 18px rgba(0,0,0,.15);padding:10px 7px;transition:transform .12s ease, filter .12s ease;}
      .tr279-tile:active{transform:scale(.98);filter:brightness(.95);}
      .tr279-tile i{font-size:42px;line-height:1;color:#fff!important;margin-bottom:14px;}
      .tr279-tile .lbl{font-size:17px;font-weight:950;line-height:1.08;color:#fff;text-shadow:0 1px 1px rgba(0,0,0,.14);}
      .tr279-tile .sub{font-size:12px;font-weight:900;line-height:1.1;color:#fff;margin-top:6px;opacity:.98;}
      .tr279-grid2 .tr279-tile{min-height:128px;}
      .tr279-grid2 .tr279-tile i{font-size:43px;margin-bottom:16px;}
      .tr279-grid2 .tr279-tile .lbl{font-size:19px;}
      .tr279-info{margin-top:25px;border:1px solid #b8d7ff;background:#eef6ff;border-radius:12px;color:#073b8e;font-size:15px;font-weight:850;line-height:1.55;padding:17px 16px;display:grid;grid-template-columns:30px 1fr;gap:10px;align-items:start;}
      .tr279-info i{font-size:24px;color:#0b46a0;margin-top:2px;}

      /* Submenú móvil conductor en cuadros, no lista */
      .tr279-subhead{height:88px;background:linear-gradient(135deg,#075d2a,#0b7837);color:#fff;display:flex;align-items:center;position:relative;padding:0 16px;}
      .tr279-subhead a{color:#fff!important;text-decoration:none;font-size:34px;width:44px;}
      .tr279-subhead .ttl{flex:1;text-align:center;font-size:18px;font-weight:950;margin-right:44px;}
      .tr279-subbody{background:#fff;padding:22px 18px 26px;}
      .tr279-subgrid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:15px;}
      .tr279-subgrid .wide{grid-column:1/-1;min-height:108px;}
      .tr279-note{margin-top:18px;border:1px solid #b8d7ff;background:#eef6ff;color:#073b8e;border-radius:12px;padding:13px 14px;font-size:13px;line-height:1.45;font-weight:850;}

      @media(max-width:390px){
        .shell{max-width:395px!important;padding:0 6px 20px!important;}
        .tr279-phone{max-width:372px!important;margin-top:5px!important;}
        .tr279-app{border-radius:16px;}
        .tr279-hero{height:122px;padding-top:12px;}
        .tr279-back{left:11px;top:34px;font-size:39px;width:39px;height:39px;}
        .tr279-config{right:9px;top:22px;font-size:13px;padding:7px 9px;border-radius:12px;}
        .tr279-bus{font-size:47px;}
        .tr279-title{font-size:16px;margin-top:5px;}
        .tr279-body{padding:25px 14px 24px;min-height:500px;}
        .tr279-section{font-size:18px;margin-bottom:14px;}
        .tr279-section.op{margin-top:26px;margin-bottom:13px;}
        .tr279-grid3{gap:10px;}
        .tr279-grid2{gap:14px;}
        .tr279-tile{min-height:102px;border-radius:11px;padding:8px 5px;}
        .tr279-tile i{font-size:34px;margin-bottom:10px;}
        .tr279-tile .lbl{font-size:14px;}
        .tr279-tile .sub{font-size:10.5px;margin-top:4px;}
        .tr279-grid2 .tr279-tile{min-height:111px;}
        .tr279-grid2 .tr279-tile i{font-size:37px;margin-bottom:12px;}
        .tr279-grid2 .tr279-tile .lbl{font-size:16px;}
        .tr279-info{font-size:12.8px;padding:13px 12px;grid-template-columns:24px 1fr;gap:8px;}
        .tr279-info i{font-size:20px;}
      }
    </style>
    """

@login_required
def tf_transporte_home_279():
    body = _transporte_ui_279_css() + r'''
    <div class="tr279-phone">
      <div class="tr279-app">
        <div class="tr279-hero">
          <a class="tr279-back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>
          <a class="tr279-config" href="{{url_for('transporte_config')}}"><i class="bi bi-gear"></i> Config.</a>
          <div class="tr279-bus"><i class="bi bi-bus-front-fill"></i></div>
          <div class="tr279-title">Módulo Transporte</div>
        </div>
        <div class="tr279-body">
          <div class="tr279-section">Módulos</div>
          <div class="tr279-grid3">
            <a class="tr279-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span class="lbl">Conductores</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span class="lbl">Buses</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i><span class="lbl">Rutas</span></a>
          </div>

          <div class="tr279-section op">Operación</div>
          <div class="tr279-grid2">
            <a class="tr279-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span class="lbl">GPS / Seguimiento</span><span class="sub">Ver ubicación</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_mobile_home')}}"><i class="bi bi-phone"></i><span class="lbl">Móvil conductor</span><span class="sub">Abordaje y GPS</span></a>
          </div>

          <div class="tr279-info"><i class="bi bi-info-circle-fill"></i><div>Así queda conectado: primero cargas conductores y buses, luego creas/cargas rutas asignando bus + conductor. El conductor entra a Móvil conductor, registra abordajes y envía GPS.</div></div>
        </div>
      </div>
    </div>
    '''
    return render_page(body)

@login_required
def tf_transporte_mobile_home_279():
    """Submenú de Móvil conductor en cuadros verdes, no en lista."""
    body = _transporte_ui_279_css() + r'''
    <div class="tr279-phone">
      <div class="tr279-app">
        <div class="tr279-subhead">
          <a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a>
          <div class="ttl">Móvil conductor</div>
        </div>
        <div class="tr279-subbody">
          <div class="tr279-section">Operación conductor</div>
          <div class="tr279-subgrid">
            <a class="tr279-tile wide" href="{{url_for('transporte_abordaje_home')}}"><i class="bi bi-people"></i><span class="lbl">Abordaje trabajadores</span><span class="sub">QR / código / DNI</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span class="lbl">GPS</span><span class="sub">Enviar ubicación</span></a>
            <a class="tr279-tile" href="{{url_for('conductor_movil_login')}}"><i class="bi bi-phone"></i><span class="lbl">PIN conductor</span><span class="sub">Ingreso real</span></a>
          </div>
          <div class="tr279-note">El conductor ingresa desde su celular, registra abordajes y envía GPS. El administrador revisa seguimiento y bases desde Transporte.</div>
        </div>
      </div>
    </div>
    '''
    return render_page(body)

# Reemplazo final: solo afecta endpoints de Transporte aprobados.
app.view_functions['transporte'] = tf_transporte_home_279
app.view_functions['transporte_mobile_home'] = tf_transporte_mobile_home_279


# ========================= PATCH 280 OMAR: TRANSPORTE MÁS DELGADO + GPS CORREGIDO =========================
def _transporte_ui_280_css():
    return _transporte_ui_279_css() + r"""
    <style>
      html,body{background:#ffffff!important;overflow-x:hidden!important}
      .shell{max-width:392px!important;padding:0 2px 18px!important;background:#ffffff!important}
      .tr279-phone{max-width:352px!important;margin:8px auto 14px!important}
      .tr279-app{border-radius:18px!important;overflow:hidden!important;box-shadow:0 10px 26px rgba(0,0,0,.10)!important;border:1px solid #e7eee7!important;background:#fff!important}
      .tr279-hero{height:132px!important;padding-top:14px!important}
      .tr279-back{left:12px!important;top:34px!important;font-size:40px!important;width:40px!important;height:40px!important}
      .tr279-config{right:10px!important;top:24px!important;font-size:13px!important;padding:7px 9px!important;border-radius:12px!important}
      .tr279-config i{font-size:15px!important}
      .tr279-bus{font-size:46px!important}
      .tr279-title{font-size:16px!important;margin-top:4px!important}
      .tr279-body{padding:24px 14px 22px!important;min-height:auto!important;border-radius:16px 16px 0 0!important}
      .tr279-section{font-size:17px!important;margin:0 0 14px!important}
      .tr279-section.op{margin-top:24px!important;margin-bottom:13px!important}
      .tr279-grid3{gap:10px!important}
      .tr279-grid2{gap:12px!important}
      .tr279-tile{min-height:96px!important;border-radius:12px!important;padding:8px 5px!important;box-shadow:0 8px 16px rgba(0,0,0,.12)!important}
      .tr279-grid2 .tr279-tile{min-height:104px!important}
      .tr279-tile i{font-size:33px!important;margin-bottom:10px!important}
      .tr279-grid2 .tr279-tile i{font-size:35px!important;margin-bottom:11px!important}
      .tr279-tile .lbl{font-size:14px!important;line-height:1.08!important}
      .tr279-grid2 .tr279-tile .lbl{font-size:15px!important}
      .tr279-tile .sub{font-size:10px!important;margin-top:4px!important}
      .tr279-info{margin-top:18px!important;padding:12px 11px!important;font-size:12.5px!important;line-height:1.45!important;grid-template-columns:22px 1fr!important;gap:8px!important;border-radius:12px!important}
      .tr279-info i{font-size:18px!important}

      .tr279-subhead{height:78px!important;padding:0 12px!important}
      .tr279-subhead a{font-size:32px!important;width:38px!important}
      .tr279-subhead .ttl{font-size:16px!important;margin-right:38px!important}
      .tr279-subbody{padding:18px 14px 20px!important}
      .tr279-subgrid{gap:12px!important}
      .tr279-subgrid .wide{min-height:88px!important}
      .tr279-note{margin-top:14px!important;padding:11px 12px!important;font-size:12px!important;line-height:1.4!important}

      .tr280-gps-kpis{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:9px;margin:12px 0 14px}
      .tr280-gps-kpi{background:linear-gradient(135deg,#109349,#0b7a38);border-radius:10px;color:#fff;text-align:center;padding:9px 4px;box-shadow:0 8px 16px rgba(0,0,0,.10)}
      .tr280-gps-kpi label{display:block;font-size:10px;font-weight:900;line-height:1.1;margin-bottom:5px;color:#f3fff6}
      .tr280-gps-kpi strong{display:block;font-size:25px;line-height:1;font-weight:950;color:#fff}
      .tr280-gps-card{border:1px solid #dfe9df;background:#fff;border-radius:13px;padding:11px 11px 12px;margin-bottom:10px;box-shadow:0 5px 14px rgba(0,0,0,.05)}
      .tr280-gps-driver{display:flex;align-items:center;gap:9px;padding-bottom:9px;margin-bottom:9px;border-bottom:1px solid #edf2ed}
      .tr280-gps-avatar{width:38px;height:38px;border-radius:999px;background:#eaf9ee;color:#08713b;display:grid;place-items:center;font-size:20px;flex:0 0 38px}
      .tr280-gps-driver b{display:block;font-size:13px;color:#102a43;line-height:1.15}
      .tr280-gps-meta{font-size:10px;color:#64748b;font-weight:800;line-height:1.25;margin-top:2px}
      .tr280-gps-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px}
      .tr280-gps-box{background:#f8fff9;border:1px solid #ddeadd;border-radius:10px;padding:8px}
      .tr280-gps-box small{display:block;font-size:9px;color:#64748b;font-weight:900;line-height:1.1;margin-bottom:3px}
      .tr280-gps-box strong{display:block;font-size:11px;color:#065f2a;line-height:1.28;font-weight:900}
      .tr280-map-btn{margin-top:9px;width:100%}
      .tr280-empty{border:1px solid #dbe7db;background:#ffffff;border-radius:14px;padding:18px 12px;text-align:center;color:#5f6b74;font-weight:800;box-shadow:0 5px 14px rgba(0,0,0,.04)}
      .tr280-empty .big{font-size:14px;color:#334155;margin-bottom:4px}
      .tr280-empty small{display:block;font-size:11px;line-height:1.35;color:#64748b;font-weight:800}
    </style>
    """

@login_required
def tf_transporte_home_280():
    body = _transporte_ui_280_css() + r"""
    <div class="tr279-phone">
      <div class="tr279-app">
        <div class="tr279-hero">
          <a class="tr279-back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>
          <a class="tr279-config" href="{{url_for('transporte_config')}}"><i class="bi bi-gear"></i> Config.</a>
          <div class="tr279-bus"><i class="bi bi-bus-front-fill"></i></div>
          <div class="tr279-title">Módulo Transporte</div>
        </div>
        <div class="tr279-body">
          <div class="tr279-section">Módulos</div>
          <div class="tr279-grid3">
            <a class="tr279-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span class="lbl">Conductores</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span class="lbl">Buses</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i><span class="lbl">Rutas</span></a>
          </div>

          <div class="tr279-section op">Operación</div>
          <div class="tr279-grid2">
            <a class="tr279-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span class="lbl">GPS / Seguimiento</span><span class="sub">Ver ubicación</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_mobile_home')}}"><i class="bi bi-phone"></i><span class="lbl">Móvil conductor</span><span class="sub">Abordaje y GPS</span></a>
          </div>

          <div class="tr279-info"><i class="bi bi-info-circle-fill"></i><div>Así queda conectado: primero cargas conductores y buses, luego creas/cargas rutas asignando bus + conductor. El conductor entra a Móvil conductor, registra abordajes y envía GPS.</div></div>
        </div>
      </div>
    </div>
    """
    return render_page(body)

@login_required
def tf_transporte_mobile_home_280():
    body = _transporte_ui_280_css() + r"""
    <div class="tr279-phone">
      <div class="tr279-app">
        <div class="tr279-subhead">
          <a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a>
          <div class="ttl">Móvil conductor</div>
        </div>
        <div class="tr279-subbody">
          <div class="tr279-section">Operación conductor</div>
          <div class="tr279-subgrid">
            <a class="tr279-tile wide" href="{{url_for('transporte_abordaje_home')}}"><i class="bi bi-people"></i><span class="lbl">Abordaje trabajadores</span><span class="sub">QR / código / DNI</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span class="lbl">GPS</span><span class="sub">Enviar ubicación</span></a>
            <a class="tr279-tile" href="{{url_for('conductor_movil_login')}}"><i class="bi bi-phone"></i><span class="lbl">PIN conductor</span><span class="sub">Ingreso real</span></a>
          </div>
          <div class="tr279-note">El conductor ingresa desde su celular, registra abordajes y envía GPS. El administrador revisa seguimiento y bases desde Transporte.</div>
        </div>
      </div>
    </div>
    """
    return render_page(body)

@login_required
def tf_transporte_mapa_general_280():
    sql = """SELECT r.*, v.placa, v.tipo, c.dni AS conductor_dni, c.nombres AS conductor, c.estado AS conductor_estado,
                    c.ultima_latitud, c.ultima_longitud, c.ultimo_gps
             FROM transporte_rutas r
             LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
             LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
             ORDER BY COALESCE(r.ultima_ubicacion, r.creado_en) DESC, r.id DESC LIMIT 40"""
    rutas = rows_to_dict(execute(sql, fetchall=True))
    total = len(rutas)
    con_gps = sum(1 for r in rutas if (r.get('latitud') or r.get('ultima_latitud')) and (r.get('longitud') or r.get('ultima_longitud')))
    sin_bus = sum(1 for r in rutas if not r.get('placa'))
    body = _transporte_ui_280_css() + r"""
    <div class="tr279-phone">
      <div class="tr279-app">
        <div class="tr279-subhead">
          <a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a>
          <div class="ttl">GPS / Seguimiento</div>
        </div>
        <div class="tr279-subbody">
          <div class="tr279-info"><i class="bi bi-info-circle-fill"></i><div><b>¿Cómo funciona?</b><br>1) Crea conductor y PIN. 2) Asigna conductor y bus a una ruta. 3) El conductor entra a Móvil conductor desde su celular. 4) Permite ubicación/GPS y toca Enviar GPS.</div></div>
          <div class="tr280-gps-kpis">
            <div class="tr280-gps-kpi"><label>Rutas</label><strong>{{total}}</strong></div>
            <div class="tr280-gps-kpi"><label>Con GPS</label><strong>{{con_gps}}</strong></div>
            <div class="tr280-gps-kpi"><label>Sin bus</label><strong>{{sin_bus}}</strong></div>
          </div>
          {% if rutas %}
            {% for r in rutas %}
              <div class="tr280-gps-card">
                <div class="tr280-gps-driver">
                  <div class="tr280-gps-avatar"><i class="bi bi-person-fill"></i></div>
                  <div>
                    <b>{{r.conductor or 'SIN CONDUCTOR'}}</b>
                    <div class="tr280-gps-meta">DNI: {{r.conductor_dni or '-'}} · Ruta: {{r.nombre or 'SIN RUTA'}} · Estado: {{r.estado or 'SIN ESTADO'}}</div>
                  </div>
                </div>
                <div class="tr280-gps-grid">
                  <div class="tr280-gps-box"><small>Bus</small><strong>{{r.placa or 'SIN BUS'}}</strong></div>
                  <div class="tr280-gps-box"><small>Última actualización</small><strong>{{r.ultima_ubicacion or r.ultimo_gps or r.creado_en or '-'}}</strong></div>
                  <div class="tr280-gps-box"><small>Latitud</small><strong>{{r.latitud or r.ultima_latitud or '-'}}</strong></div>
                  <div class="tr280-gps-box"><small>Longitud</small><strong>{{r.longitud or r.ultima_longitud or '-'}}</strong></div>
                </div>
                {% set lat = r.latitud or r.ultima_latitud %}
                {% set lon = r.longitud or r.ultima_longitud %}
                {% if lat and lon %}
                  <a class="tf-btn tr280-map-btn" href="https://www.google.com/maps?q={{lat}},{{lon}}" target="_blank"><i class="bi bi-map"></i> Ver en mapa</a>
                {% endif %}
              </div>
            {% endfor %}
          {% else %}
            <div class="tr280-empty">
              <div class="big">Sin rutas para seguimiento.</div>
              <small>Primero crea rutas y luego envía GPS desde Móvil conductor.</small>
            </div>
          {% endif %}
        </div>
      </div>
    </div>
    """
    return render_page(body, rutas=rutas, total=total, con_gps=con_gps, sin_bus=sin_bus)

# Reemplazo final solicitado por Omar: home y móvil más delgados, y GPS corregido.
app.view_functions['transporte'] = tf_transporte_home_280
app.view_functions['transporte_mobile_home'] = tf_transporte_mobile_home_280
app.view_functions['transporte_mapa_general'] = tf_transporte_mapa_general_280

# ========================= PATCH TRANSPORTE OMAR 281 =========================
# Enlace real: Móvil conductor => login DNI + PIN; credenciales desde Conductores;
# abordaje/GPS funcionan con sesión móvil conductor, sin usar ADMIN.

def _trans_user_281():
    return session.get('usuario') or session.get('conductor_nombre') or 'CONDUCTOR MOVIL'

def _trans_is_allowed_281(ruta_id=None):
    if session.get('usuario'):
        return True
    cid = session.get('conductor_id')
    if not cid:
        return False
    if ruta_id is None:
        return True
    r = row_to_dict(execute('SELECT conductor_id FROM transporte_rutas WHERE id=?', (ruta_id,), fetchone=True))
    return bool(r and str(r.get('conductor_id') or '') == str(cid))

def _trans_redirect_281(ruta_id):
    if session.get('conductor_id') and not session.get('usuario'):
        return redirect(url_for('conductor_movil_ruta', ruta_id=ruta_id))
    return redirect(url_for('transporte_ruta_detalle', ruta_id=ruta_id))

def api_trabajador_281(dni):
    if not session.get('usuario') and not session.get('conductor_id'):
        return jsonify(ok=False, msg='Sesión no válida. Inicie sesión nuevamente.'), 401
    dni = limpiar_dni(dni)
    if len(dni) != 8:
        return jsonify(ok=False, msg='DNI inválido.')
    t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    if not t:
        return jsonify(ok=False, msg='DNI no encontrado en la base de trabajadores.')
    return jsonify(ok=True, trabajador=t, sugerido='TRANSPORTE')

def api_transporte_conductor_lookup_281(dni):
    if not session.get('usuario'):
        return jsonify(ok=False, msg='Sesión ADMIN requerida.'), 401
    dni = limpiar_dni(dni)
    if len(dni) != 8:
        return jsonify(ok=False, msg='DNI inválido.')
    trabajador = row_to_dict(execute('SELECT dni, trabajador, empresa, area, cargo, actividad FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    conductor = row_to_dict(execute('SELECT * FROM transporte_conductores WHERE dni=?', (dni,), fetchone=True))
    if not trabajador and not conductor:
        return jsonify(ok=False, msg='DNI no existe en Trabajadores ni en Conductores.')
    return jsonify(ok=True, trabajador=trabajador, conductor=conductor, pin_sugerido=_pin_conductor(dni))

try:
    app.add_url_rule('/api/transporte/conductor-lookup/<dni>', 'api_transporte_conductor_lookup_281', api_transporte_conductor_lookup_281, methods=['GET'])
except Exception:
    app.view_functions['api_transporte_conductor_lookup_281'] = api_transporte_conductor_lookup_281

def _conductores_html_281():
    return _tf_css() + _tf_transport_css_275() + _tf_top('Conductores / usuarios') + r"""
      <div class="tf-info">Busque el DNI en la base de trabajadores. Si existe, se llenan los datos y se crea el acceso móvil del conductor: <b>usuario = DNI</b> y <b>clave/PIN móvil</b>.</div>
      <div class="tf-mini-note">Flujo conectado: <b>Conductores → Buses → Rutas → Móvil conductor → Abordaje/GPS</b>.</div>
      <form method="post" class="tf-form" id="frmCond281">
        <div class="row g-2">
          <div class="col-12"><label>DNI conductor</label><div class="input-group"><input name="dni" id="dniConductor281" class="form-control" maxlength="8" inputmode="numeric" required placeholder="12345678" autofocus><button class="tf-btn" type="button" onclick="buscarConductorDni281()"><i class="bi bi-search"></i> Buscar</button></div></div>
          <div class="col-12"><label>Nombre conductor</label><input name="nombres" id="nombreConductor281" class="form-control" required placeholder="Se llena desde Trabajadores"></div>
          <div class="col-6"><label>Teléfono</label><input name="telefono" id="telCond281" class="form-control"></div>
          <div class="col-6"><label>Licencia</label><select name="licencia" class="form-select">{{lic_opts|safe}}</select></div>
          <div class="col-6"><label>Categoría</label><select name="categoria" class="form-select">{{cat_opts|safe}}</select></div>
          <div class="col-6"><label>Clave/PIN móvil</label><input name="movil_pin" id="pinCond281" class="form-control" placeholder="AUTO últimos 4 DNI"></div>
          <div class="col-4"><label>Venc. lic.</label><input name="venc_licencia" type="date" class="form-control"></div>
          <div class="col-4"><label>Cert. méd.</label><input name="venc_cert_medico" type="date" class="form-control"></div>
          <div class="col-4"><label>SCTR</label><input name="venc_sctr" type="date" class="form-control"></div>
          <div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>ACTIVO</option><option>APTO</option><option>OBSERVADO</option><option>INACTIVO</option></select></div>
          <div class="col-6"><label>Acceso móvil</label><select name="movil_estado" class="form-select"><option>ACTIVO</option><option>BLOQUEADO</option></select></div>
          <div class="col-12"><label>Observación</label><input name="observacion" class="form-control"></div>
        </div>
        <div id="statusCond281" class="field-help mt-2">Ingrese DNI de 8 dígitos y pulse buscar.</div>
        <button class="tf-btn w100 mt-2"><i class="bi bi-key"></i> Crear / actualizar usuario móvil</button>
      </form>
      <form method="post" enctype="multipart/form-data" action="{{url_for('transporte_upload_conductores')}}" class="tf-mass">
        <label><i class="bi bi-cloud-arrow-up"></i> Carga masiva conductores Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required>
        <div class="tf-help-mini">Columnas sugeridas: DNI, NOMBRE, TELEFONO, LICENCIA, CATEGORIA, VENC LICENCIA, CERT MEDICO, SCTR, ESTADO, PIN.</div>
        <button class="tf-btn w100" type="submit"><i class="bi bi-upload"></i> Cargar conductores</button><a class="tf-template-btn" href="{{url_for('transporte_plantilla_conductores')}}"><i class="bi bi-file-earmark-excel"></i> Descargar plantilla</a>
      </form>
      <div class="tf-searchrow"><input id="qcond" class="tf-search" placeholder="Buscar DNI o conductor..."><a class="tf-template-btn" style="margin:0;padding:9px 11px" href="{{url_for('transporte_plantilla_conductores')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div>
      <div class="tf-kpi-green"><div class="tf-kpi"><label>Total</label><strong>{{total}}</strong></div><div class="tf-kpi"><label>Activos</label><strong>{{activos}}</strong></div><div class="tf-kpi"><label>Bloq.</label><strong>{{bloqueados}}</strong></div></div>
      <div class="tf-base-title">Base de conductores y accesos móviles</div>
      <div class="tf-tablewrap"><table id="tblcond" class="tf-table"><thead><tr><th>DNI</th><th>Conductor</th><th>Licencia</th><th>Usuario</th><th>PIN</th><th>Móvil</th><th>Activo</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.nombres}}</td><td>{{r.licencia or '-' }}<br><small>{{r.categoria or ''}}</small></td><td>{{r.movil_usuario or r.dni}}</td><td>{{r.movil_pin or '-'}}</td><td>{{badge(r.movil_estado or 'ACTIVO')|safe}}</td><td>{{switch((r.estado or '').upper() in ['ACTIVO','APTO',''], url_for('transporte_toggle_conductor', item_id=r.id))|safe}}</td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}</tbody></table></div>
      <a class="tf-btn w100 mt-2" href="{{url_for('conductor_movil_login')}}"><i class="bi bi-phone"></i> Probar ingreso móvil conductor</a>
      <script>
      async function buscarConductorDni281(){const dniInput=document.getElementById('dniConductor281'), st=document.getElementById('statusCond281'); const d=String(dniInput.value||'').replace(/\D/g,'').slice(-8); dniInput.value=d; if(d.length!==8){st.className='scan-bad mt-2';st.innerHTML='DNI inválido. Debe tener 8 dígitos.';return;} st.className='scan-ok mt-2'; st.innerHTML='Buscando DNI '+d+'...'; try{const r=await fetch('/api/transporte/conductor-lookup/'+d,{cache:'no-store'}); const j=await r.json(); if(!j.ok){st.className='scan-bad mt-2';st.innerHTML='✕ '+(j.msg||'No encontrado');return;} const t=j.trabajador||{}, c=j.conductor||{}; document.getElementById('nombreConductor281').value=c.nombres||t.trabajador||''; document.getElementById('telCond281').value=c.telefono||''; document.getElementById('pinCond281').value=c.movil_pin||j.pin_sugerido||d.slice(-4); st.className='scan-ok mt-2'; st.innerHTML='✓ Datos encontrados. Usuario móvil: <b>'+d+'</b>. Revise/guarde la clave PIN.'; if(typeof beep==='function')beep();}catch(e){st.className='scan-bad mt-2';st.innerHTML='Error consultando DNI.';}}
      (function(){const i=document.getElementById('dniConductor281'); if(i){i.addEventListener('input',()=>{const d=i.value.replace(/\D/g,'').slice(-8); i.value=d; if(d.length===8)buscarConductorDni281();});}})();
      </script>
    """ + _tf_end() + _tf_filter_script('qcond','tblcond')

@login_required
def transporte_conductores_281():
    if request.method == 'POST':
        dni = limpiar_dni(request.form.get('dni'))
        base = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True)) if len(dni) == 8 else None
        nombres = limpiar_texto(request.form.get('nombres') or (base or {}).get('trabajador'))
        if len(dni) != 8 or not nombres:
            flash('Ingrese un DNI válido. Si el conductor está en Trabajadores, pulse BUSCAR DNI para traer sus datos.', 'danger'); return redirect(url_for('transporte_conductores'))
        pin = (request.form.get('movil_pin') or _pin_conductor(dni)).strip()
        if len(pin) < 4:
            flash('La clave/PIN móvil debe tener mínimo 4 caracteres.', 'danger'); return redirect(url_for('transporte_conductores'))
        estado = limpiar_texto(request.form.get('estado') or 'ACTIVO')
        movil_estado = limpiar_texto(request.form.get('movil_estado') or ('BLOQUEADO' if estado in ('INACTIVO','BLOQUEADO') else 'ACTIVO'))
        vals = (dni, nombres, request.form.get('telefono',''), limpiar_texto(request.form.get('licencia') or 'A-IIIc'), limpiar_texto(request.form.get('categoria') or 'BUS / TRANSPORTE DE PERSONAS'), request.form.get('venc_licencia',''), request.form.get('venc_cert_medico',''), request.form.get('venc_sctr',''), estado, limpiar_texto(request.form.get('observacion'), upper=False), dni, pin, movil_estado, now_str())
        try:
            execute("""INSERT INTO transporte_conductores(dni,nombres,telefono,licencia,categoria,venc_licencia,venc_cert_medico,venc_sctr,estado,observacion,movil_usuario,movil_pin,movil_estado,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", vals, commit=True)
            flash(f'Conductor creado. Usuario móvil: {dni} | Clave/PIN: {pin}', 'success')
        except Exception:
            execute("""UPDATE transporte_conductores SET nombres=?,telefono=?,licencia=?,categoria=?,venc_licencia=?,venc_cert_medico=?,venc_sctr=?,estado=?,observacion=?,movil_usuario=?,movil_pin=?,movil_estado=? WHERE dni=?""", (vals[1],vals[2],vals[3],vals[4],vals[5],vals[6],vals[7],vals[8],vals[9],dni,pin,movil_estado,dni), commit=True)
            flash(f'Conductor actualizado. Usuario móvil: {dni} | Clave/PIN: {pin}', 'success')
        return redirect(url_for('transporte_conductores'))
    rows = rows_to_dict(execute('SELECT * FROM transporte_conductores ORDER BY id DESC LIMIT 300', fetchall=True))
    hoy = today_str(); total = len(rows)
    activos = sum(1 for r in rows if (r.get('estado') or '').upper() in ('ACTIVO','APTO',''))
    bloqueados = sum(1 for r in rows if (r.get('movil_estado') or '').upper() == 'BLOQUEADO')
    venc = sum(1 for r in rows if any(r.get(c) and str(r.get(c)) < hoy for c in ('venc_licencia','venc_cert_medico','venc_sctr')))
    return render_page(_conductores_html_281(), rows=rows, total=total, activos=activos, bloqueados=bloqueados, venc=venc, badge=_tf_badge, switch=_tf_switch, lic_opts=_option_list(LICENCIAS_PERU,'A-IIIc'), cat_opts=_option_list(CATEGORIAS_CONDUCTOR_PERU,'BUS / TRANSPORTE DE PERSONAS'))

@login_required
def tf_transporte_home_281():
    body = _transporte_ui_280_css() + r"""
    <div class="tr279-phone">
      <div class="tr279-app">
        <div class="tr279-hero">
          <a class="tr279-back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>
          <a class="tr279-config" href="{{url_for('transporte_config')}}"><i class="bi bi-gear"></i> Config.</a>
          <div class="tr279-bus"><i class="bi bi-bus-front-fill"></i></div>
          <div class="tr279-title">Módulo Transporte</div>
        </div>
        <div class="tr279-body">
          <div class="tr279-section">Módulos</div>
          <div class="tr279-grid3">
            <a class="tr279-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span class="lbl">Conductores</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span class="lbl">Buses</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i><span class="lbl">Rutas</span></a>
          </div>

          <div class="tr279-section op">Operación</div>
          <div class="tr279-grid2">
            <a class="tr279-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span class="lbl">GPS / Seguimiento</span><span class="sub">Ver ubicación</span></a>
            <a class="tr279-tile" href="{{url_for('conductor_movil_login')}}"><i class="bi bi-phone"></i><span class="lbl">Móvil conductor</span><span class="sub">Abordaje y GPS</span></a>
          </div>

          <div class="tr279-info"><i class="bi bi-info-circle-fill"></i><div><b>Flujo correcto:</b> primero crea el usuario/PIN en Conductores, luego asigna conductor + bus a la ruta. Al tocar Móvil conductor se abre directamente el login DNI + PIN, no ADMIN.</div></div>
        </div>
      </div>
    </div>
    """
    return render_page(body)

def conductor_movil_login_281():
    if request.method == 'POST':
        dni = limpiar_dni(request.form.get('dni'))
        pin = (request.form.get('pin') or '').strip()
        if len(dni) != 8 or dni.upper() == 'ADMIN':
            flash('Solo conductores: ingrese DNI de 8 dígitos. No usar ADMIN.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        if not pin:
            flash('Ingrese su PIN móvil.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        c = row_to_dict(execute("""SELECT * FROM transporte_conductores
                                 WHERE dni=? AND COALESCE(movil_pin,'')=?
                                 AND COALESCE(movil_estado,'ACTIVO')='ACTIVO'
                                 AND COALESCE(estado,'ACTIVO') NOT IN ('INACTIVO','BLOQUEADO')""", (dni, pin), fetchone=True))
        if not c:
            flash('DNI o PIN incorrecto, bloqueado o no creado en Transporte > Conductores.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        for k in ('usuario','rol','nombres'):
            session.pop(k, None)
        session['conductor_id'] = c.get('id')
        session['conductor_dni'] = c.get('dni')
        session['conductor_nombre'] = c.get('nombres') or c.get('dni')
        flash('Bienvenido conductor. Ya puede registrar abordaje y GPS.', 'success')
        return redirect(url_for('conductor_movil_panel'))
    body = r"""
    <div class="phone-wrap">
      <div class="page-card" style="border-radius:14px;overflow:hidden;margin-top:16px">
        <div class="green-hero" style="min-height:128px;border-radius:0;padding:18px 16px 38px">
          <a href="{{url_for('transporte')}}" style="position:absolute;left:18px;top:23px;color:white;font-size:38px;text-decoration:none"><i class="bi bi-chevron-left"></i></a>
          <div style="font-size:34px"><i class="bi bi-phone"></i></div>
          <div style="font-family:Georgia,serif;font-weight:900;font-size:13px;margin-top:4px">ACCESO MÓVIL CONDUCTOR</div>
        </div>
        <div class="floating-card" style="margin:-30px 12px 14px;border-radius:12px;padding:16px">
          <div class="alert alert-light" style="border:1px solid #dbe3db;color:#24405f;line-height:1.45"><b>Solo conductores:</b> ingrese DNI del conductor + PIN móvil creado en Transporte &gt; Conductores. No usar ADMIN.</div>
          <form method="post" id="frmMovilCond281">
            <label class="form-label">DNI conductor</label>
            <input name="dni" id="dniMovCond281" class="form-control mb-3" inputmode="numeric" maxlength="8" pattern="\d{8}" placeholder="Ingrese DNI" required autocomplete="username" autofocus>
            <label class="form-label">PIN móvil</label>
            <input name="pin" class="form-control mb-3" type="password" placeholder="PIN" required autocomplete="current-password">
            <button class="btn btn-green w-100" style="font-size:19px;height:47px">INGRESAR</button>
          </form>
          {% if session.get('usuario') %}
          <a class="btn btn-outline-success w-100 mt-2" href="{{url_for('transporte_conductores')}}">Crear / resetear PIN</a>
          {% endif %}
          <a class="btn btn-outline-secondary w-100 mt-2" href="{{url_for('transporte')}}">Volver</a>
        </div>
      </div>
    </div>
    <script>
      (function(){const i=document.getElementById('dniMovCond281'); if(i){i.addEventListener('input',()=>{i.value=String(i.value||'').replace(/\D/g,'').slice(-8);});}})();
    </script>
    """
    return render_page(body, title='Acceso móvil conductor')

def conductor_movil_panel_281():
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    cid = session.get('conductor_id')
    hoy = today_str()
    # Mostrar rutas de hoy y también rutas activas anteriores. Si una ruta se creó ayer y sigue PROGRAMADA/EN RUTA,
    # el conductor debe poder abrir abordaje y GPS.
    rutas = rows_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor
                                  FROM transporte_rutas r
                                  LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                                  LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                                  WHERE r.conductor_id=?
                                    AND (r.fecha>=? OR UPPER(COALESCE(r.estado,'')) IN ('PROGRAMADA','EN RUTA'))
                                  ORDER BY CASE WHEN r.fecha=? THEN 0 ELSE 1 END,
                                           r.fecha DESC, r.hora_salida ASC, r.id DESC LIMIT 50""", (cid, hoy, hoy), fetchall=True))
    body = r"""
    <div class="phone-wrap"><div class="page-card" style="border-radius:14px;overflow:hidden;margin-top:12px">
      <div class="green-hero" style="border-radius:0;min-height:118px;padding:18px 14px 32px">
        <a href="{{url_for('conductor_movil_logout')}}" style="position:absolute;left:15px;top:24px;color:white;font-size:30px"><i class="bi bi-box-arrow-left"></i></a>
        <div style="font-size:34px"><i class="bi bi-person-badge"></i></div><div style="font-weight:900;font-size:13px">{{session.get('conductor_nombre')}}</div><div style="font-size:11px;font-weight:800">{{session.get('conductor_dni')}}</div>
      </div>
      <div class="floating-card" style="margin:-24px 12px 14px;border-radius:12px;padding:13px"><b>Mis rutas asignadas</b><br><small>Toque una ruta para abrir <b>Abordaje</b> y <b>GPS</b>.</small></div>
      {% for r in rutas %}
        <a class="worker-card" style="display:block;text-decoration:none;color:inherit" href="{{url_for('conductor_movil_ruta', ruta_id=r.id)}}">
          <div class="worker-title"><div>{{r.fecha}} {{r.hora_salida or ''}}<br><b>{{r.nombre}}</b></div><div class="text-end">{{r.placa or 'SIN BUS'}}<br><b>{{r.estado}}</b></div></div>
          <div class="worker-grid"><div><label>ORIGEN</label><div class="small-value">{{r.origen or '-'}}</div></div><div><label>DESTINO</label><div class="small-value">{{r.destino or '-'}}</div></div><div><label>CAP.</label><div class="small-value">{{r.capacidad or '-'}}</div></div></div>
        </a>
      {% else %}
        <div class="worker-card text-center" style="color:#24405f">
          <b>No hay ruta asignada a este conductor.</b><br>
          <small>El login está correcto, pero el abordaje aparece dentro de una ruta. Cree una ruta en ADMIN y asigne este conductor, o use ruta rápida para prueba/emergencia.</small>
          <form method="post" action="{{url_for('conductor_movil_ruta_rapida')}}" class="mt-2">
            <button class="btn btn-green w-100" type="submit"><i class="bi bi-plus-circle"></i> Crear ruta rápida y abrir abordaje</button>
          </form>
          <div class="field-help mt-2">Ruta rápida no asigna bus. Para control completo: ADMIN &gt; Movilidad &gt; Rutas.</div>
        </div>
      {% endfor %}
    </div></div>
    """
    return render_page(body, rutas=rutas, title='Panel conductor')

@app.route('/movil/conductor/ruta-rapida', methods=['POST'])
def conductor_movil_ruta_rapida_282():
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    cid = session.get('conductor_id')
    hoy = today_str()
    # Reutiliza una ruta rápida abierta del mismo día para no duplicar.
    existe = row_to_dict(execute("""SELECT id FROM transporte_rutas
                                  WHERE conductor_id=? AND fecha=?
                                    AND nombre LIKE 'RUTA RAPIDA MOVIL%'
                                    AND UPPER(COALESCE(estado,'')) IN ('PROGRAMADA','EN RUTA')
                                  ORDER BY id DESC LIMIT 1""", (cid, hoy), fetchone=True))
    if existe:
        return redirect(url_for('conductor_movil_ruta', ruta_id=existe.get('id')))
    nombre = 'RUTA RAPIDA MOVIL ' + datetime.now().strftime('%H:%M')
    execute("""INSERT INTO transporte_rutas(fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,vehiculo_id,conductor_id,estado,creado_por,creado_en)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", (hoy, nombre, 'ORIGEN', 'DESTINO', '', datetime.now().strftime('%H:%M'), '', None, cid, 'EN RUTA', _trans_user_281(), now_str()), commit=True)
    r = row_to_dict(execute('SELECT id FROM transporte_rutas WHERE conductor_id=? AND fecha=? AND nombre=? ORDER BY id DESC LIMIT 1', (cid, hoy, nombre), fetchone=True))
    flash('Ruta rápida creada. Ya puede registrar abordaje y enviar GPS.', 'success')
    return redirect(url_for('conductor_movil_ruta', ruta_id=r.get('id')))

def conductor_movil_ruta_281(ruta_id):
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    if not _trans_is_allowed_281(ruta_id):
        flash('Esta ruta no está asignada a su usuario conductor.', 'danger')
        return redirect(url_for('conductor_movil_panel'))
    ruta = row_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor
                                FROM transporte_rutas r
                                LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                                LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                                WHERE r.id=?""", (ruta_id,), fetchone=True))
    pasajeros = rows_to_dict(execute('SELECT * FROM transporte_pasajeros WHERE ruta_id=? ORDER BY fecha_hora DESC', (ruta_id,), fetchall=True))
    ocupados = len(pasajeros); capacidad = int((ruta or {}).get('capacidad') or 0); libres = max(0, capacidad-ocupados) if capacidad else '-'
    body = r"""
    <div class="phone-wrap"><div class="page-card" style="border-radius:14px;overflow:hidden;margin-top:10px">
      <div class="green-hero" style="border-radius:0;min-height:126px;padding:17px 14px 32px">
        <a href="{{url_for('conductor_movil_panel')}}" style="position:absolute;left:15px;top:24px;color:white;font-size:32px"><i class="bi bi-chevron-left"></i></a>
        <div style="font-size:34px"><i class="bi bi-bus-front-fill"></i></div><div style="font-weight:900;font-size:13px">{{ruta.nombre}}</div><div style="font-size:11px;font-weight:800">{{ruta.origen}} → {{ruta.destino}}</div>
      </div>
      <div class="floating-card" style="margin:-24px 12px 10px;border-radius:12px;padding:12px">
        <div class="row text-center g-2"><div class="col-4"><b>{{ocupados}}</b><br><small>Abordados</small></div><div class="col-4"><b>{{libres}}</b><br><small>Libres</small></div><div class="col-4"><b>{{ruta.placa or '-'}}</b><br><small>Bus</small></div></div>
      </div>
      <form method="post" action="{{url_for('transporte_abordar', ruta_id=ruta.id)}}" class="worker-card" id="frmAbordar281">
        <label class="form-label">DNI trabajador</label>
        <div class="input-group mb-2"><input name="dni" id="dniTransporte281" class="form-control" inputmode="numeric" maxlength="8" required placeholder="Escanear o digitar DNI"><button class="btn btn-outline-success" type="button" onclick="abrirScanner&&abrirScanner('readerTrans281','dniTransporte281')"><i class="bi bi-qr-code-scan"></i></button></div>
        <div id="readerTrans281" style="display:none" class="scan-box mb-2"></div>
        <div id="transpStatus281" class="field-help mb-2">Al completar 8 dígitos se validará en Trabajadores.</div>
        <input type="hidden" name="metodo" value="QR/DNI MOVIL"><input type="hidden" name="latitud" id="latitudTrans281"><input type="hidden" name="longitud" id="longitudTrans281">
        <button class="btn btn-green w-100"><i class="bi bi-person-check"></i> REGISTRAR SUBIDA</button>
      </form>
      <div class="mx-2 mb-2 d-grid gap-2">
        <button class="btn btn-outline-success" onclick="enviarGpsTransporte281({{ruta.id}}, false)"><i class="bi bi-geo-alt"></i> Enviar GPS ahora</button>
        <button class="btn btn-outline-success" onclick="enviarGpsTransporte281({{ruta.id}}, true)"><i class="bi bi-broadcast-pin"></i> Iniciar GPS en vivo</button>
      </div>
      {% for p in pasajeros %}<div class="worker-card"><div class="worker-title"><div>{{p.hora}}<br><b>{{p.trabajador}}</b></div><div class="text-end">{{p.dni}}<br><b>{{p.metodo}}</b></div></div></div>{% else %}<div class="worker-card text-center text-muted">Aún no hay trabajadores abordados.</div>{% endfor %}
    </div></div>
    <script>
    (function(){const dni=v=>String(v||'').replace(/\D/g,'').slice(-8), i=document.getElementById('dniTransporte281'), st=document.getElementById('transpStatus281'); async function val(){const d=dni(i.value); i.value=d; if(d.length<8){st.className='field-help mb-2'; st.innerHTML='Esperando 8 dígitos...'; return;} st.className='scan-ok mb-2'; st.innerHTML='Validando DNI '+d+'...'; try{const r=await fetch('/api/trabajador/'+d,{cache:'no-store'}); const j=await r.json(); if(j.ok){st.className='scan-ok mb-2'; st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR')+' encontrado'; if(typeof beep==='function')beep();}else{st.className='scan-bad mb-2'; st.innerHTML='✕ '+j.msg;}}catch(e){st.className='scan-bad mb-2'; st.innerHTML='Error validando DNI';}} if(i){i.addEventListener('input',val); i.addEventListener('paste',()=>setTimeout(val,80));} if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{document.getElementById('latitudTrans281').value=p.coords.latitude;document.getElementById('longitudTrans281').value=p.coords.longitude;},()=>{});}})();
    let gpsWatch281=null; async function enviarGpsTransporte281(rid, live){function post(p){let fd=new FormData();fd.append('latitud',p.coords.latitude);fd.append('longitud',p.coords.longitude);document.getElementById('latitudTrans281').value=p.coords.latitude;document.getElementById('longitudTrans281').value=p.coords.longitude;return fetch('/transporte/ruta/'+rid+'/gps',{method:'POST',body:fd}).then(r=>r.json());} if(!navigator.geolocation){alert('GPS no disponible');return;} if(live){if(gpsWatch281){navigator.geolocation.clearWatch(gpsWatch281);gpsWatch281=null;alert('GPS en vivo detenido');return;} gpsWatch281=navigator.geolocation.watchPosition(p=>post(p).catch(()=>{}),()=>alert('Permite ubicación/GPS'),{enableHighAccuracy:true,maximumAge:10000,timeout:15000}); alert('GPS en vivo iniciado. Mantenga esta pantalla abierta.'); return;} navigator.geolocation.getCurrentPosition(async p=>{let j=await post(p);alert(j.msg||'GPS actualizado');},()=>alert('Permite ubicación/GPS en el navegador'),{enableHighAccuracy:true,maximumAge:10000,timeout:15000});}
    </script>
    """
    return render_page(body, ruta=ruta, pasajeros=pasajeros, ocupados=ocupados, capacidad=capacidad, libres=libres, title='Ruta conductor')

def transporte_abordar_281(ruta_id):
    if not _trans_is_allowed_281(ruta_id):
        flash('Sesión no válida o ruta no asignada al conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    ruta=row_to_dict(execute('SELECT r.*, v.capacidad FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id WHERE r.id=?', (ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.', 'danger'); return redirect(url_for('transporte') if session.get('usuario') else url_for('conductor_movil_panel'))
    dni=limpiar_dni(request.form.get('dni'))
    if len(dni)!=8:
        flash('DNI inválido. No se registró subida.', 'danger'); return _trans_redirect_281(ruta_id)
    t=row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    if not t:
        flash('DNI no existe en base trabajadores. No se registró subida.', 'danger'); return _trans_redirect_281(ruta_id)
    dup=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE ruta_id=? AND dni=?', (ruta_id,dni))
    if dup:
        flash('Este trabajador ya fue registrado en esta ruta.', 'danger'); return _trans_redirect_281(ruta_id)
    capacidad=int(ruta.get('capacidad') or 0); ocupados=scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE ruta_id=?', (ruta_id,))
    if capacidad and ocupados >= capacidad:
        flash('Capacidad completa del vehículo. No se registró subida.', 'danger'); return _trans_redirect_281(ruta_id)
    ahora=datetime.now(); fecha=ahora.strftime('%Y-%m-%d'); hora=ahora.strftime('%H:%M:%S')
    execute("""INSERT INTO transporte_pasajeros(ruta_id,dni,trabajador,empresa,area,cargo,fecha,hora,fecha_hora,metodo,latitud,longitud,registrado_por,observacion) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (ruta_id,dni,t.get('trabajador'),t.get('empresa'),t.get('area'),t.get('cargo'),fecha,hora,ahora.strftime('%Y-%m-%d %H:%M:%S'),limpiar_texto(request.form.get('metodo') or 'QR/DNI MOVIL'),request.form.get('latitud',''),request.form.get('longitud',''),_trans_user_281(),limpiar_texto(request.form.get('observacion'), upper=False)), commit=True)
    flash(f'Subida registrada: {dni} - {t.get("trabajador")}', 'success')
    return _trans_redirect_281(ruta_id)

def transporte_gps_actualizar_281(ruta_id):
    if not _trans_is_allowed_281(ruta_id):
        return jsonify(ok=False, msg='Sesión no válida o ruta no asignada al conductor.'), 401
    lat=(request.form.get('latitud') or '').strip(); lon=(request.form.get('longitud') or '').strip(); fh=now_str()
    if not lat or not lon:
        return jsonify(ok=False, msg='No llegó coordenada GPS. Permita ubicación en el navegador.'), 400
    ruta = row_to_dict(execute('SELECT r.*, v.placa FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id WHERE r.id=?', (ruta_id,), fetchone=True)) or {}
    conductor_id = ruta.get('conductor_id') or session.get('conductor_id')
    execute('INSERT INTO transporte_gps(ruta_id,latitud,longitud,fecha_hora,registrado_por,conductor_id,placa,ruta_nombre) VALUES(?,?,?,?,?,?,?,?)', (ruta_id,lat,lon,fh,_trans_user_281(),conductor_id,ruta.get('placa'),ruta.get('nombre')), commit=True)
    execute('UPDATE transporte_rutas SET latitud=?, longitud=?, ultima_ubicacion=?, estado=? WHERE id=?', (lat,lon,fh,'EN RUTA',ruta_id), commit=True)
    if conductor_id:
        execute('UPDATE transporte_conductores SET ultima_latitud=?, ultima_longitud=?, ultimo_gps=? WHERE id=?', (lat,lon,fh,conductor_id), commit=True)
    return jsonify(ok=True, msg='GPS actualizado correctamente')

def conductor_movil_logout_281():
    for k in ('conductor_id','conductor_dni','conductor_nombre'):
        session.pop(k, None)
    flash('Sesión móvil cerrada.', 'success')
    return redirect(url_for('conductor_movil_login'))

# Activar overrides finales del módulo transporte
app.view_functions['api_trabajador'] = api_trabajador_281
app.view_functions['transporte'] = tf_transporte_home_281
app.view_functions['transporte_mobile_home'] = conductor_movil_login_281
app.view_functions['transporte_conductores'] = transporte_conductores_281
app.view_functions['conductor_movil_login'] = conductor_movil_login_281
app.view_functions['conductor_movil_panel'] = conductor_movil_panel_281
app.view_functions['conductor_movil_ruta'] = conductor_movil_ruta_281
app.view_functions['conductor_movil_logout'] = conductor_movil_logout_281
app.view_functions['transporte_abordar'] = transporte_abordar_281
app.view_functions['transporte_gps_actualizar'] = transporte_gps_actualizar_281
# ======================= FIN PATCH TRANSPORTE OMAR 281 =======================


# ========================= PATCH TRANSPORTE OMAR 283 =========================
# Unifica Móvil conductor con la interfaz preferida de Abordaje trabajadores / Abordaje ruta.
# Elimina la experiencia duplicada tipo imagen 3/4: ahora el conductor ve la misma lógica visual de imagen 1/2.

def _movil_cond_283_css():
    return r'''
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}.shell{max-width:560px!important;width:100%!important;margin:0 auto!important;padding:8px 10px 28px!important;background:#fff!important}.phone-wrap{max-width:520px!important;width:100%!important;margin:0 auto!important}.mv283-card{background:#fff;border:1px solid #e3e8e3;border-radius:15px;overflow:hidden;box-shadow:0 12px 26px rgba(0,0,0,.08);margin:8px auto 16px}.mv283-head{height:70px;background:#25773a;color:#fff;display:flex;align-items:center;justify-content:center;position:relative}.mv283-head a{position:absolute;left:14px;top:50%;transform:translateY(-50%);color:#fff!important;text-decoration:none;font-size:35px;line-height:1}.mv283-head .ttl{font-size:19px;font-weight:950;letter-spacing:.2px;color:#fff;text-align:center}.mv283-body{padding:15px 15px 18px;background:#fff}.mv283-info{border:1px solid #b8d7ff;background:#eef6ff;color:#0b2e83;border-radius:10px;padding:13px 14px;font-size:15px;font-weight:900;line-height:1.45;margin-bottom:12px}.mv283-ok{border:1px solid #bbf7d0;background:#ecfdf5;color:#065f2a;border-radius:10px;padding:11px 12px;font-size:14px;font-weight:900;line-height:1.45;margin-bottom:15px}.mv283-section{font-size:16px;font-weight:950;color:#08713b;text-transform:uppercase;margin:7px 2px 10px}.mv283-route{display:grid;grid-template-columns:42px 1fr 24px;gap:10px;align-items:center;text-decoration:none;color:#102a43!important;border:1px solid #dfe7df;background:#fff;border-radius:12px;padding:15px 12px;margin:8px 0;box-shadow:0 5px 14px rgba(0,0,0,.04)}.mv283-route i.bus{font-size:27px;color:#08713b}.mv283-route .name{font-size:17px;font-weight:950;color:#0a1f44;line-height:1.1}.mv283-route .meta{font-size:12px;font-weight:950;color:#0a1f44;margin-top:5px}.mv283-route .chev{font-size:26px;color:#111}.mv283-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin:12px 0 14px}.mv283-kpi{background:#10964e;color:#fff;border-radius:8px;text-align:center;padding:10px 6px;box-shadow:0 8px 16px rgba(16,150,78,.18)}.mv283-kpi small{display:block;font-size:12px;font-weight:950;line-height:1.08;color:#effff3}.mv283-kpi b{display:block;font-size:29px;line-height:1.05;font-weight:950;color:#fff;margin-top:5px}.mv283-routebox{border:1px solid #b8d7ff;background:#eaf3ff;color:#0b2e83;border-radius:10px;padding:14px;font-size:14px;font-weight:950;line-height:1.48;margin-bottom:13px}.mv283-routebox .label{font-size:14px;color:#0b2e83;text-transform:uppercase}.mv283-routebox .route-title{font-size:16px;color:#0b2e83;font-weight:950}.mv283-form{border:1px solid #d7eadc;background:#fbfffc;border-radius:13px;padding:14px;margin-bottom:11px}.mv283-form label{font-size:13px;font-weight:950;color:#176a35;margin-bottom:6px}.mv283-form .input-group .form-control{height:45px;border-radius:10px 0 0 10px!important;font-size:16px;font-weight:850}.mv283-camera{min-width:58px;background:#08713b!important;color:#fff!important;border-color:#08713b!important;border-radius:0 10px 10px 0!important}.mv283-btn{height:46px;border-radius:10px;background:#08713b;border:1px solid #08713b;color:#fff;font-weight:950;font-size:16px;width:100%}.mv283-btn:hover{background:#065f2a;color:#fff}.mv283-outline{height:46px;border-radius:10px;background:#fff;border:1px solid #08713b;color:#08713b;font-weight:900;font-size:16px;width:100%}.mv283-tablewrap{border:1px solid #e5e7eb;border-radius:10px;overflow-x:auto;background:#fff;scrollbar-color:#08713b #e5e7eb}.mv283-tablewrap::-webkit-scrollbar{height:9px}.mv283-tablewrap::-webkit-scrollbar-thumb{background:#08713b;border-radius:999px}.mv283-tablewrap::-webkit-scrollbar-track{background:#e5e7eb}.mv283-table{width:100%;min-width:500px;border-collapse:collapse}.mv283-table th{background:#f8fafc;color:#12223b;font-size:14px;font-weight:950;padding:9px;border-bottom:1px solid #e5e7eb}.mv283-table td{font-size:13px;color:#334155;padding:9px;border-bottom:1px solid #f1f5f9}.mv283-empty{border:1px solid #e5e7eb;background:#fff;border-radius:12px;text-align:center;color:#5f6673;font-size:19px;line-height:1.45;padding:17px;margin-top:10px;box-shadow:0 5px 14px rgba(0,0,0,.04)}.mv283-quick{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:13px;text-align:center;color:#1f3b2a}.mv283-driver{background:#25773a;color:#fff;text-align:center;padding:18px 12px 28px;position:relative}.mv283-driver a{position:absolute;left:15px;top:24px;color:#fff!important;text-decoration:none;font-size:30px}.mv283-driver .ico{font-size:38px}.mv283-driver .name{font-size:15px;font-weight:950;color:#fff}.mv283-driver .dni{font-size:13px;font-weight:900;color:#fff}.mv283-driver + .mv283-body{margin-top:-16px;border-radius:15px 15px 0 0;position:relative}.mv283-mini-card{background:#fff;border-radius:11px;padding:11px 14px;margin-bottom:13px;box-shadow:0 6px 15px rgba(0,0,0,.09);font-size:15px;color:#102a43;line-height:1.5}.scan-ok,.scan-bad,.field-help{font-size:13px!important;font-weight:900!important}.scan-ok{background:#ecfdf5!important;border:1px solid #bbf7d0!important;color:#065f2a!important;border-radius:9px!important;padding:8px 10px!important}.scan-bad{background:#fee2e2!important;border:1px solid #fecaca!important;color:#991b1b!important;border-radius:9px!important;padding:8px 10px!important}.field-help{color:#4a644f!important;margin-top:6px}.flash{animation:mv283flash .3s ease}@keyframes mv283flash{0%{transform:scale(.99)}60%{transform:scale(1.01)}100%{transform:scale(1)}}@media(max-width:430px){.shell{max-width:100%!important;padding:5px 7px 22px!important}.phone-wrap{max-width:100%!important}.mv283-head{height:68px}.mv283-head .ttl{font-size:18px}.mv283-body{padding:14px 14px 16px}.mv283-info{font-size:14px}.mv283-ok{font-size:13px}.mv283-routebox{font-size:13px}.mv283-kpis{gap:8px}.mv283-kpi{padding:9px 4px}.mv283-kpi small{font-size:11px}.mv283-kpi b{font-size:28px}.mv283-empty{font-size:17px}.mv283-route{grid-template-columns:39px 1fr 20px;padding:13px 10px}.mv283-route .name{font-size:16px}.mv283-route .meta{font-size:11px}.mv283-form .input-group .form-control{font-size:15px}}
    </style>
    '''

def conductor_movil_panel_283():
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    cid = session.get('conductor_id')
    hoy = today_str()
    rutas = rows_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor
                                  FROM transporte_rutas r
                                  LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                                  LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                                  WHERE r.conductor_id=?
                                    AND (r.fecha>=? OR UPPER(COALESCE(r.estado,'')) IN ('PROGRAMADA','EN RUTA'))
                                  ORDER BY CASE WHEN r.fecha=? THEN 0 ELSE 1 END,
                                           r.fecha DESC, r.hora_salida ASC, r.id DESC LIMIT 50""", (cid, hoy, hoy), fetchall=True))
    ruta_ids = [str(r.get('id')) for r in rutas if r.get('id')]
    abordados_hoy = 0
    rutas_con_abordaje = 0
    if ruta_ids:
        marks = ','.join(['?'] * len(ruta_ids))
        abordados_hoy = scalar(f"SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=? AND ruta_id IN ({marks})", tuple([hoy] + ruta_ids))
        rutas_con_abordaje = scalar(f"SELECT COUNT(DISTINCT ruta_id) AS c FROM transporte_pasajeros WHERE fecha=? AND ruta_id IN ({marks})", tuple([hoy] + ruta_ids))
    pendientes = max(0, len(rutas) - int(rutas_con_abordaje or 0))
    body = _movil_cond_283_css() + r'''
    <div class="phone-wrap"><div class="mv283-card">
      <div class="mv283-head"><a href="{{url_for('conductor_movil_logout')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Abordaje trabajadores</div></div>
      <div class="mv283-body">
        <div class="mv283-info">Seleccione una ruta y luego escanee QR / código de barras o digite DNI manualmente.</div>
        <div class="mv283-ok"><i class="bi bi-check-circle"></i> Lectura habilitada: QR, código de barras y DNI manual. La cámara se abre dentro de cada ruta.</div>
        <div class="mv283-section">Rutas disponibles</div>
        {% for r in rutas %}
          <a class="mv283-route" href="{{url_for('conductor_movil_ruta', ruta_id=r.id)}}">
            <i class="bi bi-bus-front bus"></i>
            <div><div class="name">{{r.nombre or 'RUTA'}}</div><div class="meta">{{r.placa or 'SIN BUS'}} · {{r.hora_salida or '-'}} · {{r.conductor or session.get('conductor_nombre') or '-'}}</div></div>
            <i class="bi bi-chevron-right chev"></i>
          </a>
        {% else %}
          <div class="mv283-quick">
            <b>No hay ruta asignada a este conductor.</b><br>
            <small>El login está correcto. El abordaje aparece cuando ADMIN crea una ruta y asigna este conductor.</small>
            <form method="post" action="{{url_for('conductor_movil_ruta_rapida')}}" class="mt-2"><button class="mv283-btn" type="submit"><i class="bi bi-plus-circle"></i> Crear ruta rápida y abrir abordaje</button></form>
            <div class="field-help mt-2">Ruta rápida no asigna bus. Para control completo: ADMIN &gt; Movilidad &gt; Rutas.</div>
          </div>
        {% endfor %}
        <div class="mv283-kpis"><div class="mv283-kpi"><small>Rutas visibles</small><b>{{rutas|length}}</b></div><div class="mv283-kpi"><small>Abordados hoy</small><b>{{abordados_hoy}}</b></div><div class="mv283-kpi"><small>Pendientes</small><b>{{pendientes}}</b></div></div>
      </div>
    </div></div>
    '''
    return render_page(body, rutas=rutas, abordados_hoy=abordados_hoy, pendientes=pendientes, title='Abordaje trabajadores')

def conductor_movil_ruta_283(ruta_id):
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    if not _trans_is_allowed_281(ruta_id):
        flash('Esta ruta no está asignada a su usuario conductor.', 'danger')
        return redirect(url_for('conductor_movil_panel'))
    ruta = row_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor
                                FROM transporte_rutas r
                                LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                                LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                                WHERE r.id=?""", (ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.', 'danger')
        return redirect(url_for('conductor_movil_panel'))
    pasajeros = rows_to_dict(execute('SELECT * FROM transporte_pasajeros WHERE ruta_id=? ORDER BY fecha_hora DESC', (ruta_id,), fetchall=True))
    ocupados = len(pasajeros)
    capacidad = int((ruta or {}).get('capacidad') or 0)
    libres = max(0, capacidad - ocupados) if capacidad else 0
    body = _movil_cond_283_css() + r'''
    <div class="phone-wrap"><div class="mv283-card">
      <div class="mv283-head"><a href="{{url_for('conductor_movil_panel')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Abordaje ruta</div></div>
      <div class="mv283-body">
        <div class="mv283-routebox">
          <div class="label">Ruta</div>
          <div class="route-title">{{ruta.origen or '-'}} → {{ruta.destino or '-'}}</div>
          <div>Bus: <b>{{ruta.placa or 'SIN BUS'}}</b> | Conductor: <b>{{ruta.conductor or session.get('conductor_nombre') or '-'}}</b> | Capacidad: <b>{{capacidad or 'SIN DEFINIR'}}</b> | Ocupados: <b>{{ocupados}}</b></div>
        </div>
        <form method="post" action="{{url_for('transporte_abordar', ruta_id=ruta.id)}}" id="frmAbordar283" class="mv283-form">
          <label>DNI / QR / Código de barras</label>
          <div class="input-group"><input name="dni" id="dniTransporte283" class="form-control" inputmode="numeric" maxlength="20" required placeholder="ESCANEAR O DIGITAR DNI" autofocus><button type="button" class="btn mv283-camera" onclick="abrirScanner&&abrirScanner('readerTrans283','dniTransporte283')"><i class="bi bi-camera"></i></button></div>
          <div id="readerTrans283" class="scan-box mt-2" style="display:none"></div>
          <div id="transpStatus283" class="field-help">Al completar 8 dígitos se validará en Trabajadores.</div>
          <select name="metodo" id="metodoTransporte283" class="form-select mt-2"><option>DNI DIGITADO</option><option>QR</option><option>CODIGO DE BARRAS</option></select>
          <input type="hidden" name="latitud" id="latitudTrans283"><input type="hidden" name="longitud" id="longitudTrans283">
          <button class="mv283-btn mt-2"><i class="bi bi-person-check"></i> Registrar subida</button>
        </form>
        <div class="d-grid gap-2 mb-2"><button class="mv283-btn" type="button" onclick="enviarGpsTransporte283({{ruta.id}}, false)"><i class="bi bi-geo-alt"></i> Enviar GPS de esta ruta</button><button class="mv283-outline" type="button" onclick="enviarGpsTransporte283({{ruta.id}}, true)"><i class="bi bi-broadcast-pin"></i> Iniciar / detener GPS en vivo</button></div>
        <div class="mv283-kpis"><div class="mv283-kpi"><small>Abordados</small><b>{{ocupados}}</b></div><div class="mv283-kpi"><small>Libres</small><b>{{libres}}</b></div><div class="mv283-kpi"><small>Capacidad</small><b>{{capacidad or 0}}</b></div></div>
        <div class="mv283-tablewrap"><table class="mv283-table"><thead><tr><th>Hora</th><th>DNI</th><th>Trabajador</th><th>Método</th></tr></thead><tbody>{% for p in pasajeros %}<tr><td>{{p.hora}}</td><td>{{p.dni}}</td><td>{{p.trabajador}}</td><td>{{p.metodo}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin abordajes.</td></tr>{% endfor %}</tbody></table></div>
      </div>
    </div></div>
    <script>
    (function(){
      const input=document.getElementById('dniTransporte283'), st=document.getElementById('transpStatus283'), lat=document.getElementById('latitudTrans283'), lon=document.getElementById('longitudTrans283'), frm=document.getElementById('frmAbordar283');
      const dni=v=>{const m=String(v||'').match(/(?:^|\D)(\d{8})(?:\D|$)/); return m?m[1]:String(v||'').replace(/\D/g,'').slice(-8)};
      let last='';
      async function validar(){const d=dni(input.value); if(d.length<8){st.className='field-help';st.innerHTML='Esperando 8 dígitos...';return;} input.value=d; if(d===last)return; last=d; st.className='scan-ok flash'; st.innerHTML='Validando DNI '+d+'...'; try{let r=await fetch('/api/trabajador/'+d,{cache:'no-store',credentials:'same-origin'});let j=await r.json(); if(j.ok){st.className='scan-ok flash';st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR')+' encontrado'; if(typeof beep==='function')beep();}else{st.className='scan-bad flash';st.innerHTML='✕ '+(j.msg||'DNI no encontrado en base trabajadores');}}catch(e){st.className='scan-bad flash';st.innerHTML='Error validando DNI';}}
      if(input){input.addEventListener('input',validar);input.addEventListener('paste',()=>setTimeout(validar,80));input.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();frm.requestSubmit();}});}
      if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{lat.value=p.coords.latitude;lon.value=p.coords.longitude;},()=>{}, {enableHighAccuracy:true,maximumAge:10000,timeout:15000});}
    })();
    let gpsWatch283=null; async function enviarGpsTransporte283(rid, live){function post(p){let fd=new FormData();fd.append('latitud',p.coords.latitude);fd.append('longitud',p.coords.longitude);document.getElementById('latitudTrans283').value=p.coords.latitude;document.getElementById('longitudTrans283').value=p.coords.longitude;return fetch('/transporte/ruta/'+rid+'/gps',{method:'POST',body:fd,credentials:'same-origin'}).then(r=>r.json());} if(!navigator.geolocation){alert('GPS no disponible');return;} if(live){if(gpsWatch283){navigator.geolocation.clearWatch(gpsWatch283);gpsWatch283=null;alert('GPS en vivo detenido');return;} gpsWatch283=navigator.geolocation.watchPosition(p=>post(p).catch(()=>{}),()=>alert('Permite ubicación/GPS'),{enableHighAccuracy:true,maximumAge:10000,timeout:15000}); alert('GPS en vivo iniciado. Mantenga esta pantalla abierta.'); return;} navigator.geolocation.getCurrentPosition(async p=>{let j=await post(p);alert(j.msg||'GPS actualizado');},()=>alert('Permite ubicación/GPS en el navegador'),{enableHighAccuracy:true,maximumAge:10000,timeout:15000});}
    </script>
    '''
    return render_page(body, ruta=ruta, pasajeros=pasajeros, ocupados=ocupados, capacidad=capacidad, libres=libres, title='Abordaje ruta')

# Overrides finales: conductor móvil queda con UI única y profesional.
app.view_functions['conductor_movil_panel'] = conductor_movil_panel_283
app.view_functions['conductor_movil_ruta'] = conductor_movil_ruta_283
# ======================= FIN PATCH TRANSPORTE OMAR 283 =======================


# ========================= PATCH TRANSPORTE OMAR 284 =========================
# 1) Vista móvil conductor más compacta.
# 2) Al salir de abordaje/submódulos vuelve a Módulo Transporte, no al login genérico.
# 3) Se preserva la sesión ADMIN al probar ingreso como conductor y se refuerza la relación Conductor ↔ Bus ↔ Ruta.

def _trans_can_open_284():
    return bool(session.get('usuario') or session.get('conductor_id'))


def _movil_cond_284_css():
    return r"""
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}
      .shell{max-width:430px!important;width:100%!important;margin:0 auto!important;padding:6px 8px 24px!important;background:#fff!important}
      .phone-wrap{max-width:390px!important;width:100%!important;margin:0 auto!important}
      .mv284-card{background:#fff;border:1px solid #e4e8e4;border-radius:14px;overflow:hidden;box-shadow:0 10px 24px rgba(0,0,0,.07);margin:6px auto 12px}
      .mv284-head{height:62px;background:#25773a;color:#fff;display:flex;align-items:center;justify-content:center;position:relative}
      .mv284-head a{position:absolute;left:12px;top:50%;transform:translateY(-50%);color:#fff!important;text-decoration:none;font-size:31px;line-height:1}
      .mv284-head .ttl{font-size:16px;font-weight:950;letter-spacing:.15px;color:#fff;text-align:center}
      .mv284-body{padding:12px 12px 14px;background:#fff}
      .mv284-info{border:1px solid #b8d7ff;background:#eef6ff;color:#0b2e83;border-radius:10px;padding:10px 11px;font-size:12px;font-weight:900;line-height:1.42;margin-bottom:10px}
      .mv284-ok{border:1px solid #bbf7d0;background:#ecfdf5;color:#065f2a;border-radius:10px;padding:9px 10px;font-size:12px;font-weight:900;line-height:1.4;margin-bottom:11px}
      .mv284-section{font-size:13px;font-weight:950;color:#08713b;text-transform:uppercase;margin:6px 1px 8px}
      .mv284-route{display:grid;grid-template-columns:34px 1fr 17px;gap:9px;align-items:center;text-decoration:none;color:#102a43!important;border:1px solid #dfe7df;background:#fff;border-radius:11px;padding:12px 10px;margin:7px 0;box-shadow:0 4px 10px rgba(0,0,0,.04)}
      .mv284-route i.bus{font-size:23px;color:#08713b}
      .mv284-route .name{font-size:14px;font-weight:950;color:#0a1f44;line-height:1.1}
      .mv284-route .meta{font-size:11px;font-weight:900;color:#0a1f44;margin-top:3px}
      .mv284-route .chev{font-size:21px;color:#111}
      .mv284-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:10px 0 12px}
      .mv284-kpi{background:#10964e;color:#fff;border-radius:8px;text-align:center;padding:8px 4px;box-shadow:0 6px 12px rgba(16,150,78,.16)}
      .mv284-kpi small{display:block;font-size:10px;font-weight:950;line-height:1.08;color:#effff3}
      .mv284-kpi b{display:block;font-size:21px;line-height:1.05;font-weight:950;color:#fff;margin-top:4px}
      .mv284-routebox{border:1px solid #b8d7ff;background:#eaf3ff;color:#0b2e83;border-radius:10px;padding:10px 11px;font-size:12px;font-weight:900;line-height:1.42;margin-bottom:10px}
      .mv284-routebox .label{font-size:12px;color:#0b2e83;text-transform:uppercase}
      .mv284-routebox .route-title{font-size:14px;color:#0b2e83;font-weight:950}
      .mv284-form{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:11px;margin-bottom:10px}
      .mv284-form label{font-size:12px;font-weight:950;color:#176a35;margin-bottom:5px}
      .mv284-form .input-group .form-control{height:40px;border-radius:10px 0 0 10px!important;font-size:13px;font-weight:850}
      .mv284-camera{min-width:54px;background:#08713b!important;color:#fff!important;border-color:#08713b!important;border-radius:0 10px 10px 0!important}
      .mv284-btn{height:41px;border-radius:10px;background:#08713b;border:1px solid #08713b;color:#fff;font-weight:950;font-size:13px;width:100%}
      .mv284-btn:hover{background:#065f2a;color:#fff}
      .mv284-outline{height:41px;border-radius:10px;background:#fff;border:1px solid #08713b;color:#08713b;font-weight:900;font-size:13px;width:100%}
      .mv284-tablewrap{border:1px solid #e5e7eb;border-radius:10px;overflow-x:auto;background:#fff;scrollbar-color:#08713b #e5e7eb}
      .mv284-tablewrap::-webkit-scrollbar{height:8px}.mv284-tablewrap::-webkit-scrollbar-thumb{background:#08713b;border-radius:999px}.mv284-tablewrap::-webkit-scrollbar-track{background:#e5e7eb}
      .mv284-table{width:100%;min-width:460px;border-collapse:collapse}
      .mv284-table th{background:#f8fafc;color:#12223b;font-size:12px;font-weight:950;padding:8px;border-bottom:1px solid #e5e7eb}
      .mv284-table td{font-size:11.5px;color:#334155;padding:8px;border-bottom:1px solid #f1f5f9}
      .mv284-quick{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:12px;text-align:center;color:#1f3b2a}
      .scan-ok,.scan-bad,.field-help{font-size:12px!important;font-weight:900!important}
      .scan-ok{background:#ecfdf5!important;border:1px solid #bbf7d0!important;color:#065f2a!important;border-radius:9px!important;padding:8px 10px!important}
      .scan-bad{background:#fee2e2!important;border:1px solid #fecaca!important;color:#991b1b!important;border-radius:9px!important;padding:8px 10px!important}
      .field-help{color:#4a644f!important;margin-top:5px}.flash{animation:mv284flash .3s ease}@keyframes mv284flash{0%{transform:scale(.99)}60%{transform:scale(1.01)}100%{transform:scale(1)}}
      .tr284-rel{margin-top:16px;border:1px solid #cfe4d6;background:#f7fff8;border-radius:12px;padding:11px 12px;color:#1f3b2a;box-shadow:0 5px 12px rgba(0,0,0,.04)}
      .tr284-rel .ttl{font-size:13px;font-weight:950;color:#08713b;margin-bottom:6px;text-transform:uppercase}
      .tr284-rel .line{font-size:12px;font-weight:900;line-height:1.45}
      .tr284-rel .muted{font-size:11px;color:#486156;font-weight:800;line-height:1.35;margin-top:4px}
      .tr284-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:9px;margin-top:12px}
      .tr284-kpi{background:#0d8f47;color:#fff;border-radius:10px;text-align:center;padding:8px 5px;box-shadow:0 8px 14px rgba(13,143,71,.16)}
      .tr284-kpi small{display:block;font-size:10px;font-weight:900;line-height:1.1}.tr284-kpi b{display:block;font-size:20px;font-weight:950;line-height:1.05;margin-top:4px}
    </style>
    """


def transporte_home_284():
    if not _trans_can_open_284():
        flash('Inicie sesión para ingresar a Transporte.', 'danger')
        return redirect(url_for('login'))
    total_cond = int(scalar('SELECT COUNT(*) AS c FROM transporte_conductores') or 0)
    total_bus = int(scalar('SELECT COUNT(*) AS c FROM transporte_vehiculos') or 0)
    total_rutas = int(scalar('SELECT COUNT(*) AS c FROM transporte_rutas') or 0)
    rutas_rel = int(scalar('SELECT COUNT(*) AS c FROM transporte_rutas WHERE conductor_id IS NOT NULL AND vehiculo_id IS NOT NULL') or 0)
    sin_rel = max(0, total_rutas - rutas_rel)
    back_target = url_for('home') if session.get('usuario') else url_for('conductor_movil_panel')
    body = _transporte_ui_280_css() + _movil_cond_284_css() + r"""
    <div class="tr279-phone">
      <div class="tr279-app">
        <div class="tr279-hero">
          <a class="tr279-back" href="{{back_target}}"><i class="bi bi-chevron-left"></i></a>
          {% if session.get('usuario') %}<a class="tr279-config" href="{{url_for('transporte_config')}}"><i class="bi bi-gear"></i> Config.</a>{% endif %}
          <div class="tr279-bus"><i class="bi bi-bus-front-fill"></i></div>
          <div class="tr279-title">Módulo Transporte</div>
        </div>
        <div class="tr279-body">
          {% if session.get('usuario') %}
          <div class="tr279-section">Módulos</div>
          <div class="tr279-grid3">
            <a class="tr279-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span class="lbl">Conductores</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span class="lbl">Buses</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i><span class="lbl">Rutas</span></a>
          </div>
          {% endif %}
          <div class="tr279-section op">Operación</div>
          <div class="tr279-grid2">
            <a class="tr279-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span class="lbl">GPS / Seguimiento</span><span class="sub">Ver ubicación</span></a>
            <a class="tr279-tile" href="{{url_for('transporte_mobile_home')}}"><i class="bi bi-phone"></i><span class="lbl">Móvil conductor</span><span class="sub">Abordaje y GPS</span></a>
          </div>
          <div class="tr279-info"><i class="bi bi-info-circle-fill"></i><div>Flujo correcto: primero crea el usuario/PIN en <b>Conductores</b>, luego registra el <b>Bus</b>, después crea la <b>Ruta</b> asignando <b>conductor + bus</b>. Recién allí el conductor entra a <b>Móvil conductor</b> para abordar y enviar GPS.</div></div>
          <div class="tr284-kpis">
            <div class="tr284-kpi"><small>Conductores</small><b>{{total_cond}}</b></div>
            <div class="tr284-kpi"><small>Buses</small><b>{{total_bus}}</b></div>
            <div class="tr284-kpi"><small>Rutas</small><b>{{total_rutas}}</b></div>
          </div>
          <div class="tr284-rel">
            <div class="ttl">Relación conductor ↔ bus ↔ ruta</div>
            <div class="line"><b>{{rutas_rel}}</b> ruta(s) ya están completas con <b>conductor + bus</b>.</div>
            <div class="line"><b>{{sin_rel}}</b> ruta(s) aún requieren completar la relación.</div>
            <div class="muted">Si desea que el abordaje funcione perfecto, cada ruta debe tener conductor asignado y, de preferencia, bus/capacidad definida.</div>
          </div>
        </div>
      </div>
    </div>
    """
    return render_page(body, back_target=back_target, total_cond=total_cond, total_bus=total_bus, total_rutas=total_rutas, rutas_rel=rutas_rel, sin_rel=sin_rel, title='Módulo Transporte')


def conductor_movil_login_284():
    if request.method == 'POST':
        dni = limpiar_dni(request.form.get('dni'))
        pin = (request.form.get('pin') or '').strip()
        if len(dni) != 8 or str(dni).upper() == 'ADMIN':
            flash('Solo conductores: ingrese DNI de 8 dígitos. No usar ADMIN.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        if not pin:
            flash('Ingrese su PIN móvil.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        c = row_to_dict(execute("""SELECT * FROM transporte_conductores
                                 WHERE dni=? AND COALESCE(movil_pin,'')=?
                                 AND COALESCE(movil_estado,'ACTIVO')='ACTIVO'
                                 AND COALESCE(estado,'ACTIVO') NOT IN ('INACTIVO','BLOQUEADO')""", (dni, pin), fetchone=True))
        if not c:
            flash('DNI o PIN incorrecto, bloqueado o no creado en Transporte > Conductores.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        session['conductor_id'] = c.get('id')
        session['conductor_dni'] = c.get('dni')
        session['conductor_nombre'] = c.get('nombres') or c.get('dni')
        flash('Bienvenido conductor. Ya puede registrar abordaje y GPS.', 'success')
        return redirect(url_for('conductor_movil_panel'))
    body = r"""
    <div class="phone-wrap">
      <div class="page-card" style="border-radius:14px;overflow:hidden;margin-top:16px">
        <div class="green-hero" style="min-height:128px;border-radius:0;padding:18px 16px 38px">
          <a href="{{url_for('transporte')}}" style="position:absolute;left:18px;top:23px;color:white;font-size:38px;text-decoration:none"><i class="bi bi-chevron-left"></i></a>
          <div style="font-size:34px"><i class="bi bi-phone"></i></div>
          <div style="font-family:Georgia,serif;font-weight:900;font-size:13px;margin-top:4px">ACCESO MÓVIL CONDUCTOR</div>
        </div>
        <div class="floating-card" style="margin:-30px 12px 14px;border-radius:12px;padding:16px">
          <div class="alert alert-light" style="border:1px solid #dbe3db;color:#24405f;line-height:1.45"><b>Solo conductores:</b> ingrese DNI del conductor + PIN móvil creado en Transporte &gt; Conductores. No usar ADMIN.</div>
          <form method="post" id="frmMovilCond284">
            <label class="form-label">DNI conductor</label>
            <input name="dni" id="dniMovCond284" class="form-control mb-3" inputmode="numeric" maxlength="8" pattern="\d{8}" placeholder="Ingrese DNI" required autocomplete="username" autofocus>
            <label class="form-label">PIN móvil</label>
            <input name="pin" class="form-control mb-3" type="password" placeholder="PIN" required autocomplete="current-password">
            <button class="btn btn-green w-100" style="font-size:19px;height:47px">INGRESAR</button>
          </form>
          {% if session.get('usuario') %}
          <a class="btn btn-outline-success w-100 mt-2" href="{{url_for('transporte_conductores')}}">Crear / resetear PIN</a>
          {% endif %}
          <a class="btn btn-outline-secondary w-100 mt-2" href="{{url_for('transporte')}}">Volver</a>
        </div>
      </div>
    </div>
    <script>(function(){const i=document.getElementById('dniMovCond284'); if(i){i.addEventListener('input',()=>{i.value=String(i.value||'').replace(/\D/g,'').slice(-8);});}})();</script>
    """
    return render_page(body, title='Acceso móvil conductor')


def conductor_movil_panel_284():
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    cid = session.get('conductor_id')
    hoy = today_str()
    rutas = rows_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor
                                  FROM transporte_rutas r
                                  LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                                  LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                                  WHERE r.conductor_id=?
                                    AND (r.fecha>=? OR UPPER(COALESCE(r.estado,'')) IN ('PROGRAMADA','EN RUTA'))
                                  ORDER BY CASE WHEN r.fecha=? THEN 0 ELSE 1 END,
                                           r.fecha DESC, r.hora_salida ASC, r.id DESC LIMIT 50""", (cid, hoy, hoy), fetchall=True))
    ruta_ids = [str(r.get('id')) for r in rutas if r.get('id')]
    abordados_hoy = 0
    rutas_con_abordaje = 0
    if ruta_ids:
        marks = ','.join(['?'] * len(ruta_ids))
        abordados_hoy = scalar(f"SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=? AND ruta_id IN ({marks})", tuple([hoy] + ruta_ids))
        rutas_con_abordaje = scalar(f"SELECT COUNT(DISTINCT ruta_id) AS c FROM transporte_pasajeros WHERE fecha=? AND ruta_id IN ({marks})", tuple([hoy] + ruta_ids))
    pendientes = max(0, len(rutas) - int(rutas_con_abordaje or 0))
    body = _movil_cond_284_css() + r"""
    <div class="phone-wrap"><div class="mv284-card">
      <div class="mv284-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Abordaje trabajadores</div></div>
      <div class="mv284-body">
        <div class="mv284-info">Seleccione una ruta y luego escanee QR / código de barras o digite DNI manualmente.</div>
        <div class="mv284-ok"><i class="bi bi-check-circle"></i> Lectura habilitada: QR, código de barras y DNI manual. La cámara se abre dentro de cada ruta.</div>
        <div class="mv284-section">Rutas disponibles</div>
        {% for r in rutas %}
          <a class="mv284-route" href="{{url_for('conductor_movil_ruta', ruta_id=r.id)}}">
            <i class="bi bi-bus-front bus"></i>
            <div><div class="name">{{r.nombre or 'RUTA'}}</div><div class="meta">{{r.placa or 'SIN BUS'}} · {{r.hora_salida or '-'}} · {{r.conductor or session.get('conductor_nombre') or '-'}}</div></div>
            <i class="bi bi-chevron-right chev"></i>
          </a>
        {% else %}
          <div class="mv284-quick">
            <b>No hay ruta asignada a este conductor.</b><br>
            <small>El login está correcto. El abordaje aparece cuando ADMIN crea una ruta y asigna este conductor.</small>
            <form method="post" action="{{url_for('conductor_movil_ruta_rapida')}}" class="mt-2"><button class="mv284-btn" type="submit"><i class="bi bi-plus-circle"></i> Crear ruta rápida y abrir abordaje</button></form>
            <div class="field-help mt-2">Ruta rápida no asigna bus. Para control completo: ADMIN &gt; Movilidad &gt; Rutas.</div>
          </div>
        {% endfor %}
        <div class="mv284-kpis"><div class="mv284-kpi"><small>Rutas visibles</small><b>{{rutas|length}}</b></div><div class="mv284-kpi"><small>Abordados hoy</small><b>{{abordados_hoy}}</b></div><div class="mv284-kpi"><small>Pendientes</small><b>{{pendientes}}</b></div></div>
      </div>
    </div></div>
    """
    return render_page(body, rutas=rutas, abordados_hoy=abordados_hoy, pendientes=pendientes, title='Abordaje trabajadores')


def conductor_movil_ruta_284(ruta_id):
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    if not _trans_is_allowed_281(ruta_id):
        flash('Esta ruta no está asignada a su usuario conductor.', 'danger')
        return redirect(url_for('conductor_movil_panel'))
    ruta = row_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor
                                FROM transporte_rutas r
                                LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                                LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                                WHERE r.id=?""", (ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.', 'danger')
        return redirect(url_for('conductor_movil_panel'))
    pasajeros = rows_to_dict(execute('SELECT * FROM transporte_pasajeros WHERE ruta_id=? ORDER BY fecha_hora DESC', (ruta_id,), fetchall=True))
    ocupados = len(pasajeros)
    capacidad = int((ruta or {}).get('capacidad') or 0)
    libres = max(0, capacidad - ocupados) if capacidad else 0
    body = _movil_cond_284_css() + r"""
    <div class="phone-wrap"><div class="mv284-card">
      <div class="mv284-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Abordaje ruta</div></div>
      <div class="mv284-body">
        <div class="mv284-routebox">
          <div class="label">Ruta</div>
          <div class="route-title">{{ruta.origen or '-'}} → {{ruta.destino or '-'}}</div>
          <div>Bus: <b>{{ruta.placa or 'SIN BUS'}}</b> | Conductor: <b>{{ruta.conductor or session.get('conductor_nombre') or '-'}}</b> | Capacidad: <b>{{capacidad or 'SIN DEFINIR'}}</b> | Ocupados: <b>{{ocupados}}</b></div>
        </div>
        <form method="post" action="{{url_for('transporte_abordar', ruta_id=ruta.id)}}" id="frmAbordar284" class="mv284-form">
          <label>DNI / QR / Código de barras</label>
          <div class="input-group"><input name="dni" id="dniTransporte284" class="form-control" inputmode="numeric" maxlength="20" required placeholder="ESCANEAR O DIGITAR DNI" autofocus><button type="button" class="btn mv284-camera" onclick="abrirScanner&&abrirScanner('readerTrans284','dniTransporte284')"><i class="bi bi-camera"></i></button></div>
          <div id="readerTrans284" class="scan-box mt-2" style="display:none"></div>
          <div id="transpStatus284" class="field-help">Al completar 8 dígitos se validará en Trabajadores.</div>
          <select name="metodo" id="metodoTransporte284" class="form-select mt-2"><option>DNI DIGITADO</option><option>QR</option><option>CODIGO DE BARRAS</option></select>
          <input type="hidden" name="latitud" id="latitudTrans284"><input type="hidden" name="longitud" id="longitudTrans284">
          <button class="mv284-btn mt-2"><i class="bi bi-person-check"></i> Registrar subida</button>
        </form>
        <div class="d-grid gap-2 mb-2"><button class="mv284-btn" type="button" onclick="enviarGpsTransporte284({{ruta.id}}, false)"><i class="bi bi-geo-alt"></i> Enviar GPS de esta ruta</button><button class="mv284-outline" type="button" onclick="enviarGpsTransporte284({{ruta.id}}, true)"><i class="bi bi-broadcast-pin"></i> Iniciar / detener GPS en vivo</button></div>
        <div class="mv284-kpis"><div class="mv284-kpi"><small>Abordados</small><b>{{ocupados}}</b></div><div class="mv284-kpi"><small>Libres</small><b>{{libres}}</b></div><div class="mv284-kpi"><small>Capacidad</small><b>{{capacidad or 0}}</b></div></div>
        <div class="mv284-tablewrap"><table class="mv284-table"><thead><tr><th>Hora</th><th>DNI</th><th>Trabajador</th><th>Método</th></tr></thead><tbody>{% for p in pasajeros %}<tr><td>{{p.hora}}</td><td>{{p.dni}}</td><td>{{p.trabajador}}</td><td>{{p.metodo}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin abordajes.</td></tr>{% endfor %}</tbody></table></div>
      </div>
    </div></div>
    <script>
    (function(){
      const input=document.getElementById('dniTransporte284'), st=document.getElementById('transpStatus284'), lat=document.getElementById('latitudTrans284'), lon=document.getElementById('longitudTrans284'), frm=document.getElementById('frmAbordar284');
      const dni=v=>{const m=String(v||'').match(/(?:^|\D)(\d{8})(?:\D|$)/); return m?m[1]:String(v||'').replace(/\D/g,'').slice(-8)};
      let last='';
      async function validar(){const d=dni(input.value); if(d.length<8){st.className='field-help';st.innerHTML='Esperando 8 dígitos...';return;} input.value=d; if(d===last)return; last=d; st.className='scan-ok flash'; st.innerHTML='Validando DNI '+d+'...'; try{let r=await fetch('/api/trabajador/'+d,{cache:'no-store',credentials:'same-origin'});let j=await r.json(); if(j.ok){st.className='scan-ok flash';st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR')+' encontrado'; if(typeof beep==='function')beep();}else{st.className='scan-bad flash';st.innerHTML='✕ '+(j.msg||'DNI no encontrado en base trabajadores');}}catch(e){st.className='scan-bad flash';st.innerHTML='Error validando DNI';}}
      if(input){input.addEventListener('input',validar);input.addEventListener('paste',()=>setTimeout(validar,80));input.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();frm.requestSubmit();}});} if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{lat.value=p.coords.latitude;lon.value=p.coords.longitude;},()=>{}, {enableHighAccuracy:true,maximumAge:10000,timeout:15000});}
    })();
    let gpsWatch284=null; async function enviarGpsTransporte284(rid, live){function post(p){let fd=new FormData();fd.append('latitud',p.coords.latitude);fd.append('longitud',p.coords.longitude);document.getElementById('latitudTrans284').value=p.coords.latitude;document.getElementById('longitudTrans284').value=p.coords.longitude;return fetch('/transporte/ruta/'+rid+'/gps',{method:'POST',body:fd,credentials:'same-origin'}).then(r=>r.json());} if(!navigator.geolocation){alert('GPS no disponible');return;} if(live){if(gpsWatch284){navigator.geolocation.clearWatch(gpsWatch284);gpsWatch284=null;alert('GPS en vivo detenido');return;} gpsWatch284=navigator.geolocation.watchPosition(p=>post(p).catch(()=>{}),()=>alert('Permite ubicación/GPS'),{enableHighAccuracy:true,maximumAge:10000,timeout:15000}); alert('GPS en vivo iniciado. Mantenga esta pantalla abierta.'); return;} navigator.geolocation.getCurrentPosition(async p=>{let j=await post(p);alert(j.msg||'GPS actualizado');},()=>alert('Permite ubicación/GPS en el navegador'),{enableHighAccuracy:true,maximumAge:10000,timeout:15000});}
    </script>
    """
    return render_page(body, ruta=ruta, pasajeros=pasajeros, ocupados=ocupados, capacidad=capacidad, libres=libres, title='Abordaje ruta')


def conductor_movil_logout_284():
    for k in ('conductor_id','conductor_dni','conductor_nombre'):
        session.pop(k, None)
    flash('Sesión móvil cerrada.', 'success')
    return redirect(url_for('transporte') if session.get('usuario') else url_for('conductor_movil_login'))

app.view_functions['transporte'] = transporte_home_284
app.view_functions['conductor_movil_login'] = conductor_movil_login_284
app.view_functions['transporte_mobile_home'] = conductor_movil_login_284
app.view_functions['conductor_movil_panel'] = conductor_movil_panel_284
app.view_functions['conductor_movil_ruta'] = conductor_movil_ruta_284
app.view_functions['conductor_movil_logout'] = conductor_movil_logout_284
# ======================= FIN PATCH TRANSPORTE OMAR 284 =======================


# ========================= PATCH TRANSPORTE OMAR 285 =========================
# Flujo conductor profesional:
# El conductor entra con DNI/PIN, selecciona ruta base + bus, define hora de inicio,
# toma GPS inicial y recién comienza el registro de abordaje.

def _movil_cond_285_css():
    return r"""
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}
      .shell{max-width:430px!important;width:100%!important;margin:0 auto!important;padding:6px 8px 24px!important;background:#fff!important}
      .phone-wrap{max-width:390px!important;width:100%!important;margin:0 auto!important}
      .mv285-card{background:#fff;border:1px solid #e4e8e4;border-radius:14px;overflow:hidden;box-shadow:0 10px 24px rgba(0,0,0,.07);margin:6px auto 12px}
      .mv285-head{height:62px;background:#25773a;color:#fff;display:flex;align-items:center;justify-content:center;position:relative}
      .mv285-head a{position:absolute;left:12px;top:50%;transform:translateY(-50%);color:#fff!important;text-decoration:none;font-size:30px;line-height:1}
      .mv285-head .ttl{font-size:16px;font-weight:950;letter-spacing:.15px;color:#fff;text-align:center}
      .mv285-body{padding:12px 12px 14px;background:#fff}
      .mv285-info{border:1px solid #b8d7ff;background:#eef6ff;color:#0b2e83;border-radius:10px;padding:10px 11px;font-size:12px;font-weight:900;line-height:1.42;margin-bottom:10px}
      .mv285-ok{border:1px solid #bbf7d0;background:#ecfdf5;color:#065f2a;border-radius:10px;padding:9px 10px;font-size:12px;font-weight:900;line-height:1.4;margin-bottom:10px}
      .mv285-section{font-size:13px;font-weight:950;color:#08713b;text-transform:uppercase;margin:8px 1px 8px}
      .mv285-form{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:11px;margin-bottom:10px}
      .mv285-form label{font-size:11px;font-weight:950;color:#176a35;margin-bottom:4px}
      .mv285-form .form-control,.mv285-form .form-select{height:38px!important;border-radius:9px!important;font-size:12px!important;font-weight:850}
      .mv285-btn{height:41px;border-radius:10px;background:#08713b;border:1px solid #08713b;color:#fff;font-weight:950;font-size:13px;width:100%}
      .mv285-btn:hover{background:#065f2a;color:#fff}
      .mv285-outline{height:39px;border-radius:10px;background:#fff;border:1px solid #08713b;color:#08713b;font-weight:900;font-size:12px;width:100%}
      .mv285-grid2{display:grid;grid-template-columns:1fr 1fr;gap:8px}
      .mv285-route{display:grid;grid-template-columns:35px 1fr 18px;gap:9px;align-items:center;text-decoration:none;color:#102a43!important;border:1px solid #dfe7df;background:#fff;border-radius:11px;padding:11px 10px;margin:7px 0;box-shadow:0 4px 10px rgba(0,0,0,.04)}
      .mv285-route i.bus{font-size:22px;color:#08713b}
      .mv285-route .name{font-size:13px;font-weight:950;color:#0a1f44;line-height:1.1}
      .mv285-route .meta{font-size:10.5px;font-weight:900;color:#0a1f44;margin-top:3px;line-height:1.25}
      .mv285-route .chev{font-size:21px;color:#111}
      .mv285-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:10px 0 12px}
      .mv285-kpi{background:#10964e;color:#fff;border-radius:8px;text-align:center;padding:8px 4px;box-shadow:0 6px 12px rgba(16,150,78,.16)}
      .mv285-kpi small{display:block;font-size:10px;font-weight:950;line-height:1.08;color:#effff3}
      .mv285-kpi b{display:block;font-size:21px;line-height:1.05;font-weight:950;color:#fff;margin-top:4px}
      .mv285-pill{display:inline-flex;align-items:center;gap:5px;background:#ecfdf5;border:1px solid #bbf7d0;color:#065f2a;border-radius:999px;padding:5px 9px;font-size:10.5px;font-weight:950;margin-bottom:8px}
      .mv285-status{font-size:11px;font-weight:900;border-radius:9px;padding:8px 9px;margin-top:7px}
      .mv285-status.ok{background:#ecfdf5;border:1px solid #bbf7d0;color:#065f2a}
      .mv285-status.bad{background:#fee2e2;border:1px solid #fecaca;color:#991b1b}
      .mv285-help{font-size:10.5px;color:#4a644f;font-weight:800;line-height:1.35;margin-top:5px}
      .mv285-empty{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:12px;text-align:center;color:#1f3b2a;font-size:12px;font-weight:900;line-height:1.4}
    </style>
    """

def _bases_rutas_285():
    rows = rows_to_dict(execute("""SELECT id,nombre,origen,destino,sede,hora_salida,hora_retorno,estado,fecha
                                  FROM transporte_rutas
                                  WHERE COALESCE(nombre,'')<>'' OR COALESCE(origen,'')<>'' OR COALESCE(destino,'')<>''
                                  ORDER BY fecha DESC, id DESC LIMIT 300""", fetchall=True))
    seen, out = set(), []
    for r in rows:
        key = f"{r.get('nombre') or ''}|{r.get('origen') or ''}|{r.get('destino') or ''}|{r.get('sede') or ''}".upper()
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
        if len(out) >= 80:
            break
    return out

def _buses_activos_285():
    return rows_to_dict(execute("""SELECT id,placa,tipo,capacidad,estado
                                  FROM transporte_vehiculos
                                  WHERE COALESCE(estado,'ACTIVO') NOT IN ('INACTIVO','BLOQUEADO')
                                  ORDER BY placa LIMIT 200""", fetchall=True))

def conductor_movil_login_285():
    if request.method == 'POST':
        dni = limpiar_dni(request.form.get('dni'))
        pin = (request.form.get('pin') or '').strip()
        if len(dni) != 8 or str(dni).upper() == 'ADMIN':
            flash('Solo conductores: ingrese DNI de 8 dígitos. No usar ADMIN.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        if not pin:
            flash('Ingrese su PIN móvil.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        c = row_to_dict(execute("""SELECT * FROM transporte_conductores
                                 WHERE dni=? AND COALESCE(movil_pin,'')=?
                                 AND COALESCE(movil_estado,'ACTIVO')='ACTIVO'
                                 AND COALESCE(estado,'ACTIVO') NOT IN ('INACTIVO','BLOQUEADO')""", (dni, pin), fetchone=True))
        if not c:
            flash('DNI o PIN incorrecto, bloqueado o no creado en Transporte > Conductores.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        session['conductor_id'] = c.get('id')
        session['conductor_dni'] = c.get('dni')
        session['conductor_nombre'] = c.get('nombres') or c.get('dni')
        flash('Bienvenido conductor. Seleccione ruta, bus y hora de inicio.', 'success')
        return redirect(url_for('conductor_movil_inicio_ruta'))
    if session.get('conductor_id'):
        return redirect(url_for('conductor_movil_inicio_ruta'))
    body = r"""
    <div class="phone-wrap">
      <div class="page-card" style="border-radius:14px;overflow:hidden;margin-top:16px">
        <div class="green-hero" style="min-height:128px;border-radius:0;padding:18px 16px 38px">
          <a href="{{url_for('transporte')}}" style="position:absolute;left:18px;top:23px;color:white;font-size:38px;text-decoration:none"><i class="bi bi-chevron-left"></i></a>
          <div style="font-size:34px"><i class="bi bi-phone"></i></div>
          <div style="font-family:Georgia,serif;font-weight:900;font-size:13px;margin-top:4px">ACCESO MÓVIL CONDUCTOR</div>
        </div>
        <div class="floating-card" style="margin:-30px 12px 14px;border-radius:12px;padding:16px">
          <div class="alert alert-light" style="border:1px solid #dbe3db;color:#24405f;line-height:1.45"><b>Solo conductores:</b> ingrese DNI del conductor + PIN móvil creado en Transporte &gt; Conductores. No usar ADMIN.</div>
          <form method="post" id="frmMovilCond285">
            <label class="form-label">DNI conductor</label>
            <input name="dni" id="dniMovCond285" class="form-control mb-3" inputmode="numeric" maxlength="8" pattern="\d{8}" placeholder="Ingrese DNI" required autocomplete="username" autofocus>
            <label class="form-label">PIN móvil</label>
            <input name="pin" class="form-control mb-3" type="password" placeholder="PIN" required autocomplete="current-password">
            <button class="btn btn-green w-100" style="font-size:19px;height:47px">INGRESAR</button>
          </form>
          {% if session.get('usuario') %}
          <a class="btn btn-outline-success w-100 mt-2" href="{{url_for('transporte_conductores')}}">Crear / resetear PIN</a>
          {% endif %}
          <a class="btn btn-outline-secondary w-100 mt-2" href="{{url_for('transporte')}}">Volver</a>
        </div>
      </div>
    </div>
    <script>(function(){const i=document.getElementById('dniMovCond285'); if(i){i.addEventListener('input',()=>{i.value=String(i.value||'').replace(/\D/g,'').slice(-8);});}})();</script>
    """
    return render_page(body, title='Acceso móvil conductor')

def conductor_movil_inicio_ruta_285():
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))

    cid = session.get('conductor_id')
    if request.method == 'POST':
        base_id = request.form.get('base_ruta_id') or ''
        bus_id = request.form.get('vehiculo_id') or ''
        hora_inicio = request.form.get('hora_inicio') or datetime.now().strftime('%H:%M')
        km_inicio_raw = request.form.get('km_inicio') or 0
        lat = (request.form.get('latitud') or '').strip()
        lon = (request.form.get('longitud') or '').strip()

        if not base_id:
            flash('Seleccione una ruta base.', 'danger')
            return redirect(url_for('conductor_movil_inicio_ruta'))
        if not bus_id:
            flash('Seleccione el bus asignado.', 'danger')
            return redirect(url_for('conductor_movil_inicio_ruta'))

        base = row_to_dict(execute('SELECT * FROM transporte_rutas WHERE id=?', (base_id,), fetchone=True))
        bus = row_to_dict(execute('SELECT * FROM transporte_vehiculos WHERE id=?', (bus_id,), fetchone=True))
        if not base:
            flash('Ruta base no encontrada.', 'danger')
            return redirect(url_for('conductor_movil_inicio_ruta'))
        if not bus:
            flash('Bus no encontrado.', 'danger')
            return redirect(url_for('conductor_movil_inicio_ruta'))

        try:
            km_inicio = float(km_inicio_raw or 0)
        except Exception:
            km_inicio = 0

        creado = now_str()
        nombre = limpiar_texto(base.get('nombre') or 'RUTA')
        origen = limpiar_texto(base.get('origen'))
        destino = limpiar_texto(base.get('destino'))
        sede = limpiar_texto(base.get('sede'))
        execute("""INSERT INTO transporte_rutas(
                    fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,
                    vehiculo_id,conductor_id,estado,latitud,longitud,ultima_ubicacion,
                    km_inicio,creado_por,creado_en)
                  VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (today_str(), nombre, origen, destino, sede, hora_inicio, base.get('hora_retorno') or '',
                 bus_id, cid, 'EN RUTA', lat, lon, creado if lat and lon else '', km_inicio, _trans_user_281(), creado),
                commit=True)
        nueva = row_to_dict(execute('SELECT id FROM transporte_rutas WHERE conductor_id=? AND creado_en=? ORDER BY id DESC LIMIT 1', (cid, creado), fetchone=True))
        nueva_id = nueva.get('id') if nueva else None
        if nueva_id and lat and lon:
            execute("""INSERT INTO transporte_gps(ruta_id,latitud,longitud,fecha_hora,registrado_por,conductor_id,placa,ruta_nombre)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (nueva_id, lat, lon, creado, _trans_user_281(), cid, bus.get('placa'), nombre), commit=True)
            execute('UPDATE transporte_conductores SET ultima_latitud=?, ultima_longitud=?, ultimo_gps=? WHERE id=?',
                    (lat, lon, creado, cid), commit=True)
        flash('Ruta iniciada. Ya puede registrar abordaje de trabajadores.', 'success')
        return redirect(url_for('conductor_movil_ruta', ruta_id=nueva_id)) if nueva_id else redirect(url_for('conductor_movil_inicio_ruta'))

    bases = _bases_rutas_285()
    buses = _buses_activos_285()
    hoy = today_str()
    activas = rows_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor,
                                      (SELECT COUNT(*) FROM transporte_pasajeros p WHERE p.ruta_id=r.id) AS ocupados
                                      FROM transporte_rutas r
                                      LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                                      LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                                      WHERE r.conductor_id=?
                                        AND (r.fecha>=? OR UPPER(COALESCE(r.estado,'')) IN ('EN RUTA','PROGRAMADA'))
                                      ORDER BY CASE WHEN UPPER(COALESCE(r.estado,''))='EN RUTA' THEN 0 ELSE 1 END,
                                               r.fecha DESC, r.id DESC LIMIT 30""", (cid, hoy), fetchall=True))
    body = _movil_cond_285_css() + r"""
    <div class="phone-wrap"><div class="mv285-card">
      <div class="mv285-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Inicio ruta conductor</div></div>
      <div class="mv285-body">
        <div class="mv285-pill"><i class="bi bi-person-badge"></i> {{session.get('conductor_nombre')}} · {{session.get('conductor_dni')}}</div>
        <div class="mv285-info">Antes de abordar, seleccione <b>ruta base</b>, <b>bus</b>, hora de inicio y tome GPS inicial. Luego pulse <b>Comenzar registro</b>.</div>

        <div class="mv285-section">Comenzar nueva ruta</div>
        <form method="post" class="mv285-form" id="frmInicioRuta285">
          <label>Ruta base</label>
          <select name="base_ruta_id" id="baseRuta285" class="form-select" required>
            <option value="">Seleccionar ruta...</option>
            {% for r in bases %}
              <option value="{{r.id}}">{{r.nombre or 'RUTA'}} · {{r.origen or '-'}} → {{r.destino or '-'}}</option>
            {% endfor %}
          </select>
          <div class="mv285-help">La ruta base trae origen, destino y sede. Se creará una ruta operativa nueva para este viaje.</div>

          <div class="mv285-grid2 mt-2">
            <div>
              <label>Bus</label>
              <select name="vehiculo_id" class="form-select" required>
                <option value="">Bus...</option>
                {% for b in buses %}
                  <option value="{{b.id}}">{{b.placa}} · Cap. {{b.capacidad or 0}}</option>
                {% endfor %}
              </select>
            </div>
            <div>
              <label>Hora inicio</label>
              <input name="hora_inicio" type="time" class="form-control" value="{{hora_actual}}" required>
            </div>
          </div>

          <div class="mv285-grid2 mt-2">
            <div>
              <label>Km inicio</label>
              <input name="km_inicio" type="number" step="0.01" class="form-control" placeholder="Opcional">
            </div>
            <div>
              <label>GPS inicial</label>
              <button type="button" class="mv285-outline" onclick="tomarGpsInicio285()"><i class="bi bi-geo-alt"></i> Tomar GPS</button>
            </div>
          </div>

          <input type="hidden" name="latitud" id="latInicio285">
          <input type="hidden" name="longitud" id="lonInicio285">
          <div id="gpsStatus285" class="mv285-status bad">GPS pendiente. Pulse “Tomar GPS”.</div>

          <button class="mv285-btn mt-2" type="submit"><i class="bi bi-play-circle"></i> Comenzar registro</button>
        </form>

        <div class="mv285-section">Rutas en proceso</div>
        {% for r in activas %}
          <a class="mv285-route" href="{{url_for('conductor_movil_ruta', ruta_id=r.id)}}">
            <i class="bi bi-bus-front bus"></i>
            <div>
              <div class="name">{{r.nombre or 'RUTA'}} · {{r.placa or 'SIN BUS'}}</div>
              <div class="meta">{{r.origen or '-'}} → {{r.destino or '-'}} · Inicio {{r.hora_salida or '-'}} · Abordados {{r.ocupados or 0}}</div>
            </div>
            <i class="bi bi-chevron-right chev"></i>
          </a>
        {% else %}
          <div class="mv285-empty">Aún no tiene rutas iniciadas. Complete el formulario superior para comenzar.</div>
        {% endfor %}

        <div class="mv285-kpis">
          <div class="mv285-kpi"><small>Rutas</small><b>{{activas|length}}</b></div>
          <div class="mv285-kpi"><small>Buses</small><b>{{buses|length}}</b></div>
          <div class="mv285-kpi"><small>Bases</small><b>{{bases|length}}</b></div>
        </div>
      </div>
    </div></div>
    <script>
      function tomarGpsInicio285(){
        const st=document.getElementById('gpsStatus285'), lat=document.getElementById('latInicio285'), lon=document.getElementById('lonInicio285');
        if(!navigator.geolocation){st.className='mv285-status bad';st.innerHTML='GPS no disponible en este navegador.';return;}
        st.className='mv285-status ok'; st.innerHTML='Tomando GPS...';
        navigator.geolocation.getCurrentPosition(p=>{
          lat.value=p.coords.latitude; lon.value=p.coords.longitude;
          st.className='mv285-status ok'; st.innerHTML='✓ GPS capturado: '+Number(p.coords.latitude).toFixed(6)+', '+Number(p.coords.longitude).toFixed(6);
        },()=>{
          st.className='mv285-status bad'; st.innerHTML='No se pudo tomar GPS. Permita ubicación desde el candado del navegador.';
        },{enableHighAccuracy:true,maximumAge:10000,timeout:15000});
      }
    </script>
    """
    return render_page(body, bases=bases, buses=buses, activas=activas, hora_actual=datetime.now().strftime('%H:%M'), title='Inicio ruta conductor')

def conductor_movil_panel_285():
    return conductor_movil_inicio_ruta_285()

def conductor_movil_ruta_285(ruta_id):
    # Misma experiencia compacta del 284, pero la flecha vuelve al inicio de ruta del conductor.
    if not session.get('conductor_id'):
        flash('Inicie sesión como conductor.', 'danger')
        return redirect(url_for('conductor_movil_login'))
    if not _trans_is_allowed_281(ruta_id):
        flash('Esta ruta no está asignada a su usuario conductor.', 'danger')
        return redirect(url_for('conductor_movil_inicio_ruta'))
    ruta = row_to_dict(execute("""SELECT r.*, v.placa, v.capacidad, c.nombres AS conductor
                                FROM transporte_rutas r
                                LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                                LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                                WHERE r.id=?""", (ruta_id,), fetchone=True))
    if not ruta:
        flash('Ruta no encontrada.', 'danger')
        return redirect(url_for('conductor_movil_inicio_ruta'))
    pasajeros = rows_to_dict(execute('SELECT * FROM transporte_pasajeros WHERE ruta_id=? ORDER BY fecha_hora DESC', (ruta_id,), fetchall=True))
    ocupados = len(pasajeros)
    capacidad = int((ruta or {}).get('capacidad') or 0)
    libres = max(0, capacidad - ocupados) if capacidad else 0
    body = _movil_cond_284_css() + r"""
    <div class="phone-wrap"><div class="mv284-card">
      <div class="mv284-head"><a href="{{url_for('conductor_movil_inicio_ruta')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Abordaje ruta</div></div>
      <div class="mv284-body">
        <div class="mv284-routebox">
          <div class="label">Ruta</div>
          <div class="route-title">{{ruta.origen or '-'}} → {{ruta.destino or '-'}}</div>
          <div>Bus: <b>{{ruta.placa or 'SIN BUS'}}</b> | Conductor: <b>{{ruta.conductor or session.get('conductor_nombre') or '-'}}</b> | Inicio: <b>{{ruta.hora_salida or '-'}}</b> | Ocupados: <b>{{ocupados}}</b></div>
        </div>
        <form method="post" action="{{url_for('transporte_abordar', ruta_id=ruta.id)}}" id="frmAbordar284" class="mv284-form">
          <label>DNI / QR / Código de barras</label>
          <div class="input-group"><input name="dni" id="dniTransporte284" class="form-control" inputmode="numeric" maxlength="20" required placeholder="ESCANEAR O DIGITAR DNI" autofocus><button type="button" class="btn mv284-camera" onclick="abrirScanner&&abrirScanner('readerTrans284','dniTransporte284')"><i class="bi bi-camera"></i></button></div>
          <div id="readerTrans284" class="scan-box mt-2" style="display:none"></div>
          <div id="transpStatus284" class="field-help">Al completar 8 dígitos se validará en Trabajadores.</div>
          <select name="metodo" id="metodoTransporte284" class="form-select mt-2"><option>DNI DIGITADO</option><option>QR</option><option>CODIGO DE BARRAS</option></select>
          <input type="hidden" name="latitud" id="latitudTrans284"><input type="hidden" name="longitud" id="longitudTrans284">
          <button class="mv284-btn mt-2"><i class="bi bi-person-check"></i> Registrar subida</button>
        </form>
        <div class="d-grid gap-2 mb-2"><button class="mv284-btn" type="button" onclick="enviarGpsTransporte284({{ruta.id}}, false)"><i class="bi bi-geo-alt"></i> Enviar GPS de esta ruta</button><button class="mv284-outline" type="button" onclick="enviarGpsTransporte284({{ruta.id}}, true)"><i class="bi bi-broadcast-pin"></i> Iniciar / detener GPS en vivo</button></div>
        <div class="mv284-kpis"><div class="mv284-kpi"><small>Abordados</small><b>{{ocupados}}</b></div><div class="mv284-kpi"><small>Libres</small><b>{{libres}}</b></div><div class="mv284-kpi"><small>Capacidad</small><b>{{capacidad or 0}}</b></div></div>
        <div class="mv284-tablewrap"><table class="mv284-table"><thead><tr><th>Hora</th><th>DNI</th><th>Trabajador</th><th>Método</th></tr></thead><tbody>{% for p in pasajeros %}<tr><td>{{p.hora}}</td><td>{{p.dni}}</td><td>{{p.trabajador}}</td><td>{{p.metodo}}</td></tr>{% else %}<tr><td colspan="4" class="text-center text-muted">Sin abordajes.</td></tr>{% endfor %}</tbody></table></div>
      </div>
    </div></div>
    <script>
    (function(){
      const input=document.getElementById('dniTransporte284'), st=document.getElementById('transpStatus284'), lat=document.getElementById('latitudTrans284'), lon=document.getElementById('longitudTrans284'), frm=document.getElementById('frmAbordar284');
      const dni=v=>{const m=String(v||'').match(/(?:^|\D)(\d{8})(?:\D|$)/); return m?m[1]:String(v||'').replace(/\D/g,'').slice(-8)};
      let last='';
      async function validar(){const d=dni(input.value); if(d.length<8){st.className='field-help';st.innerHTML='Esperando 8 dígitos...';return;} input.value=d; if(d===last)return; last=d; st.className='scan-ok flash'; st.innerHTML='Validando DNI '+d+'...'; try{let r=await fetch('/api/trabajador/'+d,{cache:'no-store',credentials:'same-origin'});let j=await r.json(); if(j.ok){st.className='scan-ok flash';st.innerHTML='✓ '+(j.trabajador.trabajador||'TRABAJADOR')+' encontrado'; if(typeof beep==='function')beep();}else{st.className='scan-bad flash';st.innerHTML='✕ '+(j.msg||'DNI no encontrado en base trabajadores');}}catch(e){st.className='scan-bad flash';st.innerHTML='Error validando DNI';}}
      if(input){input.addEventListener('input',validar);input.addEventListener('paste',()=>setTimeout(validar,80));input.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();frm.requestSubmit();}});}
      if(navigator.geolocation){navigator.geolocation.getCurrentPosition(p=>{lat.value=p.coords.latitude;lon.value=p.coords.longitude;},()=>{}, {enableHighAccuracy:true,maximumAge:10000,timeout:15000});}
    })();
    let gpsWatch284=null; async function enviarGpsTransporte284(rid, live){function post(p){let fd=new FormData();fd.append('latitud',p.coords.latitude);fd.append('longitud',p.coords.longitude);document.getElementById('latitudTrans284').value=p.coords.latitude;document.getElementById('longitudTrans284').value=p.coords.longitude;return fetch('/transporte/ruta/'+rid+'/gps',{method:'POST',body:fd,credentials:'same-origin'}).then(r=>r.json());} if(!navigator.geolocation){alert('GPS no disponible');return;} if(live){if(gpsWatch284){navigator.geolocation.clearWatch(gpsWatch284);gpsWatch284=null;alert('GPS en vivo detenido');return;} gpsWatch284=navigator.geolocation.watchPosition(p=>post(p).catch(()=>{}),()=>alert('Permite ubicación/GPS'),{enableHighAccuracy:true,maximumAge:10000,timeout:15000}); alert('GPS en vivo iniciado. Mantenga esta pantalla abierta.'); return;} navigator.geolocation.getCurrentPosition(async p=>{let j=await post(p);alert(j.msg||'GPS actualizado');},()=>alert('Permite ubicación/GPS en el navegador'),{enableHighAccuracy:true,maximumAge:10000,timeout:15000});}
    </script>
    """
    return render_page(body, ruta=ruta, pasajeros=pasajeros, ocupados=ocupados, capacidad=capacidad, libres=libres, title='Abordaje ruta')

try:
    app.add_url_rule('/movil/conductor/iniciar-ruta', 'conductor_movil_inicio_ruta', conductor_movil_inicio_ruta_285, methods=['GET','POST'])
except Exception:
    app.view_functions['conductor_movil_inicio_ruta'] = conductor_movil_inicio_ruta_285

# Overrides finales 285
app.view_functions['conductor_movil_login'] = conductor_movil_login_285
app.view_functions['transporte_mobile_home'] = conductor_movil_login_285
app.view_functions['conductor_movil_panel'] = conductor_movil_panel_285
app.view_functions['conductor_movil_ruta'] = conductor_movil_ruta_285
# ======================= FIN PATCH TRANSPORTE OMAR 285 =======================


# ========================= PATCH TRANSPORTE OMAR 286 =========================
# Ajustes solicitados:
# 1) En Rutas solo se registra base de rutas: sin bus, sin conductor, sin hora salida.
# 2) Antes de Inicio ruta conductor siempre va login DNI conductor / PIN.
# 3) Portada Transporte compacta tipo menú clásico, con recuadros verdes pequeños.
# 4) Se agrega submódulo Reportes al módulo Transporte.

def _transport_286_css():
    return r"""
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}
      .shell{max-width:430px!important;width:100%!important;margin:0 auto!important;padding:6px 8px 26px!important;background:#fff!important}
      .tr286-phone{max-width:390px;margin:0 auto}
      .tr286-app{background:#fff;border:1px solid #e4e8e4;border-radius:13px;overflow:hidden;box-shadow:0 10px 24px rgba(0,0,0,.07)}
      .tr286-hero{height:122px;background:#08713b;color:#fff;position:relative;display:flex;flex-direction:column;align-items:center;justify-content:center}
      .tr286-back{position:absolute;left:14px;top:26px;color:#fff!important;font-size:31px;text-decoration:none;line-height:1}
      .tr286-config{position:absolute;right:11px;top:18px;border:1px solid rgba(255,255,255,.75);color:#fff!important;text-decoration:none;border-radius:12px;padding:6px 9px;font-size:11px;font-weight:950}
      .tr286-bus{font-size:38px;line-height:1;margin-bottom:5px;color:#fff}
      .tr286-title{font-size:16px;font-weight:950;letter-spacing:.25px;text-transform:uppercase;color:#fff}
      .tr286-body{padding:18px 14px 18px}
      .tr286-section{font-size:15px;font-weight:950;color:#08713b;text-transform:uppercase;margin:4px 0 10px;letter-spacing:.3px}
      .tr286-section.op{margin-top:17px}
      .tr286-grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}
      .tr286-grid2{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}
      .tr286-tile{height:78px;background:#08713b;border-radius:10px;color:#fff!important;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 7px 13px rgba(0,0,0,.12);padding:5px}
      .tr286-tile i{font-size:25px;color:#fff;margin-bottom:5px;line-height:1}
      .tr286-tile .lbl{font-size:11.5px;font-weight:950;line-height:1.05;color:#fff}
      .tr286-tile .sub{font-size:8.5px;font-weight:900;color:#eaffee;margin-top:3px;line-height:1}
      .tr286-grid2 .tr286-tile{height:82px}
      .tr286-info{display:grid;grid-template-columns:20px 1fr;gap:8px;border:1px solid #b8d7ff;background:#eef6ff;border-radius:12px;padding:11px;margin-top:16px;color:#0b2e83;font-size:11.2px;font-weight:900;line-height:1.38}
      .tr286-info i{font-size:17px;color:#0b5ed7;margin-top:1px}
      .tr286-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:9px;margin-top:12px}
      .tr286-kpi{background:#10964e;color:#fff;border-radius:9px;text-align:center;padding:7px 4px;box-shadow:0 6px 12px rgba(16,150,78,.16)}
      .tr286-kpi small{display:block;font-size:9.2px;font-weight:950;line-height:1.05;color:#effff3}
      .tr286-kpi b{display:block;font-size:20px;line-height:1.05;font-weight:950;color:#fff;margin-top:4px}
      .tr286-note{margin-top:12px;border:1px solid #cfe4d6;background:#f7fff8;border-radius:12px;padding:10px 11px;color:#1f3b2a;box-shadow:0 5px 12px rgba(0,0,0,.04)}
      .tr286-note .ttl{font-size:12px;font-weight:950;color:#08713b;margin-bottom:5px;text-transform:uppercase}
      .tr286-note .line{font-size:10.7px;font-weight:850;line-height:1.35}

      .rt286-head{height:66px;background:#25773a;color:#fff;display:flex;align-items:center;justify-content:center;position:relative}
      .rt286-head a{position:absolute;left:12px;top:50%;transform:translateY(-50%);color:#fff!important;text-decoration:none;font-size:31px;line-height:1}
      .rt286-head .ttl{font-size:17px;font-weight:950;color:#fff}
      .rt286-body{padding:14px 14px 17px;background:#fff}
      .rt286-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:9px;margin-bottom:13px}
      .rt286-kpi{background:#10964e;color:#fff;border-radius:9px;text-align:center;padding:8px 4px;box-shadow:0 6px 12px rgba(16,150,78,.16)}
      .rt286-kpi small{display:block;font-size:10px;font-weight:950}.rt286-kpi b{display:block;font-size:22px;font-weight:950;line-height:1;margin-top:4px}
      .rt286-search{display:grid;grid-template-columns:1fr 112px;gap:8px;margin-bottom:11px}
      .rt286-search input{height:42px;border:1px solid #dfe7df;border-radius:10px;padding:8px 11px;font-size:12px;font-weight:850}
      .rt286-btnline{height:42px;border:1px solid #bbf7d0;background:#ecfdf5;color:#08713b;border-radius:10px;text-decoration:none;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:950;gap:5px}
      .rt286-title{font-size:13px;font-weight:950;color:#08713b;text-transform:uppercase;margin:10px 0 8px}
      .rt286-tablewrap{border:1px solid #e5e7eb;border-radius:10px;overflow:auto;background:#fff;scrollbar-color:#08713b #e5e7eb;margin-bottom:12px}
      .rt286-tablewrap::-webkit-scrollbar{height:8px}.rt286-tablewrap::-webkit-scrollbar-thumb{background:#08713b;border-radius:999px}.rt286-tablewrap::-webkit-scrollbar-track{background:#e5e7eb}
      .rt286-table{width:100%;min-width:470px;border-collapse:collapse}
      .rt286-table th{background:#f8fafc;color:#12223b;font-size:11px;font-weight:950;padding:8px;border-bottom:1px solid #e5e7eb}
      .rt286-table td{font-size:11px;color:#334155;padding:8px;border-bottom:1px solid #f1f5f9;font-weight:750}
      .rt286-form,.rt286-upload{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:12px;margin-top:12px}
      .rt286-form label,.rt286-upload label{font-size:11px;font-weight:950;color:#176a35;margin-bottom:4px}
      .rt286-form .form-control,.rt286-form .form-select,.rt286-upload .form-control{height:38px!important;border-radius:9px!important;font-size:12px!important;font-weight:850}
      .rt286-form .row{--bs-gutter-x:.55rem;--bs-gutter-y:.55rem}
      .rt286-btn{height:41px;border-radius:10px;background:#08713b;border:1px solid #08713b;color:#fff;font-weight:950;font-size:13px;width:100%}
      .rt286-help{font-size:10.5px;color:#08713b;font-weight:850;line-height:1.35;margin:7px 0}
      .rt286-muted{font-size:10.5px;color:#4a644f;font-weight:850;line-height:1.35}

      .rp286-card{border:1px solid #e4e8e4;border-radius:13px;overflow:hidden;background:#fff;box-shadow:0 10px 24px rgba(0,0,0,.07)}
      .rp286-body{padding:14px}
      .rp286-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}
      .rp286-tile{height:84px;background:#08713b;color:#fff!important;border-radius:10px;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 7px 13px rgba(0,0,0,.12)}
      .rp286-tile i{font-size:26px;margin-bottom:6px;color:#fff}.rp286-tile b{font-size:12px;line-height:1.1;color:#fff}.rp286-tile small{font-size:9px;color:#eaffee;font-weight:900}
    </style>
    """

def _bases_rutas_285():
    # Base de rutas para el conductor: no muestra rutas operativas EN RUTA creadas al iniciar viaje.
    rows = rows_to_dict(execute("""SELECT id,nombre,origen,destino,sede,hora_retorno,estado,fecha
                                  FROM transporte_rutas
                                  WHERE UPPER(COALESCE(estado,'PROGRAMADA')) NOT IN ('EN RUTA','LLEGÓ','LLEGO','RETORNO','CERRADA','FINALIZADA','INACTIVA')
                                    AND (COALESCE(nombre,'')<>'' OR COALESCE(origen,'')<>'' OR COALESCE(destino,'')<>'')
                                  ORDER BY fecha DESC, id DESC LIMIT 400""", fetchall=True))
    seen, out = set(), []
    for r in rows:
        key = f"{r.get('nombre') or ''}|{r.get('origen') or ''}|{r.get('destino') or ''}|{r.get('sede') or ''}".upper()
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
        if len(out) >= 100:
            break
    return out

def transporte_home_286():
    if not (session.get('usuario') or session.get('conductor_id')):
        flash('Inicie sesión para ingresar a Transporte.', 'danger')
        return redirect(url_for('login'))
    total_cond = int(scalar('SELECT COUNT(*) AS c FROM transporte_conductores') or 0)
    total_bus = int(scalar('SELECT COUNT(*) AS c FROM transporte_vehiculos') or 0)
    total_rutas = int(scalar("""SELECT COUNT(*) AS c FROM transporte_rutas
                                WHERE UPPER(COALESCE(estado,'PROGRAMADA')) NOT IN ('EN RUTA','LLEGÓ','LLEGO','RETORNO','CERRADA','FINALIZADA','INACTIVA')""") or 0)
    viajes_hoy = int(scalar("""SELECT COUNT(*) AS c FROM transporte_rutas
                               WHERE fecha=? AND UPPER(COALESCE(estado,'')) IN ('EN RUTA','PROGRAMADA') AND conductor_id IS NOT NULL""", (today_str(),)) or 0)
    abordajes_hoy = int(scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=?', (today_str(),)) or 0)
    back_target = url_for('home') if session.get('usuario') else url_for('conductor_movil_inicio_ruta')
    body = _transport_286_css() + r"""
    <div class="tr286-phone">
      <div class="tr286-app">
        <div class="tr286-hero">
          <a class="tr286-back" href="{{back_target}}"><i class="bi bi-chevron-left"></i></a>
          {% if session.get('usuario') %}<a class="tr286-config" href="{{url_for('transporte_config')}}"><i class="bi bi-gear"></i> Config.</a>{% endif %}
          <div class="tr286-bus"><i class="bi bi-bus-front-fill"></i></div>
          <div class="tr286-title">Módulo Transporte</div>
        </div>
        <div class="tr286-body">
          {% if session.get('usuario') %}
          <div class="tr286-section">Módulos</div>
          <div class="tr286-grid3">
            <a class="tr286-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span class="lbl">Conductores</span></a>
            <a class="tr286-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span class="lbl">Buses</span></a>
            <a class="tr286-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-geo-alt"></i><span class="lbl">Rutas</span></a>
          </div>
          {% endif %}

          <div class="tr286-section op">Operación</div>
          <div class="tr286-grid3">
            <a class="tr286-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-pin-map"></i><span class="lbl">GPS</span><span class="sub">Seguimiento</span></a>
            <a class="tr286-tile" href="{{url_for('transporte_mobile_home')}}"><i class="bi bi-phone"></i><span class="lbl">Móvil</span><span class="sub">Conductor</span></a>
            <a class="tr286-tile" href="{{url_for('transporte_reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i><span class="lbl">Reportes</span><span class="sub">Transporte</span></a>
          </div>

          <div class="tr286-info"><i class="bi bi-info-circle-fill"></i><div>Base correcta: <b>Rutas</b> solo guarda origen/destino/sede. El <b>conductor</b> elige ruta base, bus, hora inicio y GPS desde su móvil antes de abordar.</div></div>

          <div class="tr286-kpis">
            <div class="tr286-kpi"><small>Conductores</small><b>{{total_cond}}</b></div>
            <div class="tr286-kpi"><small>Buses</small><b>{{total_bus}}</b></div>
            <div class="tr286-kpi"><small>Rutas base</small><b>{{total_rutas}}</b></div>
          </div>

          <div class="tr286-note">
            <div class="ttl">Resumen operativo</div>
            <div class="line"><b>{{viajes_hoy}}</b> viaje(s) activos/programados hoy con conductor.</div>
            <div class="line"><b>{{abordajes_hoy}}</b> abordaje(s) registrados hoy.</div>
          </div>
        </div>
      </div>
    </div>
    """
    return render_page(body, back_target=back_target, total_cond=total_cond, total_bus=total_bus, total_rutas=total_rutas, viajes_hoy=viajes_hoy, abordajes_hoy=abordajes_hoy, title='Módulo Transporte')

@login_required
def transporte_rutas_286():
    if request.method == 'POST':
        fecha = request.form.get('fecha') or today_str()
        nombre = limpiar_texto(request.form.get('nombre') or 'RUTA')
        origen = limpiar_texto(request.form.get('origen'))
        destino = limpiar_texto(request.form.get('destino'))
        sede = limpiar_texto(request.form.get('sede'))
        estado = limpiar_texto(request.form.get('estado') or 'PROGRAMADA')
        if not nombre or not origen or not destino:
            flash('Ingrese ruta, origen y destino.', 'danger')
            return redirect(url_for('transporte_rutas'))
        execute("""INSERT INTO transporte_rutas(fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,vehiculo_id,conductor_id,estado,creado_por,creado_en)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (fecha,nombre,origen,destino,sede,'','',None,None,estado,session.get('usuario'),now_str()),
                commit=True)
        flash('Ruta base registrada. El bus, conductor, hora y GPS se seleccionan desde Móvil conductor.', 'success')
        return redirect(url_for('transporte_rutas'))

    rutas = rows_to_dict(execute("""SELECT *
                                  FROM transporte_rutas
                                  WHERE UPPER(COALESCE(estado,'PROGRAMADA')) NOT IN ('EN RUTA','LLEGÓ','LLEGO','RETORNO','FINALIZADA')
                                  ORDER BY fecha DESC, id DESC LIMIT 250""", fetchall=True))
    total = len(rutas)
    activas = sum(1 for r in rutas if (r.get('estado') or '').upper() in ('PROGRAMADA','ACTIVA','BASE',''))
    cerradas = sum(1 for r in rutas if (r.get('estado') or '').upper() in ('CERRADA','INACTIVA'))
    body = _transport_286_css() + r"""
    <div class="tr286-phone"><div class="tr286-app">
      <div class="rt286-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Rutas</div></div>
      <div class="rt286-body">
        <div class="rt286-kpis">
          <div class="rt286-kpi"><small>Total rutas</small><b>{{total}}</b></div>
          <div class="rt286-kpi"><small>Activas</small><b>{{activas}}</b></div>
          <div class="rt286-kpi"><small>Cerradas</small><b>{{cerradas}}</b></div>
        </div>

        <div class="rt286-search">
          <input id="qruta286" placeholder="BUSCAR RUTA...">
          <a class="rt286-btnline" href="{{url_for('transporte_plantilla_rutas')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a>
        </div>

        <div class="rt286-title">Base de rutas</div>
        <div class="rt286-muted mb-2">Aquí solo se registra la base: ruta, origen, destino y sede. Bus, conductor, hora inicio y GPS se registran desde el móvil del conductor.</div>

        <div class="rt286-tablewrap">
          <table id="tblruta286" class="rt286-table">
            <thead><tr><th>Código</th><th>Ruta</th><th>Origen</th><th>Destino</th><th>Sede</th><th>Estado</th><th>Activo</th></tr></thead>
            <tbody>
            {% for r in rutas %}
              <tr>
                <td>R-{{'%03d'%r.id}}</td>
                <td>{{r.nombre or '-'}}</td>
                <td>{{r.origen or '-'}}</td>
                <td>{{r.destino or '-'}}</td>
                <td>{{r.sede or '-'}}</td>
                <td>{{badge(r.estado or 'PROGRAMADA')|safe}}</td>
                <td>{{switch((r.estado or '').upper() not in ['INACTIVA','CERRADA'], url_for('transporte_toggle_ruta', item_id=r.id))|safe}}</td>
              </tr>
            {% else %}
              <tr><td colspan="7" class="text-center text-muted">Sin rutas base.</td></tr>
            {% endfor %}
            </tbody>
          </table>
        </div>

        <form method="post" enctype="multipart/form-data" action="{{url_for('transporte_upload_rutas')}}" class="rt286-upload">
          <label><i class="bi bi-cloud-arrow-up"></i> Carga masiva rutas Excel</label>
          <input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required>
          <div class="rt286-help">Columnas sugeridas: FECHA, RUTA, ORIGEN, DESTINO, SEDE, ESTADO. Ya no se carga bus, conductor ni salida aquí.</div>
          <button class="rt286-btn" type="submit"><i class="bi bi-upload"></i> Cargar rutas</button>
        </form>

        <form method="post" class="rt286-form">
          <div class="row">
            <div class="col-6"><label>Fecha</label><input name="fecha" type="date" value="{{hoy}}" class="form-control"></div>
            <div class="col-6"><label>Ruta</label><input name="nombre" class="form-control" placeholder="R-001" required></div>
            <div class="col-6"><label>Origen</label><input name="origen" class="form-control" required></div>
            <div class="col-6"><label>Destino</label><input name="destino" class="form-control" required></div>
            <div class="col-6"><label>Sede/Fundo</label><input name="sede" class="form-control"></div>
            <div class="col-6"><label>Estado</label><select name="estado" class="form-select"><option>PROGRAMADA</option><option>ACTIVA</option><option>CERRADA</option><option>INACTIVA</option></select></div>
          </div>
          <button class="rt286-btn mt-2">+ Guardar ruta base</button>
        </form>
      </div>
    </div></div>
    <script>
      (function(){const q=document.getElementById('qruta286'),t=document.getElementById('tblruta286'); if(!q||!t)return; q.addEventListener('input',()=>{const s=q.value.toUpperCase(); t.querySelectorAll('tbody tr').forEach(r=>{r.style.display=r.innerText.toUpperCase().includes(s)?'':'none';});});})();
    </script>
    """
    return render_page(body, rutas=rutas, hoy=today_str(), total=total, activas=activas, cerradas=cerradas, badge=_tf_badge, switch=_tf_switch, title='Rutas')

def _importar_rutas_excel(file_storage):
    ok = bad = 0
    for row in _iter_excel_upload(file_storage):
        fecha = _excel_date(_valor(row, ['FECHA','FECHA RUTA'])) or today_str()
        nombre = limpiar_texto(_valor(row, ['RUTA','NOMBRE','NOMBRE RUTA']) or 'RUTA')
        origen = limpiar_texto(_valor(row, ['ORIGEN','PARADERO ORIGEN']))
        destino = limpiar_texto(_valor(row, ['DESTINO','PARADERO DESTINO']))
        sede = limpiar_texto(_valor(row, ['SEDE','FUNDO']))
        estado = limpiar_texto(_valor(row, ['ESTADO']) or 'PROGRAMADA')
        if not nombre or not origen or not destino:
            bad += 1
            continue
        execute("""INSERT INTO transporte_rutas(fecha,nombre,origen,destino,sede,hora_salida,hora_retorno,vehiculo_id,conductor_id,estado,creado_por,creado_en)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (fecha,nombre,origen,destino,sede,'','',None,None,estado,session.get('usuario'),now_str()),
                commit=True)
        ok += 1
    return ok, 0, bad

@login_required
def transporte_plantilla_rutas_286():
    headers = ['FECHA','RUTA','ORIGEN','DESTINO','SEDE','ESTADO']
    example = [today_str(),'R-001','TRUJILLO','PAIJAN','SEDE NORTE','PROGRAMADA']
    return _tpl_xlsx(headers, example, 'plantilla_rutas_base.xlsx', 'RUTAS')

def conductor_movil_login_286():
    # Siempre que se ingrese al botón Móvil conductor se muestra login DNI/PIN.
    # Luego del POST se abre Inicio ruta conductor.
    if request.method == 'POST':
        dni = limpiar_dni(request.form.get('dni'))
        pin = (request.form.get('pin') or '').strip()
        if len(dni) != 8 or str(dni).upper() == 'ADMIN':
            flash('Solo conductores: ingrese DNI de 8 dígitos. No usar ADMIN.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        if not pin:
            flash('Ingrese su PIN móvil.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        c = row_to_dict(execute("""SELECT * FROM transporte_conductores
                                 WHERE dni=? AND COALESCE(movil_pin,'')=?
                                 AND COALESCE(movil_estado,'ACTIVO')='ACTIVO'
                                 AND COALESCE(estado,'ACTIVO') NOT IN ('INACTIVO','BLOQUEADO')""", (dni, pin), fetchone=True))
        if not c:
            flash('DNI o PIN incorrecto, bloqueado o no creado en Transporte > Conductores.', 'danger')
            return redirect(url_for('conductor_movil_login'))
        session['conductor_id'] = c.get('id')
        session['conductor_dni'] = c.get('dni')
        session['conductor_nombre'] = c.get('nombres') or c.get('dni')
        flash('Bienvenido conductor. Seleccione ruta, bus y hora de inicio.', 'success')
        return redirect(url_for('conductor_movil_inicio_ruta'))
    body = r"""
    <div class="phone-wrap">
      <div class="page-card" style="border-radius:14px;overflow:hidden;margin-top:16px">
        <div class="green-hero" style="min-height:128px;border-radius:0;padding:18px 16px 38px">
          <a href="{{url_for('transporte')}}" style="position:absolute;left:18px;top:23px;color:white;font-size:38px;text-decoration:none"><i class="bi bi-chevron-left"></i></a>
          <div style="font-size:34px"><i class="bi bi-phone"></i></div>
          <div style="font-family:Georgia,serif;font-weight:900;font-size:13px;margin-top:4px">ACCESO MÓVIL CONDUCTOR</div>
        </div>
        <div class="floating-card" style="margin:-30px 12px 14px;border-radius:12px;padding:16px">
          <div class="alert alert-light" style="border:1px solid #dbe3db;color:#24405f;line-height:1.45"><b>Solo conductores:</b> ingrese DNI del conductor + PIN móvil creado en Transporte &gt; Conductores. No usar ADMIN.</div>
          <form method="post" id="frmMovilCond286">
            <label class="form-label">DNI conductor</label>
            <input name="dni" id="dniMovCond286" class="form-control mb-3" inputmode="numeric" maxlength="8" pattern="\d{8}" placeholder="Ingrese DNI" required autocomplete="username" autofocus>
            <label class="form-label">PIN móvil</label>
            <input name="pin" class="form-control mb-3" type="password" placeholder="PIN" required autocomplete="current-password">
            <button class="btn btn-green w-100" style="font-size:19px;height:47px">INGRESAR</button>
          </form>
          {% if session.get('usuario') %}
          <a class="btn btn-outline-success w-100 mt-2" href="{{url_for('transporte_conductores')}}">Crear / resetear PIN</a>
          {% endif %}
          <a class="btn btn-outline-secondary w-100 mt-2" href="{{url_for('transporte')}}">Volver</a>
        </div>
      </div>
    </div>
    <script>(function(){const i=document.getElementById('dniMovCond286'); if(i){i.addEventListener('input',()=>{i.value=String(i.value||'').replace(/\D/g,'').slice(-8);});}})();</script>
    """
    return render_page(body, title='Acceso móvil conductor')

@login_required
def transporte_reportes_286():
    hoy = today_str()
    rutas_hoy = int(scalar('SELECT COUNT(*) AS c FROM transporte_rutas WHERE fecha=?', (hoy,)) or 0)
    abordaron = int(scalar('SELECT COUNT(*) AS c FROM transporte_pasajeros WHERE fecha=?', (hoy,)) or 0)
    gps_hoy = int(scalar("SELECT COUNT(*) AS c FROM transporte_gps WHERE substr(fecha_hora,1,10)=?", (hoy,)) or 0)
    conductores = int(scalar('SELECT COUNT(*) AS c FROM transporte_conductores') or 0)
    buses = int(scalar('SELECT COUNT(*) AS c FROM transporte_vehiculos') or 0)
    body = _transport_286_css() + r"""
    <div class="tr286-phone"><div class="rp286-card">
      <div class="rt286-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Reportes transporte</div></div>
      <div class="rp286-body">
        <div class="tr286-kpis">
          <div class="tr286-kpi"><small>Rutas hoy</small><b>{{rutas_hoy}}</b></div>
          <div class="tr286-kpi"><small>Abordajes</small><b>{{abordaron}}</b></div>
          <div class="tr286-kpi"><small>GPS hoy</small><b>{{gps_hoy}}</b></div>
        </div>
        <div class="tr286-kpis">
          <div class="tr286-kpi"><small>Conductores</small><b>{{conductores}}</b></div>
          <div class="tr286-kpi"><small>Buses</small><b>{{buses}}</b></div>
          <div class="tr286-kpi"><small>Fecha</small><b style="font-size:13px">{{hoy[5:]}}</b></div>
        </div>
        <div class="tr286-section op">Reportes</div>
        <div class="rp286-grid">
          <a class="rp286-tile" href="{{url_for('transporte_reporte_abordajes')}}"><i class="bi bi-people"></i><b>Abordajes</b><small>Resumen / Excel</small></a>
          <a class="rp286-tile" href="{{url_for('exportar_transporte_pasajeros')}}"><i class="bi bi-file-earmark-excel"></i><b>Excel abordajes</b><small>Descarga total</small></a>
          <a class="rp286-tile" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-geo-alt"></i><b>GPS</b><small>Seguimiento</small></a>
          <a class="rp286-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-signpost"></i><b>Rutas base</b><small>Base maestra</small></a>
        </div>
      </div>
    </div></div>
    """
    return render_page(body, hoy=hoy, rutas_hoy=rutas_hoy, abordaron=abordaron, gps_hoy=gps_hoy, conductores=conductores, buses=buses, title='Reportes transporte')

try:
    app.add_url_rule('/transporte/reportes', 'transporte_reportes', transporte_reportes_286, methods=['GET'])
except Exception:
    app.view_functions['transporte_reportes'] = transporte_reportes_286

# Overrides finales 286
app.view_functions['transporte'] = transporte_home_286
app.view_functions['transporte_rutas'] = transporte_rutas_286
app.view_functions['transporte_plantilla_rutas'] = transporte_plantilla_rutas_286
app.view_functions['conductor_movil_login'] = conductor_movil_login_286
app.view_functions['transporte_mobile_home'] = conductor_movil_login_286
# ======================= FIN PATCH TRANSPORTE OMAR 286 =======================


# ========================= PATCH TRANSPORTE OMAR 288 =========================
# Reportes transporte mejorado: filtros por fecha/rango + búsqueda + KPIs + 4 reportes:
# Abordajes, Buses, Rutas y Conductores.

def _reporte_transporte_288_css():
    return r"""
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}
      .shell{max-width:430px!important;width:100%!important;margin:0 auto!important;padding:6px 8px 26px!important;background:#fff!important}
      .rp288-phone{max-width:390px;margin:0 auto}
      .rp288-card{border:1px solid #e4e8e4;border-radius:14px;overflow:hidden;background:#fff;box-shadow:0 10px 24px rgba(0,0,0,.07)}
      .rp288-head{height:68px;background:#25773a;color:#fff;display:flex;align-items:center;justify-content:center;position:relative}
      .rp288-head a{position:absolute;left:12px;top:50%;transform:translateY(-50%);color:#fff!important;text-decoration:none;font-size:31px;line-height:1}
      .rp288-head .ttl{font-size:17px;font-weight:950;color:#fff;letter-spacing:.2px}
      .rp288-body{padding:14px;background:#fff}
      .rp288-filter{background:#fff;border:1px solid #e5ece5;border-radius:14px;padding:12px;box-shadow:0 8px 18px rgba(0,0,0,.06);margin-bottom:12px}
      .rp288-dategrid{display:grid;grid-template-columns:1fr 1fr;gap:9px;margin-bottom:10px}
      .rp288-datebox{border:1px solid #dce5dc;border-radius:10px;background:#fff;height:46px;display:flex;align-items:center;justify-content:space-between;padding:6px 9px;color:#142033}
      .rp288-datebox span{display:block;font-size:9.5px;color:#5e6b72;font-weight:850;line-height:1}
      .rp288-datebox input{border:0;outline:0;width:100%;font-size:12px;font-weight:950;color:#142033;background:transparent;padding:0;margin-top:2px}
      .rp288-datebox i{font-size:16px;color:#142033;margin-left:6px}
      .rp288-searchrow{display:grid;grid-template-columns:1fr 52px;gap:8px;margin-bottom:10px}
      .rp288-searchrow input{height:42px;border:1px solid #dce5dc;border-radius:10px;padding:8px 11px;font-size:12px;font-weight:850;color:#142033}
      .rp288-searchrow button{height:42px;border:0;border-radius:10px;background:#08713b;color:#fff;font-size:22px;display:grid;place-items:center}
      .rp288-export{height:42px;border:1px solid #08713b;background:#fff;color:#08713b;border-radius:10px;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:9px;font-size:13px;font-weight:900;margin-bottom:9px}
      .rp288-export i{font-size:20px}.rp288-help{display:flex;align-items:center;gap:8px;color:#4a5565;font-size:11px;font-weight:800;line-height:1.35}
      .rp288-help i{color:#08713b;font-size:18px;flex:0 0 auto}
      .rp288-kpibox{background:#fff;border:1px solid #e5ece5;border-radius:14px;padding:11px;box-shadow:0 8px 18px rgba(0,0,0,.06);margin-bottom:14px}
      .rp288-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:9px}
      .rp288-kpi{background:#fff;border:1px solid #dfe7df;border-radius:12px;min-height:86px;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:6px 3px}
      .rp288-kpi i{font-size:24px;color:#08713b;margin-bottom:5px;line-height:1}
      .rp288-kpi small{display:block;font-size:10.5px;color:#1f2937;font-weight:800;line-height:1.05}
      .rp288-kpi b{display:block;color:#08713b;font-size:24px;font-weight:950;line-height:1;margin-top:5px}
      .rp288-section{display:flex;align-items:center;gap:8px;margin:3px 0 10px;color:#1f2937;font-size:18px;font-weight:950}
      .rp288-section:before{content:'';width:4px;height:24px;background:#10a05b;border-radius:999px;display:block}
      .rp288-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}
      .rp288-tile{height:98px;background:#08713b;color:#fff!important;border-radius:12px;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 8px 15px rgba(0,0,0,.14);padding:8px}
      .rp288-tile i{font-size:31px;margin-bottom:7px;color:#fff;line-height:1}.rp288-tile b{font-size:14px;line-height:1.08;color:#fff;font-weight:950}.rp288-tile small{font-size:10px;color:#eaffee;font-weight:850;line-height:1.1;margin-top:3px}
      .rp288-results{margin-top:12px;border:1px solid #e5e7eb;border-radius:12px;background:#fff;overflow:auto;max-height:220px}
      .rp288-results table{width:100%;min-width:520px;border-collapse:collapse}.rp288-results th{background:#f8fafc;color:#12223b;font-size:11px;font-weight:950;padding:8px;border-bottom:1px solid #e5e7eb}.rp288-results td{font-size:10.5px;color:#334155;padding:8px;border-bottom:1px solid #f1f5f9;font-weight:750}
    </style>
    """

def _rp288_params():
    hoy = today_str()
    desde = request.args.get('desde') or hoy
    hasta = request.args.get('hasta') or hoy
    q = (request.args.get('q') or '').strip()
    if hasta < desde:
        desde, hasta = hasta, desde
    return desde, hasta, q

def _rp288_like_clause(q):
    if not q:
        return "", []
    like = f"%{q.upper()}%"
    clause = """ AND (
        UPPER(COALESCE(p.dni,'')) LIKE ? OR
        UPPER(COALESCE(p.trabajador,'')) LIKE ? OR
        UPPER(COALESCE(r.nombre,'')) LIKE ? OR
        UPPER(COALESCE(r.origen,'')) LIKE ? OR
        UPPER(COALESCE(r.destino,'')) LIKE ? OR
        UPPER(COALESCE(v.placa,'')) LIKE ? OR
        UPPER(COALESCE(c.nombres,'')) LIKE ?
    )"""
    return clause, [like]*7

def _rp288_rows(desde, hasta, q, limit=80):
    clause, params_q = _rp288_like_clause(q)
    sql = f"""SELECT p.fecha, p.hora, p.dni, p.trabajador, p.metodo,
                     r.nombre AS ruta, r.origen, r.destino, v.placa, c.nombres AS conductor
              FROM transporte_pasajeros p
              LEFT JOIN transporte_rutas r ON r.id=p.ruta_id
              LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
              LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
              WHERE p.fecha BETWEEN ? AND ? {clause}
              ORDER BY p.fecha DESC, p.hora DESC, p.id DESC LIMIT {int(limit)}"""
    return rows_to_dict(execute(sql, tuple([desde, hasta] + params_q), fetchall=True))

def transporte_reportes_288():
    if not session.get('usuario'):
        flash('Inicie sesión para ver reportes.', 'danger')
        return redirect(url_for('login'))
    desde, hasta, q = _rp288_params()
    clause, params_q = _rp288_like_clause(q)

    abordajes = int(scalar(f"""SELECT COUNT(*) AS c FROM transporte_pasajeros p
                              LEFT JOIN transporte_rutas r ON r.id=p.ruta_id
                              LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                              LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                              WHERE p.fecha BETWEEN ? AND ? {clause}""", tuple([desde, hasta] + params_q)) or 0)
    rutas = int(scalar("SELECT COUNT(*) AS c FROM transporte_rutas WHERE fecha BETWEEN ? AND ?", (desde, hasta)) or 0)
    gps_total = int(scalar("SELECT COUNT(*) AS c FROM transporte_gps WHERE substr(fecha_hora,1,10) BETWEEN ? AND ?", (desde, hasta)) or 0)
    conductores = int(scalar('SELECT COUNT(*) AS c FROM transporte_conductores') or 0)
    buses = int(scalar('SELECT COUNT(*) AS c FROM transporte_vehiculos') or 0)
    cap = float(scalar("""SELECT COALESCE(SUM(COALESCE(v.capacidad,0)),0) AS c
                         FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                         WHERE r.fecha BETWEEN ? AND ?""", (desde, hasta)) or 0)
    ocupacion = int(round((abordajes / cap) * 100, 0)) if cap else 0
    gps_pct = int(round((gps_total / rutas) * 100, 0)) if rutas else 0
    rows = _rp288_rows(desde, hasta, q, limit=60)

    body = _reporte_transporte_288_css() + r"""
    <div class="rp288-phone"><div class="rp288-card">
      <div class="rp288-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Reportes transporte</div></div>
      <div class="rp288-body">
        <form class="rp288-filter" method="get" action="{{url_for('transporte_reportes')}}">
          <div class="rp288-dategrid">
            <label class="rp288-datebox"><div><span>Desde</span><input type="date" name="desde" value="{{desde}}"></div><i class="bi bi-calendar3"></i></label>
            <label class="rp288-datebox"><div><span>Hasta</span><input type="date" name="hasta" value="{{hasta}}"></div><i class="bi bi-calendar3"></i></label>
          </div>
          <div class="rp288-searchrow"><input name="q" value="{{q}}" placeholder="DNI / trabajador / ruta / bus"><button type="submit"><i class="bi bi-search"></i></button></div>
          <a class="rp288-export" href="{{url_for('exportar_transporte_reportes_288', desde=desde, hasta=hasta, q=q)}}"><i class="bi bi-file-earmark-excel"></i> EXPORTAR EXCEL</a>
          <div class="rp288-help"><i class="bi bi-info-circle"></i><span>Selecciona un rango de fechas para ver indicadores y reportes filtrados.</span></div>
        </form>

        <div class="rp288-kpibox"><div class="rp288-kpis">
          <div class="rp288-kpi"><i class="bi bi-signpost-split"></i><small>Rutas</small><b>{{rutas}}</b></div>
          <div class="rp288-kpi"><i class="bi bi-people"></i><small>Abordajes</small><b>{{abordajes}}</b></div>
          <div class="rp288-kpi"><i class="bi bi-geo-alt"></i><small>GPS</small><b>{{gps_pct}}%</b></div>
          <div class="rp288-kpi"><i class="bi bi-person-badge"></i><small>Conductores</small><b>{{conductores}}</b></div>
          <div class="rp288-kpi"><i class="bi bi-bus-front"></i><small>Buses</small><b>{{buses}}</b></div>
          <div class="rp288-kpi"><i class="bi bi-pie-chart"></i><small>Ocupación</small><b>{{ocupacion}}%</b></div>
        </div></div>

        <div class="rp288-section">Reportes</div>
        <div class="rp288-grid">
          <a class="rp288-tile" href="{{url_for('transporte_reporte_abordajes')}}"><i class="bi bi-people"></i><b>Abordajes</b><small>Resumen / Excel</small></a>
          <a class="rp288-tile" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><b>Buses</b><small>Flota / estado</small></a>
          <a class="rp288-tile" href="{{url_for('transporte_rutas')}}"><i class="bi bi-signpost-split"></i><b>Rutas</b><small>Base maestra</small></a>
          <a class="rp288-tile" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-badge"></i><b>Conductores</b><small>Lista / estado</small></a>
        </div>

        <div class="rp288-results">
          <table><thead><tr><th>Fecha</th><th>DNI</th><th>Trabajador</th><th>Ruta</th><th>Bus</th><th>Método</th></tr></thead><tbody>
          {% for r in rows %}<tr><td>{{r.fecha}} {{r.hora}}</td><td>{{r.dni}}</td><td>{{r.trabajador}}</td><td>{{r.ruta or '-'}}<br><small>{{r.origen or ''}} → {{r.destino or ''}}</small></td><td>{{r.placa or '-'}}</td><td>{{r.metodo or '-'}}</td></tr>
          {% else %}<tr><td colspan="6" class="text-center text-muted">Sin datos para el rango seleccionado.</td></tr>{% endfor %}
          </tbody></table>
        </div>
      </div>
    </div></div>
    """
    return render_page(body, desde=desde, hasta=hasta, q=q, rutas=rutas, abordajes=abordajes, gps_pct=gps_pct, conductores=conductores, buses=buses, ocupacion=ocupacion, rows=rows, title='Reportes transporte')

def exportar_transporte_reportes_288():
    if not session.get('usuario'):
        flash('Inicie sesión para exportar.', 'danger')
        return redirect(url_for('login'))
    desde, hasta, q = _rp288_params()
    rows = _rp288_rows(desde, hasta, q, limit=100000)
    data = []
    for r in rows:
        data.append({
            'fecha': r.get('fecha'), 'hora': r.get('hora'), 'dni': r.get('dni'), 'trabajador': r.get('trabajador'),
            'ruta': r.get('ruta'), 'origen': r.get('origen'), 'destino': r.get('destino'),
            'bus': r.get('placa'), 'conductor': r.get('conductor'), 'metodo': r.get('metodo')
        })
    return excel_response(['fecha','hora','dni','trabajador','ruta','origen','destino','bus','conductor','metodo'], data, f'reporte_transporte_{desde}_a_{hasta}.xlsx', 'REPORTE')

try:
    app.add_url_rule('/transporte/reportes/exportar', 'exportar_transporte_reportes_288', exportar_transporte_reportes_288, methods=['GET'])
except Exception:
    app.view_functions['exportar_transporte_reportes_288'] = exportar_transporte_reportes_288

app.view_functions['transporte_reportes'] = transporte_reportes_288
# ======================= FIN PATCH TRANSPORTE OMAR 288 =======================


# ========================= PATCH TRANSPORTE OMAR 289 =========================
# Reportes transporte miniatura + reportes reales de Abordajes, Buses, Rutas y Conductores.
# Configuraciones depuradas: se eliminan opciones repetidas/no funcionales que enviaban a Transporte.

def _reporte_transporte_289_css():
    return r"""
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}
      .shell{max-width:390px!important;width:100%!important;margin:0 auto!important;padding:4px 7px 22px!important;background:#fff!important}
      .rp289-phone{max-width:355px;margin:0 auto}
      .rp289-card{border:1px solid #e4e8e4;border-radius:13px;overflow:hidden;background:#fff;box-shadow:0 8px 18px rgba(0,0,0,.065)}
      .rp289-head{height:58px;background:#25773a;color:#fff;display:flex;align-items:center;justify-content:center;position:relative}
      .rp289-head a{position:absolute;left:10px;top:50%;transform:translateY(-50%);color:#fff!important;text-decoration:none;font-size:28px;line-height:1}
      .rp289-head .ttl{font-size:15px;font-weight:950;color:#fff;letter-spacing:.15px}
      .rp289-body{padding:10px;background:#fff}
      .rp289-filter{background:#fff;border:1px solid #e5ece5;border-radius:12px;padding:10px;box-shadow:0 6px 14px rgba(0,0,0,.045);margin-bottom:10px}
      .rp289-dategrid{display:grid;grid-template-columns:1fr 1fr;gap:7px;margin-bottom:8px}
      .rp289-datebox{border:1px solid #dce5dc;border-radius:9px;background:#fff;height:39px;display:flex;align-items:center;justify-content:space-between;padding:5px 7px;color:#142033;min-width:0}
      .rp289-datebox span{display:block;font-size:8px;color:#5e6b72;font-weight:850;line-height:1}.rp289-datebox input{border:0;outline:0;width:100%;font-size:10px;font-weight:950;color:#142033;background:transparent;padding:0;margin-top:2px}.rp289-datebox i{font-size:13px;color:#142033;margin-left:3px}
      .rp289-searchrow{display:grid;grid-template-columns:1fr 43px;gap:7px;margin-bottom:8px}.rp289-searchrow input{height:38px;border:1px solid #dce5dc;border-radius:9px;padding:7px 10px;font-size:10.5px;font-weight:850;color:#142033}.rp289-searchrow button{height:38px;border:0;border-radius:9px;background:#08713b;color:#fff;font-size:19px;display:grid;place-items:center}
      .rp289-export{height:38px;border:1px solid #08713b;background:#fff;color:#08713b;border-radius:9px;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:7px;font-size:11.5px;font-weight:900;margin-bottom:7px}.rp289-export i{font-size:17px}
      .rp289-help{display:flex;align-items:flex-start;gap:7px;color:#4a5565;font-size:9.8px;font-weight:800;line-height:1.28}.rp289-help i{color:#08713b;font-size:15px;flex:0 0 auto;margin-top:1px}
      .rp289-kpibox{background:#fff;border:1px solid #e5ece5;border-radius:12px;padding:9px;box-shadow:0 6px 14px rgba(0,0,0,.045);margin-bottom:12px}.rp289-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:7px}.rp289-kpi{background:#fff;border:1px solid #dfe7df;border-radius:10px;min-height:68px;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:5px 2px}.rp289-kpi i{font-size:19px;color:#08713b;margin-bottom:4px;line-height:1}.rp289-kpi small{display:block;font-size:8.9px;color:#1f2937;font-weight:850;line-height:1.05}.rp289-kpi b{display:block;color:#08713b;font-size:18px;font-weight:950;line-height:1;margin-top:4px}
      .rp289-section{display:flex;align-items:center;gap:7px;margin:2px 0 9px;color:#1f2937;font-size:16px;font-weight:950}.rp289-section:before{content:'';width:4px;height:22px;background:#10a05b;border-radius:999px;display:block}
      .rp289-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:8px}.rp289-tile{height:76px;background:#08713b;color:#fff!important;border-radius:10px;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 6px 12px rgba(0,0,0,.13);padding:6px}.rp289-tile i{font-size:25px;margin-bottom:4px;color:#fff;line-height:1}.rp289-tile b{font-size:12px;line-height:1.06;color:#fff;font-weight:950}.rp289-tile small{font-size:8.8px;color:#eaffee;font-weight:850;line-height:1.05;margin-top:2px}
      .rp289-results{margin-top:10px;border:1px solid #e5e7eb;border-radius:10px;background:#fff;overflow:auto;max-height:160px}.rp289-results table{width:100%;min-width:480px;border-collapse:collapse}.rp289-results th{background:#f8fafc;color:#12223b;font-size:10px;font-weight:950;padding:7px;border-bottom:1px solid #e5e7eb}.rp289-results td{font-size:9.5px;color:#334155;padding:7px;border-bottom:1px solid #f1f5f9;font-weight:750}
      .rp289-mini-title{font-size:12px;font-weight:950;color:#08713b;text-transform:uppercase;margin:8px 0 7px}
    </style>
    """

def _rp289_sql_filter(alias_cols, q):
    if not q:
        return "", []
    like = f"%{q.upper()}%"
    parts = [f"UPPER(COALESCE({c},'')) LIKE ?" for c in alias_cols]
    return " AND (" + " OR ".join(parts) + ")", [like] * len(parts)

def _rp289_main_metrics(desde, hasta, q):
    clause, params_q = _rp288_like_clause(q)
    abordajes = int(scalar(f"""SELECT COUNT(*) AS c FROM transporte_pasajeros p
                              LEFT JOIN transporte_rutas r ON r.id=p.ruta_id
                              LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                              LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                              WHERE p.fecha BETWEEN ? AND ? {clause}""", tuple([desde, hasta] + params_q)) or 0)
    rutas = int(scalar("SELECT COUNT(*) AS c FROM transporte_rutas WHERE fecha BETWEEN ? AND ?", (desde, hasta)) or 0)
    gps_total = int(scalar("SELECT COUNT(*) AS c FROM transporte_gps WHERE substr(fecha_hora,1,10) BETWEEN ? AND ?", (desde, hasta)) or 0)
    conductores = int(scalar('SELECT COUNT(*) AS c FROM transporte_conductores') or 0)
    buses = int(scalar('SELECT COUNT(*) AS c FROM transporte_vehiculos') or 0)
    cap = float(scalar("""SELECT COALESCE(SUM(COALESCE(v.capacidad,0)),0) AS c
                         FROM transporte_rutas r LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                         WHERE r.fecha BETWEEN ? AND ?""", (desde, hasta)) or 0)
    ocupacion = int(round((abordajes / cap) * 100, 0)) if cap else 0
    gps_pct = int(round((gps_total / rutas) * 100, 0)) if rutas else 0
    return rutas, abordajes, gps_pct, conductores, buses, ocupacion

def transporte_reportes_289():
    if not session.get('usuario'):
        flash('Inicie sesión para ver reportes.', 'danger')
        return redirect(url_for('login'))
    desde, hasta, q = _rp288_params()
    rutas, abordajes, gps_pct, conductores, buses, ocupacion = _rp289_main_metrics(desde, hasta, q)
    rows = _rp288_rows(desde, hasta, q, limit=35)
    body = _reporte_transporte_289_css() + r"""
    <div class="rp289-phone"><div class="rp289-card">
      <div class="rp289-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Reportes transporte</div></div>
      <div class="rp289-body">
        <form class="rp289-filter" method="get" action="{{url_for('transporte_reportes')}}">
          <div class="rp289-dategrid">
            <label class="rp289-datebox"><div><span>Desde</span><input type="date" name="desde" value="{{desde}}"></div><i class="bi bi-calendar3"></i></label>
            <label class="rp289-datebox"><div><span>Hasta</span><input type="date" name="hasta" value="{{hasta}}"></div><i class="bi bi-calendar3"></i></label>
          </div>
          <div class="rp289-searchrow"><input name="q" value="{{q}}" placeholder="DNI / trabajador / ruta / bus"><button type="submit"><i class="bi bi-search"></i></button></div>
          <a class="rp289-export" href="{{url_for('exportar_transporte_reportes_288', desde=desde, hasta=hasta, q=q)}}"><i class="bi bi-file-earmark-excel"></i> EXPORTAR EXCEL</a>
          <div class="rp289-help"><i class="bi bi-info-circle"></i><span>Selecciona un rango de fechas para ver indicadores y reportes filtrados.</span></div>
        </form>
        <div class="rp289-kpibox"><div class="rp289-kpis">
          <div class="rp289-kpi"><i class="bi bi-signpost-split"></i><small>Rutas</small><b>{{rutas}}</b></div>
          <div class="rp289-kpi"><i class="bi bi-people"></i><small>Abordajes</small><b>{{abordajes}}</b></div>
          <div class="rp289-kpi"><i class="bi bi-geo-alt"></i><small>GPS</small><b>{{gps_pct}}%</b></div>
          <div class="rp289-kpi"><i class="bi bi-person-badge"></i><small>Conductores</small><b>{{conductores}}</b></div>
          <div class="rp289-kpi"><i class="bi bi-bus-front"></i><small>Buses</small><b>{{buses}}</b></div>
          <div class="rp289-kpi"><i class="bi bi-pie-chart"></i><small>Ocupación</small><b>{{ocupacion}}%</b></div>
        </div></div>
        <div class="rp289-section">Reportes</div>
        <div class="rp289-grid">
          <a class="rp289-tile" href="{{url_for('reporte_abordajes_transporte_289', desde=desde, hasta=hasta, q=q)}}"><i class="bi bi-people"></i><b>Abordajes</b><small>Resumen / Excel</small></a>
          <a class="rp289-tile" href="{{url_for('reporte_buses_transporte_289', desde=desde, hasta=hasta, q=q)}}"><i class="bi bi-bus-front"></i><b>Buses</b><small>Flota / estado</small></a>
          <a class="rp289-tile" href="{{url_for('reporte_rutas_transporte_289', desde=desde, hasta=hasta, q=q)}}"><i class="bi bi-signpost-split"></i><b>Rutas</b><small>Base / viajes</small></a>
          <a class="rp289-tile" href="{{url_for('reporte_conductores_transporte_289', desde=desde, hasta=hasta, q=q)}}"><i class="bi bi-person-badge"></i><b>Conductores</b><small>Lista / estado</small></a>
        </div>
        <div class="rp289-results"><table><thead><tr><th>Fecha</th><th>DNI</th><th>Trabajador</th><th>Ruta</th><th>Bus</th><th>Método</th></tr></thead><tbody>
          {% for r in rows %}<tr><td>{{r.fecha}} {{r.hora}}</td><td>{{r.dni}}</td><td>{{r.trabajador}}</td><td>{{r.ruta or '-'}}<br><small>{{r.origen or ''}} → {{r.destino or ''}}</small></td><td>{{r.placa or '-'}}</td><td>{{r.metodo or '-'}}</td></tr>
          {% else %}<tr><td colspan="6" class="text-center text-muted">Sin datos para el rango seleccionado.</td></tr>{% endfor %}
        </tbody></table></div>
      </div>
    </div></div>
    """
    return render_page(body, desde=desde, hasta=hasta, q=q, rutas=rutas, abordajes=abordajes, gps_pct=gps_pct, conductores=conductores, buses=buses, ocupacion=ocupacion, rows=rows, title='Reportes transporte')

def _rp289_filter_form(title, endpoint, desde, hasta, q, export_endpoint=None):
    export_link = ''
    if export_endpoint:
        export_link = """<a class=\"rp289-export\" href=\"{{url_for('""" + export_endpoint + """', desde=desde, hasta=hasta, q=q)}}\"><i class=\"bi bi-file-earmark-excel\"></i> EXPORTAR EXCEL</a>"""
    return r"""
      <div class="rp289-head"><a href="{{url_for('transporte_reportes', desde=desde, hasta=hasta, q=q)}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">""" + title + r"""</div></div>
      <div class="rp289-body">
      <form class="rp289-filter" method="get" action="{{url_for('""" + endpoint + r"""')}}">
        <div class="rp289-dategrid">
          <label class="rp289-datebox"><div><span>Desde</span><input type="date" name="desde" value="{{desde}}"></div><i class="bi bi-calendar3"></i></label>
          <label class="rp289-datebox"><div><span>Hasta</span><input type="date" name="hasta" value="{{hasta}}"></div><i class="bi bi-calendar3"></i></label>
        </div>
        <div class="rp289-searchrow"><input name="q" value="{{q}}" placeholder="Buscar..."><button type="submit"><i class="bi bi-search"></i></button></div>
        """ + export_link + r"""
      </form>
    """

def reporte_abordajes_transporte_289():
    desde, hasta, q = _rp288_params()
    rows = _rp288_rows(desde, hasta, q, limit=200)
    body = _reporte_transporte_289_css() + r"""<div class="rp289-phone"><div class="rp289-card">""" + _rp289_filter_form('Reporte abordajes','reporte_abordajes_transporte_289',desde,hasta,q,'exportar_transporte_reportes_288') + r"""
      <div class="rp289-mini-title">Abordajes filtrados</div>
      <div class="rp289-results" style="max-height:430px"><table><thead><tr><th>Fecha</th><th>DNI</th><th>Trabajador</th><th>Ruta</th><th>Bus</th><th>Conductor</th></tr></thead><tbody>
      {% for r in rows %}<tr><td>{{r.fecha}} {{r.hora}}</td><td>{{r.dni}}</td><td>{{r.trabajador}}</td><td>{{r.ruta or '-'}}</td><td>{{r.placa or '-'}}</td><td>{{r.conductor or '-'}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin abordajes.</td></tr>{% endfor %}
      </tbody></table></div></div></div></div>"""
    return render_page(body, desde=desde, hasta=hasta, q=q, rows=rows, title='Reporte abordajes')

def reporte_buses_transporte_289():
    desde, hasta, q = _rp288_params()
    clause, params = _rp289_sql_filter(['v.placa','v.tipo','v.empresa_transportista','v.estado'], q)
    rows = rows_to_dict(execute(f"""SELECT v.placa, v.tipo, v.capacidad, v.empresa_transportista, v.estado,
                                      COUNT(DISTINCT r.id) AS rutas,
                                      COUNT(p.id) AS abordajes,
                                      MAX(r.ultima_ubicacion) AS ultimo_gps
                               FROM transporte_vehiculos v
                               LEFT JOIN transporte_rutas r ON r.vehiculo_id=v.id AND r.fecha BETWEEN ? AND ?
                               LEFT JOIN transporte_pasajeros p ON p.ruta_id=r.id
                               WHERE 1=1 {clause}
                               GROUP BY v.id, v.placa, v.tipo, v.capacidad, v.empresa_transportista, v.estado
                               ORDER BY v.placa LIMIT 200""", tuple([desde,hasta]+params), fetchall=True))
    body = _reporte_transporte_289_css() + r"""<div class="rp289-phone"><div class="rp289-card">""" + _rp289_filter_form('Reporte buses','reporte_buses_transporte_289',desde,hasta,q,None) + r"""
      <div class="rp289-mini-title">Buses / flota / estado</div>
      <div class="rp289-results" style="max-height:430px"><table><thead><tr><th>Bus</th><th>Tipo</th><th>Cap.</th><th>Estado</th><th>Rutas</th><th>Abordajes</th></tr></thead><tbody>
      {% for r in rows %}<tr><td>{{r.placa}}</td><td>{{r.tipo or '-'}}</td><td>{{r.capacidad or 0}}</td><td>{{r.estado or '-'}}</td><td>{{r.rutas or 0}}</td><td>{{r.abordajes or 0}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin buses.</td></tr>{% endfor %}
      </tbody></table></div></div></div></div>"""
    return render_page(body, desde=desde, hasta=hasta, q=q, rows=rows, title='Reporte buses')

def reporte_rutas_transporte_289():
    desde, hasta, q = _rp288_params()
    clause, params = _rp289_sql_filter(['r.nombre','r.origen','r.destino','r.sede','v.placa','c.nombres','r.estado'], q)
    rows = rows_to_dict(execute(f"""SELECT r.fecha, r.nombre, r.origen, r.destino, r.estado, v.placa, c.nombres AS conductor,
                                      COUNT(p.id) AS abordajes,
                                      COUNT(g.id) AS gps
                               FROM transporte_rutas r
                               LEFT JOIN transporte_vehiculos v ON v.id=r.vehiculo_id
                               LEFT JOIN transporte_conductores c ON c.id=r.conductor_id
                               LEFT JOIN transporte_pasajeros p ON p.ruta_id=r.id
                               LEFT JOIN transporte_gps g ON g.ruta_id=r.id
                               WHERE r.fecha BETWEEN ? AND ? {clause}
                               GROUP BY r.id, r.fecha, r.nombre, r.origen, r.destino, r.estado, v.placa, c.nombres
                               ORDER BY r.fecha DESC, r.id DESC LIMIT 200""", tuple([desde,hasta]+params), fetchall=True))
    body = _reporte_transporte_289_css() + r"""<div class="rp289-phone"><div class="rp289-card">""" + _rp289_filter_form('Reporte rutas','reporte_rutas_transporte_289',desde,hasta,q,None) + r"""
      <div class="rp289-mini-title">Rutas base / viajes</div>
      <div class="rp289-results" style="max-height:430px"><table><thead><tr><th>Fecha</th><th>Ruta</th><th>Origen → Destino</th><th>Bus</th><th>Conductor</th><th>Abord.</th></tr></thead><tbody>
      {% for r in rows %}<tr><td>{{r.fecha}}</td><td>{{r.nombre or '-'}}</td><td>{{r.origen or '-'}} → {{r.destino or '-'}}</td><td>{{r.placa or '-'}}</td><td>{{r.conductor or '-'}}</td><td>{{r.abordajes or 0}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin rutas.</td></tr>{% endfor %}
      </tbody></table></div></div></div></div>"""
    return render_page(body, desde=desde, hasta=hasta, q=q, rows=rows, title='Reporte rutas')

def reporte_conductores_transporte_289():
    desde, hasta, q = _rp288_params()
    clause, params = _rp289_sql_filter(['c.dni','c.nombres','c.estado','c.licencia','c.movil_estado'], q)
    rows = rows_to_dict(execute(f"""SELECT c.dni, c.nombres, c.licencia, c.estado, c.movil_estado,
                                      COUNT(DISTINCT r.id) AS rutas,
                                      COUNT(p.id) AS abordajes,
                                      MAX(c.ultimo_gps) AS ultimo_gps
                               FROM transporte_conductores c
                               LEFT JOIN transporte_rutas r ON r.conductor_id=c.id AND r.fecha BETWEEN ? AND ?
                               LEFT JOIN transporte_pasajeros p ON p.ruta_id=r.id
                               WHERE 1=1 {clause}
                               GROUP BY c.id, c.dni, c.nombres, c.licencia, c.estado, c.movil_estado
                               ORDER BY c.nombres LIMIT 200""", tuple([desde,hasta]+params), fetchall=True))
    body = _reporte_transporte_289_css() + r"""<div class="rp289-phone"><div class="rp289-card">""" + _rp289_filter_form('Reporte conductores','reporte_conductores_transporte_289',desde,hasta,q,None) + r"""
      <div class="rp289-mini-title">Conductores / estado</div>
      <div class="rp289-results" style="max-height:430px"><table><thead><tr><th>DNI</th><th>Conductor</th><th>Licencia</th><th>Estado</th><th>Rutas</th><th>Abord.</th></tr></thead><tbody>
      {% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.nombres}}</td><td>{{r.licencia or '-'}}</td><td>{{r.estado or '-'}} / {{r.movil_estado or 'ACTIVO'}}</td><td>{{r.rutas or 0}}</td><td>{{r.abordajes or 0}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin conductores.</td></tr>{% endfor %}
      </tbody></table></div></div></div></div>"""
    return render_page(body, desde=desde, hasta=hasta, q=q, rows=rows, title='Reporte conductores')

def transporte_config_289():
    if not session.get('usuario'):
        flash('Inicie sesión para configurar transporte.', 'danger')
        return redirect(url_for('login'))
    css = r"""
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}.shell{max-width:390px!important;width:100%!important;margin:0 auto!important;padding:4px 7px 22px!important}.cfg289-phone{max-width:355px;margin:0 auto}.cfg289-card{border:1px solid #e4e8e4;border-radius:13px;overflow:hidden;background:#fff;box-shadow:0 8px 18px rgba(0,0,0,.065)}.cfg289-head{height:58px;background:#25773a;color:#fff;display:flex;align-items:center;justify-content:center;position:relative}.cfg289-head a{position:absolute;left:10px;top:50%;transform:translateY(-50%);color:#fff!important;text-decoration:none;font-size:28px;line-height:1}.cfg289-head .ttl{font-size:15px;font-weight:950;color:#fff}.cfg289-list{padding:12px;background:#fff}.cfg289-item{display:grid;grid-template-columns:30px 1fr 16px;align-items:center;gap:8px;text-decoration:none;color:#102a43;border:1px solid #e4ece4;border-radius:11px;padding:10px 10px;margin-bottom:8px;background:#fff;box-shadow:0 4px 10px rgba(0,0,0,.035)}.cfg289-item i{font-size:20px;color:#08713b;text-align:center}.cfg289-item .chev{font-size:18px;color:#111}.cfg289-item b{display:block;font-size:12px;color:#102a43;line-height:1.1}.cfg289-item small{display:block;font-size:9.5px;color:#557064;font-weight:800;line-height:1.15;margin-top:2px}.cfg289-note{border:1px solid #b8d7ff;background:#eef6ff;color:#0b2e83;border-radius:11px;padding:9px 10px;font-size:10.5px;font-weight:850;line-height:1.35;margin:2px 0 10px}
    </style>
    """
    body = css + r"""
    <div class="cfg289-phone"><div class="cfg289-card">
      <div class="cfg289-head"><a href="{{url_for('transporte')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Configuraciones</div></div>
      <div class="cfg289-list">
        <div class="cfg289-note"><b>Configuración depurada:</b> solo quedan accesos útiles. Se retiraron opciones repetidas o sin pantalla propia.</div>
        <a class="cfg289-item" href="{{url_for('transporte_pin_conductor')}}"><i class="bi bi-phone"></i><span><b>Acceso móvil conductor</b><small>PIN, estado y relación conductor - móvil</small></span><i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg289-item" href="{{url_for('transporte_conductores')}}"><i class="bi bi-person-vcard"></i><span><b>Conductores</b><small>Base, licencias y credenciales</small></span><i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg289-item" href="{{url_for('transporte_vehiculos')}}"><i class="bi bi-bus-front"></i><span><b>Buses / vehículos</b><small>Flota, capacidad y requisitos</small></span><i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg289-item" href="{{url_for('transporte_rutas')}}"><i class="bi bi-signpost-split"></i><span><b>Rutas base</b><small>Origen, destino y sede</small></span><i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg289-item" href="{{url_for('transporte_carga_masiva')}}"><i class="bi bi-cloud-arrow-up"></i><span><b>Carga masiva / plantillas</b><small>Excel de conductores, buses y rutas</small></span><i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg289-item" href="{{url_for('transporte_mapa_general')}}"><i class="bi bi-geo-alt"></i><span><b>GPS / seguimiento</b><small>Última ubicación de rutas</small></span><i class="bi bi-chevron-right chev"></i></a>
        <a class="cfg289-item" href="{{url_for('transporte_reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i><span><b>Reportes transporte</b><small>Abordajes, buses, rutas y conductores</small></span><i class="bi bi-chevron-right chev"></i></a>
      </div>
    </div></div>
    """
    return render_page(body, title='Configuraciones transporte')

try:
    app.add_url_rule('/transporte/reportes/abordajes', 'reporte_abordajes_transporte_289', reporte_abordajes_transporte_289, methods=['GET'])
except Exception:
    app.view_functions['reporte_abordajes_transporte_289'] = reporte_abordajes_transporte_289
try:
    app.add_url_rule('/transporte/reportes/buses', 'reporte_buses_transporte_289', reporte_buses_transporte_289, methods=['GET'])
except Exception:
    app.view_functions['reporte_buses_transporte_289'] = reporte_buses_transporte_289
try:
    app.add_url_rule('/transporte/reportes/rutas', 'reporte_rutas_transporte_289', reporte_rutas_transporte_289, methods=['GET'])
except Exception:
    app.view_functions['reporte_rutas_transporte_289'] = reporte_rutas_transporte_289
try:
    app.add_url_rule('/transporte/reportes/conductores', 'reporte_conductores_transporte_289', reporte_conductores_transporte_289, methods=['GET'])
except Exception:
    app.view_functions['reporte_conductores_transporte_289'] = reporte_conductores_transporte_289

app.view_functions['transporte_reportes'] = transporte_reportes_289
app.view_functions['transporte_config'] = transporte_config_289
# ======================= FIN PATCH TRANSPORTE OMAR 289 =======================

# ========================= PATCH CONTRATACIÓN OMAR 290 =========================
# Módulo integrado al Tareo Móvil con interfaz clásica miniatura.
# Se toma la lógica base del sistema de contratación: requerimiento -> postulante -> evaluación médica -> inducción -> indumentaria -> fotocheck -> firma/contrato.

def _contratacion_init_290():
    idtype = "SERIAL PRIMARY KEY" if is_pg() else "INTEGER PRIMARY KEY AUTOINCREMENT"
    execute(f"""CREATE TABLE IF NOT EXISTS contratacion_requerimientos(
        id {idtype}, fecha TEXT, codigo TEXT, empresa TEXT, area TEXT, cargo TEXT, actividad TEXT,
        cantidad INTEGER DEFAULT 0, fecha_ingreso TEXT, tipo_contrato TEXT, regimen_laboral TEXT,
        estado TEXT DEFAULT 'ABIERTO', creado_por TEXT, creado_en TEXT)""", commit=True)
    execute(f"""CREATE TABLE IF NOT EXISTS contratacion_ingresos(
        id {idtype}, requerimiento_id INTEGER, requerimiento TEXT, dni TEXT, nombres TEXT, telefono TEXT, correo TEXT,
        empresa TEXT, area TEXT, cargo TEXT, actividad TEXT, tipo_contrato TEXT, regimen_laboral TEXT,
        fecha_inicio TEXT, fecha_fin TEXT, basico REAL DEFAULT 0, estado TEXT DEFAULT 'POSTULANTE',
        medica_estado TEXT DEFAULT 'PENDIENTE', induccion_estado TEXT DEFAULT 'PENDIENTE',
        indumentaria_estado TEXT DEFAULT 'PENDIENTE', fotocheck_estado TEXT DEFAULT 'PENDIENTE',
        firma_estado TEXT DEFAULT 'PENDIENTE', observacion TEXT, creado_por TEXT, creado_en TEXT)""", commit=True)
    try:
        conn=get_conn(); cur=conn.cursor()
        for col, ddl in [('codigo','TEXT'),('fecha_ingreso','TEXT'),('tipo_contrato','TEXT'),('regimen_laboral','TEXT')]:
            _add_column_if_missing(cur, 'contratacion_requerimientos', col, ddl)
        for col, ddl in [('requerimiento_id','INTEGER'),('requerimiento','TEXT'),('telefono','TEXT'),('correo','TEXT'),('tipo_contrato','TEXT'),('regimen_laboral','TEXT'),('fecha_inicio','TEXT'),('fecha_fin','TEXT'),('basico','REAL DEFAULT 0'),('medica_estado',"TEXT DEFAULT 'PENDIENTE'"),('induccion_estado',"TEXT DEFAULT 'PENDIENTE'"),('indumentaria_estado',"TEXT DEFAULT 'PENDIENTE'"),('fotocheck_estado',"TEXT DEFAULT 'PENDIENTE'"),('firma_estado',"TEXT DEFAULT 'PENDIENTE'"),('observacion','TEXT')]:
            _add_column_if_missing(cur, 'contratacion_ingresos', col, ddl)
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print('Migración contratación 290:', e)

def _contratacion_css_290():
    return r'''
    <style>
      .ct290-phone{max-width:390px;margin:0 auto}.ct290-app{background:#fff;border:1px solid #e4e8e4;border-radius:13px;overflow:hidden;box-shadow:0 10px 24px rgba(0,0,0,.07)}
      .ct290-head{height:118px;background:#08713b;color:#fff;position:relative;display:flex;flex-direction:column;align-items:center;justify-content:center}.ct290-head a{position:absolute;left:14px;top:26px;color:#fff!important;text-decoration:none;font-size:31px;line-height:1}.ct290-head .ico{font-size:35px;line-height:1;margin-bottom:5px}.ct290-head .ttl{font-size:16px;font-weight:950;letter-spacing:.2px;text-transform:uppercase;color:#fff}
      .ct290-body{padding:16px 14px 18px}.ct290-section{font-size:14px;font-weight:950;color:#08713b;text-transform:uppercase;margin:5px 0 10px;letter-spacing:.2px}.ct290-grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}.ct290-grid2{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}.ct290-tile{height:78px;background:#08713b;border-radius:10px;color:#fff!important;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 7px 13px rgba(0,0,0,.12);padding:5px}.ct290-tile i{font-size:24px;color:#fff;margin-bottom:5px;line-height:1}.ct290-tile .lbl{font-size:10.8px;font-weight:950;line-height:1.05;color:#fff}.ct290-tile .sub{font-size:8px;font-weight:900;color:#eaffee;margin-top:3px;line-height:1}.ct290-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:12px 0}.ct290-kpi{background:#10964e;color:#fff;border-radius:9px;text-align:center;padding:8px 4px;box-shadow:0 6px 12px rgba(16,150,78,.16)}.ct290-kpi small{display:block;font-size:9px;font-weight:950;line-height:1.05}.ct290-kpi b{display:block;font-size:21px;line-height:1.05;font-weight:950;margin-top:4px}.ct290-info{border:1px solid #b8d7ff;background:#eef6ff;color:#0b2e83;border-radius:12px;padding:10px 11px;font-size:11px;font-weight:900;line-height:1.36;margin:12px 0}
      .ct290-card{border:1px solid #e3e8e3;background:#fff;border-radius:12px;box-shadow:0 5px 14px rgba(0,0,0,.06);padding:11px;margin:9px 0}.ct290-mini{font-size:10.5px;color:#486156;font-weight:850;line-height:1.35}.ct290-form{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:11px;margin:10px 0}.ct290-form label{font-size:10.5px;font-weight:950;color:#176a35;margin-bottom:4px}.ct290-form .form-control,.ct290-form .form-select{height:37px!important;border-radius:9px!important;font-size:12px!important;font-weight:850}.ct290-btn{height:39px;border-radius:10px;background:#08713b;border:1px solid #08713b;color:#fff!important;font-weight:950;font-size:12px;width:100%;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:5px}.ct290-outline{height:39px;border-radius:10px;background:#fff;border:1px solid #08713b;color:#08713b!important;font-weight:900;font-size:12px;width:100%;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:5px}.ct290-search{height:38px;border:1px solid #dfe7df;border-radius:10px;padding:8px 11px;font-size:12px;font-weight:850;width:100%;margin-bottom:9px}.ct290-tablewrap{border:1px solid #e5e7eb;border-radius:10px;overflow:auto;background:#fff;scrollbar-color:#08713b #e5e7eb}.ct290-tablewrap::-webkit-scrollbar{height:8px}.ct290-tablewrap::-webkit-scrollbar-thumb{background:#08713b;border-radius:999px}.ct290-tablewrap::-webkit-scrollbar-track{background:#e5e7eb}.ct290-table{width:100%;min-width:590px;border-collapse:collapse}.ct290-table th{background:#f8fafc;color:#12223b;font-size:10.5px;font-weight:950;padding:8px;border-bottom:1px solid #e5e7eb}.ct290-table td{font-size:10.5px;color:#334155;padding:8px;border-bottom:1px solid #f1f5f9;font-weight:750}.ct290-badge{display:inline-block;border-radius:999px;background:#eaf8ee;color:#08713b;padding:4px 8px;font-size:9px;font-weight:950}.ct290-badge.bad{background:#fee2e2;color:#991b1b}.ct290-badge.warn{background:#fff7ed;color:#9a3412}.ct290-row{display:grid;grid-template-columns:1fr 1fr;gap:8px}.ct290-row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px}@media(max-width:430px){.ct290-phone{max-width:360px}.ct290-body{padding:14px 12px}.ct290-tile{height:72px}.ct290-tile i{font-size:22px}.ct290-tile .lbl{font-size:10px}.ct290-table{min-width:560px}}
    </style>'''

def _ct_badge_290(v):
    s=(v or 'PENDIENTE').upper()
    cls='bad' if s in ('NO APTO','RECHAZADO','BLOQUEADO','OBSERVADO') else ('warn' if s in ('PENDIENTE','POSTULANTE','EN PROCESO') else '')
    return f"<span class='ct290-badge {cls}'>{s}</span>"

def _ct_req_label_290(r):
    return f"REQ-{int(r.get('id') or 0):03d} · {r.get('cargo') or '-'} · {r.get('area') or '-'}"

def _ct_req_options_290(reqs, selected=''):
    html=['<option value="">Seleccione requerimiento...</option>']
    for r in reqs:
        sel='selected' if str(r.get('id'))==str(selected) else ''
        html.append(f"<option value='{r.get('id')}' {sel}>{_ct_req_label_290(r)}</option>")
    return ''.join(html)

@login_required
def contratacion_home_290():
    _contratacion_init_290()
    total_req=int(scalar('SELECT COUNT(*) AS c FROM contratacion_requerimientos') or 0)
    abiertos=int(scalar("SELECT COUNT(*) AS c FROM contratacion_requerimientos WHERE UPPER(COALESCE(estado,'ABIERTO')) NOT IN ('CERRADO','CANCELADO')") or 0)
    postulantes=int(scalar('SELECT COUNT(*) AS c FROM contratacion_ingresos') or 0)
    med_ok=int(scalar("SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE UPPER(COALESCE(medica_estado,'')) IN ('APTO','OK','COMPLETO')") or 0)
    firma=int(scalar("SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE UPPER(COALESCE(firma_estado,'')) IN ('FIRMADO','COMPLETO','OK')") or 0)
    body=_contratacion_css_290()+r'''
    <div class="ct290-phone"><div class="ct290-app">
      <div class="ct290-head"><a href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-person-plus"></i></div><div class="ttl">Módulo contratación</div></div>
      <div class="ct290-body">
        <div class="ct290-section">Gestión</div>
        <div class="ct290-grid3">
          <a class="ct290-tile" href="{{url_for('contratacion_requerimientos')}}"><i class="bi bi-clipboard2-plus"></i><span class="lbl">Requerim.</span><span class="sub">Cupos</span></a>
          <a class="ct290-tile" href="{{url_for('contratacion_postulantes')}}"><i class="bi bi-person-lines-fill"></i><span class="lbl">Postulantes</span><span class="sub">DNI / datos</span></a>
          <a class="ct290-tile" href="{{url_for('contratacion_etapa', etapa='medica')}}"><i class="bi bi-heart-pulse"></i><span class="lbl">Médica</span><span class="sub">Aptitud</span></a>
          <a class="ct290-tile" href="{{url_for('contratacion_etapa', etapa='induccion')}}"><i class="bi bi-play-btn"></i><span class="lbl">Inducción</span><span class="sub">Asistencia</span></a>
          <a class="ct290-tile" href="{{url_for('contratacion_etapa', etapa='indumentaria')}}"><i class="bi bi-bag-check"></i><span class="lbl">Indument.</span><span class="sub">EPP</span></a>
          <a class="ct290-tile" href="{{url_for('contratacion_etapa', etapa='fotocheck')}}"><i class="bi bi-person-badge"></i><span class="lbl">Fotocheck</span><span class="sub">Carnet</span></a>
        </div>
        <div class="ct290-section" style="margin-top:16px">Operación</div>
        <div class="ct290-grid2">
          <a class="ct290-tile" href="{{url_for('contratacion_etapa', etapa='firma')}}"><i class="bi bi-pen"></i><span class="lbl">Firma / contrato</span><span class="sub">Documentos</span></a>
          <a class="ct290-tile" href="{{url_for('contratacion_reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i><span class="lbl">Reportes</span><span class="sub">Control</span></a>
        </div>
        <div class="ct290-info"><i class="bi bi-info-circle"></i> Flujo: Requerimiento → Postulantes → Evaluación médica → Inducción → Indumentaria → Fotocheck → Firma/contrato.</div>
        <div class="ct290-kpis"><div class="ct290-kpi"><small>Req.</small><b>{{total_req}}</b></div><div class="ct290-kpi"><small>Abiertos</small><b>{{abiertos}}</b></div><div class="ct290-kpi"><small>Postul.</small><b>{{postulantes}}</b></div></div>
        <div class="ct290-kpis"><div class="ct290-kpi"><small>Méd. OK</small><b>{{med_ok}}</b></div><div class="ct290-kpi"><small>Firmados</small><b>{{firma}}</b></div><div class="ct290-kpi"><small>Hoy</small><b style="font-size:13px">{{hoy[5:]}}</b></div></div>
      </div>
    </div></div>'''
    return render_page(body,total_req=total_req,abiertos=abiertos,postulantes=postulantes,med_ok=med_ok,firma=firma,hoy=today_str(),title='Contratación')

@login_required
def contratacion_requerimientos_290():
    _contratacion_init_290()
    if request.method=='POST':
        fecha=request.form.get('fecha') or today_str(); empresa=limpiar_texto(request.form.get('empresa') or 'AQUANQA'); area=limpiar_texto(request.form.get('area')); cargo=limpiar_texto(request.form.get('cargo')); actividad=limpiar_texto(request.form.get('actividad'))
        try: cantidad=int(request.form.get('cantidad') or 0)
        except Exception: cantidad=0
        if not area or not cargo or cantidad<=0:
            flash('Ingrese área, cargo y cantidad.', 'danger'); return redirect(url_for('contratacion_requerimientos'))
        estado=limpiar_texto(request.form.get('estado') or 'ABIERTO'); tipo=limpiar_texto(request.form.get('tipo_contrato') or 'TEMPORAL'); regimen=limpiar_texto(request.form.get('regimen_laboral') or 'AGRARIO')
        execute('''INSERT INTO contratacion_requerimientos(fecha,codigo,empresa,area,cargo,actividad,cantidad,fecha_ingreso,tipo_contrato,regimen_laboral,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''', (fecha,'',empresa,area,cargo,actividad,cantidad,request.form.get('fecha_ingreso') or fecha,tipo,regimen,estado,session.get('usuario'),now_str()), commit=True)
        rid=scalar('SELECT MAX(id) AS id FROM contratacion_requerimientos'); execute('UPDATE contratacion_requerimientos SET codigo=? WHERE id=?', (f'REQ-{int(rid or 0):03d}',rid), commit=True)
        flash('Requerimiento creado correctamente.', 'success'); return redirect(url_for('contratacion_requerimientos'))
    reqs=rows_to_dict(execute('''SELECT r.*, (SELECT COUNT(*) FROM contratacion_ingresos i WHERE i.requerimiento_id=r.id) AS registrados FROM contratacion_requerimientos r ORDER BY r.id DESC LIMIT 150''', fetchall=True))
    body=_contratacion_css_290()+r'''
    <div class="ct290-phone"><div class="ct290-app"><div class="ct290-head"><a href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-clipboard2-plus"></i></div><div class="ttl">Requerimientos</div></div><div class="ct290-body">
      <div class="ct290-kpis"><div class="ct290-kpi"><small>Total</small><b>{{reqs|length}}</b></div><div class="ct290-kpi"><small>Abiertos</small><b>{{abiertos}}</b></div><div class="ct290-kpi"><small>Cupos</small><b>{{cupos}}</b></div></div>
      <form method="post" class="ct290-form"><div class="ct290-row"><div><label>Fecha</label><input class="form-control" type="date" name="fecha" value="{{hoy}}"></div><div><label>Cantidad</label><input class="form-control" type="number" name="cantidad" value="1" min="1"></div></div><div class="ct290-row mt-2"><div><label>Empresa</label><input class="form-control" name="empresa" placeholder="AQUANQA"></div><div><label>Área</label><input class="form-control" name="area" required></div></div><div class="ct290-row mt-2"><div><label>Cargo</label><input class="form-control" name="cargo" required></div><div><label>Actividad</label><input class="form-control" name="actividad"></div></div><div class="ct290-row mt-2"><div><label>Fecha ingreso</label><input class="form-control" type="date" name="fecha_ingreso" value="{{hoy}}"></div><div><label>Estado</label><select name="estado" class="form-select"><option>ABIERTO</option><option>CERRADO</option><option>CANCELADO</option></select></div></div><div class="ct290-row mt-2"><div><label>Tipo contrato</label><select name="tipo_contrato" class="form-select"><option>TEMPORAL</option><option>INTERMITENTE</option><option>INDETERMINADO</option></select></div><div><label>Régimen</label><select name="regimen_laboral" class="form-select"><option>AGRARIO</option><option>GENERAL</option></select></div></div><button class="ct290-btn mt-2"><i class="bi bi-plus-circle"></i> Crear requerimiento</button></form>
      <input class="ct290-search" id="qreq290" placeholder="Buscar requerimiento..."><div class="ct290-tablewrap"><table id="tblreq290" class="ct290-table"><thead><tr><th>Código</th><th>Área</th><th>Cargo</th><th>Cupos</th><th>Registrados</th><th>Estado</th></tr></thead><tbody>{% for r in reqs %}<tr><td>{{r.codigo or ('REQ-%03d'%r.id)}}</td><td>{{r.area or '-'}}</td><td>{{r.cargo or '-'}}</td><td>{{r.cantidad or 0}}</td><td>{{r.registrados or 0}}</td><td>{{badge(r.estado)|safe}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin requerimientos.</td></tr>{% endfor %}</tbody></table></div>
    </div></div></div><script>(function(){const q=document.getElementById('qreq290'),t=document.getElementById('tblreq290'); if(q&&t)q.addEventListener('input',()=>{const s=q.value.toUpperCase();t.querySelectorAll('tbody tr').forEach(r=>r.style.display=r.innerText.toUpperCase().includes(s)?'':'none')});})();</script>'''
    abiertos=sum(1 for r in reqs if (r.get('estado') or '').upper() not in ('CERRADO','CANCELADO')); cupos=sum(int(r.get('cantidad') or 0) for r in reqs)
    return render_page(body, reqs=reqs, abiertos=abiertos, cupos=cupos, hoy=today_str(), badge=_ct_badge_290, title='Requerimientos')

@login_required
def contratacion_postulantes_290():
    _contratacion_init_290()
    req_id=request.args.get('req') or request.form.get('requerimiento_id') or ''
    reqs=rows_to_dict(execute("SELECT * FROM contratacion_requerimientos WHERE UPPER(COALESCE(estado,'ABIERTO')) NOT IN ('CANCELADO') ORDER BY id DESC LIMIT 200", fetchall=True))
    if request.method=='POST':
        dni=limpiar_dni(request.form.get('dni')); req=row_to_dict(execute('SELECT * FROM contratacion_requerimientos WHERE id=?', (req_id,), fetchone=True)) if req_id else None
        if not req: flash('Seleccione requerimiento.', 'danger'); return redirect(url_for('contratacion_postulantes'))
        if len(dni)!=8: flash('DNI inválido.', 'danger'); return redirect(url_for('contratacion_postulantes', req=req_id))
        if scalar('SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE requerimiento_id=? AND dni=?', (req_id,dni)):
            flash('Este DNI ya está registrado en el requerimiento.', 'danger'); return redirect(url_for('contratacion_postulantes', req=req_id))
        ocupados=int(scalar('SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE requerimiento_id=?', (req_id,)) or 0)
        if int(req.get('cantidad') or 0) and ocupados >= int(req.get('cantidad') or 0): flash('Cupo completo para este requerimiento.', 'danger'); return redirect(url_for('contratacion_postulantes', req=req_id))
        base=row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True)) or {}; nombres=limpiar_texto(request.form.get('nombres') or base.get('trabajador'))
        if not nombres: flash('Ingrese nombres del postulante.', 'danger'); return redirect(url_for('contratacion_postulantes', req=req_id))
        execute('''INSERT INTO contratacion_ingresos(requerimiento_id,requerimiento,dni,nombres,telefono,correo,empresa,area,cargo,actividad,tipo_contrato,regimen_laboral,fecha_inicio,fecha_fin,basico,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (req_id, req.get('codigo') or f'REQ-{int(req_id):03d}', dni, nombres, request.form.get('telefono',''), request.form.get('correo',''), req.get('empresa'), req.get('area'), req.get('cargo'), req.get('actividad'), req.get('tipo_contrato'), req.get('regimen_laboral'), req.get('fecha_ingreso') or today_str(), request.form.get('fecha_fin',''), float(request.form.get('basico') or 0), 'POSTULANTE', session.get('usuario'), now_str()), commit=True)
        flash('Postulante registrado correctamente.', 'success'); return redirect(url_for('contratacion_postulantes', req=req_id))
    posts=rows_to_dict(execute('''SELECT * FROM contratacion_ingresos WHERE (?='' OR requerimiento_id=?) ORDER BY id DESC LIMIT 250''', (str(req_id),str(req_id)), fetchall=True))
    body=_contratacion_css_290()+r'''
    <div class="ct290-phone"><div class="ct290-app"><div class="ct290-head"><a href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-person-lines-fill"></i></div><div class="ttl">Postulantes</div></div><div class="ct290-body"><form method="get" class="ct290-form"><label>Requerimiento</label><select class="form-select" name="req" onchange="this.form.submit()">{{req_opts|safe}}</select></form>
      <form method="post" class="ct290-form" id="frmPost290"><input type="hidden" name="requerimiento_id" value="{{req_id}}"><div class="ct290-row"><div><label>DNI</label><input class="form-control" name="dni" id="dniPost290" maxlength="8" inputmode="numeric" required></div><div><label>Nombres</label><input class="form-control" name="nombres" id="nomPost290" required></div></div><div class="ct290-row mt-2"><div><label>Teléfono</label><input class="form-control" name="telefono"></div><div><label>Correo</label><input class="form-control" name="correo" type="email"></div></div><div class="ct290-row mt-2"><div><label>Básico</label><input class="form-control" name="basico" type="number" step="0.01"></div><div><label>Fin contrato</label><input class="form-control" name="fecha_fin" type="date"></div></div><div id="stPost290" class="ct290-mini mt-2">Digite DNI de 8 dígitos para buscar en trabajadores.</div><button class="ct290-btn mt-2"><i class="bi bi-person-plus"></i> Registrar postulante</button></form>
      <input class="ct290-search" id="qpost290" placeholder="Buscar DNI / postulante..."><div class="ct290-tablewrap"><table id="tblpost290" class="ct290-table"><thead><tr><th>DNI</th><th>Postulante</th><th>Req.</th><th>Cargo</th><th>Médica</th><th>Firma</th></tr></thead><tbody>{% for p in posts %}<tr><td>{{p.dni}}</td><td>{{p.nombres}}</td><td>{{p.requerimiento}}</td><td>{{p.cargo or '-'}}</td><td>{{badge(p.medica_estado)|safe}}</td><td>{{badge(p.firma_estado)|safe}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin postulantes.</td></tr>{% endfor %}</tbody></table></div>
    </div></div></div><script>(function(){const q=document.getElementById('qpost290'),t=document.getElementById('tblpost290'); if(q&&t)q.addEventListener('input',()=>{const s=q.value.toUpperCase();t.querySelectorAll('tbody tr').forEach(r=>r.style.display=r.innerText.toUpperCase().includes(s)?'':'none')}); const i=document.getElementById('dniPost290'),n=document.getElementById('nomPost290'),st=document.getElementById('stPost290'); async function buscar(){const d=String(i.value||'').replace(/\D/g,'').slice(-8);i.value=d;if(d.length<8)return;st.innerHTML='Buscando DNI '+d+'...';try{let r=await fetch('/api/trabajador/'+d,{cache:'no-store'});let j=await r.json();if(j.ok){n.value=j.trabajador.trabajador||'';st.innerHTML='✓ Datos encontrados en trabajadores.';if(typeof beep==='function')beep();}else{st.innerHTML='No está en trabajadores. Puede digitar nombre manualmente.';}}catch(e){st.innerHTML='No se pudo consultar trabajadores.';}} if(i){i.addEventListener('input',buscar);}})();</script>'''
    return render_page(body, reqs=reqs, posts=posts, req_id=req_id, req_opts=_ct_req_options_290(reqs, req_id), badge=_ct_badge_290, title='Postulantes')

_ETAPAS_290={'medica':('Evaluación médica','heart-pulse','medica_estado',['PENDIENTE','APTO','NO APTO','OBSERVADO']),'induccion':('Inducción','play-btn','induccion_estado',['PENDIENTE','ASISTIÓ','NO ASISTIÓ','OBSERVADO']),'indumentaria':('Indumentaria','bag-check','indumentaria_estado',['PENDIENTE','ENTREGADO','OBSERVADO']),'fotocheck':('Fotocheck','person-badge','fotocheck_estado',['PENDIENTE','GENERADO','IMPRESO','ENTREGADO']),'firma':('Firma / contrato','pen','firma_estado',['PENDIENTE','GENERADO','ENVIADO','FIRMADO','OBSERVADO'])}

@login_required
def contratacion_etapa_290(etapa):
    _contratacion_init_290()
    if etapa not in _ETAPAS_290: return redirect(url_for('contratacion_home'))
    titulo, icono, col, estados=_ETAPAS_290[etapa]; req_id=request.args.get('req') or request.form.get('requerimiento_id') or ''
    reqs=rows_to_dict(execute('SELECT * FROM contratacion_requerimientos ORDER BY id DESC LIMIT 200', fetchall=True))
    if request.method=='POST':
        dni=limpiar_dni(request.form.get('dni')); estado=limpiar_texto(request.form.get('estado') or 'PENDIENTE'); obs=limpiar_texto(request.form.get('observacion'), upper=False)
        where='dni=?' + (' AND requerimiento_id=?' if req_id else ''); params=[dni] + ([req_id] if req_id else [])
        if len(dni)!=8: flash('DNI inválido.', 'danger'); return redirect(url_for('contratacion_etapa', etapa=etapa, req=req_id))
        if not scalar(f'SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE {where}', tuple(params)):
            flash('El DNI no está registrado como postulante en el requerimiento.', 'danger'); return redirect(url_for('contratacion_etapa', etapa=etapa, req=req_id))
        execute(f"UPDATE contratacion_ingresos SET {col}=?, observacion=? WHERE {where}", tuple([estado,obs]+params), commit=True)
        flash(f'{titulo}: estado actualizado para {dni}.', 'success'); return redirect(url_for('contratacion_etapa', etapa=etapa, req=req_id))
    posts=rows_to_dict(execute('''SELECT * FROM contratacion_ingresos WHERE (?='' OR requerimiento_id=?) ORDER BY id DESC LIMIT 250''', (str(req_id),str(req_id)), fetchall=True))
    pendientes=sum(1 for p in posts if (p.get(col) or '').upper() in ('PENDIENTE','POSTULANTE','')); ok=sum(1 for p in posts if (p.get(col) or '').upper() not in ('PENDIENTE','POSTULANTE','','NO APTO','OBSERVADO','NO ASISTIÓ'))
    opts=''.join([f'<option>{e}</option>' for e in estados])
    body=_contratacion_css_290()+r'''
    <div class="ct290-phone"><div class="ct290-app"><div class="ct290-head"><a href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-{{icono}}"></i></div><div class="ttl">{{titulo}}</div></div><div class="ct290-body"><div class="ct290-kpis"><div class="ct290-kpi"><small>Total</small><b>{{posts|length}}</b></div><div class="ct290-kpi"><small>OK</small><b>{{ok}}</b></div><div class="ct290-kpi"><small>Pend.</small><b>{{pendientes}}</b></div></div>
      <form method="get" class="ct290-form"><label>Requerimiento</label><select class="form-select" name="req" onchange="this.form.submit()">{{req_opts|safe}}</select></form><form method="post" class="ct290-form"><input type="hidden" name="requerimiento_id" value="{{req_id}}"><div class="ct290-row"><div><label>DNI postulante</label><input class="form-control" name="dni" maxlength="8" inputmode="numeric" required></div><div><label>Estado</label><select name="estado" class="form-select">{{opts|safe}}</select></div></div><div class="mt-2"><label>Observación</label><input class="form-control" name="observacion"></div><button class="ct290-btn mt-2"><i class="bi bi-check-circle"></i> Actualizar estado</button></form>
      <input class="ct290-search" id="qet290" placeholder="Buscar DNI / trabajador..."><div class="ct290-tablewrap"><table id="tblet290" class="ct290-table"><thead><tr><th>DNI</th><th>Postulante</th><th>Req.</th><th>Cargo</th><th>Estado etapa</th></tr></thead><tbody>{% for p in posts %}<tr><td>{{p.dni}}</td><td>{{p.nombres}}</td><td>{{p.requerimiento}}</td><td>{{p.cargo or '-'}}</td><td>{{badge(p[col])|safe}}</td></tr>{% else %}<tr><td colspan="5" class="text-center text-muted">Sin postulantes.</td></tr>{% endfor %}</tbody></table></div></div></div></div><script>(function(){const q=document.getElementById('qet290'),t=document.getElementById('tblet290'); if(q&&t)q.addEventListener('input',()=>{const s=q.value.toUpperCase();t.querySelectorAll('tbody tr').forEach(r=>r.style.display=r.innerText.toUpperCase().includes(s)?'':'none')});})();</script>'''
    return render_page(body, etapa=etapa,titulo=titulo,icono=icono,posts=posts,req_id=req_id,req_opts=_ct_req_options_290(reqs, req_id),opts=opts,badge=_ct_badge_290,col=col,ok=ok,pendientes=pendientes,title=titulo)

@login_required
def contratacion_reportes_290():
    _contratacion_init_290()
    total_req=int(scalar('SELECT COUNT(*) AS c FROM contratacion_requerimientos') or 0); total_post=int(scalar('SELECT COUNT(*) AS c FROM contratacion_ingresos') or 0)
    med_ok=int(scalar("SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE UPPER(COALESCE(medica_estado,'')) IN ('APTO','OK','COMPLETO')") or 0); indu_ok=int(scalar("SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE UPPER(COALESCE(induccion_estado,'')) IN ('ASISTIÓ','ASISTIO','OK','COMPLETO')") or 0); foto_ok=int(scalar("SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE UPPER(COALESCE(fotocheck_estado,'')) IN ('GENERADO','IMPRESO','ENTREGADO','OK')") or 0); firm_ok=int(scalar("SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE UPPER(COALESCE(firma_estado,'')) IN ('FIRMADO','OK','COMPLETO')") or 0)
    rows=rows_to_dict(execute('SELECT * FROM contratacion_ingresos ORDER BY id DESC LIMIT 250', fetchall=True))
    body=_contratacion_css_290()+r'''<div class="ct290-phone"><div class="ct290-app"><div class="ct290-head"><a href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-file-earmark-bar-graph"></i></div><div class="ttl">Reportes contratación</div></div><div class="ct290-body"><div class="ct290-kpis"><div class="ct290-kpi"><small>Req.</small><b>{{total_req}}</b></div><div class="ct290-kpi"><small>Post.</small><b>{{total_post}}</b></div><div class="ct290-kpi"><small>Méd.</small><b>{{med_ok}}</b></div></div><div class="ct290-kpis"><div class="ct290-kpi"><small>Induc.</small><b>{{indu_ok}}</b></div><div class="ct290-kpi"><small>Foto.</small><b>{{foto_ok}}</b></div><div class="ct290-kpi"><small>Firma</small><b>{{firm_ok}}</b></div></div><a class="ct290-outline mb-2" href="{{url_for('contratacion_export_postulantes')}}"><i class="bi bi-file-earmark-excel"></i> Exportar Excel</a><input class="ct290-search" id="qrep290" placeholder="Buscar DNI / trabajador / cargo..."><div class="ct290-tablewrap"><table id="tblrep290" class="ct290-table"><thead><tr><th>DNI</th><th>Postulante</th><th>Cargo</th><th>Médica</th><th>Inducción</th><th>Firma</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.nombres}}</td><td>{{r.cargo or '-'}}</td><td>{{badge(r.medica_estado)|safe}}</td><td>{{badge(r.induccion_estado)|safe}}</td><td>{{badge(r.firma_estado)|safe}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin datos.</td></tr>{% endfor %}</tbody></table></div></div></div></div><script>(function(){const q=document.getElementById('qrep290'),t=document.getElementById('tblrep290'); if(q&&t)q.addEventListener('input',()=>{const s=q.value.toUpperCase();t.querySelectorAll('tbody tr').forEach(r=>r.style.display=r.innerText.toUpperCase().includes(s)?'':'none')});})();</script>'''
    return render_page(body,total_req=total_req,total_post=total_post,med_ok=med_ok,indu_ok=indu_ok,foto_ok=foto_ok,firm_ok=firm_ok,rows=rows,badge=_ct_badge_290,title='Reportes contratación')

@login_required
def contratacion_export_postulantes_290():
    _contratacion_init_290(); rows=rows_to_dict(execute('SELECT dni,nombres,requerimiento,empresa,area,cargo,actividad,medica_estado,induccion_estado,indumentaria_estado,fotocheck_estado,firma_estado,estado,creado_en FROM contratacion_ingresos ORDER BY id DESC', fetchall=True))
    headers=['dni','nombres','requerimiento','empresa','area','cargo','actividad','medica_estado','induccion_estado','indumentaria_estado','fotocheck_estado','firma_estado','estado','creado_en']; return excel_response(headers, rows, 'contratacion_postulantes.xlsx', 'CONTRATACION')

def home_290():
    if not session.get('usuario'): return redirect(url_for('inicio'))
    hojas = rows_to_dict(execute('SELECT * FROM hojas_tareo ORDER BY fecha DESC, id DESC LIMIT 8', fetchall=True))
    body = """
    <div class="desktop-grid"><div class="phone-wrap"><div class="green-hero" style="min-height:220px"><div class="green-top"><a class="text-white text-decoration-none" href="{{url_for('soporte')}}"><i class="bi bi-headset"></i> Soporte</a>{% if session.get('rol')=='admin' %}<a class="text-white text-decoration-none" href="{{url_for('configuraciones')}}"><i class="bi bi-gear"></i> Config.</a>{% else %}<span></span>{% endif %}</div><div class="avatar"><i class="bi bi-person-circle"></i></div><div class="login-name">{{ session.get('nombres','USUARIO') }}</div><div class="white-input mt-3"></div></div><div class="top-actions"><a class="tile text-decoration-none" href="{{url_for('hojas_tareo')}}"><i class="bi bi-list-check"></i>TAREO</a><a class="tile text-decoration-none" href="{{url_for('asistencia_modulo')}}"><i class="bi bi-fingerprint"></i>ASIST.</a><a class="tile text-decoration-none" href="{{url_for('documentos_firma')}}"><i class="bi bi-pen"></i>FIRMAS</a><a class="tile text-decoration-none" href="{{url_for('transporte')}}"><i class="bi bi-bus-front"></i>TRANSP.</a><a class="tile text-decoration-none" href="{{url_for('contratacion_home')}}"><i class="bi bi-person-plus"></i>CONTRAT.</a><a class="tile text-decoration-none" href="{{url_for('reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i>REPORTES<br>TAREO</a><a class="tile text-decoration-none" href="{{url_for('sincronizacion')}}"><i class="bi bi-arrow-repeat"></i>SINC.</a></div><div class="leaf"></div><div class="bottom-sync"><i class="bi bi-arrow-repeat"></i> Sincronizar Tablas Maestras<br>Actualizado hasta: {{ now }}</div><a href="{{url_for('logout')}}" class="bottom-out"><i class="bi bi-box-arrow-right"></i></a></div><div class="desk-panel"><h1 class="header-title">TAREO MÓVIL – GRUPO DE COSECHA</h1><div class="card-pro p-4 mb-3"><div class="d-flex justify-content-between align-items-center"><div><h4 class="fw-bold text-success mb-1">Hojas recientes</h4><div class="text-muted small">Crea una hoja y registra labores, trabajadores y avances.</div></div><a class="btn btn-green" href="{{url_for('crear_hoja')}}"><i class="bi bi-plus-lg"></i> Crear hoja</a></div></div><div class="card-pro p-3"><div class="table-responsive"><table class="table list-table"><thead><tr><th>Fecha</th><th>Grupo</th><th>Subgrupo</th><th>Labor</th><th>Responsable</th><th>Estado</th><th></th></tr></thead><tbody>{% for h in hojas %}<tr><td>{{h.fecha}}</td><td>{{h.grupo}}</td><td>{{h.subgrupo}}</td><td>{{h.labor}}</td><td>{{h.responsable}}</td><td><span class="status-pill">{{h.estado}}</span></td><td><a class="btn btn-sm btn-green" href="{{url_for('detalle_hoja', hoja_id=h.id)}}">Abrir</a></td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted py-4">Sin hojas creadas.</td></tr>{% endfor %}</tbody></table></div></div></div></div>"""
    return render_page(body, hojas=hojas, now=now_str())

try:
    app.add_url_rule('/contratacion', 'contratacion_home', contratacion_home_290, methods=['GET'])
    app.add_url_rule('/contratacion/requerimientos', 'contratacion_requerimientos', contratacion_requerimientos_290, methods=['GET','POST'])
    app.add_url_rule('/contratacion/postulantes', 'contratacion_postulantes', contratacion_postulantes_290, methods=['GET','POST'])
    app.add_url_rule('/contratacion/etapa/<etapa>', 'contratacion_etapa', contratacion_etapa_290, methods=['GET','POST'])
    app.add_url_rule('/contratacion/reportes', 'contratacion_reportes', contratacion_reportes_290, methods=['GET'])
    app.add_url_rule('/contratacion/export/postulantes', 'contratacion_export_postulantes', contratacion_export_postulantes_290, methods=['GET'])
except Exception as e:
    print('Rutas contratación 290:', e)
app.view_functions['home'] = home_290
# ======================= FIN PATCH CONTRATACIÓN OMAR 290 =======================

# ========================= PATCH BOLETAS OMAR 291 =========================
# Módulo Boleta integrado al proyecto actual con interfaz clásica/miniatura.
# Lógica adaptada desde el portal de boletas/documentos: tipos de pago, carga,
# listado por DNI-periodo, descarga, detección por DNI en nombre de archivo y reportes.

BOLETA_TIPOS_CANON_291 = [
    ('NORMAL','Boletas normal','bi-file-earmark-text'),
    ('CTS','Boletas CTS','bi-file-earmark-check'),
    ('GRATIFICACIÓN','Boletas gratificación','bi-gift'),
    ('UTILIDAD','Boletas utilidades','bi-cash-coin'),
    ('VACACIONES','Boletas vacaciones','bi-calendar-check'),
    ('LIQUIDACIÓN','Boletas liquidación','bi-file-earmark-medical'),
]
BOLETA_EXT_ALLOWED_291 = {'.pdf','.png','.jpg','.jpeg','.webp','.doc','.docx','.xls','.xlsx'}
BOLETA_DIR_291 = os.path.join(PERSIST_DIR, 'boletas_documentos')
BOLETA_AUTO_DIR_291 = os.path.join(PERSIST_DIR, 'BOLETAS_UPLOAD_AUTO')
os.makedirs(BOLETA_DIR_291, exist_ok=True)
os.makedirs(BOLETA_AUTO_DIR_291, exist_ok=True)

def _boleta_init_291():
    idtype = 'SERIAL PRIMARY KEY' if is_pg() else 'INTEGER PRIMARY KEY AUTOINCREMENT'
    execute(f'''CREATE TABLE IF NOT EXISTS boleta_documentos(
        id {idtype}, dni TEXT NOT NULL, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        tipo TEXT NOT NULL, periodo TEXT, detalle TEXT, archivo_nombre TEXT, archivo_path TEXT,
        extension TEXT, estado TEXT DEFAULT 'CARGADO', fecha_subida TEXT, uploaded_by TEXT,
        fecha_lectura TEXT, observacion TEXT)''', commit=True)

def _bt_css_291():
    return r'''
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}.shell{max-width:430px!important;width:100%!important;margin:0 auto!important;padding:7px 8px 26px!important;background:#fff!important}.bt291-phone{max-width:390px;margin:0 auto}.bt291-app{background:#fff;border:1px solid #e4e8e4;border-radius:13px;overflow:hidden;box-shadow:0 10px 24px rgba(0,0,0,.07)}
      .bt291-head{height:120px;background:#08713b;color:#fff;position:relative;display:flex;flex-direction:column;align-items:center;justify-content:center}.bt291-head a.back{position:absolute;left:14px;top:25px;color:#fff!important;text-decoration:none;font-size:31px;line-height:1}.bt291-head .ico{font-size:38px;line-height:1;margin-bottom:5px;color:#fff}.bt291-head .ttl{font-size:16px;font-weight:950;text-transform:uppercase;color:#fff;letter-spacing:.25px}.bt291-body{padding:15px 14px 17px;background:#fff}.bt291-section{font-size:14px;font-weight:950;color:#08713b;text-transform:uppercase;margin:4px 0 10px;letter-spacing:.25px}.bt291-grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}.bt291-grid2{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}.bt291-tile{height:82px;background:#08713b;border-radius:10px;color:#fff!important;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 7px 13px rgba(0,0,0,.12);padding:5px}.bt291-tile i{font-size:25px;color:#fff;margin-bottom:5px;line-height:1}.bt291-tile .lbl{font-size:11.3px;font-weight:950;line-height:1.05;color:#fff}.bt291-tile .sub{font-size:8.5px;font-weight:900;color:#eaffee;margin-top:3px;line-height:1}
      .bt291-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:10px 0 12px}.bt291-kpi{background:#10964e;color:#fff;border-radius:9px;text-align:center;padding:7px 4px;box-shadow:0 6px 12px rgba(16,150,78,.16)}.bt291-kpi small{display:block;font-size:9px;font-weight:950;line-height:1.05;color:#effff3}.bt291-kpi b{display:block;font-size:20px;line-height:1.05;font-weight:950;color:#fff;margin-top:4px}.bt291-info{display:grid;grid-template-columns:20px 1fr;gap:8px;border:1px solid #b8d7ff;background:#eef6ff;border-radius:12px;padding:10px;margin-top:13px;color:#0b2e83;font-size:11px;font-weight:900;line-height:1.35}.bt291-info i{font-size:17px;color:#0b5ed7;margin-top:1px}
      .bt291-filter{border:1px solid #e1ebe3;background:#fff;border-radius:13px;padding:11px;margin-bottom:12px;box-shadow:0 7px 16px rgba(0,0,0,.06)}.bt291-filter .row{--bs-gutter-x:.55rem;--bs-gutter-y:.55rem}.bt291-filter label,.bt291-form label{font-size:10.5px;font-weight:950;color:#176a35;margin-bottom:4px}.bt291-filter .form-control,.bt291-filter .form-select,.bt291-form .form-control,.bt291-form .form-select{height:37px!important;border-radius:9px!important;font-size:12px!important;font-weight:850}.bt291-btn{height:39px;border-radius:10px;background:#08713b;border:1px solid #08713b;color:#fff;font-weight:950;font-size:12.5px;width:100%;display:flex;align-items:center;justify-content:center;gap:6px;text-decoration:none}.bt291-outline{height:39px;border-radius:10px;background:#fff;border:1px solid #08713b;color:#08713b!important;font-weight:950;font-size:12px;width:100%;display:flex;align-items:center;justify-content:center;gap:6px;text-decoration:none}.bt291-form{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:12px;margin-bottom:12px}.bt291-help{font-size:10.5px;color:#4a644f;font-weight:850;line-height:1.35;margin-top:5px}.bt291-search{height:39px;border:1px solid #dfe7df;border-radius:10px;padding:8px 11px;font-size:12px;font-weight:850;width:100%;margin-bottom:10px}.bt291-tablewrap{border:1px solid #e5e7eb;border-radius:10px;overflow:auto;background:#fff;scrollbar-color:#08713b #e5e7eb}.bt291-tablewrap::-webkit-scrollbar{height:8px}.bt291-tablewrap::-webkit-scrollbar-thumb{background:#08713b;border-radius:999px}.bt291-table{width:100%;min-width:650px;border-collapse:collapse}.bt291-table th{background:#f8fafc;color:#12223b;font-size:11px;font-weight:950;padding:8px;border-bottom:1px solid #e5e7eb}.bt291-table td{font-size:11px;color:#334155;padding:8px;border-bottom:1px solid #f1f5f9;font-weight:750}.bt291-doc{display:grid;grid-template-columns:36px 1fr 28px;gap:8px;align-items:center;border:1px solid #e3e8e3;border-radius:11px;background:#fff;padding:9px 10px;margin:7px 0;text-decoration:none;color:#102a43!important;box-shadow:0 4px 10px rgba(0,0,0,.04)}.bt291-doc i{font-size:25px;color:#08713b}.bt291-doc b{font-size:12px;color:#0a1f44}.bt291-doc small{font-size:10px;color:#496455;font-weight:850;line-height:1.25}.bt291-mini{font-size:10px;font-weight:900;color:#08713b;text-transform:uppercase;margin:9px 0 6px}
    </style>
    '''

def _bt_norm_tipo_291(v):
    t = limpiar_texto(v or 'NORMAL')
    t = t.replace('GRATIFICACION','GRATIFICACIÓN').replace('LIQUIDACION','LIQUIDACIÓN')
    if t not in [x[0] for x in BOLETA_TIPOS_CANON_291]: t='NORMAL'
    return t

def _bt_safe_filename_291(name):
    name = os.path.basename(str(name or 'archivo'))
    name = re.sub(r'[^A-Za-z0-9_.\- ]+', '_', name).strip() or 'archivo'
    return name[:120]

def _bt_find_worker_291(dni):
    dni = limpiar_dni(dni)
    t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    if not t: return {'dni':dni,'trabajador':'','empresa':'','area':'','cargo':''}
    return {'dni':dni,'trabajador':t.get('trabajador') or t.get('nombres') or '', 'empresa':t.get('empresa') or '', 'area':t.get('area') or '', 'cargo':t.get('cargo') or ''}

def _bt_save_file_291(file_storage, dni, tipo, periodo):
    raw = file_storage.filename or 'documento.pdf'
    safe = _bt_safe_filename_291(raw)
    ext = os.path.splitext(safe)[1].lower()
    if ext not in BOLETA_EXT_ALLOWED_291:
        raise ValueError('Formato no permitido. Use PDF, imagen, Word o Excel.')
    carpeta = os.path.join(BOLETA_DIR_291, _bt_norm_tipo_291(tipo), re.sub(r'[^A-Za-z0-9_\-]+','_', periodo or today_str()))
    os.makedirs(carpeta, exist_ok=True)
    fname = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{limpiar_dni(dni)}_{safe}"
    path = os.path.join(carpeta, fname)
    file_storage.save(path)
    return path, safe, ext

def _bt_insert_doc_291(dni, tipo, periodo, detalle, observacion, path, original, ext, uploaded_by):
    w = _bt_find_worker_291(dni)
    execute('''INSERT INTO boleta_documentos(dni,trabajador,empresa,area,cargo,tipo,periodo,detalle,archivo_nombre,archivo_path,extension,estado,fecha_subida,uploaded_by,observacion)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (w['dni'], w['trabajador'], w['empresa'], w['area'], w['cargo'], _bt_norm_tipo_291(tipo), periodo, detalle, original, path, ext, 'CARGADO', now_str(), uploaded_by, observacion), commit=True)

def _bt_tipo_options_291(sel=''):
    sel = _bt_norm_tipo_291(sel or 'NORMAL')
    return ''.join([f'<option value="{t}" {"selected" if t==sel else ""}>{lbl}</option>' for t,lbl,ico in BOLETA_TIPOS_CANON_291])

def _bt_query_docs_291(desde=None, hasta=None, tipo=None, buscar=None, limit=300):
    _boleta_init_291()
    where=[]; params=[]
    if desde:
        where.append('substr(fecha_subida,1,10)>=?'); params.append(desde)
    if hasta:
        where.append('substr(fecha_subida,1,10)<=?'); params.append(hasta)
    if tipo:
        where.append('tipo=?'); params.append(_bt_norm_tipo_291(tipo))
    if buscar:
        b = '%' + str(buscar).upper() + '%'
        where.append('(UPPER(dni) LIKE ? OR UPPER(trabajador) LIKE ? OR UPPER(tipo) LIKE ? OR UPPER(periodo) LIKE ? OR UPPER(archivo_nombre) LIKE ?)')
        params += [b,b,b,b,b]
    sql='SELECT * FROM boleta_documentos'
    if where: sql += ' WHERE ' + ' AND '.join(where)
    sql += ' ORDER BY id DESC LIMIT %d' % int(limit or 300)
    return rows_to_dict(execute(sql, tuple(params), fetchall=True))

@login_required
def boletas_home_291():
    _boleta_init_291()
    total=int(scalar('SELECT COUNT(*) AS c FROM boleta_documentos') or 0)
    cargadas_hoy=int(scalar('SELECT COUNT(*) AS c FROM boleta_documentos WHERE substr(fecha_subida,1,10)=?', (today_str(),)) or 0)
    trabajadores=int(scalar('SELECT COUNT(DISTINCT dni) AS c FROM boleta_documentos') or 0)
    rows=rows_to_dict(execute('SELECT tipo, COUNT(*) AS c FROM boleta_documentos GROUP BY tipo ORDER BY c DESC', fetchall=True))
    conteos={r.get('tipo'):r.get('c') for r in rows}
    body=_bt_css_291()+r'''
    <div class="bt291-phone"><div class="bt291-app"><div class="bt291-head"><a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-file-earmark-text"></i></div><div class="ttl">Módulo Boletas</div></div><div class="bt291-body">
      <div class="bt291-kpis"><div class="bt291-kpi"><small>Total</small><b>{{total}}</b></div><div class="bt291-kpi"><small>Hoy</small><b>{{cargadas_hoy}}</b></div><div class="bt291-kpi"><small>DNI</small><b>{{trabajadores}}</b></div></div>
      <div class="bt291-section">Documentos de pago</div><div class="bt291-grid3">{% for t,lbl,ico in tipos %}<a class="bt291-tile" href="{{url_for('boletas_listar', tipo=t)}}"><i class="bi {{ico}}"></i><span class="lbl">{{t}}</span><span class="sub">{{conteos.get(t,0)}} docs.</span></a>{% endfor %}</div>
      <div class="bt291-section" style="margin-top:16px">Operación</div><div class="bt291-grid3"><a class="bt291-tile" href="{{url_for('boletas_subir')}}"><i class="bi bi-cloud-arrow-up"></i><span class="lbl">Subir</span><span class="sub">PDF / docs</span></a><a class="bt291-tile" href="{{url_for('boletas_detectar')}}"><i class="bi bi-search"></i><span class="lbl">Detectar</span><span class="sub">PDF por DNI</span></a><a class="bt291-tile" href="{{url_for('boletas_reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i><span class="lbl">Reportes</span><span class="sub">Excel</span></a></div>
      <div class="bt291-info"><i class="bi bi-info-circle-fill"></i><div>Tipos tomados del portal de boletas: Normal, CTS, Gratificación, Utilidad, Vacaciones y Liquidación. Cada archivo queda vinculado por DNI, periodo y tipo.</div></div>
    </div></div></div>'''
    return render_page(body,total=total,cargadas_hoy=cargadas_hoy,trabajadores=trabajadores,tipos=BOLETA_TIPOS_CANON_291,conteos=conteos,title='Módulo Boletas')

@login_required
def boletas_subir_291():
    _boleta_init_291()
    if request.method == 'POST':
        dni=limpiar_dni(request.form.get('dni'))
        tipo=_bt_norm_tipo_291(request.form.get('tipo'))
        periodo=limpiar_texto(request.form.get('periodo') or datetime.now().strftime('%Y-%m'), upper=False)
        detalle=limpiar_texto(request.form.get('detalle'), upper=False)
        obs=limpiar_texto(request.form.get('observacion'), upper=False)
        f=request.files.get('archivo')
        if len(dni)!=8:
            flash('Ingrese DNI válido de 8 dígitos.', 'danger'); return redirect(url_for('boletas_subir'))
        if not f or not f.filename:
            flash('Seleccione archivo de boleta/documento.', 'danger'); return redirect(url_for('boletas_subir'))
        try:
            path, original, ext = _bt_save_file_291(f, dni, tipo, periodo)
            _bt_insert_doc_291(dni, tipo, periodo, detalle, obs, path, original, ext, session.get('usuario'))
            flash('Boleta/documento cargado correctamente.', 'success')
            return redirect(url_for('boletas_listar', tipo=tipo, buscar=dni))
        except Exception as e:
            flash(str(e), 'danger')
            return redirect(url_for('boletas_subir'))
    body=_bt_css_291()+r'''
    <div class="bt291-phone"><div class="bt291-app"><div class="bt291-head"><a class="back" href="{{url_for('boletas_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-cloud-arrow-up"></i></div><div class="ttl">Subir boleta</div></div><div class="bt291-body"><form class="bt291-form" method="post" enctype="multipart/form-data"><div class="row"><div class="col-6"><label>DNI</label><input class="form-control" name="dni" maxlength="8" inputmode="numeric" required autofocus></div><div class="col-6"><label>Tipo</label><select class="form-select" name="tipo">{{tipo_opts|safe}}</select></div><div class="col-6"><label>Periodo</label><input class="form-control" name="periodo" value="{{periodo}}" placeholder="2026-06"></div><div class="col-6"><label>Archivo</label><input class="form-control" name="archivo" type="file" required></div><div class="col-12"><label>Detalle</label><input class="form-control" name="detalle" placeholder="Boleta mensual / CTS / grati..."></div><div class="col-12"><label>Observación</label><input class="form-control" name="observacion"></div></div><button class="bt291-btn mt-2"><i class="bi bi-upload"></i> Guardar boleta</button><div class="bt291-help">Se acepta PDF, imagen, Word o Excel. El trabajador se jala desde la base Trabajadores si existe el DNI.</div></form><a class="bt291-outline" href="{{url_for('boletas_plantilla')}}"><i class="bi bi-file-earmark-excel"></i> Descargar plantilla control</a></div></div></div>'''
    return render_page(body,tipo_opts=_bt_tipo_options_291(),periodo=datetime.now().strftime('%Y-%m'),title='Subir boleta')

@login_required
def boletas_listar_291(tipo=None):
    _boleta_init_291()
    desde=request.args.get('desde') or ''
    hasta=request.args.get('hasta') or ''
    buscar=request.args.get('buscar') or ''
    tipo = tipo or request.args.get('tipo') or ''
    docs=_bt_query_docs_291(desde,hasta,tipo,buscar,300)
    body=_bt_css_291()+r'''
    <div class="bt291-phone"><div class="bt291-app"><div class="bt291-head"><a class="back" href="{{url_for('boletas_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-folder2-open"></i></div><div class="ttl">Boletas {{tipo or ''}}</div></div><div class="bt291-body"><form class="bt291-filter" method="get"><div class="row"><div class="col-6"><label>Desde</label><input type="date" name="desde" value="{{desde}}" class="form-control"></div><div class="col-6"><label>Hasta</label><input type="date" name="hasta" value="{{hasta}}" class="form-control"></div><div class="col-12"><label>Buscar</label><input name="buscar" value="{{buscar}}" class="form-control" placeholder="DNI / trabajador / periodo"></div><div class="col-6"><button class="bt291-btn"><i class="bi bi-search"></i> Buscar</button></div><div class="col-6"><a class="bt291-outline" href="{{url_for('boletas_subir')}}"><i class="bi bi-plus-circle"></i> Subir</a></div></div></form><div class="bt291-mini">Documentos encontrados: {{docs|length}}</div>{% for d in docs %}<a class="bt291-doc" href="{{url_for('boletas_archivo', doc_id=d.id)}}" target="_blank"><i class="bi bi-file-earmark-text"></i><div><b>{{d.tipo}} · {{d.periodo or '-'}}</b><br><small>{{d.dni}} · {{d.trabajador or 'NO ENCONTRADO'}}<br>{{d.archivo_nombre}}</small></div><i class="bi bi-chevron-right"></i></a>{% else %}<div class="bt291-form text-center text-muted">Sin boletas para el filtro.</div>{% endfor %}</div></div></div>'''
    return render_page(body,docs=docs,desde=desde,hasta=hasta,buscar=buscar,tipo=tipo,title='Boletas')

@login_required
def boletas_archivo_291(doc_id):
    _boleta_init_291()
    d=row_to_dict(execute('SELECT * FROM boleta_documentos WHERE id=?', (doc_id,), fetchone=True))
    if not d or not d.get('archivo_path') or not os.path.exists(d.get('archivo_path')):
        flash('Archivo no encontrado.', 'danger'); return redirect(url_for('boletas_home'))
    try:
        execute("UPDATE boleta_documentos SET fecha_lectura=? WHERE id=? AND COALESCE(fecha_lectura,'')=''", (now_str(), doc_id), commit=True)
    except Exception:
        pass
    return send_file(d.get('archivo_path'), as_attachment=False, download_name=d.get('archivo_nombre') or os.path.basename(d.get('archivo_path')))

@login_required
def boletas_detectar_291():
    _boleta_init_291()
    registrados=omitidos=0
    for root, dirs, files in os.walk(BOLETA_AUTO_DIR_291):
        for fn in files:
            ext=os.path.splitext(fn)[1].lower()
            if ext not in BOLETA_EXT_ALLOWED_291: continue
            m=re.search(r'(\d{8})', fn)
            if not m:
                omitidos+=1; continue
            dni=m.group(1)
            src=os.path.join(root,fn)
            exists=scalar('SELECT COUNT(*) AS c FROM boleta_documentos WHERE archivo_path=?', (src,))
            if exists:
                omitidos+=1; continue
            periodo_match=re.search(r'(20\d{2}[-_ ]?(?:0[1-9]|1[0-2]))', fn)
            periodo=(periodo_match.group(1).replace('_','-').replace(' ','-') if periodo_match else datetime.now().strftime('%Y-%m'))
            tipo='NORMAL'
            up=fn.upper()
            for cand in ['CTS','GRATIFICACION','GRATIFICACIÓN','UTILIDAD','VACACIONES','LIQUIDACION','LIQUIDACIÓN']:
                if cand in up:
                    tipo=_bt_norm_tipo_291(cand); break
            _bt_insert_doc_291(dni,tipo,periodo,'Detectado automáticamente','Archivo detectado en carpeta BOLETAS_UPLOAD_AUTO',src,fn,ext,session.get('usuario'))
            registrados+=1
    flash(f'Detección finalizada. Registrados: {registrados}. Omitidos: {omitidos}. Carpeta: {BOLETA_AUTO_DIR_291}', 'success')
    return redirect(url_for('boletas_home'))

@login_required
def boletas_reportes_291():
    _boleta_init_291()
    desde=request.args.get('desde') or today_str()
    hasta=request.args.get('hasta') or today_str()
    buscar=request.args.get('buscar') or ''
    docs=_bt_query_docs_291(desde,hasta,None,buscar,250)
    total=len(docs); leidos=sum(1 for d in docs if d.get('fecha_lectura'))
    dni_unicos=len(set(d.get('dni') for d in docs if d.get('dni')))
    tipos=len(set(d.get('tipo') for d in docs if d.get('tipo')))
    body=_bt_css_291()+r'''
    <div class="bt291-phone"><div class="bt291-app"><div class="bt291-head"><a class="back" href="{{url_for('boletas_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-file-earmark-bar-graph"></i></div><div class="ttl">Reportes boletas</div></div><div class="bt291-body"><form class="bt291-filter" method="get"><div class="row"><div class="col-6"><label>Desde</label><input type="date" name="desde" value="{{desde}}" class="form-control"></div><div class="col-6"><label>Hasta</label><input type="date" name="hasta" value="{{hasta}}" class="form-control"></div><div class="col-8"><label>Búsqueda</label><input name="buscar" value="{{buscar}}" class="form-control" placeholder="DNI / trabajador / periodo"></div><div class="col-4 d-flex align-items-end"><button class="bt291-btn"><i class="bi bi-search"></i></button></div><div class="col-12"><a class="bt291-outline" href="{{url_for('boletas_exportar')}}?desde={{desde}}&hasta={{hasta}}&buscar={{buscar}}"><i class="bi bi-file-earmark-excel"></i> Exportar Excel</a></div></div></form><div class="bt291-kpis"><div class="bt291-kpi"><small>Total</small><b>{{total}}</b></div><div class="bt291-kpi"><small>DNI</small><b>{{dni_unicos}}</b></div><div class="bt291-kpi"><small>Leídos</small><b>{{leidos}}</b></div></div><div class="bt291-kpis"><div class="bt291-kpi"><small>Tipos</small><b>{{tipos}}</b></div><div class="bt291-kpi"><small>Desde</small><b style="font-size:12px">{{desde[5:]}}</b></div><div class="bt291-kpi"><small>Hasta</small><b style="font-size:12px">{{hasta[5:]}}</b></div></div><div class="bt291-tablewrap"><table class="bt291-table"><thead><tr><th>Fecha</th><th>DNI</th><th>Trabajador</th><th>Tipo</th><th>Periodo</th><th>Archivo</th></tr></thead><tbody>{% for d in docs %}<tr><td>{{d.fecha_subida[:10]}}</td><td>{{d.dni}}</td><td>{{d.trabajador or '-'}}</td><td>{{d.tipo}}</td><td>{{d.periodo or '-'}}</td><td>{{d.archivo_nombre}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin datos.</td></tr>{% endfor %}</tbody></table></div></div></div></div>'''
    return render_page(body,docs=docs,desde=desde,hasta=hasta,buscar=buscar,total=total,dni_unicos=dni_unicos,leidos=leidos,tipos=tipos,title='Reportes boletas')

@login_required
def boletas_exportar_291():
    _boleta_init_291()
    desde=request.args.get('desde') or ''
    hasta=request.args.get('hasta') or ''
    buscar=request.args.get('buscar') or ''
    rows=_bt_query_docs_291(desde,hasta,None,buscar,5000)
    headers=['fecha_subida','dni','trabajador','empresa','area','cargo','tipo','periodo','detalle','archivo_nombre','estado','fecha_lectura','uploaded_by','observacion']
    return excel_response(headers, rows, 'reporte_boletas.xlsx', 'BOLETAS')

@login_required
def boletas_plantilla_291():
    wb=Workbook(); ws=wb.active; ws.title='BOLETAS'
    headers=['DNI','TIPO','PERIODO','DETALLE','OBSERVACION','ARCHIVO']
    ws.append(headers); ws.append(['74324033','NORMAL',datetime.now().strftime('%Y-%m'),'Boleta mensual','','BOLETA_74324033.pdf'])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=max(14,min(32,max(len(str(c.value or '')) for c in col)+2))
    out=BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name='plantilla_control_boletas.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def home_291():
    if not session.get('usuario'): return redirect(url_for('inicio'))
    hojas = rows_to_dict(execute('SELECT * FROM hojas_tareo ORDER BY fecha DESC, id DESC LIMIT 8', fetchall=True))
    body = """
    <div class="desktop-grid"><div class="phone-wrap"><div class="green-hero" style="min-height:220px"><div class="green-top"><a class="text-white text-decoration-none" href="{{url_for('soporte')}}"><i class="bi bi-headset"></i> Soporte</a>{% if session.get('rol')=='admin' %}<a class="text-white text-decoration-none" href="{{url_for('configuraciones')}}"><i class="bi bi-gear"></i> Config.</a>{% else %}<span></span>{% endif %}</div><div class="avatar"><i class="bi bi-person-circle"></i></div><div class="login-name">{{ session.get('nombres','USUARIO') }}</div><div class="white-input mt-3"></div></div><div class="top-actions"><a class="tile text-decoration-none" href="{{url_for('hojas_tareo')}}"><i class="bi bi-list-check"></i>TAREO</a><a class="tile text-decoration-none" href="{{url_for('asistencia_modulo')}}"><i class="bi bi-fingerprint"></i>ASIST.</a><a class="tile text-decoration-none" href="{{url_for('documentos_firma')}}"><i class="bi bi-pen"></i>FIRMAS</a><a class="tile text-decoration-none" href="{{url_for('transporte')}}"><i class="bi bi-bus-front"></i>TRANSP.</a><a class="tile text-decoration-none" href="{{url_for('contratacion_home')}}"><i class="bi bi-person-plus"></i>CONTRAT.</a><a class="tile text-decoration-none" href="{{url_for('boletas_home')}}"><i class="bi bi-file-earmark-text"></i>BOLETA</a><a class="tile text-decoration-none" href="{{url_for('reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i>REPORTES<br>TAREO</a><a class="tile text-decoration-none" href="{{url_for('sincronizacion')}}"><i class="bi bi-arrow-repeat"></i>SINC.</a></div><div class="leaf"></div><div class="bottom-sync"><i class="bi bi-arrow-repeat"></i> Sincronizar Tablas Maestras<br>Actualizado hasta: {{ now }}</div><a href="{{url_for('logout')}}" class="bottom-out"><i class="bi bi-box-arrow-right"></i></a></div><div class="desk-panel"><h1 class="header-title">TAREO MÓVIL – GRUPO DE COSECHA</h1><div class="card-pro p-4 mb-3"><div class="d-flex justify-content-between align-items-center"><div><h4 class="fw-bold text-success mb-1">Hojas recientes</h4><div class="text-muted small">Crea una hoja y registra labores, trabajadores y avances.</div></div><a class="btn btn-green" href="{{url_for('crear_hoja')}}"><i class="bi bi-plus-lg"></i> Crear hoja</a></div></div><div class="card-pro p-3"><div class="table-responsive"><table class="table list-table"><thead><tr><th>Fecha</th><th>Grupo</th><th>Subgrupo</th><th>Labor</th><th>Responsable</th><th>Estado</th><th></th></tr></thead><tbody>{% for h in hojas %}<tr><td>{{h.fecha}}</td><td>{{h.grupo}}</td><td>{{h.subgrupo}}</td><td>{{h.labor}}</td><td>{{h.responsable}}</td><td><span class="status-pill">{{h.estado}}</span></td><td><a class="btn btn-sm btn-green" href="{{url_for('detalle_hoja', hoja_id=h.id)}}">Abrir</a></td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted py-4">Sin hojas creadas.</td></tr>{% endfor %}</tbody></table></div></div></div></div>"""
    return render_page(body, hojas=hojas, now=now_str())

try:
    app.add_url_rule('/boletas', 'boletas_home', boletas_home_291, methods=['GET'])
    app.add_url_rule('/boletas/subir', 'boletas_subir', boletas_subir_291, methods=['GET','POST'])
    app.add_url_rule('/boletas/listar', 'boletas_listar', boletas_listar_291, methods=['GET'])
    app.add_url_rule('/boletas/listar/<tipo>', 'boletas_listar', boletas_listar_291, methods=['GET'])
    app.add_url_rule('/boletas/archivo/<int:doc_id>', 'boletas_archivo', boletas_archivo_291, methods=['GET'])
    app.add_url_rule('/boletas/detectar', 'boletas_detectar', boletas_detectar_291, methods=['GET'])
    app.add_url_rule('/boletas/reportes', 'boletas_reportes', boletas_reportes_291, methods=['GET'])
    app.add_url_rule('/boletas/exportar', 'boletas_exportar', boletas_exportar_291, methods=['GET'])
    app.add_url_rule('/boletas/plantilla', 'boletas_plantilla', boletas_plantilla_291, methods=['GET'])
except Exception as e:
    print('Rutas boletas 291:', e)
app.view_functions['home'] = home_291
# ======================= FIN PATCH BOLETAS OMAR 291 =======================

# ========================= PATCH OMAR 292: PORTAL RRHH INTEGRADO =========================
# Integra: Boletas Admin/Usuario, Solicitud de vacaciones, Renovación,
# configuración por módulo y firma de contratación con cámara/firma.


def _is_admin_292():
    return str(session.get('rol') or '').lower() == 'admin'


def _user_dni_292():
    d = session.get('dni') or session.get('usuario') or ''
    return limpiar_dni(d)


def _deny_admin_292(msg='Opción disponible solo para ADMINISTRADOR.'):
    flash(msg, 'danger')
    return redirect(url_for('home'))


def _rrhh_css_292():
    return r'''
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}.shell{max-width:430px!important;width:100%!important;margin:0 auto!important;padding:6px 8px 26px!important;background:#fff!important}.phone-wrap{max-width:390px!important;margin:0 auto!important}.rr292-app{background:#fff;border:1px solid #e4e8e4;border-radius:13px;overflow:hidden;box-shadow:0 10px 24px rgba(0,0,0,.07)}
      .rr292-head{height:78px;background:#25773a;color:#fff;position:relative;display:flex;align-items:center;justify-content:center}.rr292-head .back{position:absolute;left:13px;top:50%;transform:translateY(-50%);font-size:31px;color:#fff!important;text-decoration:none;line-height:1}.rr292-head .cfg{position:absolute;right:10px;top:14px;color:#fff!important;text-decoration:none;font-size:12px;font-weight:950;border:1px solid rgba(255,255,255,.65);border-radius:12px;padding:5px 8px}.rr292-head .ttl{font-size:17px;font-weight:950;text-align:center;color:#fff;line-height:1.1}.rr292-body{padding:13px 13px 16px;background:#fff}.rr292-section{font-size:14px;font-weight:950;color:#08713b;text-transform:uppercase;margin:8px 0 9px;letter-spacing:.2px}.rr292-grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:9px}.rr292-grid2{display:grid;grid-template-columns:repeat(2,1fr);gap:9px}.rr292-tile{height:78px;background:#08713b;border-radius:10px;color:#fff!important;text-decoration:none;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;box-shadow:0 7px 13px rgba(0,0,0,.12);padding:5px}.rr292-tile i{font-size:24px;color:#fff;margin-bottom:4px;line-height:1}.rr292-tile .lbl{font-size:11px;font-weight:950;line-height:1.04;color:#fff}.rr292-tile .sub{font-size:8.5px;font-weight:850;color:#eaffee;margin-top:3px;line-height:1.05}.rr292-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:0 0 12px}.rr292-kpi{background:#10964e;color:#fff;border-radius:9px;text-align:center;padding:8px 4px;box-shadow:0 6px 12px rgba(16,150,78,.16)}.rr292-kpi small{display:block;font-size:9.5px;font-weight:950;line-height:1.05}.rr292-kpi b{display:block;font-size:21px;line-height:1.05;font-weight:950;margin-top:4px;color:#fff}.rr292-form{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:11px;margin-bottom:11px}.rr292-form label{font-size:11px;font-weight:950;color:#176a35;margin-bottom:4px}.rr292-form .form-control,.rr292-form .form-select{height:38px!important;border-radius:9px!important;font-size:12px!important;font-weight:850}.rr292-btn{height:41px;border-radius:10px;background:#08713b;border:1px solid #08713b;color:#fff!important;font-weight:950;font-size:13px;width:100%;display:flex;align-items:center;justify-content:center;gap:6px;text-decoration:none}.rr292-outline{height:40px;border-radius:10px;background:#fff;border:1px solid #08713b;color:#08713b!important;font-weight:900;font-size:12px;width:100%;display:flex;align-items:center;justify-content:center;gap:6px;text-decoration:none}.rr292-list{border:1px solid #e5e7eb;border-radius:10px;overflow:auto;background:#fff;scrollbar-color:#08713b #e5e7eb}.rr292-table{width:100%;min-width:540px;border-collapse:collapse}.rr292-table th{background:#f8fafc;color:#12223b;font-size:11px;font-weight:950;padding:8px;border-bottom:1px solid #e5e7eb}.rr292-table td{font-size:11px;color:#334155;padding:8px;border-bottom:1px solid #f1f5f9;font-weight:750}.rr292-info{display:grid;grid-template-columns:20px 1fr;gap:8px;border:1px solid #b8d7ff;background:#eef6ff;border-radius:12px;padding:10px;margin:10px 0;color:#0b2e83;font-size:11px;font-weight:900;line-height:1.35}.rr292-info i{font-size:17px;color:#0b5ed7}.rr292-rowlink{display:grid;grid-template-columns:32px 1fr 18px;gap:10px;align-items:center;text-decoration:none;color:#102a43!important;border:1px solid #dfe7df;background:#fff;border-radius:11px;padding:12px 10px;margin:7px 0}.rr292-rowlink i{font-size:22px;color:#08713b}.rr292-rowlink b{font-size:12.5px;color:#0a1f44}.rr292-rowlink small{font-size:10px;color:#486156;font-weight:800}.rr292-badge{display:inline-block;border-radius:999px;padding:3px 7px;font-size:9px;font-weight:950;background:#dcfce7;color:#166534}.rr292-danger{background:#fee2e2!important;color:#991b1b!important}.rr292-canvas{height:170px;border:2px dashed #86c995;border-radius:12px;background:#fff;touch-action:none;width:100%}.rr292-video{width:100%;height:180px;background:#f1f5f9;border:1px dashed #86c995;border-radius:12px;object-fit:cover}.rr292-photo{max-width:100%;border-radius:10px;border:1px solid #d7eadc;display:none}.top-actions{justify-content:center!important}.top-actions .tile{width:74px!important;height:70px!important;font-size:9px!important}.top-actions .tile i{font-size:25px!important;margin-bottom:4px!important}.green-hero{min-height:210px!important}.avatar{width:76px!important;height:76px!important;font-size:41px!important}.white-input{height:34px!important}
    </style>
    '''


def _ensure_rrhh_292():
    _boleta_init_291()
    _contratacion_init_290()
    idtype = 'SERIAL PRIMARY KEY' if is_pg() else 'INTEGER PRIMARY KEY AUTOINCREMENT'
    execute(f'''CREATE TABLE IF NOT EXISTS vacaciones_saldos(
        id {idtype}, dni TEXT, trabajador TEXT, empresa TEXT, area TEXT, jefe TEXT,
        periodo_inicio TEXT, periodo_fin TEXT, dias_ganados REAL DEFAULT 0,
        dias_gozados REAL DEFAULT 0, saldo REAL DEFAULT 0, fecha_carga TEXT, uploaded_by TEXT)''', commit=True)
    execute(f'''CREATE TABLE IF NOT EXISTS vacaciones_solicitudes(
        id {idtype}, dni TEXT, trabajador TEXT, fecha_inicio TEXT, fecha_fin TEXT,
        dias REAL DEFAULT 0, motivo TEXT, estado TEXT DEFAULT 'PENDIENTE', fecha_solicitud TEXT,
        comentario_jefe TEXT, comentario_gh TEXT, creado_por TEXT)''', commit=True)
    execute(f'''CREATE TABLE IF NOT EXISTS renovaciones(
        id {idtype}, dni TEXT, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        contrato_actual_inicio TEXT, contrato_actual_fin TEXT, nuevo_inicio TEXT, nuevo_fin TEXT,
        tipo_contrato TEXT, basico REAL DEFAULT 0, estado TEXT DEFAULT 'PENDIENTE', observacion TEXT,
        creado_por TEXT, creado_en TEXT, actualizado_en TEXT)''', commit=True)
    execute(f'''CREATE TABLE IF NOT EXISTS contratacion_firmas_bio(
        id {idtype}, dni TEXT, trabajador TEXT, documento TEXT, metodo TEXT,
        firma_path TEXT, foto_path TEXT, fecha TEXT, hora TEXT, fecha_hora TEXT,
        registrado_por TEXT, observacion TEXT)''', commit=True)
    execute(f'''CREATE TABLE IF NOT EXISTS contratacion_maestros(
        id {idtype}, tipo TEXT, valor TEXT, estado TEXT DEFAULT 'ACTIVO', creado_en TEXT, creado_por TEXT)''', commit=True)


def _find_worker_any_292(dni):
    dni = limpiar_dni(dni)
    r = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    if not r:
        return {'dni': dni, 'trabajador': '', 'empresa':'', 'area':'', 'cargo':'', 'actividad':''}
    return {'dni': r.get('dni') or dni, 'trabajador': r.get('trabajador') or r.get('nombre') or '', 'empresa': r.get('empresa') or '', 'area': r.get('area') or '', 'cargo': r.get('cargo') or '', 'actividad': r.get('actividad') or ''}


def login_292():
    if request.method == 'POST':
        usuario = (request.form.get('usuario') or '').strip()
        password = request.form.get('password') or ''
        modo = request.form.get('modo') or 'usuario'
        row = row_to_dict(execute('SELECT * FROM usuarios WHERE usuario=?', (usuario,), fetchone=True))
        if row and row.get('estado','ACTIVO') == 'ACTIVO' and check_password_hash(row['password_hash'], password):
            if modo == 'admin' and row.get('rol') != 'admin':
                flash('Este acceso es solo para administradores.', 'danger'); return redirect(url_for('login'))
            session['usuario']=usuario; session['rol']=row.get('rol','operador'); session['nombres']=row.get('nombres') or usuario; session['dni']=limpiar_dni(usuario)
            return redirect(url_for('home'))
        # Acceso usuario trabajador por DNI. Claves simples demo: DNI, últimos 4 o 1234.
        dni = limpiar_dni(usuario)
        if modo == 'usuario' and len(dni) == 8:
            w = _find_worker_any_292(dni)
            if w and w.get('dni') and password in (dni, dni[-4:], '1234'):
                session['usuario']=dni; session['rol']='usuario'; session['nombres']=w.get('trabajador') or dni; session['dni']=dni
                return redirect(url_for('home'))
        flash('Usuario/clave incorrecta o usuario inactivo.', 'danger')
    body = r'''
    <div class="phone-wrap">
      <div class="green-hero" style="min-height:245px;border-radius:0 0 22px 22px">
        <div class="green-top"><span><i class="bi bi-headset"></i> Soporte</span><span><i class="bi bi-gear"></i> Config.</span></div>
        <div class="splash-logo" style="width:96px;height:96px;font-size:42px;margin:14px auto 6px"><i class="bi bi-clipboard2-data"></i></div>
        <div class="splash-title">TAREO MOVIL</div><div class="login-name">INICIAR SESIÓN</div>
      </div>
      <form method="post" class="login-form"><div class="floating-card">
        <div class="role-toggle mb-2"><label><input type="radio" name="modo" value="usuario" checked><span>USUARIO</span></label><label><input type="radio" name="modo" value="admin"><span>ADMINISTRADOR</span></label></div>
        <input class="form-control white-input mb-2" name="usuario" required autofocus placeholder="Usuario / DNI">
        <input class="form-control white-input mb-3" name="password" type="password" required placeholder="Clave / PIN">
        <button class="btn btn-green w-100"><i class="bi bi-box-arrow-in-right me-1"></i> INGRESAR</button>
        <div class="text-center small mt-2 text-muted">Admin demo: admin / admin123 · Usuario demo: DNI / últimos 4</div>
      </div></form>
      <div class="d-flex justify-content-center gap-3 mt-4"><div class="tile"><i class="bi bi-list-check"></i>TAREO</div><div class="tile"><i class="bi bi-file-earmark-text"></i>BOLETA</div><div class="tile"><i class="bi bi-calendar-check"></i>VACAC.</div></div>
      <div class="leaf"></div>
    </div>'''
    return render_page(body, title='Login Tareo Móvil', now=now_str())


def home_292():
    if not session.get('usuario'):
        return redirect(url_for('inicio'))
    hojas = rows_to_dict(execute('SELECT * FROM hojas_tareo ORDER BY fecha DESC, id DESC LIMIT 8', fetchall=True))
    is_admin = _is_admin_292()
    body = _rrhh_css_292() + r'''
    <div class="desktop-grid"><div class="phone-wrap"><div class="green-hero" style="min-height:220px"><div class="green-top"><a class="text-white text-decoration-none" href="{{url_for('soporte')}}"><i class="bi bi-headset"></i> Soporte</a>{% if is_admin %}<a class="text-white text-decoration-none" href="{{url_for('configuraciones')}}"><i class="bi bi-gear"></i> Config.</a>{% else %}<span></span>{% endif %}</div><div class="avatar"><i class="bi bi-person-circle"></i></div><div class="login-name">{{ session.get('nombres','USUARIO') }}</div><div class="white-input mt-3"></div></div>
      <div class="top-actions">
        <a class="tile text-decoration-none" href="{{url_for('hojas_tareo')}}"><i class="bi bi-list-check"></i>TAREO</a>
        <a class="tile text-decoration-none" href="{{url_for('asistencia_modulo')}}"><i class="bi bi-fingerprint"></i>ASIST.</a>
        <a class="tile text-decoration-none" href="{{url_for('documentos_firma')}}"><i class="bi bi-pen"></i>FIRMAS</a>
        {% if is_admin %}<a class="tile text-decoration-none" href="{{url_for('transporte')}}"><i class="bi bi-bus-front"></i>TRANSP.</a>{% endif %}
        {% if is_admin %}<a class="tile text-decoration-none" href="{{url_for('contratacion_home')}}"><i class="bi bi-person-plus"></i>CONTRAT.</a>{% endif %}
        <a class="tile text-decoration-none" href="{{url_for('boletas_home')}}"><i class="bi bi-file-earmark-text"></i>BOLETA</a>
        <a class="tile text-decoration-none" href="{{url_for('vacaciones_home')}}"><i class="bi bi-calendar-check"></i>VACAC.</a>
        <a class="tile text-decoration-none" href="{{url_for('renovacion_home')}}"><i class="bi bi-arrow-repeat"></i>RENOV.</a>
        {% if is_admin %}<a class="tile text-decoration-none" href="{{url_for('reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i>REPORTES<br>TAREO</a>{% endif %}
        {% if is_admin %}<a class="tile text-decoration-none" href="{{url_for('sincronizacion')}}"><i class="bi bi-arrow-repeat"></i>SINC.</a>{% endif %}
      </div><div class="leaf"></div><div class="bottom-sync"><i class="bi bi-arrow-repeat"></i> Actualizado hasta: {{ now }}</div><a href="{{url_for('logout')}}" class="bottom-out"><i class="bi bi-box-arrow-right"></i></a></div>
      <div class="desk-panel"><h1 class="header-title">TAREO MÓVIL – GRUPO DE COSECHA</h1><div class="card-pro p-4 mb-3"><h4 class="fw-bold text-success mb-1">Panel integrado RR.HH.</h4><div class="text-muted small">Administrador: acceso total. Usuario: boletas, vacaciones, renovaciones y firmas asignadas.</div></div></div></div>'''
    return render_page(body, hojas=hojas, now=now_str(), is_admin=is_admin)


# ========================= BOLETAS 292 =========================
def _bt_query_docs_role_292(desde=None, hasta=None, tipo=None, buscar=None, limit=300):
    if _is_admin_292():
        return _bt_query_docs_291(desde,hasta,tipo,buscar,limit)
    dni = _user_dni_292()
    b = buscar or dni
    rows = _bt_query_docs_291(desde,hasta,tipo,b,limit)
    return [r for r in rows if limpiar_dni(r.get('dni')) == dni]


@login_required
def boletas_home_292():
    _boleta_init_291()
    if _is_admin_292():
        total=int(scalar('SELECT COUNT(*) AS c FROM boleta_documentos') or 0)
        cargadas_hoy=int(scalar('SELECT COUNT(*) AS c FROM boleta_documentos WHERE substr(fecha_subida,1,10)=?', (today_str(),)) or 0)
        trabajadores=int(scalar('SELECT COUNT(DISTINCT dni) AS c FROM boleta_documentos') or 0)
        rows=rows_to_dict(execute('SELECT tipo, COUNT(*) AS c FROM boleta_documentos GROUP BY tipo ORDER BY c DESC', fetchall=True)); conteos={r.get('tipo'):r.get('c') for r in rows}
    else:
        dni=_user_dni_292(); total=int(scalar('SELECT COUNT(*) AS c FROM boleta_documentos WHERE dni=?', (dni,)) or 0); cargadas_hoy=int(scalar('SELECT COUNT(*) AS c FROM boleta_documentos WHERE dni=? AND substr(fecha_subida,1,10)=?', (dni,today_str())) or 0); trabajadores=1 if dni else 0; conteos={}
    body=_rrhh_css_292()+r'''
    <div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>{% if is_admin %}<a class="cfg" href="{{url_for('boletas_config')}}"><i class="bi bi-gear"></i> Config.</a>{% endif %}<div class="ttl">Módulo Boletas</div></div><div class="rr292-body">
      <div class="rr292-kpis"><div class="rr292-kpi"><small>Total</small><b>{{total}}</b></div><div class="rr292-kpi"><small>Hoy</small><b>{{cargadas_hoy}}</b></div><div class="rr292-kpi"><small>DNI</small><b>{{trabajadores}}</b></div></div>
      {% if is_admin %}<div class="rr292-info"><i class="bi bi-info-circle-fill"></i><div>El administrador carga boletas en PDF por DNI. El usuario ingresa con su DNI/PIN y solo visualiza sus boletas.</div></div>{% else %}<div class="rr292-info"><i class="bi bi-info-circle-fill"></i><div>Visualización limitada: solo documentos vinculados al DNI {{dni}}.</div></div>{% endif %}
      <div class="rr292-section">Documentos de pago</div><div class="rr292-grid3">{% for t,lbl,ico in tipos %}<a class="rr292-tile" href="{{url_for('boletas_listar', tipo=t)}}"><i class="bi {{ico}}"></i><span class="lbl">{{t}}</span><span class="sub">{{conteos.get(t,'')}}</span></a>{% endfor %}</div>
      <div class="rr292-section" style="margin-top:16px">Operación</div><div class="rr292-grid3">{% if is_admin %}<a class="rr292-tile" href="{{url_for('boletas_subir')}}"><i class="bi bi-cloud-arrow-up"></i><span class="lbl">Subir PDF</span><span class="sub">Admin</span></a><a class="rr292-tile" href="{{url_for('boletas_detectar')}}"><i class="bi bi-search"></i><span class="lbl">Detectar</span><span class="sub">DNI</span></a><a class="rr292-tile" href="{{url_for('boletas_reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i><span class="lbl">Reportes</span><span class="sub">Excel</span></a>{% else %}<a class="rr292-tile" href="{{url_for('boletas_listar')}}"><i class="bi bi-folder2-open"></i><span class="lbl">Mis boletas</span><span class="sub">Ver/descargar</span></a>{% endif %}</div>
    </div></div></div>'''
    return render_page(body,total=total,cargadas_hoy=cargadas_hoy,trabajadores=trabajadores,tipos=BOLETA_TIPOS_CANON_291,conteos=conteos,is_admin=_is_admin_292(),dni=_user_dni_292(),title='Módulo Boletas')


@login_required
def boletas_subir_292():
    if not _is_admin_292(): return _deny_admin_292()
    _boleta_init_291()
    if request.method == 'POST':
        dni=limpiar_dni(request.form.get('dni')); tipo=_bt_norm_tipo_291(request.form.get('tipo')); periodo=limpiar_texto(request.form.get('periodo') or datetime.now().strftime('%Y-%m'), upper=False); detalle=limpiar_texto(request.form.get('detalle'), upper=False); obs=limpiar_texto(request.form.get('observacion'), upper=False); f=request.files.get('archivo')
        if len(dni)!=8: flash('Ingrese DNI válido de 8 dígitos.', 'danger'); return redirect(url_for('boletas_subir'))
        if not f or not f.filename: flash('Seleccione archivo PDF de boleta.', 'danger'); return redirect(url_for('boletas_subir'))
        if not f.filename.lower().endswith('.pdf'): flash('Las boletas deben cargarse en PDF.', 'danger'); return redirect(url_for('boletas_subir'))
        try:
            path, original, ext = _bt_save_file_291(f, dni, tipo, periodo); _bt_insert_doc_291(dni, tipo, periodo, detalle, obs, path, original, ext, session.get('usuario'))
            flash('Boleta PDF cargada correctamente.', 'success'); return redirect(url_for('boletas_listar', tipo=tipo, buscar=dni))
        except Exception as e:
            flash(str(e), 'danger'); return redirect(url_for('boletas_subir'))
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('boletas_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Subir boleta PDF</div></div><div class="rr292-body"><form class="rr292-form" method="post" enctype="multipart/form-data"><div class="row"><div class="col-6"><label>DNI</label><input class="form-control" name="dni" maxlength="8" inputmode="numeric" required autofocus></div><div class="col-6"><label>Tipo</label><select class="form-select" name="tipo">{{tipo_opts|safe}}</select></div><div class="col-6"><label>Periodo</label><input class="form-control" name="periodo" value="{{periodo}}" placeholder="2026-06"></div><div class="col-6"><label>PDF</label><input class="form-control" name="archivo" type="file" accept="application/pdf,.pdf" required></div><div class="col-12"><label>Detalle</label><input class="form-control" name="detalle" placeholder="Boleta mensual / CTS / grati..."></div><div class="col-12"><label>Observación</label><input class="form-control" name="observacion"></div></div><button class="rr292-btn mt-2"><i class="bi bi-upload"></i> Guardar boleta</button></form><a class="rr292-outline" href="{{url_for('boletas_plantilla')}}"><i class="bi bi-file-earmark-excel"></i> Descargar plantilla control</a></div></div></div>'''
    return render_page(body,tipo_opts=_bt_tipo_options_291(),periodo=datetime.now().strftime('%Y-%m'),title='Subir boleta')


@login_required
def boletas_listar_292(tipo=None):
    _boleta_init_291(); desde=request.args.get('desde') or ''; hasta=request.args.get('hasta') or ''; buscar=request.args.get('buscar') or ''
    docs=_bt_query_docs_role_292(desde,hasta,tipo,buscar,300)
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('boletas_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Boletas {{tipo or ''}}</div></div><div class="rr292-body"><form class="rr292-form" method="get"><div class="row"><div class="col-6"><label>Desde</label><input type="date" name="desde" value="{{desde}}" class="form-control"></div><div class="col-6"><label>Hasta</label><input type="date" name="hasta" value="{{hasta}}" class="form-control"></div>{% if is_admin %}<div class="col-12"><label>Buscar</label><input name="buscar" value="{{buscar}}" class="form-control" placeholder="DNI / trabajador / periodo"></div>{% endif %}<div class="col-6"><button class="rr292-btn"><i class="bi bi-search"></i> Buscar</button></div>{% if is_admin %}<div class="col-6"><a class="rr292-outline" href="{{url_for('boletas_subir')}}"><i class="bi bi-plus-circle"></i> Subir</a></div>{% endif %}</div></form><div class="rr292-info"><i class="bi bi-file-earmark-text"></i><div>Documentos encontrados: <b>{{docs|length}}</b></div></div>{% for d in docs %}<a class="rr292-rowlink" href="{{url_for('boletas_archivo', doc_id=d.id)}}" target="_blank"><i class="bi bi-filetype-pdf"></i><div><b>{{d.tipo}} · {{d.periodo or '-'}}</b><br><small>{{d.dni}} · {{d.trabajador or 'NO ENCONTRADO'}}<br>{{d.archivo_nombre}}</small></div><i class="bi bi-chevron-right"></i></a>{% else %}<div class="rr292-form text-center text-muted">Sin boletas para el filtro.</div>{% endfor %}</div></div></div>'''
    return render_page(body,docs=docs,desde=desde,hasta=hasta,buscar=buscar,tipo=tipo or '',is_admin=_is_admin_292(),title='Boletas')


@login_required
def boletas_reportes_292():
    if not _is_admin_292(): return _deny_admin_292()
    return boletas_reportes_291()


@login_required
def boletas_config_292():
    if not _is_admin_292(): return _deny_admin_292()
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('boletas_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Config. Boletas</div></div><div class="rr292-body"><div class="rr292-section">Configuraciones</div><a class="rr292-rowlink" href="{{url_for('boletas_subir')}}"><i class="bi bi-cloud-arrow-up"></i><div><b>Carga individual PDF</b><br><small>Subir boleta por DNI y periodo</small></div><i class="bi bi-chevron-right"></i></a><a class="rr292-rowlink" href="{{url_for('boletas_detectar')}}"><i class="bi bi-search"></i><div><b>Detección automática</b><br><small>Carpeta BOLETAS_UPLOAD_AUTO por DNI</small></div><i class="bi bi-chevron-right"></i></a><a class="rr292-rowlink" href="{{url_for('boletas_plantilla')}}"><i class="bi bi-file-earmark-excel"></i><div><b>Plantilla de control</b><br><small>Excel para seguimiento</small></div><i class="bi bi-chevron-right"></i></a><a class="rr292-rowlink" href="{{url_for('boletas_reportes')}}"><i class="bi bi-bar-chart"></i><div><b>Reportes</b><br><small>Control de lectura y carga</small></div><i class="bi bi-chevron-right"></i></a></div></div></div>'''
    return render_page(body,title='Config. Boletas')


# ========================= VACACIONES 292 =========================
@login_required
def vacaciones_home_292():
    _ensure_rrhh_292(); dni=_user_dni_292(); is_admin=_is_admin_292()
    if request.method == 'POST' and not is_admin:
        fi=request.form.get('fecha_inicio') or ''; ff=request.form.get('fecha_fin') or ''; dias=float(request.form.get('dias') or 0); motivo=limpiar_texto(request.form.get('motivo'), upper=False)
        w=_find_worker_any_292(dni); execute('''INSERT INTO vacaciones_solicitudes(dni,trabajador,fecha_inicio,fecha_fin,dias,motivo,estado,fecha_solicitud,creado_por) VALUES(?,?,?,?,?,?,?,?,?)''', (dni,w.get('trabajador'),fi,ff,dias,motivo,'PENDIENTE',now_str(),session.get('usuario')), commit=True)
        flash('Solicitud de vacaciones registrada.', 'success'); return redirect(url_for('vacaciones_home'))
    if is_admin:
        solicitudes=rows_to_dict(execute('SELECT * FROM vacaciones_solicitudes ORDER BY id DESC LIMIT 80', fetchall=True)); total=len(solicitudes); pend=sum(1 for s in solicitudes if (s.get('estado') or '').upper()=='PENDIENTE'); aprob=sum(1 for s in solicitudes if (s.get('estado') or '').upper()=='APROBADO')
    else:
        solicitudes=rows_to_dict(execute('SELECT * FROM vacaciones_solicitudes WHERE dni=? ORDER BY id DESC LIMIT 40', (dni,), fetchall=True)); total=len(solicitudes); pend=sum(1 for s in solicitudes if (s.get('estado') or '').upper()=='PENDIENTE'); aprob=sum(1 for s in solicitudes if (s.get('estado') or '').upper()=='APROBADO')
    saldo=float(scalar('SELECT COALESCE(SUM(saldo),0) AS s FROM vacaciones_saldos WHERE dni=?', (dni,)) or 0) if not is_admin else float(scalar('SELECT COALESCE(SUM(saldo),0) AS s FROM vacaciones_saldos') or 0)
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>{% if is_admin %}<a class="cfg" href="{{url_for('vacaciones_config')}}"><i class="bi bi-gear"></i> Config.</a>{% endif %}<div class="ttl">Solicitud vacaciones</div></div><div class="rr292-body"><div class="rr292-kpis"><div class="rr292-kpi"><small>Saldo</small><b>{{saldo|round(1)}}</b></div><div class="rr292-kpi"><small>Pend.</small><b>{{pend}}</b></div><div class="rr292-kpi"><small>Aprob.</small><b>{{aprob}}</b></div></div>{% if not is_admin %}<form method="post" class="rr292-form"><div class="row"><div class="col-6"><label>Inicio</label><input type="date" name="fecha_inicio" class="form-control" required></div><div class="col-6"><label>Fin</label><input type="date" name="fecha_fin" class="form-control" required></div><div class="col-4"><label>Días</label><input type="number" step="0.5" name="dias" class="form-control" required></div><div class="col-8"><label>Motivo</label><input name="motivo" class="form-control" placeholder="Vacaciones"></div></div><button class="rr292-btn mt-2"><i class="bi bi-send"></i> Solicitar vacaciones</button></form>{% else %}<div class="rr292-grid3 mb-2"><a class="rr292-tile" href="{{url_for('vacaciones_config')}}"><i class="bi bi-upload"></i><span class="lbl">Saldos</span><span class="sub">Carga masiva</span></a><a class="rr292-tile" href="{{url_for('vacaciones_exportar')}}"><i class="bi bi-file-earmark-excel"></i><span class="lbl">Exportar</span><span class="sub">Solicitudes</span></a><a class="rr292-tile" href="{{url_for('vacaciones_home')}}"><i class="bi bi-check2-square"></i><span class="lbl">Aprobar</span><span class="sub">Estados</span></a></div>{% endif %}<div class="rr292-section">Solicitudes</div><div class="rr292-list"><table class="rr292-table"><thead><tr><th>DNI</th><th>Trabajador</th><th>Inicio</th><th>Fin</th><th>Días</th><th>Estado</th>{% if is_admin %}<th>Acción</th>{% endif %}</tr></thead><tbody>{% for s in solicitudes %}<tr><td>{{s.dni}}</td><td>{{s.trabajador or '-'}}</td><td>{{s.fecha_inicio}}</td><td>{{s.fecha_fin}}</td><td>{{s.dias}}</td><td><span class="rr292-badge">{{s.estado}}</span></td>{% if is_admin %}<td><a href="{{url_for('vacaciones_estado', sol_id=s.id, estado='APROBADO')}}">Aprobar</a> · <a href="{{url_for('vacaciones_estado', sol_id=s.id, estado='RECHAZADO')}}">Rechazar</a></td>{% endif %}</tr>{% else %}<tr><td colspan="7" class="text-center text-muted">Sin solicitudes.</td></tr>{% endfor %}</tbody></table></div></div></div></div>'''
    return render_page(body,solicitudes=solicitudes,saldo=saldo,pend=pend,aprob=aprob,is_admin=is_admin,title='Vacaciones')


@login_required
def vacaciones_estado_292(sol_id, estado):
    if not _is_admin_292(): return _deny_admin_292()
    estado = estado if estado in ('APROBADO','RECHAZADO','PENDIENTE') else 'PENDIENTE'
    execute('UPDATE vacaciones_solicitudes SET estado=?, comentario_gh=? WHERE id=?', (estado, f'Actualizado por {session.get("usuario")} {now_str()}', sol_id), commit=True)
    flash('Estado actualizado.', 'success'); return redirect(url_for('vacaciones_home'))


@login_required
def vacaciones_config_292():
    if not _is_admin_292(): return _deny_admin_292()
    _ensure_rrhh_292(); cargados=0
    if request.method=='POST':
        f=request.files.get('archivo')
        if not f or not f.filename: flash('Seleccione Excel de saldos.', 'danger'); return redirect(url_for('vacaciones_config'))
        try:
            for row in _iter_excel_upload(f):
                dni=limpiar_dni(_valor(row,['DNI','DOCUMENTO']));
                if len(dni)!=8: continue
                w=_find_worker_any_292(dni); saldo=float(_valor(row,['SALDO','DIAS SALDO','DÍAS SALDO']) or 0); ganados=float(_valor(row,['DIAS GANADOS','DÍAS GANADOS']) or saldo); gozados=float(_valor(row,['DIAS GOZADOS','DÍAS GOZADOS']) or 0)
                execute('INSERT INTO vacaciones_saldos(dni,trabajador,empresa,area,jefe,periodo_inicio,periodo_fin,dias_ganados,dias_gozados,saldo,fecha_carga,uploaded_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)', (dni,w.get('trabajador'),w.get('empresa'),w.get('area'),'',_valor(row,['I_PERIODO','PERIODO INICIO']) or '',_valor(row,['F_PERIODO','PERIODO FIN']) or '',ganados,gozados,saldo,now_str(),session.get('usuario')), commit=True); cargados+=1
            flash(f'Saldos cargados: {cargados}.', 'success')
        except Exception as e: flash(str(e),'danger')
        return redirect(url_for('vacaciones_config'))
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('vacaciones_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Config. vacaciones</div></div><div class="rr292-body"><form method="post" enctype="multipart/form-data" class="rr292-form"><label>Carga masiva de saldos Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="field-help mt-1">Columnas: DNI, TRABAJADOR, I_PERIODO, F_PERIODO, DIAS GANADOS, DIAS GOZADOS, SALDO.</div><button class="rr292-btn mt-2"><i class="bi bi-upload"></i> Cargar saldos</button></form><a class="rr292-rowlink" href="{{url_for('vacaciones_exportar')}}"><i class="bi bi-file-earmark-excel"></i><div><b>Exportar solicitudes</b><br><small>Reporte de vacaciones</small></div><i class="bi bi-chevron-right"></i></a></div></div></div>'''
    return render_page(body,title='Config. vacaciones')


@login_required
def vacaciones_exportar_292():
    _ensure_rrhh_292()
    if _is_admin_292(): rows=rows_to_dict(execute('SELECT * FROM vacaciones_solicitudes ORDER BY id DESC', fetchall=True))
    else: rows=rows_to_dict(execute('SELECT * FROM vacaciones_solicitudes WHERE dni=? ORDER BY id DESC', (_user_dni_292(),), fetchall=True))
    headers=['id','dni','trabajador','fecha_inicio','fecha_fin','dias','motivo','estado','fecha_solicitud','comentario_jefe','comentario_gh']
    return excel_response(headers, rows, 'reporte_vacaciones.xlsx', 'VACACIONES')


# ========================= RENOVACION 292 =========================
@login_required
def renovacion_home_292():
    _ensure_rrhh_292(); is_admin=_is_admin_292(); dni=_user_dni_292()
    if request.method=='POST':
        if not is_admin: return _deny_admin_292()
        d=limpiar_dni(request.form.get('dni')); w=_find_worker_any_292(d)
        execute('''INSERT INTO renovaciones(dni,trabajador,empresa,area,cargo,contrato_actual_inicio,contrato_actual_fin,nuevo_inicio,nuevo_fin,tipo_contrato,basico,estado,observacion,creado_por,creado_en,actualizado_en) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (d,w.get('trabajador'),w.get('empresa'),w.get('area'),w.get('cargo'),request.form.get('actual_inicio',''),request.form.get('actual_fin',''),request.form.get('nuevo_inicio',''),request.form.get('nuevo_fin',''),request.form.get('tipo_contrato','RENOVACIÓN'),float(request.form.get('basico') or 0),'PENDIENTE',request.form.get('observacion',''),session.get('usuario'),now_str(),now_str()), commit=True)
        flash('Renovación registrada.', 'success'); return redirect(url_for('renovacion_home'))
    rows=rows_to_dict(execute('SELECT * FROM renovaciones ORDER BY id DESC LIMIT 120', fetchall=True)) if is_admin else rows_to_dict(execute('SELECT * FROM renovaciones WHERE dni=? ORDER BY id DESC LIMIT 80', (dni,), fetchall=True))
    total=len(rows); pend=sum(1 for r in rows if (r.get('estado') or '').upper()=='PENDIENTE'); aprob=sum(1 for r in rows if (r.get('estado') or '').upper()=='APROBADO')
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a>{% if is_admin %}<a class="cfg" href="{{url_for('renovacion_config')}}"><i class="bi bi-gear"></i> Config.</a>{% endif %}<div class="ttl">Renovaciones</div></div><div class="rr292-body"><div class="rr292-kpis"><div class="rr292-kpi"><small>Total</small><b>{{total}}</b></div><div class="rr292-kpi"><small>Pend.</small><b>{{pend}}</b></div><div class="rr292-kpi"><small>Aprob.</small><b>{{aprob}}</b></div></div>{% if is_admin %}<form method="post" class="rr292-form"><div class="row"><div class="col-6"><label>DNI</label><input name="dni" maxlength="8" class="form-control" required></div><div class="col-6"><label>Tipo</label><select name="tipo_contrato" class="form-select"><option>RENOVACIÓN</option><option>INTERMITENTE</option><option>INDETERMINADO</option><option>TEMPORAL</option></select></div><div class="col-6"><label>Actual fin</label><input type="date" name="actual_fin" class="form-control"></div><div class="col-6"><label>Nuevo inicio</label><input type="date" name="nuevo_inicio" class="form-control"></div><div class="col-6"><label>Nuevo fin</label><input type="date" name="nuevo_fin" class="form-control"></div><div class="col-6"><label>Básico</label><input type="number" step="0.01" name="basico" class="form-control"></div><div class="col-12"><label>Observación</label><input name="observacion" class="form-control"></div></div><button class="rr292-btn mt-2"><i class="bi bi-save"></i> Registrar renovación</button></form>{% endif %}<div class="rr292-list"><table class="rr292-table"><thead><tr><th>DNI</th><th>Trabajador</th><th>Nuevo inicio</th><th>Nuevo fin</th><th>Tipo</th><th>Estado</th>{% if is_admin %}<th>Acción</th>{% endif %}</tr></thead><tbody>{% for r in rows %}<tr><td>{{r.dni}}</td><td>{{r.trabajador or '-'}}</td><td>{{r.nuevo_inicio}}</td><td>{{r.nuevo_fin}}</td><td>{{r.tipo_contrato}}</td><td><span class="rr292-badge">{{r.estado}}</span></td>{% if is_admin %}<td><a href="{{url_for('renovacion_estado', ren_id=r.id, estado='APROBADO')}}">Aprobar</a> · <a href="{{url_for('renovacion_estado', ren_id=r.id, estado='RECHAZADO')}}">Rechazar</a></td>{% endif %}</tr>{% else %}<tr><td colspan="7" class="text-center text-muted">Sin renovaciones.</td></tr>{% endfor %}</tbody></table></div></div></div></div>'''
    return render_page(body,rows=rows,total=total,pend=pend,aprob=aprob,is_admin=is_admin,title='Renovaciones')


@login_required
def renovacion_estado_292(ren_id, estado):
    if not _is_admin_292(): return _deny_admin_292()
    estado=estado if estado in ('APROBADO','RECHAZADO','PENDIENTE') else 'PENDIENTE'
    execute('UPDATE renovaciones SET estado=?, actualizado_en=? WHERE id=?', (estado,now_str(),ren_id), commit=True)
    flash('Renovación actualizada.', 'success'); return redirect(url_for('renovacion_home'))


@login_required
def renovacion_config_292():
    if not _is_admin_292(): return _deny_admin_292()
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('renovacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Config. renovación</div></div><div class="rr292-body"><a class="rr292-rowlink" href="{{url_for('renovacion_exportar')}}"><i class="bi bi-file-earmark-excel"></i><div><b>Exportar renovaciones</b><br><small>Control y auditoría</small></div><i class="bi bi-chevron-right"></i></a><div class="rr292-info"><i class="bi bi-info-circle"></i><div>La renovación usa la base trabajadores y permite aprobar/rechazar antes de generar contrato.</div></div></div></div></div>'''
    return render_page(body,title='Config. renovación')


@login_required
def renovacion_exportar_292():
    _ensure_rrhh_292(); rows=rows_to_dict(execute('SELECT * FROM renovaciones ORDER BY id DESC', fetchall=True)) if _is_admin_292() else rows_to_dict(execute('SELECT * FROM renovaciones WHERE dni=? ORDER BY id DESC', (_user_dni_292(),), fetchall=True))
    headers=['id','dni','trabajador','empresa','area','cargo','contrato_actual_inicio','contrato_actual_fin','nuevo_inicio','nuevo_fin','tipo_contrato','basico','estado','observacion','creado_por','creado_en','actualizado_en']
    return excel_response(headers, rows, 'reporte_renovaciones.xlsx', 'RENOVACIONES')


# ========================= CONTRATACION MEJORAS 292 =========================
@login_required
def contratacion_home_292():
    if not _is_admin_292(): return _deny_admin_292()
    _ensure_rrhh_292(); total_req=int(scalar('SELECT COUNT(*) AS c FROM contratacion_requerimientos') or 0); postulantes=int(scalar('SELECT COUNT(*) AS c FROM contratacion_ingresos') or 0); firmados=int(scalar('SELECT COUNT(*) AS c FROM contratacion_firmas_bio') or 0)
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><a class="cfg" href="{{url_for('contratacion_config')}}"><i class="bi bi-gear"></i> Config.</a><div class="ttl">Contratación</div></div><div class="rr292-body"><div class="rr292-kpis"><div class="rr292-kpi"><small>Req.</small><b>{{total_req}}</b></div><div class="rr292-kpi"><small>Post.</small><b>{{postulantes}}</b></div><div class="rr292-kpi"><small>Firmas</small><b>{{firmados}}</b></div></div><div class="rr292-section">Flujo</div><div class="rr292-grid3"><a class="rr292-tile" href="{{url_for('contratacion_requerimientos')}}"><i class="bi bi-clipboard2-plus"></i><span class="lbl">Requerim.</span><span class="sub">Cupos</span></a><a class="rr292-tile" href="{{url_for('contratacion_postulantes')}}"><i class="bi bi-person-lines-fill"></i><span class="lbl">Postul.</span><span class="sub">DNI</span></a><a class="rr292-tile" href="{{url_for('contratacion_etapa', etapa='medica')}}"><i class="bi bi-heart-pulse"></i><span class="lbl">Médica</span><span class="sub">Aptitud</span></a><a class="rr292-tile" href="{{url_for('contratacion_etapa', etapa='induccion')}}"><i class="bi bi-play-btn"></i><span class="lbl">Inducción</span><span class="sub">Asistencia</span></a><a class="rr292-tile" href="{{url_for('contratacion_etapa', etapa='indumentaria')}}"><i class="bi bi-bag-check"></i><span class="lbl">Indument.</span><span class="sub">EPP</span></a><a class="rr292-tile" href="{{url_for('contratacion_etapa', etapa='fotocheck')}}"><i class="bi bi-person-badge"></i><span class="lbl">Fotocheck</span><span class="sub">Carnet</span></a><a class="rr292-tile" href="{{url_for('contratacion_firma_bio')}}"><i class="bi bi-camera"></i><span class="lbl">Firma facial</span><span class="sub">Cámara/firma</span></a><a class="rr292-tile" href="{{url_for('contratacion_reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i><span class="lbl">Reportes</span><span class="sub">Control</span></a><a class="rr292-tile" href="{{url_for('contratacion_config')}}"><i class="bi bi-sliders"></i><span class="lbl">Config.</span><span class="sub">Maestros</span></a></div></div></div></div>'''
    return render_page(body,total_req=total_req,postulantes=postulantes,firmados=firmados,title='Contratación')


@login_required
def contratacion_firma_bio_292():
    if not _is_admin_292(): return _deny_admin_292()
    _ensure_rrhh_292()
    if request.method=='POST':
        dni=limpiar_dni(request.form.get('dni')); doc=request.form.get('documento') or 'CONTRATO DE TRABAJO'; metodo=request.form.get('metodo') or 'FIRMA DIGITAL + FACIAL'; obs=request.form.get('observacion') or ''; firma_data=request.form.get('firma_data') or ''; foto_data=request.form.get('foto_data') or ''
        if len(dni)!=8: flash('DNI inválido.', 'danger'); return redirect(url_for('contratacion_firma_bio'))
        w=_find_worker_any_292(dni); base=os.path.join(FIRMA_DIR, 'contratacion'); os.makedirs(base, exist_ok=True)
        def save_data(data,prefix):
            if not data or ',' not in data: return ''
            raw=base64.b64decode(data.split(',',1)[1]); path=os.path.join(base, f'{prefix}_{dni}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'); open(path,'wb').write(raw); return path
        firma_path=save_data(firma_data,'firma'); foto_path=save_data(foto_data,'rostro')
        execute('''INSERT INTO contratacion_firmas_bio(dni,trabajador,documento,metodo,firma_path,foto_path,fecha,hora,fecha_hora,registrado_por,observacion) VALUES(?,?,?,?,?,?,?,?,?,?,?)''', (dni,w.get('trabajador'),doc,metodo,firma_path,foto_path,today_str(),datetime.now().strftime('%H:%M'),now_str(),session.get('usuario'),obs), commit=True)
        execute("UPDATE contratacion_ingresos SET firma_estado='FIRMADO', observacion=? WHERE dni=?", (obs,dni), commit=True)
        flash('Firma facial/digital registrada.', 'success'); return redirect(url_for('contratacion_firma_bio'))
    firmas=rows_to_dict(execute('SELECT * FROM contratacion_firmas_bio ORDER BY id DESC LIMIT 30', fetchall=True))
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Firma facial / digital</div></div><div class="rr292-body"><form method="post" class="rr292-form" id="frmFirma292"><label>DNI trabajador</label><div class="input-group"><input name="dni" class="form-control" maxlength="8" inputmode="numeric" required placeholder="ESCANEAR O DIGITAR"><button type="button" class="btn btn-green" onclick="abrirScanner&&abrirScanner('readerFirma292','dniFirma292')"><i class="bi bi-qr-code-scan"></i></button></div><input id="dniFirma292" style="display:none"><div id="readerFirma292" class="scan-box mt-2" style="display:none"></div><label class="mt-2">Documento</label><select name="documento" class="form-select"><option>CONTRATO DE TRABAJO</option><option>RENOVACIÓN DE CONTRATO</option><option>DECLARACIÓN JURADA</option><option>POLÍTICAS / RIT</option></select><label class="mt-2">Método</label><select name="metodo" class="form-select"><option>RECONOCIMIENTO FACIAL + FIRMA DIGITAL</option><option>FIRMA DIGITAL EN PANTALLA</option><option>RECONOCIMIENTO FACIAL</option></select><label class="mt-2">Reconocimiento facial</label><video id="videoFirma292" class="rr292-video" autoplay playsinline muted></video><canvas id="fotoCanvas292" style="display:none"></canvas><img id="fotoPreview292" class="rr292-photo mt-2"><div class="row mt-2"><div class="col-6"><button type="button" class="rr292-outline" onclick="startCam292()"><i class="bi bi-camera-video"></i> Cámara</button></div><div class="col-6"><button type="button" class="rr292-outline" onclick="takePhoto292()"><i class="bi bi-camera"></i> Capturar</button></div></div><label class="mt-2">Firma en pantalla</label><canvas id="firmaCanvas292" class="rr292-canvas"></canvas><div class="row mt-2"><div class="col-6"><button type="button" class="rr292-outline" onclick="clearFirma292()">Limpiar</button></div><div class="col-6"><button class="rr292-btn"><i class="bi bi-save"></i> Guardar</button></div></div><label class="mt-2">Observación</label><input name="observacion" class="form-control" placeholder="Opcional"><input type="hidden" name="firma_data" id="firmaData292"><input type="hidden" name="foto_data" id="fotoData292"></form><div class="rr292-section">Últimas firmas</div>{% for f in firmas %}<div class="rr292-rowlink"><i class="bi bi-person-check"></i><div><b>{{f.dni}} · {{f.trabajador or '-'}}</b><br><small>{{f.documento}} · {{f.fecha_hora}}</small></div><i class="bi bi-check2-circle"></i></div>{% else %}<div class="rr292-form text-center text-muted">Sin firmas.</div>{% endfor %}</div></div></div><script>let stream292=null;async function startCam292(){try{stream292=await navigator.mediaDevices.getUserMedia({video:{facingMode:'user'},audio:false});document.getElementById('videoFirma292').srcObject=stream292;}catch(e){alert('Permita cámara en el navegador.');}}function takePhoto292(){const v=document.getElementById('videoFirma292'),c=document.getElementById('fotoCanvas292'),img=document.getElementById('fotoPreview292');if(!v.videoWidth){alert('Active la cámara primero.');return;}c.width=v.videoWidth;c.height=v.videoHeight;c.getContext('2d').drawImage(v,0,0);let data=c.toDataURL('image/png');document.getElementById('fotoData292').value=data;img.src=data;img.style.display='block';}const canvas=document.getElementById('firmaCanvas292'),ctx=canvas.getContext('2d');function resizeFirma292(){canvas.width=canvas.clientWidth;canvas.height=170;ctx.lineWidth=3;ctx.lineCap='round';ctx.strokeStyle='#08713b';}resizeFirma292();let drawing=false;function pos(e){let r=canvas.getBoundingClientRect(),t=e.touches?e.touches[0]:e;return{x:t.clientX-r.left,y:t.clientY-r.top};}canvas.addEventListener('mousedown',e=>{drawing=true;let p=pos(e);ctx.beginPath();ctx.moveTo(p.x,p.y);});canvas.addEventListener('mousemove',e=>{if(!drawing)return;let p=pos(e);ctx.lineTo(p.x,p.y);ctx.stroke();document.getElementById('firmaData292').value=canvas.toDataURL('image/png');});canvas.addEventListener('mouseup',()=>drawing=false);canvas.addEventListener('touchstart',e=>{drawing=true;let p=pos(e);ctx.beginPath();ctx.moveTo(p.x,p.y);e.preventDefault();});canvas.addEventListener('touchmove',e=>{if(!drawing)return;let p=pos(e);ctx.lineTo(p.x,p.y);ctx.stroke();document.getElementById('firmaData292').value=canvas.toDataURL('image/png');e.preventDefault();});canvas.addEventListener('touchend',()=>drawing=false);function clearFirma292(){ctx.clearRect(0,0,canvas.width,canvas.height);document.getElementById('firmaData292').value='';}</script>'''
    return render_page(body,firmas=firmas,title='Firma facial contratación')


@login_required
def contratacion_config_292():
    if not _is_admin_292(): return _deny_admin_292()
    _ensure_rrhh_292(); cargados=0
    if request.method=='POST':
        action=request.form.get('action')
        if action=='maestro':
            execute('INSERT INTO contratacion_maestros(tipo,valor,estado,creado_en,creado_por) VALUES(?,?,?,?,?)', (request.form.get('tipo'), limpiar_texto(request.form.get('valor')), 'ACTIVO', now_str(), session.get('usuario')), commit=True); flash('Dato maestro guardado.', 'success'); return redirect(url_for('contratacion_config'))
        f=request.files.get('archivo')
        if not f or not f.filename: flash('Seleccione Excel de trabajadores.', 'danger'); return redirect(url_for('contratacion_config'))
        try:
            for row in _iter_excel_upload(f):
                dni=limpiar_dni(_valor(row,['DNI','DOCUMENTO']));
                if len(dni)!=8: continue
                trabajador=limpiar_texto(_valor(row,['TRABAJADOR','NOMBRES','APELLIDOS Y NOMBRES'])); empresa=limpiar_texto(_valor(row,['EMPRESA'])); area=limpiar_texto(_valor(row,['AREA','ÁREA'])); cargo=limpiar_texto(_valor(row,['CARGO','PUESTO'])); actividad=limpiar_texto(_valor(row,['ACTIVIDAD','LABOR'])); planilla=limpiar_texto(_valor(row,['PLANILLA','REGIMEN','RÉGIMEN']))
                if scalar('SELECT COUNT(*) AS c FROM trabajadores WHERE dni=?', (dni,)): execute('UPDATE trabajadores SET trabajador=?,empresa=?,area=?,cargo=?,actividad=?,planilla=?,estado=?,fecha_carga=? WHERE dni=?', (trabajador,empresa,area,cargo,actividad,planilla,'ACTIVO',today_str(),dni), commit=True)
                else: execute('INSERT INTO trabajadores(dni,trabajador,empresa,area,cargo,actividad,planilla,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?,?)', (dni,trabajador,empresa,area,cargo,actividad,planilla,'ACTIVO',today_str()), commit=True)
                cargados+=1
            flash(f'Trabajadores cargados/actualizados: {cargados}.', 'success')
        except Exception as e: flash(str(e), 'danger')
        return redirect(url_for('contratacion_config'))
    maestros=rows_to_dict(execute('SELECT * FROM contratacion_maestros ORDER BY tipo,valor LIMIT 120', fetchall=True))
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Config. contratación</div></div><div class="rr292-body"><form method="post" enctype="multipart/form-data" class="rr292-form"><label>Carga masiva trabajadores Excel</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="field-help mt-1">Columnas: DNI, TRABAJADOR, EMPRESA, AREA, CARGO, ACTIVIDAD, PLANILLA.</div><button class="rr292-btn mt-2"><i class="bi bi-upload"></i> Cargar trabajadores</button></form><form method="post" class="rr292-form"><input type="hidden" name="action" value="maestro"><div class="row"><div class="col-6"><label>Tipo maestro</label><select name="tipo" class="form-select"><option>EMPRESA</option><option>AREA</option><option>CARGO</option><option>ACTIVIDAD</option><option>SEDE</option><option>REGIMEN</option><option>MODALIDAD</option></select></div><div class="col-6"><label>Valor</label><input name="valor" class="form-control" required></div></div><button class="rr292-btn mt-2"><i class="bi bi-save"></i> Guardar dato maestro</button></form><div class="rr292-section">Datos maestros</div><div class="rr292-list"><table class="rr292-table"><thead><tr><th>Tipo</th><th>Valor</th><th>Estado</th></tr></thead><tbody>{% for m in maestros %}<tr><td>{{m.tipo}}</td><td>{{m.valor}}</td><td>{{m.estado}}</td></tr>{% else %}<tr><td colspan="3" class="text-center text-muted">Sin datos maestros.</td></tr>{% endfor %}</tbody></table></div></div></div></div>'''
    return render_page(body,maestros=maestros,title='Config. contratación')


# ========================= REGISTRO RUTAS 292 =========================
try:
    app.add_url_rule('/boletas/config', 'boletas_config', boletas_config_292, methods=['GET'])
except Exception: app.view_functions['boletas_config']=boletas_config_292
try:
    app.add_url_rule('/vacaciones', 'vacaciones_home', vacaciones_home_292, methods=['GET','POST'])
    app.add_url_rule('/vacaciones/config', 'vacaciones_config', vacaciones_config_292, methods=['GET','POST'])
    app.add_url_rule('/vacaciones/exportar', 'vacaciones_exportar', vacaciones_exportar_292, methods=['GET'])
    app.add_url_rule('/vacaciones/estado/<int:sol_id>/<estado>', 'vacaciones_estado', vacaciones_estado_292, methods=['GET'])
except Exception as e:
    print('Rutas vacaciones 292:', e)
try:
    app.add_url_rule('/renovacion', 'renovacion_home', renovacion_home_292, methods=['GET','POST'])
    app.add_url_rule('/renovacion/config', 'renovacion_config', renovacion_config_292, methods=['GET'])
    app.add_url_rule('/renovacion/exportar', 'renovacion_exportar', renovacion_exportar_292, methods=['GET'])
    app.add_url_rule('/renovacion/estado/<int:ren_id>/<estado>', 'renovacion_estado', renovacion_estado_292, methods=['GET'])
except Exception as e:
    print('Rutas renovación 292:', e)
try:
    app.add_url_rule('/contratacion/firma-biometrica', 'contratacion_firma_bio', contratacion_firma_bio_292, methods=['GET','POST'])
    app.add_url_rule('/contratacion/config', 'contratacion_config', contratacion_config_292, methods=['GET','POST'])
except Exception as e:
    print('Rutas contratación extra 292:', e)

app.view_functions['login'] = login_292
app.view_functions['home'] = home_292
app.view_functions['boletas_home'] = boletas_home_292
app.view_functions['boletas_subir'] = boletas_subir_292
app.view_functions['boletas_listar'] = boletas_listar_292
app.view_functions['boletas_reportes'] = boletas_reportes_292
app.view_functions['contratacion_home'] = contratacion_home_292
# ======================= FIN PATCH OMAR 292 =======================

# ========================= PATCH OMAR 293 =========================
# Accesos por módulo, boletas masivas PDF, contratación con nuevos/reingresantes,
# carga masiva de trabajadores, datos maestros y validación básica por flujo.

def _ensure_rrhh_293():
    _ensure_rrhh_292()
    try:
        conn=get_conn(); cur=conn.cursor()
        for col, ddl in [
            ('tipo_ingreso','TEXT'),('fecha_nacimiento','TEXT'),('direccion','TEXT'),('distrito','TEXT'),('provincia','TEXT'),('departamento','TEXT'),
            ('dni_validado','INTEGER DEFAULT 0'),('fuente_datos','TEXT')
        ]:
            _add_column_if_missing(cur, 'contratacion_ingresos', col, ddl)
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print('Migración 293:', e)


def _module_targets_293():
    return {
        'tareo':('Tareo','bi-list-check','hojas_tareo','hojas_tareo'),
        'asistencia':('Asistencia','bi-fingerprint','asistencia_modulo','asistencia_modulo'),
        'transporte':('Transporte','bi-bus-front','transporte','transporte_mobile_home'),
        'contratacion':('Contratación','bi-person-plus','contratacion_home','contratacion_usuario_home'),
        'boleta':('Boleta','bi-file-earmark-text','boletas_home','boletas_home'),
        'vacaciones':('Vacaciones','bi-calendar-check','vacaciones_home','vacaciones_home'),
        'renovacion':('Renovación','bi-arrow-repeat','renovacion_home','renovacion_home')
    }


def modulo_acceso_293(modulo):
    if not session.get('usuario'):
        return redirect(url_for('login'))
    data=_module_targets_293().get(modulo)
    if not data:
        flash('Módulo no encontrado.', 'danger'); return redirect(url_for('home'))
    label, icon, ep_admin, ep_user = data
    if request.method=='POST':
        modo=request.form.get('modo') or 'usuario'
        usuario=(request.form.get('usuario') or '').strip()
        clave=request.form.get('password') or ''
        if modo=='admin':
            row=row_to_dict(execute('SELECT * FROM usuarios WHERE usuario=?', (usuario,), fetchone=True))
            ok=bool(row and row.get('rol')=='admin' and row.get('estado','ACTIVO')=='ACTIVO' and check_password_hash(row['password_hash'], clave))
            if not ok and _is_admin_292() and clave in ('','admin123'):
                ok=True
            if ok:
                session['module_role']='admin'; session['module_name']=modulo
                return redirect(url_for(ep_admin))
            flash('Acceso administrador incorrecto.', 'danger')
            return redirect(url_for('modulo_acceso', modulo=modulo))
        dni=limpiar_dni(usuario)
        if len(dni)==8 and clave in (dni, dni[-4:], '1234'):
            w=_find_worker_any_292(dni)
            session['module_role']='usuario'; session['module_name']=modulo; session['dni']=dni; session['nombres']=w.get('trabajador') or dni
            if not _is_admin_292():
                session['usuario']=dni; session['rol']='usuario'
            return redirect(url_for(ep_user))
        flash('Acceso usuario incorrecto. Use DNI y PIN/clave.', 'danger')
    body = r'''
    <div class="phone-wrap">
      <div class="green-hero" style="min-height:245px;border-radius:0 0 22px 22px">
        <div class="green-top"><a class="text-white text-decoration-none" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><span><i class="bi bi-gear"></i> Config.</span></div>
        <div class="splash-logo" style="width:112px;height:112px;font-size:48px;margin:14px auto 8px"><i class="bi {{icon}}"></i></div>
        <div class="splash-title">{{label|upper}}</div><div class="login-name">INICIAR SESIÓN</div>
      </div>
      <form method="post" class="floating-card login-form" style="margin-top:-18px">
        <div class="role-toggle mb-2"><label><input type="radio" name="modo" value="usuario" checked><span>USUARIO</span></label><label><input type="radio" name="modo" value="admin"><span>ADMINISTRADOR</span></label></div>
        <input name="usuario" class="form-control mb-2" placeholder="DNI usuario / admin" autofocus>
        <input name="password" type="password" class="form-control mb-3" placeholder="PIN / clave">
        <button class="btn btn-green w-100" style="height:46px;font-size:17px"><i class="bi bi-box-arrow-in-right"></i> INGRESAR</button>
        <div class="text-center text-muted mt-2">Usuario demo: DNI / últimos 4 · Admin: admin / admin123</div>
      </form>
    </div>'''
    return render_page(body,label=label,icon=icon,title='Acceso '+label)


def home_293():
    if not session.get('usuario'):
        return redirect(url_for('inicio'))
    is_admin=_is_admin_292()
    body=r'''
    <div class="desktop-grid"><div class="phone-wrap"><div class="green-hero" style="min-height:220px"><div class="green-top"><a class="text-white text-decoration-none" href="{{url_for('soporte')}}"><i class="bi bi-headset"></i> Soporte</a>{% if is_admin %}<a class="text-white text-decoration-none" href="{{url_for('configuraciones')}}"><i class="bi bi-gear"></i> Config.</a>{% else %}<span></span>{% endif %}</div><div class="avatar"><i class="bi bi-person-circle"></i></div><div class="login-name">{{ session.get('nombres','USUARIO') }}</div><div class="white-input mt-3"></div></div>
      <div class="top-actions">
        <a class="tile text-decoration-none" href="{{url_for('modulo_acceso', modulo='tareo')}}"><i class="bi bi-list-check"></i>TAREO</a>
        <a class="tile text-decoration-none" href="{{url_for('modulo_acceso', modulo='asistencia')}}"><i class="bi bi-fingerprint"></i>ASIST.</a>
        <a class="tile text-decoration-none" href="{{url_for('modulo_acceso', modulo='transporte')}}"><i class="bi bi-bus-front"></i>TRANSP.</a>
        <a class="tile text-decoration-none" href="{{url_for('modulo_acceso', modulo='contratacion')}}"><i class="bi bi-person-plus"></i>CONTRAT.</a>
        <a class="tile text-decoration-none" href="{{url_for('modulo_acceso', modulo='boleta')}}"><i class="bi bi-file-earmark-text"></i>BOLETA</a>
        <a class="tile text-decoration-none" href="{{url_for('modulo_acceso', modulo='vacaciones')}}"><i class="bi bi-calendar-check"></i>VACAC.</a>
        <a class="tile text-decoration-none" href="{{url_for('modulo_acceso', modulo='renovacion')}}"><i class="bi bi-arrow-repeat"></i>RENOV.</a>
        {% if is_admin %}<a class="tile text-decoration-none" href="{{url_for('reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i>REPORTES<br>TAREO</a>{% endif %}
        {% if is_admin %}<a class="tile text-decoration-none" href="{{url_for('sincronizacion')}}"><i class="bi bi-arrow-repeat"></i>SINC.</a>{% endif %}
      </div><div class="leaf"></div><div class="bottom-sync"><i class="bi bi-arrow-repeat"></i> Actualizado hasta: {{ now }}</div><a href="{{url_for('logout')}}" class="bottom-out"><i class="bi bi-box-arrow-right"></i></a></div>
      <div class="desk-panel"><h1 class="header-title">TAREO MÓVIL – GRUPO DE COSECHA</h1><div class="card-pro p-4 mb-3"><h4 class="fw-bold text-success mb-1">Panel integrado RR.HH.</h4><div class="text-muted small">Cada módulo solicita acceso ADMINISTRADOR / USUARIO. Firma se maneja dentro de Contratación y Renovación.</div></div></div></div>'''
    return render_page(body, now=now_str(), is_admin=is_admin)


def _bt_parse_period_293(name, default_period):
    m=re.search(r'(20\d{2})[-_ ]?(0[1-9]|1[0-2])', str(name or ''))
    return f'{m.group(1)}-{m.group(2)}' if m else (default_period or datetime.now().strftime('%Y-%m'))


def boletas_subir_293():
    if not _is_admin_292(): return _deny_admin_292()
    _boleta_init_291()
    if request.method=='POST':
        tipo=_bt_norm_tipo_291(request.form.get('tipo'))
        periodo_default=limpiar_texto(request.form.get('periodo') or datetime.now().strftime('%Y-%m'), upper=False)
        detalle=limpiar_texto(request.form.get('detalle') or 'Carga masiva PDF', upper=False)
        obs=limpiar_texto(request.form.get('observacion'), upper=False)
        files=request.files.getlist('archivos') or request.files.getlist('archivo')
        ok=bad=0; errores=[]
        for f in files:
            if not f or not f.filename: continue
            original=f.filename
            if not original.lower().endswith('.pdf'):
                bad+=1; errores.append(f'{original}: no es PDF'); continue
            dni=limpiar_dni(original)
            if len(dni)!=8:
                bad+=1; errores.append(f'{original}: no contiene DNI'); continue
            try:
                periodo=_bt_parse_period_293(original, periodo_default)
                path, original_saved, ext = _bt_save_file_291(f, dni, tipo, periodo)
                _bt_insert_doc_291(dni,tipo,periodo,detalle,obs,path,original_saved,ext,session.get('usuario'))
                ok+=1
            except Exception as e:
                bad+=1; errores.append(f'{original}: {e}')
        flash(f'Carga masiva finalizada. Correctos: {ok}. Observados: {bad}.', 'success' if ok else 'danger')
        if errores: flash(' | '.join(errores[:4]), 'warning')
        return redirect(url_for('boletas_listar', tipo=tipo))
    body=_rrhh_css_292()+r'''
    <div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('boletas_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Carga masiva boletas PDF</div></div><div class="rr292-body">
      <div class="rr292-info"><i class="bi bi-info-circle-fill"></i><div>Seleccione varios PDF. El DNI debe estar en el nombre: <b>BOLETA_74324033_2026-06.pdf</b></div></div>
      <form class="rr292-form" method="post" enctype="multipart/form-data"><div class="row"><div class="col-6"><label>Tipo</label><select class="form-select" name="tipo">{{tipo_opts|safe}}</select></div><div class="col-6"><label>Periodo default</label><input class="form-control" name="periodo" value="{{periodo}}"></div><div class="col-12"><label>PDFs masivos</label><input class="form-control" name="archivos" type="file" accept="application/pdf,.pdf" multiple required></div><div class="col-12"><label>Detalle</label><input class="form-control" name="detalle"></div><div class="col-12"><label>Observación</label><input class="form-control" name="observacion"></div></div><button class="rr292-btn mt-2"><i class="bi bi-upload"></i> Cargar PDFs masivos</button></form>
      <a class="rr292-outline" href="{{url_for('boletas_plantilla')}}"><i class="bi bi-file-earmark-excel"></i> Descargar plantilla control</a>
    </div></div></div>'''
    return render_page(body,tipo_opts=_bt_tipo_options_291(),periodo=datetime.now().strftime('%Y-%m'),title='Carga masiva boletas')


def boletas_config_293():
    if not _is_admin_292(): return _deny_admin_292()
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('boletas_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Config. Boletas</div></div><div class="rr292-body"><div class="rr292-section">Configuraciones</div><a class="rr292-rowlink" href="{{url_for('boletas_subir')}}"><i class="bi bi-cloud-arrow-up"></i><div><b>Carga masiva PDF</b><br><small>Varios PDF por nombre con DNI</small></div><i class="bi bi-chevron-right"></i></a><a class="rr292-rowlink" href="{{url_for('boletas_detectar')}}"><i class="bi bi-search"></i><div><b>Detección automática</b><br><small>Carpeta BOLETAS_UPLOAD_AUTO</small></div><i class="bi bi-chevron-right"></i></a><a class="rr292-rowlink" href="{{url_for('boletas_plantilla')}}"><i class="bi bi-file-earmark-excel"></i><div><b>Plantilla control</b><br><small>Excel de seguimiento</small></div><i class="bi bi-chevron-right"></i></a><a class="rr292-rowlink" href="{{url_for('boletas_reportes')}}"><i class="bi bi-bar-chart"></i><div><b>Reportes</b><br><small>Control de carga</small></div><i class="bi bi-chevron-right"></i></a></div></div></div>'''
    return render_page(body,title='Config. Boletas')


def _ct_worker_293(dni):
    r=row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (limpiar_dni(dni),), fetchone=True))
    if not r: return {}
    return {'dni':r.get('dni'),'nombres':r.get('trabajador') or r.get('nombre') or '', 'empresa':r.get('empresa') or '', 'area':r.get('area') or '', 'cargo':r.get('cargo') or '', 'actividad':r.get('actividad') or '', 'correo':r.get('correo') or '', 'telefono':r.get('celular') or '', 'direccion':r.get('direccion') or '', 'distrito':r.get('distrito') or '', 'provincia':r.get('provincia') or '', 'departamento':r.get('departamento') or '', 'fecha_nacimiento':r.get('fecha_nacimiento') or ''}


def _ct_tipo_293(dni):
    d=limpiar_dni(dni)
    return 'REINGRESANTE' if int(scalar('SELECT COUNT(*) AS c FROM trabajadores WHERE dni=?',(d,)) or 0) or int(scalar('SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE dni=?',(d,)) or 0) else 'NUEVO'


def api_contratacion_dni_293(dni):
    _ensure_rrhh_293(); d=limpiar_dni(dni)
    if len(d)!=8: return jsonify(ok=False,msg='DNI inválido')
    return jsonify(ok=True,dni=d,tipo=_ct_tipo_293(d),trabajador=_ct_worker_293(d))


def contratacion_postulantes_293():
    if not _is_admin_292(): return _deny_admin_292()
    _ensure_rrhh_293()
    reqs=rows_to_dict(execute('SELECT * FROM contratacion_requerimientos ORDER BY id DESC LIMIT 200', fetchall=True))
    req_id=request.values.get('req') or request.form.get('requerimiento_id') or (str(reqs[0]['id']) if reqs else '')
    req=row_to_dict(execute('SELECT * FROM contratacion_requerimientos WHERE id=?',(req_id,), fetchone=True)) if req_id else None
    if request.method=='POST':
        if not req: flash('Seleccione requerimiento.', 'danger'); return redirect(url_for('contratacion_postulantes'))
        dni=limpiar_dni(request.form.get('dni'))
        if len(dni)!=8: flash('DNI inválido.', 'danger'); return redirect(url_for('contratacion_postulantes', req=req_id))
        if int(scalar('SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE requerimiento_id=? AND dni=?',(req_id,dni)) or 0):
            flash('El DNI ya está registrado en este requerimiento.', 'danger'); return redirect(url_for('contratacion_postulantes', req=req_id))
        cupo=int(req.get('cantidad') or 0); usados=int(scalar('SELECT COUNT(*) AS c FROM contratacion_ingresos WHERE requerimiento_id=?',(req_id,)) or 0)
        if cupo and usados>=cupo:
            flash('El requerimiento ya completó su cupo.', 'danger'); return redirect(url_for('contratacion_postulantes', req=req_id))
        base=_ct_worker_293(dni); tipo=_ct_tipo_293(dni)
        nombres=limpiar_texto(request.form.get('nombres') or base.get('nombres'))
        execute('''INSERT INTO contratacion_ingresos(requerimiento_id,requerimiento,dni,nombres,telefono,correo,empresa,area,cargo,actividad,tipo_contrato,regimen_laboral,fecha_inicio,fecha_fin,basico,estado,medica_estado,induccion_estado,indumentaria_estado,fotocheck_estado,firma_estado,observacion,creado_por,creado_en,tipo_ingreso,fecha_nacimiento,direccion,distrito,provincia,departamento,dni_validado,fuente_datos) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (req_id,_ct_req_label_290(req),dni,nombres,request.form.get('telefono') or base.get('telefono'),request.form.get('correo') or base.get('correo'),request.form.get('empresa') or req.get('empresa') or base.get('empresa'),request.form.get('area') or req.get('area') or base.get('area'),request.form.get('cargo') or req.get('cargo') or base.get('cargo'),request.form.get('actividad') or req.get('actividad') or base.get('actividad'),request.form.get('tipo_contrato') or req.get('tipo_contrato'),request.form.get('regimen_laboral') or req.get('regimen_laboral'),request.form.get('fecha_inicio') or req.get('fecha_ingreso'),request.form.get('fecha_fin'),float(request.form.get('basico') or 0),'POSTULANTE','PENDIENTE','PENDIENTE','PENDIENTE','PENDIENTE','PENDIENTE',request.form.get('observacion'),session.get('usuario'),now_str(),tipo,request.form.get('fecha_nacimiento') or base.get('fecha_nacimiento'),request.form.get('direccion') or base.get('direccion'),request.form.get('distrito') or base.get('distrito'),request.form.get('provincia') or base.get('provincia'),request.form.get('departamento') or base.get('departamento'),1 if base else 0,'TRABAJADORES' if base else 'DIGITADO'), commit=True)
        flash(f'Postulante {tipo} registrado.', 'success')
        return redirect(url_for('contratacion_postulantes', req=req_id))
    posts=rows_to_dict(execute('SELECT * FROM contratacion_ingresos WHERE (?="" OR requerimiento_id=?) ORDER BY id DESC LIMIT 250',(str(req_id),str(req_id)), fetchall=True))
    body=_contratacion_css_290()+r'''
    <div class="ct290-phone"><div class="ct290-app"><div class="ct290-head"><a href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-person-lines-fill"></i></div><div class="ttl">Postulantes</div></div><div class="ct290-body">
      <form method="get" class="ct290-form"><label>Requerimiento</label><select class="form-select" name="req" onchange="this.form.submit()">{{req_opts|safe}}</select></form>
      <div id="ctDniStatus293" class="ct290-info">Digite DNI. Se detecta NUEVO / REINGRESANTE y se jalan datos de Trabajadores.</div>
      <form method="post" class="ct290-form"><input type="hidden" name="requerimiento_id" value="{{req_id}}"><div class="ct290-row"><div><label>DNI</label><input id="dniPost293" name="dni" maxlength="8" class="form-control" required autofocus></div><div><label>Tipo</label><input id="tipoIng293" class="form-control" readonly></div></div><label class="mt-2">Nombres</label><input id="nomPost293" name="nombres" class="form-control" required><div class="ct290-row mt-2"><div><label>Teléfono</label><input id="telPost293" name="telefono" class="form-control"></div><div><label>Correo</label><input id="corPost293" name="correo" class="form-control"></div></div><div class="ct290-row mt-2"><div><label>Empresa</label><input id="empPost293" name="empresa" class="form-control" value="{{req.empresa if req else ''}}"></div><div><label>Área</label><input id="areaPost293" name="area" class="form-control" value="{{req.area if req else ''}}"></div></div><div class="ct290-row mt-2"><div><label>Cargo</label><input id="cargoPost293" name="cargo" class="form-control" value="{{req.cargo if req else ''}}"></div><div><label>Actividad</label><input id="actPost293" name="actividad" class="form-control" value="{{req.actividad if req else ''}}"></div></div><div class="ct290-row mt-2"><div><label>Inicio</label><input type="date" name="fecha_inicio" class="form-control" value="{{req.fecha_ingreso if req else hoy}}"></div><div><label>Fin</label><input type="date" name="fecha_fin" class="form-control"></div></div><div class="ct290-row mt-2"><div><label>Tipo contrato</label><select name="tipo_contrato" class="form-select"><option>TEMPORAL</option><option>INTERMITENTE</option><option>INDETERMINADO</option></select></div><div><label>Régimen</label><select name="regimen_laboral" class="form-select"><option>AGRARIO</option><option>GENERAL</option></select></div></div><label class="mt-2">Dirección</label><input id="dirPost293" name="direccion" class="form-control"><button class="ct290-btn mt-2"><i class="bi bi-person-plus"></i> Registrar postulante</button></form>
      <input class="ct290-search" id="qpost293" placeholder="Buscar..."><div class="ct290-tablewrap"><table id="tblpost293" class="ct290-table"><thead><tr><th>DNI</th><th>Tipo</th><th>Postulante</th><th>Cargo</th><th>Médica</th><th>Inducción</th><th>Firma</th></tr></thead><tbody>{% for p in posts %}<tr><td>{{p.dni}}</td><td>{{p.tipo_ingreso or '-'}}</td><td>{{p.nombres}}</td><td>{{p.cargo or '-'}}</td><td>{{badge(p.medica_estado)|safe}}</td><td>{{badge(p.induccion_estado)|safe}}</td><td>{{badge(p.firma_estado)|safe}}</td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted">Sin postulantes.</td></tr>{% endfor %}</tbody></table></div></div></div></div>
    <script>(function(){const d=document.getElementById('dniPost293'),st=document.getElementById('ctDniStatus293');async function look(){let v=(d.value||'').replace(/\D/g,'').slice(-8);d.value=v;if(v.length<8)return;let r=await fetch('/api/contratacion/dni/'+v,{cache:'no-store'});let j=await r.json();if(j.ok){let t=j.trabajador||{};document.getElementById('tipoIng293').value=j.tipo;document.getElementById('nomPost293').value=t.nombres||'';document.getElementById('telPost293').value=t.telefono||'';document.getElementById('corPost293').value=t.correo||'';if(t.empresa)document.getElementById('empPost293').value=t.empresa;if(t.area)document.getElementById('areaPost293').value=t.area;if(t.cargo)document.getElementById('cargoPost293').value=t.cargo;if(t.actividad)document.getElementById('actPost293').value=t.actividad;document.getElementById('dirPost293').value=t.direccion||'';st.innerHTML='<b>'+j.tipo+'</b> detectado. Datos cargados.'; if(typeof beep==='function')beep();}}d&&d.addEventListener('input',look);const q=document.getElementById('qpost293'),t=document.getElementById('tblpost293');q&&q.addEventListener('input',()=>{const s=q.value.toUpperCase();t.querySelectorAll('tbody tr').forEach(r=>r.style.display=r.innerText.toUpperCase().includes(s)?'':'none')});})();</script>'''
    return render_page(body,req=req,req_id=req_id,req_opts=_ct_req_options_290(reqs,req_id),posts=posts,badge=_ct_badge_290,hoy=today_str(),title='Postulantes')


def _ct_stage_ok_293(row, etapa):
    if not row: return False,'DNI no registrado en Postulantes.'
    norm=lambda x:(x or '').upper().replace('Í','I')
    if etapa=='medica': return True,''
    if norm(row.get('medica_estado')) not in ('APTO','HABILITADO','OK','COMPLETO'): return False,'Primero debe estar APTO en Evaluación Médica.'
    if etapa=='induccion': return True,''
    if norm(row.get('induccion_estado')) not in ('ASISTIO','OK','COMPLETO'): return False,'Primero debe completar Inducción.'
    if etapa=='indumentaria': return True,''
    if norm(row.get('indumentaria_estado')) not in ('ENTREGADO','OK','COMPLETO'): return False,'Primero debe completar Indumentaria/EPP.'
    if etapa=='firma': return True,''
    if norm(row.get('firma_estado')) not in ('FIRMADO','OK','COMPLETO'): return False,'Primero debe completar Firma Digital / Facial.'
    return True,''


def contratacion_etapa_293(etapa):
    if not _is_admin_292(): return _deny_admin_292()
    _ensure_rrhh_293(); meta=_ETAPAS_290.get(etapa)
    if not meta: return redirect(url_for('contratacion_home'))
    titulo,icono,col,estados=meta
    reqs=rows_to_dict(execute('SELECT * FROM contratacion_requerimientos ORDER BY id DESC LIMIT 200', fetchall=True))
    req_id=request.values.get('req') or request.form.get('requerimiento_id') or ''
    if request.method=='POST':
        dni=limpiar_dni(request.form.get('dni')); row=row_to_dict(execute('SELECT * FROM contratacion_ingresos WHERE dni=? AND (?="" OR requerimiento_id=?) ORDER BY id DESC LIMIT 1',(dni,str(req_id),str(req_id)), fetchone=True))
        ok,msg=_ct_stage_ok_293(row, etapa)
        if not ok:
            flash('BLOQUEADO: '+msg, 'danger'); return redirect(url_for('contratacion_etapa', etapa=etapa, req=req_id))
        execute(f'UPDATE contratacion_ingresos SET {col}=?, observacion=? WHERE id=?', (request.form.get('estado'), request.form.get('observacion'), row.get('id')), commit=True)
        flash(titulo+' actualizado.', 'success'); return redirect(url_for('contratacion_etapa', etapa=etapa, req=req_id))
    posts=rows_to_dict(execute('SELECT * FROM contratacion_ingresos WHERE (?="" OR requerimiento_id=?) ORDER BY id DESC LIMIT 250',(str(req_id),str(req_id)), fetchall=True))
    opts=''.join(f'<option>{e}</option>' for e in estados)
    body=_contratacion_css_290()+r'''<div class="ct290-phone"><div class="ct290-app"><div class="ct290-head"><a href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ico"><i class="bi bi-{{icono}}"></i></div><div class="ttl">{{titulo}}</div></div><div class="ct290-body"><div class="ct290-info">Flujo bloqueado: Postulantes → Médica → Inducción → Indumentaria → Firma → Fotocheck.</div><form method="get" class="ct290-form"><label>Requerimiento</label><select class="form-select" name="req" onchange="this.form.submit()">{{req_opts|safe}}</select></form><form method="post" class="ct290-form"><input type="hidden" name="requerimiento_id" value="{{req_id}}"><div class="ct290-row"><div><label>DNI</label><input name="dni" maxlength="8" class="form-control" required></div><div><label>Estado</label><select name="estado" class="form-select">{{opts|safe}}</select></div></div><label class="mt-2">Observación</label><input name="observacion" class="form-control"><button class="ct290-btn mt-2">Actualizar</button></form><div class="ct290-tablewrap"><table class="ct290-table"><thead><tr><th>DNI</th><th>Postulante</th><th>Tipo</th><th>Médica</th><th>Inducción</th><th>Indument.</th><th>Firma</th><th>Fotocheck</th></tr></thead><tbody>{% for p in posts %}<tr><td>{{p.dni}}</td><td>{{p.nombres}}</td><td>{{p.tipo_ingreso or '-'}}</td><td>{{badge(p.medica_estado)|safe}}</td><td>{{badge(p.induccion_estado)|safe}}</td><td>{{badge(p.indumentaria_estado)|safe}}</td><td>{{badge(p.firma_estado)|safe}}</td><td>{{badge(p.fotocheck_estado)|safe}}</td></tr>{% else %}<tr><td colspan="8" class="text-center text-muted">Sin postulantes.</td></tr>{% endfor %}</tbody></table></div></div></div></div>'''
    return render_page(body,titulo=titulo,icono=icono,posts=posts,req_id=req_id,req_opts=_ct_req_options_290(reqs,req_id),opts=opts,badge=_ct_badge_290,title=titulo)


def contratacion_plantilla_trabajadores_293():
    headers=['DNI','TRABAJADOR','EMPRESA','AREA','CARGO','ACTIVIDAD','CORREO','TELEFONO','FECHA NACIMIENTO','DIRECCION','DISTRITO','PROVINCIA','DEPARTAMENTO']
    example=['74324033','JOSE GARCIA','AQUANQA','CAMPO','OPERARIO','COSECHA','demo@empresa.com','999999999','1990-01-01','AV. EJEMPLO','TRUJILLO','TRUJILLO','LA LIBERTAD']
    return _tpl_xlsx(headers, example, 'plantilla_trabajadores_contratacion.xlsx', 'TRABAJADORES')


def contratacion_carga_trabajadores_293():
    if not _is_admin_292(): return _deny_admin_292()
    if request.method=='POST':
        f=request.files.get('archivo')
        if not f or not f.filename:
            flash('Seleccione Excel.', 'danger'); return redirect(url_for('contratacion_carga_trabajadores'))
        ok=bad=0
        for row in _iter_excel_upload(f):
            dni=limpiar_dni(row.get('DNI') or row.get('DOCUMENTO')); nom=limpiar_texto(row.get('TRABAJADOR') or row.get('NOMBRES') or row.get('APELLIDOS Y NOMBRES'))
            if len(dni)!=8 or not nom: bad+=1; continue
            if int(scalar('SELECT COUNT(*) AS c FROM trabajadores WHERE dni=?',(dni,)) or 0):
                execute('UPDATE trabajadores SET trabajador=?, empresa=?, area=?, cargo=?, actividad=?, fecha_carga=? WHERE dni=?',(nom,row.get('EMPRESA') or '',row.get('AREA') or '',row.get('CARGO') or '',row.get('ACTIVIDAD') or '',today_str(),dni), commit=True)
            else:
                execute('INSERT INTO trabajadores(dni,trabajador,empresa,area,cargo,actividad,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?)',(dni,nom,row.get('EMPRESA') or '',row.get('AREA') or '',row.get('CARGO') or '',row.get('ACTIVIDAD') or '','ACTIVO',today_str()), commit=True)
            ok+=1
        flash(f'Carga trabajadores finalizada. Correctos: {ok}. Observados: {bad}.', 'success')
        return redirect(url_for('contratacion_config'))
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('contratacion_config')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Carga trabajadores</div></div><div class="rr292-body"><form class="rr292-form" method="post" enctype="multipart/form-data"><label>Excel trabajadores</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><button class="rr292-btn mt-2">Cargar trabajadores</button></form><a class="rr292-outline" href="{{url_for('contratacion_plantilla_trabajadores')}}">Descargar plantilla</a></div></div></div>'''
    return render_page(body,title='Carga trabajadores')


def contratacion_config_293():
    if not _is_admin_292(): return _deny_admin_292()
    _ensure_rrhh_293()
    if request.method=='POST':
        valor=limpiar_texto(request.form.get('valor')); tipo=request.form.get('tipo') or 'GENERAL'
        if valor: execute('INSERT INTO contratacion_maestros(tipo,valor,estado,creado_en,creado_por) VALUES(?,?,?,?,?)',(tipo,valor,'ACTIVO',now_str(),session.get('usuario')), commit=True)
        flash('Dato maestro guardado.', 'success'); return redirect(url_for('contratacion_config'))
    maestros=rows_to_dict(execute('SELECT * FROM contratacion_maestros ORDER BY tipo, valor LIMIT 100', fetchall=True))
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('contratacion_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Config. Contratación</div></div><div class="rr292-body"><a class="rr292-rowlink" href="{{url_for('contratacion_carga_trabajadores')}}"><i class="bi bi-cloud-arrow-up"></i><div><b>Carga masiva trabajadores</b><br><small>Base para nuevos/reingresantes</small></div><i class="bi bi-chevron-right"></i></a><a class="rr292-rowlink" href="{{url_for('contratacion_plantilla_trabajadores')}}"><i class="bi bi-file-earmark-excel"></i><div><b>Plantilla trabajadores</b><br><small>DNI y datos completos</small></div><i class="bi bi-chevron-right"></i></a><form method="post" class="rr292-form"><div class="row"><div class="col-6"><label>Tipo</label><select name="tipo" class="form-select"><option>EMPRESA</option><option>AREA</option><option>CARGO</option><option>ACTIVIDAD</option><option>SEDE</option><option>REGIMEN</option><option>MODALIDAD</option><option>PLANTILLA</option></select></div><div class="col-6"><label>Valor</label><input name="valor" class="form-control"></div></div><button class="rr292-btn mt-2">Agregar maestro</button></form>{% for m in maestros %}<div class="rr292-rowlink"><i class="bi bi-database"></i><div><b>{{m.tipo}}</b><br><small>{{m.valor}}</small></div><span></span></div>{% endfor %}</div></div></div>'''
    return render_page(body,maestros=maestros,title='Config. Contratación')


def contratacion_usuario_home_293():
    _ensure_rrhh_293(); dni=_user_dni_292()
    rows=rows_to_dict(execute('SELECT * FROM contratacion_ingresos WHERE dni=? ORDER BY id DESC LIMIT 50',(dni,), fetchall=True))
    body=_rrhh_css_292()+r'''<div class="phone-wrap"><div class="rr292-app"><div class="rr292-head"><a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Mi contratación</div></div><div class="rr292-body">{% for r in rows %}<div class="rr292-rowlink"><i class="bi bi-person-lines-fill"></i><div><b>{{r.requerimiento or r.cargo or 'Proceso'}}</b><br><small>Médica: {{r.medica_estado}} · Inducción: {{r.induccion_estado}} · Firma: {{r.firma_estado}}</small></div><i class="bi bi-chevron-right"></i></div>{% else %}<div class="rr292-form text-center text-muted">No tiene procesos registrados.</div>{% endfor %}</div></div></div>'''
    return render_page(body, rows=rows, title='Mi contratación')


def renovacion_firma_293():
    if not _is_admin_292(): return _deny_admin_292()
    return redirect(url_for('contratacion_firma_bio'))

try: app.add_url_rule('/acceso/<modulo>', 'modulo_acceso', modulo_acceso_293, methods=['GET','POST'])
except Exception: app.view_functions['modulo_acceso']=modulo_acceso_293
try: app.add_url_rule('/api/contratacion/dni/<dni>', 'api_contratacion_dni', api_contratacion_dni_293, methods=['GET'])
except Exception: app.view_functions['api_contratacion_dni']=api_contratacion_dni_293
try: app.add_url_rule('/contratacion/plantilla-trabajadores', 'contratacion_plantilla_trabajadores', contratacion_plantilla_trabajadores_293, methods=['GET'])
except Exception: app.view_functions['contratacion_plantilla_trabajadores']=contratacion_plantilla_trabajadores_293
try: app.add_url_rule('/contratacion/carga-trabajadores', 'contratacion_carga_trabajadores', contratacion_carga_trabajadores_293, methods=['GET','POST'])
except Exception: app.view_functions['contratacion_carga_trabajadores']=contratacion_carga_trabajadores_293
try: app.add_url_rule('/contratacion/usuario', 'contratacion_usuario_home', contratacion_usuario_home_293, methods=['GET'])
except Exception: app.view_functions['contratacion_usuario_home']=contratacion_usuario_home_293
try: app.add_url_rule('/renovacion/firma', 'renovacion_firma', renovacion_firma_293, methods=['GET'])
except Exception: app.view_functions['renovacion_firma']=renovacion_firma_293

app.view_functions['home']=home_293
app.view_functions['boletas_subir']=boletas_subir_293
app.view_functions['boletas_config']=boletas_config_293
app.view_functions['contratacion_postulantes']=contratacion_postulantes_293
app.view_functions['contratacion_etapa']=contratacion_etapa_293
app.view_functions['contratacion_config']=contratacion_config_293
# ======================= FIN PATCH OMAR 293 =======================

# ========================= PATCH VACACIONES OMAR 294 =========================
# Módulo vacaciones robusto: usuario limitado ve saldo y solicita vacaciones;
# administrador carga trabajadores/saldos masivos, aprueba/rechaza y exporta.

from math import ceil


def _vac_css_294():
    return """
    <style>
      html,body{background:#fff!important;overflow-x:hidden!important}
      .shell{max-width:430px!important;width:100%!important;margin:0 auto!important;padding:8px 8px 26px!important;background:#fff!important}
      .vac294-phone{max-width:390px;margin:0 auto}.vac294-app{background:#fff;border:1px solid #e4e8e4;border-radius:13px;overflow:hidden;box-shadow:0 10px 24px rgba(0,0,0,.07)}
      .vac294-head{height:70px;background:#25773a;color:#fff;display:flex;align-items:center;justify-content:center;position:relative}
      .vac294-head a.back{position:absolute;left:12px;top:50%;transform:translateY(-50%);color:#fff!important;text-decoration:none;font-size:31px;line-height:1}
      .vac294-head a.cfg{position:absolute;right:10px;top:16px;color:#fff!important;text-decoration:none;border:1px solid rgba(255,255,255,.65);border-radius:11px;padding:5px 8px;font-size:11px;font-weight:950}
      .vac294-head .ttl{font-size:17px;font-weight:950;color:#fff;text-align:center;letter-spacing:.2px}.vac294-body{padding:13px;background:#fff}
      .vac294-info{display:grid;grid-template-columns:20px 1fr;gap:8px;border:1px solid #b8d7ff;background:#eef6ff;border-radius:11px;padding:10px;margin-bottom:12px;color:#0b2e83;font-size:11.5px;font-weight:900;line-height:1.35}.vac294-info i{font-size:17px;margin-top:1px;color:#0b5ed7}
      .vac294-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:0 0 12px}.vac294-kpi{background:#10964e;color:#fff;border-radius:9px;text-align:center;padding:8px 4px;box-shadow:0 7px 13px rgba(16,150,78,.16)}.vac294-kpi small{display:block;font-size:9.4px;font-weight:950;line-height:1.05;color:#effff3}.vac294-kpi b{display:block;font-size:21px;line-height:1.05;font-weight:950;color:#fff;margin-top:4px}
      .vac294-section{font-size:13px;font-weight:950;color:#08713b;text-transform:uppercase;margin:11px 1px 8px;letter-spacing:.15px}.vac294-form,.vac294-card{border:1px solid #d7eadc;background:#fbfffc;border-radius:12px;padding:12px;margin-bottom:12px}.vac294-form label{font-size:11px;font-weight:950;color:#176a35;margin-bottom:4px}.vac294-form .form-control,.vac294-form .form-select{height:39px!important;border-radius:9px!important;font-size:12px!important;font-weight:850}.vac294-row{display:grid;grid-template-columns:1fr 1fr;gap:8px}
      .vac294-btn{height:42px;border-radius:10px;background:#08713b;border:1px solid #08713b;color:#fff;font-weight:950;font-size:13px;width:100%;display:flex;align-items:center;justify-content:center;gap:6px;text-decoration:none}.vac294-btn:hover{background:#065f2a;color:#fff}.vac294-outline{height:40px;border-radius:10px;background:#fff;border:1px solid #08713b;color:#08713b;font-weight:900;font-size:12px;width:100%;display:flex;align-items:center;justify-content:center;gap:6px;text-decoration:none}
      .vac294-period{display:grid;grid-template-columns:1fr 64px;gap:7px;align-items:center;border:1px solid #e5e7eb;background:#fff;border-radius:10px;padding:9px 10px;margin:7px 0}.vac294-period b{font-size:12px;color:#0a1f44}.vac294-period small{display:block;font-size:10.5px;color:#56715b;font-weight:800}.vac294-period .saldo{text-align:center;color:#08713b;font-size:20px;font-weight:950;line-height:1}.vac294-period .saldo small{font-size:9px;color:#08713b;text-transform:uppercase}
      .vac294-list{border:1px solid #e5e7eb;border-radius:10px;overflow:auto;background:#fff;scrollbar-color:#08713b #e5e7eb}.vac294-list::-webkit-scrollbar{height:8px}.vac294-list::-webkit-scrollbar-thumb{background:#08713b;border-radius:999px}.vac294-list::-webkit-scrollbar-track{background:#e5e7eb}.vac294-table{width:100%;min-width:580px;border-collapse:collapse}.vac294-table th{background:#f8fafc;color:#12223b;font-size:11px;font-weight:950;padding:8px;border-bottom:1px solid #e5e7eb}.vac294-table td{font-size:11px;color:#334155;padding:8px;border-bottom:1px solid #f1f5f9;font-weight:750;vertical-align:middle}
      .vac294-badge{display:inline-block;border-radius:999px;background:#dcfce7;color:#166534;padding:4px 8px;font-size:9.5px;font-weight:950;white-space:nowrap}.vac294-badge.pend{background:#fef3c7;color:#92400e}.vac294-badge.rech{background:#fee2e2;color:#991b1b}.vac294-badge.aprob{background:#dcfce7;color:#166534}
      .vac294-actions{display:grid;grid-template-columns:repeat(2,1fr);gap:8px;margin-bottom:12px}.vac294-actions .vac294-btn,.vac294-actions .vac294-outline{height:54px;flex-direction:column;gap:2px;font-size:11px;line-height:1.1}.vac294-actions i{font-size:20px}.vac294-help{font-size:10.5px;color:#4a644f;font-weight:850;line-height:1.35;margin-top:5px}.vac294-ok{border:1px solid #bbf7d0;background:#ecfdf5;color:#065f2a;border-radius:9px;padding:8px 9px;font-size:11px;font-weight:900;line-height:1.35;margin-top:7px}.vac294-bad{border:1px solid #fecaca;background:#fee2e2;color:#991b1b;border-radius:9px;padding:8px 9px;font-size:11px;font-weight:900;line-height:1.35;margin-top:7px}
    </style>
    """


def _vac_parse_date_294(v):
    if not v: return None
    if hasattr(v, 'date'): return v.date()
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try: return datetime.strptime(s[:10], fmt).date()
        except Exception: pass
    try: return datetime.fromisoformat(s[:10]).date()
    except Exception: return None


def _vac_date_iso_294(v):
    d = _vac_parse_date_294(v)
    return d.isoformat() if d else ''


def _vac_float_294(v, default=0.0):
    try:
        if v is None or v == '': return default
        return float(str(v).replace(',', '.'))
    except Exception:
        return default


def _vac_period_text_294(r):
    pi = r.get('periodo_inicio') or ''
    pf = r.get('periodo_fin') or ''
    if pi or pf: return f'{pi}/{pf}'.strip('/')
    return r.get('periodo') or 'Periodo'


def _vac_ensure_294():
    _ensure_rrhh_292()
    for ddl in [
        'ALTER TABLE vacaciones_saldos ADD COLUMN periodo TEXT',
        'ALTER TABLE vacaciones_saldos ADD COLUMN jefe_dni TEXT',
        'ALTER TABLE vacaciones_saldos ADD COLUMN fecha_ingreso TEXT',
        'ALTER TABLE vacaciones_solicitudes ADD COLUMN periodo_detalle TEXT',
        'ALTER TABLE vacaciones_solicitudes ADD COLUMN periodo_ids TEXT',
        'ALTER TABLE vacaciones_solicitudes ADD COLUMN fecha_jefe TEXT',
        'ALTER TABLE vacaciones_solicitudes ADD COLUMN fecha_gh TEXT',
        'ALTER TABLE vacaciones_solicitudes ADD COLUMN dias_gozar REAL DEFAULT 0',
    ]:
        try: execute(ddl, commit=True)
        except Exception: pass
    try:
        execute('CREATE INDEX IF NOT EXISTS idx_vac294_saldos_dni ON vacaciones_saldos(dni)', commit=True)
        execute('CREATE INDEX IF NOT EXISTS idx_vac294_sol_dni ON vacaciones_solicitudes(dni)', commit=True)
    except Exception: pass


def _vac_user_dni_294(): return limpiar_dni(session.get('dni') or session.get('usuario') or '')
def _vac_is_admin_294(): return session.get('rol') == 'admin'


def _vac_reserved_by_period_294(dni):
    usados = {}
    rows = rows_to_dict(execute("""SELECT id,dias,periodo_ids,estado FROM vacaciones_solicitudes
                                  WHERE dni=? AND UPPER(COALESCE(estado,'')) IN ('PENDIENTE','PENDIENTE JEFE','PENDIENTE GH')""", (dni,), fetchall=True))
    for s in rows:
        ids = [x.strip() for x in str(s.get('periodo_ids') or '').split(',') if x.strip().isdigit()]
        if not ids: continue
        dias = _vac_float_294(s.get('dias') or s.get('dias_gozar'))
        por = dias / max(1, len(ids))
        for pid in ids: usados[int(pid)] = usados.get(int(pid), 0.0) + por
    return usados


def _vac_saldos_dni_294(dni):
    rows = rows_to_dict(execute('SELECT * FROM vacaciones_saldos WHERE dni=? ORDER BY periodo_inicio DESC, periodo_fin DESC, id DESC', (dni,), fetchall=True))
    usados = _vac_reserved_by_period_294(dni)
    for r in rows:
        base = _vac_float_294(r.get('saldo'))
        usado = usados.get(int(r.get('id') or 0), 0.0)
        r['saldo_disponible'] = max(0.0, round(base - usado, 2))
        r['periodo_txt'] = _vac_period_text_294(r)
    return rows


def _vac_badge_294(estado):
    e = (estado or 'PENDIENTE').upper()
    cls = 'pend'
    if 'APROB' in e: cls = 'aprob'
    elif 'RECH' in e or 'ANUL' in e: cls = 'rech'
    return f'<span class="vac294-badge {cls}">{h(e)}</span>'


def _vac_find_worker_294(dni):
    w = _find_worker_any_292(dni)
    return w if w else {'dni': dni, 'trabajador': '', 'empresa':'', 'area':'', 'cargo':'', 'actividad':''}


def _vac_upsert_trabajador_294(row):
    dni = limpiar_dni(_valor(row, ['DNI','DOCUMENTO','DOC','NUMERO DOCUMENTO']))
    if len(dni) != 8: return False
    trabajador = limpiar_texto(_valor(row, ['TRABAJADOR','NOMBRES','NOMBRE','APELLIDOS Y NOMBRES','APELLIDOS NOMBRES']), upper=True)
    empresa = limpiar_texto(_valor(row, ['EMPRESA']), upper=True)
    area = limpiar_texto(_valor(row, ['AREA','ÁREA']), upper=True)
    cargo = limpiar_texto(_valor(row, ['CARGO','PUESTO']), upper=True)
    actividad = limpiar_texto(_valor(row, ['ACTIVIDAD','LABOR']), upper=True)
    planilla = limpiar_texto(_valor(row, ['PLANILLA','REGIMEN','RÉGIMEN']), upper=True)
    existe = row_to_dict(execute('SELECT id FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    if existe:
        execute("""UPDATE trabajadores SET trabajador=?, empresa=?, area=?, cargo=?, actividad=?, planilla=?, estado='ACTIVO', fecha_carga=? WHERE dni=?""", (trabajador, empresa, area, cargo, actividad, planilla, now_str(), dni), commit=True)
    else:
        execute("""INSERT INTO trabajadores(dni,trabajador,empresa,area,cargo,actividad,planilla,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?,?)""", (dni, trabajador, empresa, area, cargo, actividad, planilla, 'ACTIVO', now_str()), commit=True)
    return True


def _vac_upsert_saldo_294(row):
    dni = limpiar_dni(_valor(row, ['DNI','DOCUMENTO','DOC','NUMERO DOCUMENTO']))
    if len(dni) != 8: return False
    w = _vac_find_worker_294(dni)
    trabajador = limpiar_texto(_valor(row, ['TRABAJADOR','NOMBRES','NOMBRE','APELLIDOS Y NOMBRES']) or w.get('trabajador'), upper=True)
    empresa = limpiar_texto(_valor(row, ['EMPRESA']) or w.get('empresa'), upper=True)
    area = limpiar_texto(_valor(row, ['AREA','ÁREA']) or w.get('area'), upper=True)
    jefe = limpiar_texto(_valor(row, ['JEFE','JEFE NOMBRE','RESPONSABLE']), upper=True)
    jefe_dni = limpiar_dni(_valor(row, ['JEFE DNI','DNI JEFE']))
    fecha_ingreso = _vac_date_iso_294(_valor(row, ['FECHA INGRESO','FECHA_INGRESO','INGRESO']))
    pi = str(_valor(row, ['I_PERIODO','PERIODO INICIO','INICIO PERIODO','PERIODO_INICIO']) or '').strip()
    pf = str(_valor(row, ['F_PERIODO','PERIODO FIN','FIN PERIODO','PERIODO_FIN']) or '').strip()
    periodo = str(_valor(row, ['PERIODO']) or f'{pi}/{pf}'.strip('/')).strip()
    saldo = _vac_float_294(_valor(row, ['SALDO','DIAS SALDO','DÍAS SALDO','DIAS DISPONIBLES','DISPONIBLE']))
    ganados = _vac_float_294(_valor(row, ['DIAS GANADOS','DÍAS GANADOS','GANADOS']), saldo)
    gozados = _vac_float_294(_valor(row, ['DIAS GOZADOS','DÍAS GOZADOS','GOZADOS']), max(0, ganados - saldo))
    existe = row_to_dict(execute("SELECT id FROM vacaciones_saldos WHERE dni=? AND COALESCE(periodo_inicio,'')=? AND COALESCE(periodo_fin,'')=? LIMIT 1", (dni, pi, pf), fetchone=True))
    if existe:
        execute("""UPDATE vacaciones_saldos SET trabajador=?, empresa=?, area=?, jefe=?, jefe_dni=?, fecha_ingreso=?, periodo=?, dias_ganados=?, dias_gozados=?, saldo=?, fecha_carga=?, uploaded_by=? WHERE id=?""", (trabajador, empresa, area, jefe, jefe_dni, fecha_ingreso, periodo, ganados, gozados, saldo, now_str(), session.get('usuario'), existe.get('id')), commit=True)
    else:
        execute("""INSERT INTO vacaciones_saldos(dni,trabajador,empresa,area,jefe,jefe_dni,fecha_ingreso,periodo_inicio,periodo_fin,periodo,dias_ganados,dias_gozados,saldo,fecha_carga,uploaded_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (dni, trabajador, empresa, area, jefe, jefe_dni, fecha_ingreso, pi, pf, periodo, ganados, gozados, saldo, now_str(), session.get('usuario')), commit=True)
    return True


@login_required
def vacaciones_home_294():
    _vac_ensure_294()
    is_admin = _vac_is_admin_294()
    if is_admin:
        solicitudes = rows_to_dict(execute('SELECT * FROM vacaciones_solicitudes ORDER BY id DESC LIMIT 120', fetchall=True))
        total_saldos = int(scalar('SELECT COUNT(*) AS c FROM vacaciones_saldos') or 0)
        total_trab = int(scalar('SELECT COUNT(*) AS c FROM trabajadores') or 0)
        total_sol = len(solicitudes)
        pend = sum(1 for s in solicitudes if 'PEND' in (s.get('estado') or '').upper())
        aprob = sum(1 for s in solicitudes if 'APROB' in (s.get('estado') or '').upper())
        body = _vac_css_294() + r'''
        <div class="vac294-phone"><div class="vac294-app"><div class="vac294-head"><a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><a class="cfg" href="{{url_for('vacaciones_config')}}"><i class="bi bi-gear"></i> Config.</a><div class="ttl">Vacaciones Admin</div></div><div class="vac294-body"><div class="vac294-info"><i class="bi bi-info-circle-fill"></i><div>Administrador: carga trabajadores/saldos, revisa solicitudes y aprueba o rechaza. El usuario solo ve saldo y solicita vacaciones.</div></div><div class="vac294-kpis"><div class="vac294-kpi"><small>Saldos</small><b>{{total_saldos}}</b></div><div class="vac294-kpi"><small>Solicitudes</small><b>{{total_sol}}</b></div><div class="vac294-kpi"><small>Pend.</small><b>{{pend}}</b></div></div><div class="vac294-kpis"><div class="vac294-kpi"><small>Aprob.</small><b>{{aprob}}</b></div><div class="vac294-kpi"><small>Trabaj.</small><b>{{total_trab}}</b></div><div class="vac294-kpi"><small>Hoy</small><b style="font-size:13px">{{hoy[5:]}}</b></div></div><div class="vac294-actions"><a class="vac294-btn" href="{{url_for('vacaciones_config')}}"><i class="bi bi-cloud-arrow-up"></i><span>Carga masiva</span></a><a class="vac294-outline" href="{{url_for('vacaciones_exportar')}}"><i class="bi bi-file-earmark-excel"></i><span>Exportar</span></a></div><div class="vac294-section">Solicitudes por aprobar</div><div class="vac294-list"><table class="vac294-table"><thead><tr><th>DNI</th><th>Trabajador</th><th>Inicio</th><th>Fin</th><th>Días</th><th>Periodo</th><th>Estado</th><th>Acción</th></tr></thead><tbody>{% for s in solicitudes %}<tr><td>{{s.dni}}</td><td>{{s.trabajador or '-'}}</td><td>{{s.fecha_inicio}}</td><td>{{s.fecha_fin}}</td><td>{{s.dias}}</td><td>{{s.periodo_detalle or '-'}}</td><td>{{badge(s.estado)|safe}}</td><td><a href="{{url_for('vacaciones_estado', sol_id=s.id, estado='APROBADO')}}">Aprobar</a> · <a href="{{url_for('vacaciones_estado', sol_id=s.id, estado='RECHAZADO')}}">Rechazar</a></td></tr>{% else %}<tr><td colspan="8" class="text-center text-muted">Sin solicitudes.</td></tr>{% endfor %}</tbody></table></div></div></div></div>'''
        return render_page(body, solicitudes=solicitudes, total_saldos=total_saldos, total_trab=total_trab, total_sol=total_sol, pend=pend, aprob=aprob, hoy=today_str(), badge=_vac_badge_294, title='Vacaciones Admin')
    dni = _vac_user_dni_294()
    if request.method == 'POST':
        periodo_id = request.form.get('periodo_id') or ''
        dias_gozar = _vac_float_294(request.form.get('dias_gozar') or request.form.get('dias'))
        fi = _vac_parse_date_294(request.form.get('fecha_inicio'))
        motivo = limpiar_texto(request.form.get('motivo') or 'VACACIONES', upper=True)
        hoy_d = date.today()
        if not periodo_id.isdigit(): flash('Seleccione el periodo con saldo que usará.', 'danger'); return redirect(url_for('vacaciones_home'))
        if not fi: flash('Seleccione una fecha de inicio válida.', 'danger'); return redirect(url_for('vacaciones_home'))
        if fi < hoy_d: flash('La fecha de inicio no puede ser anterior a hoy.', 'danger'); return redirect(url_for('vacaciones_home'))
        if dias_gozar <= 0: flash('Digite los días a gozar.', 'danger'); return redirect(url_for('vacaciones_home'))
        saldo_row = row_to_dict(execute('SELECT * FROM vacaciones_saldos WHERE id=? AND dni=?', (int(periodo_id), dni), fetchone=True))
        if not saldo_row: flash('El periodo seleccionado no pertenece a su DNI.', 'danger'); return redirect(url_for('vacaciones_home'))
        saldos = _vac_saldos_dni_294(dni); saldo_sel = next((r for r in saldos if int(r.get('id') or 0) == int(periodo_id)), None); disponible = _vac_float_294((saldo_sel or {}).get('saldo_disponible'))
        if dias_gozar > disponible: flash(f'No puede solicitar {dias_gozar:g} día(s). Su saldo disponible del periodo es {disponible:g}.', 'danger'); return redirect(url_for('vacaciones_home'))
        dias_cal = int(ceil(dias_gozar)); ff = date.fromordinal(fi.toordinal() + max(0, dias_cal - 1))
        w = _vac_find_worker_294(dni); periodo_detalle = _vac_period_text_294(saldo_row)
        execute("""INSERT INTO vacaciones_solicitudes(dni,trabajador,jefe_dni,fecha_inicio,fecha_fin,dias,dias_gozar,motivo,estado,fecha_solicitud,periodo_detalle,periodo_ids,creado_por) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""", (dni, w.get('trabajador'), saldo_row.get('jefe_dni') or '', fi.isoformat(), ff.isoformat(), dias_gozar, dias_gozar, motivo, 'PENDIENTE', now_str(), periodo_detalle, str(periodo_id), dni), commit=True)
        flash('Solicitud de vacaciones registrada. Queda pendiente de aprobación.', 'success')
        return redirect(url_for('vacaciones_home'))
    saldos = _vac_saldos_dni_294(dni)
    solicitudes = rows_to_dict(execute('SELECT * FROM vacaciones_solicitudes WHERE dni=? ORDER BY id DESC LIMIT 80', (dni,), fetchall=True))
    total_saldo = round(sum(_vac_float_294(r.get('saldo_disponible')) for r in saldos), 2)
    pend = sum(1 for s in solicitudes if 'PEND' in (s.get('estado') or '').upper()); aprob = sum(1 for s in solicitudes if 'APROB' in (s.get('estado') or '').upper())
    body = _vac_css_294() + r'''
    <div class="vac294-phone"><div class="vac294-app"><div class="vac294-head"><a class="back" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Mis vacaciones</div></div><div class="vac294-body"><div class="vac294-info"><i class="bi bi-info-circle-fill"></i><div>Acceso usuario: solo puede ver su saldo y solicitar vacaciones. No puede cargar trabajadores ni saldos.</div></div><div class="vac294-kpis"><div class="vac294-kpi"><small>Saldo disp.</small><b>{{total_saldo|round(1)}}</b></div><div class="vac294-kpi"><small>Pend.</small><b>{{pend}}</b></div><div class="vac294-kpi"><small>Aprob.</small><b>{{aprob}}</b></div></div><div class="vac294-section">Solicitar vacaciones</div><form method="post" class="vac294-form" id="frmVac294"><label>Periodo con saldo</label><select name="periodo_id" id="periodoVac294" class="form-select" required><option value="">Seleccione periodo...</option>{% for s in saldos %}<option value="{{s.id}}" data-saldo="{{s.saldo_disponible}}">{{s.periodo_txt}} · Saldo {{s.saldo_disponible}}</option>{% endfor %}</select><div class="vac294-row mt-2"><div><label>Días a gozar</label><input name="dias_gozar" id="diasVac294" type="number" step="1" min="1" class="form-control" placeholder="Ej. 5" required></div><div><label>Fecha inicio</label><input name="fecha_inicio" id="inicioVac294" type="date" min="{{hoy}}" value="{{hoy}}" class="form-control" required></div></div><div class="vac294-row mt-2"><div><label>Fecha fin</label><input name="fecha_fin_mostrar" id="finVac294" type="date" class="form-control" readonly></div><div><label>Motivo</label><input name="motivo" class="form-control" value="VACACIONES"></div></div><div id="msgVac294" class="vac294-ok">La fecha fin se calcula automáticamente según los días a gozar.</div><button class="vac294-btn mt-2"><i class="bi bi-send"></i> Enviar solicitud</button></form><div class="vac294-section">Mis saldos</div>{% for s in saldos %}<div class="vac294-period"><div><b>{{s.periodo_txt}}</b><small>Ganados {{s.dias_ganados or 0}} · Gozados {{s.dias_gozados or 0}}</small></div><div class="saldo">{{s.saldo_disponible|round(1)}}<small>días</small></div></div>{% else %}<div class="vac294-card text-center text-muted">No tiene saldos cargados. Solicite a RR.HH. la carga de su saldo.</div>{% endfor %}<div class="vac294-section">Mis solicitudes</div><div class="vac294-list"><table class="vac294-table"><thead><tr><th>Inicio</th><th>Fin</th><th>Días</th><th>Periodo</th><th>Estado</th><th>Fecha solicitud</th></tr></thead><tbody>{% for s in solicitudes %}<tr><td>{{s.fecha_inicio}}</td><td>{{s.fecha_fin}}</td><td>{{s.dias}}</td><td>{{s.periodo_detalle or '-'}}</td><td>{{badge(s.estado)|safe}}</td><td>{{s.fecha_solicitud}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin solicitudes.</td></tr>{% endfor %}</tbody></table></div></div></div></div><script>function calcFinVac294(){const ini=document.getElementById('inicioVac294'),dias=document.getElementById('diasVac294'),fin=document.getElementById('finVac294'),sel=document.getElementById('periodoVac294'),msg=document.getElementById('msgVac294');if(!ini||!dias||!fin)return;let d=parseInt(dias.value||'0',10);let f=ini.value;if(!f||d<=0){fin.value='';return;}let max=parseFloat(sel.options[sel.selectedIndex]?.dataset?.saldo||'0');if(max&&d>max){msg.className='vac294-bad';msg.textContent='Los días a gozar superan el saldo disponible del periodo.';}else{msg.className='vac294-ok';msg.textContent='Fecha fin calculada automáticamente.';}let dt=new Date(f+'T00:00:00');dt.setDate(dt.getDate()+d-1);fin.value=dt.toISOString().slice(0,10);}['inicioVac294','diasVac294','periodoVac294'].forEach(id=>{let e=document.getElementById(id);if(e)e.addEventListener('input',calcFinVac294);if(e)e.addEventListener('change',calcFinVac294);});calcFinVac294();</script>'''
    return render_page(body, saldos=saldos, solicitudes=solicitudes, total_saldo=total_saldo, pend=pend, aprob=aprob, hoy=today_str(), badge=_vac_badge_294, title='Mis vacaciones')


@login_required
def vacaciones_estado_294(sol_id, estado):
    _vac_ensure_294()
    if not _vac_is_admin_294(): return _deny_admin_292()
    estado = (estado or '').upper()
    if estado not in ('APROBADO','RECHAZADO','ANULADO'): estado = 'PENDIENTE'
    sol = row_to_dict(execute('SELECT * FROM vacaciones_solicitudes WHERE id=?', (sol_id,), fetchone=True))
    if not sol: flash('Solicitud no encontrada.', 'danger'); return redirect(url_for('vacaciones_home'))
    old = (sol.get('estado') or '').upper()
    if estado == 'APROBADO' and 'APROB' not in old:
        dias_rest = _vac_float_294(sol.get('dias') or sol.get('dias_gozar'))
        ids = [x.strip() for x in str(sol.get('periodo_ids') or '').split(',') if x.strip().isdigit()]
        for pid in ids:
            if dias_rest <= 0: break
            saldo = row_to_dict(execute('SELECT * FROM vacaciones_saldos WHERE id=?', (int(pid),), fetchone=True))
            if not saldo: continue
            disp = _vac_float_294(saldo.get('saldo')); usar = min(disp, dias_rest)
            execute('UPDATE vacaciones_saldos SET dias_gozados=COALESCE(dias_gozados,0)+?, saldo=MAX(COALESCE(saldo,0)-?,0) WHERE id=?', (usar, usar, int(pid)), commit=True)
            dias_rest -= usar
    execute('UPDATE vacaciones_solicitudes SET estado=?, comentario_gh=?, fecha_gh=? WHERE id=?', (estado, f'Actualizado por {session.get("usuario")} {now_str()}', now_str(), sol_id), commit=True)
    flash('Estado actualizado.', 'success'); return redirect(url_for('vacaciones_home'))


@login_required
def vacaciones_config_294():
    _vac_ensure_294()
    if not _vac_is_admin_294(): return _deny_admin_292()
    if request.method == 'POST':
        action = request.form.get('action') or 'saldos'; f = request.files.get('archivo')
        if not f or not f.filename: flash('Seleccione un Excel.', 'danger'); return redirect(url_for('vacaciones_config'))
        ok = bad = 0
        try:
            for row in _iter_excel_upload(f):
                res = _vac_upsert_trabajador_294(row) if action == 'trabajadores' else _vac_upsert_saldo_294(row)
                if res: ok += 1
                else: bad += 1
            flash(f'Carga completada. Correctos: {ok}. Omitidos/error: {bad}.', 'success')
        except Exception as e: flash(f'Error leyendo Excel: {e}', 'danger')
        return redirect(url_for('vacaciones_config'))
    ult_saldos = rows_to_dict(execute('SELECT * FROM vacaciones_saldos ORDER BY id DESC LIMIT 50', fetchall=True))
    body = _vac_css_294() + r'''
    <div class="vac294-phone"><div class="vac294-app"><div class="vac294-head"><a class="back" href="{{url_for('vacaciones_home')}}"><i class="bi bi-chevron-left"></i></a><div class="ttl">Config. vacaciones</div></div><div class="vac294-body"><div class="vac294-info"><i class="bi bi-info-circle-fill"></i><div>Configuración solo administrador: primero puede cargar trabajadores y luego cargar saldos vacacionales por DNI.</div></div><form method="post" enctype="multipart/form-data" class="vac294-form"><input type="hidden" name="action" value="trabajadores"><label><i class="bi bi-people"></i> Carga masiva de trabajadores</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="vac294-help">Columnas: DNI, TRABAJADOR, EMPRESA, AREA, CARGO, ACTIVIDAD, PLANILLA.</div><div class="vac294-row mt-2"><button class="vac294-btn"><i class="bi bi-upload"></i> Cargar trabajadores</button><a class="vac294-outline" href="{{url_for('vacaciones_plantilla_trabajadores')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div></form><form method="post" enctype="multipart/form-data" class="vac294-form"><input type="hidden" name="action" value="saldos"><label><i class="bi bi-calendar-check"></i> Carga masiva de saldos</label><input type="file" name="archivo" accept=".xlsx,.xlsm" class="form-control" required><div class="vac294-help">Columnas: DNI, TRABAJADOR, I_PERIODO, F_PERIODO, DIAS GANADOS, DIAS GOZADOS, SALDO, JEFE DNI.</div><div class="vac294-row mt-2"><button class="vac294-btn"><i class="bi bi-upload"></i> Cargar saldos</button><a class="vac294-outline" href="{{url_for('vacaciones_plantilla_saldos')}}"><i class="bi bi-file-earmark-excel"></i> Plantilla</a></div></form><div class="vac294-actions"><a class="vac294-outline" href="{{url_for('vacaciones_exportar')}}"><i class="bi bi-download"></i><span>Exportar solicitudes</span></a><a class="vac294-outline" href="{{url_for('vacaciones_home')}}"><i class="bi bi-check2-square"></i><span>Aprobaciones</span></a></div><div class="vac294-section">Últimos saldos cargados</div><div class="vac294-list"><table class="vac294-table"><thead><tr><th>DNI</th><th>Trabajador</th><th>Periodo</th><th>Ganados</th><th>Gozados</th><th>Saldo</th></tr></thead><tbody>{% for s in ult_saldos %}<tr><td>{{s.dni}}</td><td>{{s.trabajador or '-'}}</td><td>{{s.periodo or (s.periodo_inicio ~ '/' ~ s.periodo_fin)}}</td><td>{{s.dias_ganados}}</td><td>{{s.dias_gozados}}</td><td>{{s.saldo}}</td></tr>{% else %}<tr><td colspan="6" class="text-center text-muted">Sin saldos cargados.</td></tr>{% endfor %}</tbody></table></div></div></div></div>'''
    return render_page(body, ult_saldos=ult_saldos, title='Config. vacaciones')


@login_required
def vacaciones_plantilla_trabajadores_294():
    if not _vac_is_admin_294(): return _deny_admin_292()
    headers = ['DNI','TRABAJADOR','EMPRESA','AREA','CARGO','ACTIVIDAD','PLANILLA']
    rows = [{'DNI':'74324033','TRABAJADOR':'JOSE GARCIA','EMPRESA':'AQUANQA','AREA':'COSECHA','CARGO':'OPERARIO','ACTIVIDAD':'ARANDANO','PLANILLA':'AGRARIO'}]
    return excel_response(headers, rows, 'plantilla_trabajadores_vacaciones.xlsx', 'TRABAJADORES')

@login_required
def vacaciones_plantilla_saldos_294():
    if not _vac_is_admin_294(): return _deny_admin_292()
    headers = ['DNI','TRABAJADOR','EMPRESA','AREA','FECHA INGRESO','I_PERIODO','F_PERIODO','DIAS GANADOS','DIAS GOZADOS','SALDO','JEFE DNI']
    rows = [{'DNI':'74324033','TRABAJADOR':'JOSE GARCIA','EMPRESA':'AQUANQA','AREA':'COSECHA','FECHA INGRESO':'2025-01-01','I_PERIODO':'2025','F_PERIODO':'2026','DIAS GANADOS':30,'DIAS GOZADOS':0,'SALDO':30,'JEFE DNI':''}]
    return excel_response(headers, rows, 'plantilla_saldos_vacaciones.xlsx', 'SALDOS')

@login_required
def vacaciones_exportar_294():
    _vac_ensure_294()
    rows = rows_to_dict(execute('SELECT * FROM vacaciones_solicitudes ORDER BY id DESC', fetchall=True)) if _vac_is_admin_294() else rows_to_dict(execute('SELECT * FROM vacaciones_solicitudes WHERE dni=? ORDER BY id DESC', (_vac_user_dni_294(),), fetchall=True))
    headers = ['id','dni','trabajador','fecha_inicio','fecha_fin','dias','periodo_detalle','motivo','estado','fecha_solicitud','comentario_gh']
    return excel_response(headers, rows, 'reporte_vacaciones.xlsx', 'VACACIONES')

for rule, endpoint, view, methods in [
    ('/vacaciones', 'vacaciones_home', vacaciones_home_294, ['GET','POST']),
    ('/vacaciones/config', 'vacaciones_config', vacaciones_config_294, ['GET','POST']),
    ('/vacaciones/exportar', 'vacaciones_exportar', vacaciones_exportar_294, ['GET']),
    ('/vacaciones/estado/<int:sol_id>/<estado>', 'vacaciones_estado', vacaciones_estado_294, ['GET']),
    ('/vacaciones/plantilla/trabajadores', 'vacaciones_plantilla_trabajadores', vacaciones_plantilla_trabajadores_294, ['GET']),
    ('/vacaciones/plantilla/saldos', 'vacaciones_plantilla_saldos', vacaciones_plantilla_saldos_294, ['GET']),
]:
    try:
        app.add_url_rule(rule, endpoint, view, methods=methods)
    except Exception:
        app.view_functions[endpoint] = view

# ======================= FIN PATCH VACACIONES OMAR 294 =======================


if __name__ == '__main__':
    port = int(os.getenv('PORT', '5000'))
    app.run(host='0.0.0.0', port=port, debug=False)
